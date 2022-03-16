def edmfr_prog3(para1,para2,para3,para4,para5,para6,para7,para8,para9,para10,para11,para12):


    import pandas as pd
    import pickle, win32com, csv, shutil
    from win32com import client
    import tkinter as tk
    from tkinter import messagebox, ttk
    import pythoncom, importlib
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
    from subprocess import check_output
    import copy
    import subprocess, os, webbrowser, datetime, time, types
    from openpyxl.utils import get_column_letter


    def col_bg_col(ws):
        for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
            for cell in rows:
                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type="solid")


    def auto_format_cell_width1(ws):
        for letter in range(1, ws.max_column):
            maximum_value = 0
            for cell in ws[get_column_letter(letter)]:
                val_to_check = len(str(cell.value))
                # print(str(cell.value))
                if val_to_check > maximum_value:
                    if val_to_check < 50:
                        maximum_value = val_to_check
                    else:
                        maximum_value = 50
            ws.column_dimensions[get_column_letter(letter)].width = maximum_value + 1


    def xcel_wraptext(ws):
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                # alignment = copy.copy(cell.alignment)
                # alignment.wrapText = True
                # Alignment.wrap_text
                # cell.column = alignment
                # ws.cell(row=cell.row, column=col).alignment = alignment

    def walk_reload(module: types.ModuleType) -> None:
        if hasattr(module, "__all__"):
            for submodule_name in module.__all__:
                walk_reload(getattr(module, submodule_name))
        importlib.reload(module)

    def process_exists(process_name):
        call = 'TASKLIST', '/FI', 'imagename eq %s' % process_name
        processes = []
        for process in check_output(call).splitlines()[3:]:
            process = process.decode()
            processes.append(process.split())
        return processes

    NORM_FONT= ("Verdana", 10)
    def popupmsg(msg):
        popup = tk.Tk()
        popup.wm_title("!")
        label = ttk.Label(popup, text=msg, font=NORM_FONT)
        label.pack(side="top", fill="x", pady=10)
        B1 = ttk.Button(popup, text="Ok", command = popup.destroy)
        B1.pack()
        popup.mainloop()



    # Your command

    # cmd = "showmessage.py"
    # cmd = "EDMFR_ALSDpick.py"
    # edmfrprog_in == 1

    # if edmfrprog_in == 1:
        # popupmsg("Please click ok and wait for a while until output report opens")

        # importlib.reload(EDMFR_ALSDpick)

    #     edmfrprog_in == 2
    # else:
    #     importlib.reload(EDMFR_ALSDpick)


    # cmd = "RADADeskApp2.py"


    # Starting process
    # process = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    # stdout, stderr = process.communicate()
    # encoding = 'utf-8'
    #
    # file1 = open("errorlog_file.txt", "w")
    # # file1.write("Error log: \n")
    # file1.writelines(str(stderr, encoding))
    # file1.close()

    def recursive_copy(src, dest):
        """
        Copy each file from src dir to dest dir, including sub-directories.
        """
        for item in os.listdir(src):
            file_path = os.path.join(src, item)

            # if item is a file, copy it
            if os.path.isfile(file_path):
                shutil.copy(file_path, dest)

            # else if item is a folder, recurse
            elif os.path.isdir(file_path):
                new_dest = os.path.join(dest, item)
                os.mkdir(new_dest)
                recursive_copy(file_path, new_dest)

    def popupmsg(msg):
        popup = tk.Tk()
        popup.wm_title("!")
        label = ttk.Label(popup, text=msg, font=NORM_FONT)
        label.pack(side="top", fill="x", pady=10)
        B1 = ttk.Button(popup, text="Ok", command = popup.destroy)
        B1.pack()
        popup.mainloop()

    def purgedir(parent):
        for root, dirs, files in os.walk(parent):
            for item in files:
                # Delete subordinate files
                filespec = os.path.join(root, item)
                # if filespec.endswith('.bak'):
                os.unlink(filespec)
            for item in dirs:
                # Recursively perform this operation for subordinate directories
                purgedir(os.path.join(root, item))



    # import EDMFR_ALSDpick

    # walk_reload(EDMFR_ALSDpick)


    dir = os.getcwd()
    SITE_ROOT1 = os.getcwd()
    username = os.environ["USERNAME"]
    ossum_path1 = 'C:\\Users\\' + username
    osssum_dir = 'OSSUM'
    ossum_path = ossum_path1 +'\\' + osssum_dir
    alspick_dir = 'ALSpickle'
    alspick_path = ossum_path  +'\\' + alspick_dir

    EDMFR_dir = 'EDMFR'
    EDMFR_path = ossum_path +'\\' + EDMFR_dir
    datestring = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M')
    EDMFR_dt_path =EDMFR_path + '\\' + datestring
    ALSdownloadpath = 'C:\\Users\\' + username + '\\ALSdownload'
    # rada_dt_path = os.path.join(rada_path, datestring)
    # os.mkdir(datestring)
    # os.chdir(ossum_path1)
    # print(EDMFR_dt_path)
    # if os.path.exists(ossum_path1):
    #     if not os.path.exists(alsdwnd_path):
    #         os.makedirs(alsdwnd_path)


    if os.path.exists(ossum_path1):
        if not os.path.exists(alspick_path):
            os.makedirs(alspick_path)

    if os.path.exists(ossum_path1):
        if not os.path.exists(EDMFR_dt_path):
            os.makedirs(EDMFR_dt_path)

    if os.path.exists(ossum_path1):
        if not os.path.exists(ALSdownloadpath):
            os.makedirs(ALSdownloadpath)
        # webbrowser.open(ossum_path1)
        #
        # try:
        #     os.makedirs(rada_dt_path)
        #     print("Directory '%s' created successfully" )
        # except OSError as error:
        #     print("Directory '%s' can not be created" )



    ######
    lists = os.listdir(ALSdownloadpath)
    sub1 = '.xls'
    files_crf1 = [mystr for mystr in lists if mystr.endswith(sub1)]
    # files_crf1 = [mystr for mystr in lists if sub1 in mystr]
    df_filenames = pd.DataFrame()
    list_filenames = []

    # path_sp = SITE_ROOT1 + '\Source Documents'
    # path_sp = PARENT_ROOT + '\Source Documents'
    os.chdir(para7)
    lists = os.listdir(para7)

    df_als_sp = pd.read_excel("SharepointALSdwn.xlsm")
    df_ALS_SP_u = df_als_sp.sort_values(by=['ALS Version'])
    df_ALS_SP_u.sort_values('ALS Version').groupby('Study').tail(1)
    df_ALS_SP_u = df_ALS_SP_u.sort_values('ALS Version').groupby('Study').tail(1)
    df_ALS_SP_u = df_ALS_SP_u[df_ALS_SP_u['Name'].str.contains(".xls$")]
    # df_ALS_SP_u = df_ALS_SP_u[~df_ALS_SP_u['Name'].str.endswith('.xls')]
    df_ALS_SP_u = df_ALS_SP_u.filter(items=['Study', 'ALS Version', 'Name'])
    # df_ALS_SP_u = df_ALS_SP_u.rename(columns={'Modified By': "FPFV"})
    df_als_sp_1 = df_als_sp.loc[:, ('Study', 'ALS Version')]
    df_als_studies_sp = df_als_sp_1.groupby("Study").max().reset_index()
    df_als_sp_dup = df_als_sp[df_als_sp.duplicated(subset=['Study', 'ALS Version'], keep=False)]
    df_als_sp_nul = df_als_sp.loc[((df_als_sp["Study"].isna()) | (df_als_sp["Study"].isna()))]

    # sdfdf
    # if any('EXCEL.EXE' in sl for sl in xlprocess):
    #     popupmsg("'EXCEL.EXE' is in progress, please in task manager")
    #     exit()

    # pythoncom.CoInitialize()
    # excelApp = win32com.client.dynamic.Dispatch('Excel.Application')
    # excelApp.Application.Quit()
    # pythoncom.CoUninitialize()

    os.chdir(para8)
    lists = os.listdir(para8)
    sub = 'Consolidated_ALS_Selected_Tabs(CAST).xlsx'
    # sub = '.pickle'
    # sub1 = 'xls'
    # uat = 'uat'
    files = [mystr for mystr in lists if sub in mystr]
    lists_als = files
    # print(alspick_path)

    if os.path.exists(para8):
        DATABASE1 = "Consolidated_ALS_Selected_Tabs(CAST).xlsx"
        DATABASE = para8 + "\\" + DATABASE1
        if os.path.isfile(DATABASE):
            pass

    if (len(lists_als) > 0) :
        df_CRFDraft = pd.read_excel(DATABASE,sheet_name='CRFDraft')
        df_alsforms = pd.read_excel(DATABASE,sheet_name='Forms')

        df_alsforms = df_alsforms.merge(df_CRFDraft.loc[:, ['ProjectName', 'DraftName']], left_on=['Study ID'],
                                                right_on=['ProjectName'], how='left', suffixes=['', '_'],
                                                indicator=True)
        df_alsforms = df_alsforms.loc[df_alsforms['_merge'] == 'both']

        mycols = set(df_alsforms.columns)
        mycols.remove('_merge')

        print("df_alsforms",df_alsforms)
        # df_alsforms = pd.read_excel('alsfilesdfforms.xlsx')
        if len(df_alsforms) >0:
            os.chdir(dir)
            df1 = df_alsforms.loc[ :,('OID','Study ID','DraftFormActive','DraftName')]
            df_alsforms1 = df1.loc[(df_alsforms["DraftFormActive"] == True)]
            print("TEST")

    # df_rulefiles = pd.read_csv('C:/Users/pillibh2/Desktop/Rulefile/rulefiledump/Prod_PREDRM_RULEFILE_Data.csv')
    # xlapp = client.DispatchEx("Excel.Application")
    # xlapp.Quit()

    # SITE_ROOT1 = os.getcwd()
    # path_ibow = SITE_ROOT1 + '\EDMFR\iBOW'
    # path_sp = PARENT_ROOT + '\Source Documents'
    # os.chdir(path_ibow)
    # lists = os.listdir(path_ibow)
    list_filenames.extend(para2.filename)

    if para2:
    # if len(lists) == 1:
        df_ibow = pd.read_excel(para2)
        print("TEST_ibow",df_ibow )
    #
    # print(path_sp)
    # print(lists)
    # os.chdir(path_sp)
    # filename = path_sp + '\\' + 'DM Book of Work.xlsx'
    # print(filename)
    # #
    # #         xl = client.Dispatch('Excel.Application')
    # #         wb = xl.Workbooks.Open(Filename=filename, ReadOnly=1)
    # xlapp = client.DispatchEx("Excel.Application")
    # # wbs = xlapp.Workbooks
    # wb = xlapp.Workbooks.Open(Filename=filename, ReadOnly=True)
    # wb.RefreshAll()
    # xlapp.CalculateUntilAsyncQueriesDone()
    # xlapp.DisplayAlerts = False
    # wb.Save()
    # filename = path_sp + '\\' + 'StudyMasterlist.xlsx'
    # wb = xlapp.Workbooks.Open(Filename=filename, ReadOnly=1)
    # # wb = xlapp.Workbooks.Open('DM Book of Work.xlsx')
    # wb.RefreshAll()
    # xlapp.CalculateUntilAsyncQueriesDone()
    # xlapp.DisplayAlerts = False
    # wb.Save()
    # filename = path_sp + '\\' + 'Rulefile tracker.xlsx'
    # wb = xlapp.Workbooks.Open(Filename=filename, ReadOnly=1)
    # # wb = xlapp.Workbooks.Open('DM Book of Work.xlsx')
    # wb.RefreshAll()
    # xlapp.CalculateUntilAsyncQueriesDone()
    # xlapp.DisplayAlerts = False
    # wb.Save()
    # xlapp.Quit()
    #
    # df_dmbow = pd.read_excel('DM Book of Work.xlsx',sheet_name='DMBOW')
    # df_studymasterlist = pd.read_excel('StudyMasterlist.xlsx',sheet_name='Studymasterlist')
    # df_rftracker = pd.read_excel('Rulefile tracker.xlsx',sheet_name='Rulefile tracker')

    # print(df_ibow.columns)
    # labels1= ['Trial','DB Phase','FPFV Planned','Upcoming Milestone']
    # labels2 = []
    #
    # # [labels2.add(lst) for lst in labels1 if not (lst in df_ibow.columns)]
    #
    # for lst in labels1:
    #     if not lst in df_ibow.columns:
    #         print(lst)
    #         labels2.add(lst)
    #
    # print("labels2",labels2)
    df_dmbow = df_ibow.filter(items=['Trial','DB Phase','FPFV Planned','Upcoming Milestone', 'FPFV Actual',
                                     'Lead DBD', 'Transition DBD'])
    df_dbd = df_ibow.filter(items=['Trial',
                                     'Lead DBD', 'Transition DBD', 'Rule File Developer'])
    df_dbd1 = df_ibow.filter(items=['Trial',
                                     'Lead DBD', 'Transition DBD'])
    df_dmbow = df_dmbow.rename(columns={"FPFV Planned": "FPFV"})

    df_studymasterlist = df_ibow.filter(items=['Trial','DB Phase','Lead DBD','Transition DBD','Planned DB Go-Live Date','EDC', 'Actual DB Go-Live Date', 'GDO DM Sourcing'])
    df_studymasterlist = df_studymasterlist.loc[((df_studymasterlist['EDC'] == 'Rave EDC -  Inhouse') |
                                                 ((df_studymasterlist['EDC'] == 'CRO - EDC') &
                                                  (df_studymasterlist['GDO DM Sourcing'].isin(
                                                      ['InHouse', 'TCO Mixed Model']))))]
    df_studymasterlist = df_studymasterlist.rename(columns={"Trial": "Study", "DB Phase": "Study Status",
                                                            "Planned DB Go-Live Date": "DB Go-Live Date"})

    # df_studymasterlist.to_excel(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS\df_studymasterlistnew.xlsx')

    df_rftracker = df_ibow.filter(items=['Trial','Old Rule File Developer','Rule File Developer','Study LSP','EDC', 'GDO DM Sourcing'])
    df_rftracker = df_rftracker.loc[((df_rftracker['EDC'] == 'Rave EDC -  Inhouse') |
                                                 ((df_rftracker['EDC'] == 'CRO - EDC') &
                                                  (df_rftracker['GDO DM Sourcing'].isin(
                                                      ['InHouse', 'TCO Mixed Model']))))]
    df_rftracker = df_rftracker.rename(columns={"Trial": "Study", "Study LSP": "LSP"})

    # path_crf = SITE_ROOT1 + '\EDMFR\CONSD RF'
    # os.chdir(path_crf)
    # lists = os.listdir(path_crf)
    list_filenames.append(para1.filename)
    sub = '.xlsx'
    files_crf = [mystr for mystr in lists if sub in mystr]

    df_rulefiles=pd.DataFrame()
    if para1:
    # if len(files_crf) == 1 :
        df_rulefiles = pd.read_excel(para1)
        # df_rulefiles = pd.read_excel(files_crf[0])
    # # df_rulefiles = pd.read_csv(files_als[0])

    # path_temp = SITE_ROOT1 + '\\EDMFR\\templates'
    # print(path_temp)
    # lists = os.listdir(path_temp)
    lists = os.listdir(para9)
    os.chdir(para9)
    reader = csv.reader(open(lists[0]))
    # path_mt = SITE_ROOT1 + '\\EDMFR\\tempmapfile'

    # print(path_mt)
    # os.chdir(path_mt)
    lists = os.listdir(para12)
    # lists = os.listdir(path_mt)
    list_filenames.extend(lists)
    print("lists",lists)
    # if len(lists)>0:
    # if para12:
    # if para12:
    if len(lists)>0:
    #     reader1 = csv.reader(open(para10))
        os.chdir(para12)
        print("para3 s",para3)
        print("para3 s",para3.filename)
        # reader1 = csv.reader(open(para3.filename))
        reader1 = csv.reader(open(lists[0]))
        os.chdir(para10)
        # path_mp = SITE_ROOT1 + '\\EDMFR\\mappingfile'

        f = open("headercomb_mapping_file.csv", "w")
        writer = csv.writer(f)
        a = 1
        for row in reader:
            if a ==1:
                writer.writerow(row)
                a=2

        for row in reader1:

            writer.writerow(row)
        f.close()

        # dfcomb=pd.read_csv("headercomb_mapping_file.csv")
        lists = os.listdir(para10)
        # sub = '.xlsx'
        # files_mp = [mystr for mystr in lists if sub in mystr]
        print("if para10:",para10)
        # if para10:
        if len(lists) == 1 :
            print("para3 read",para3)
            os.chdir(para10)
            print("para3 read",para10)
            # df_dtsmaptable = pd.read_csv(para10, engine='python')
            df_dtsmaptable = pd.read_csv(lists[0], engine='python')
            print("df_dtsmaptable",df_dtsmaptable)
            print("df_dtsmaptable",len(df_dtsmaptable))
        df_dts1 = df_dtsmaptable[df_dtsmaptable["c12"]=='OC' ]
        df_dts1 = df_dts1[df_dts1["tables_info"].notnull()]
        # df_dts1.to_excel(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS\df_dts1new.xlsx')


        df_dts1['OID'] = df_dts1.tables_info.str.extract('PI_DATA_TYPE,(.*?),', expand=True)

        df_dts1.loc[(~(df_dts1['OID'] == df_dts1[
            'table1'])), 'LSH setup'] = 'It is listed in missing DTS as Table1,2 is mismatch with column OID or PI_DATA_TYPE table name in LSH, the same has informed to LSPs for update'

        df_dts1.dropna(subset=['OID'], how='all', inplace=True)

    df_dmbow = df_dmbow.loc[:,['Trial','DB Phase','FPFV','Upcoming Milestone','FPFV Actual']]
    df_dmbow=df_dmbow.rename(columns={"Trial": "Study"})
    #print(df_dmbow)

    df_masterlist1 = df_studymasterlist[['Study','Study Status','Lead DBD','Transition DBD','DB Go-Live Date','Actual DB Go-Live Date']]

    df_rftracker1 = df_rftracker.loc[:,['Study','Old Rule File Developer','Rule File Developer','LSP']]
    df_masterlist1=df_masterlist1.merge(df_rftracker1, left_on=['Study'],right_on=['Study'], how='left', suffixes=['', '_'], indicator=True)

    df_masterlist1 = df_masterlist1[['Study', 'Study Status', 'Lead DBD','Transition DBD','Old Rule File Developer','Rule File Developer', 'DB Go-Live Date','LSP','Actual DB Go-Live Date']]

    df_masterlist1=df_masterlist1.merge(df_dmbow, left_on=['Study'],right_on=['Study'], how='left', suffixes=['', '_'], indicator=True)
    # print(df_masterlist1.columns)
    df_masterlist1 = df_masterlist1[['Study', 'Study Status', 'Lead DBD','Transition DBD','Old Rule File Developer','Rule File Developer', 'DB Go-Live Date', 'FPFV', 'FPFV Actual','LSP', 'Actual DB Go-Live Date']]

    ntreqfrms = ["SAERF", "SAE_DOS", "SAE_TEST", "SAE_DEATH", "SAEINFO", "INV_REV", "AUTO_01", "AUTO_02", "AUTO_03", "AUTO_04", "AUTO_05", "AUTO_06","AUTO_07", "AUTO_08", "AUTO_09", "AUTO_10", "TRANSMIT", "OPG001"]

    if len(df_alsforms1) >0:
        df_alsforms2 = df_alsforms1[['Study ID', 'OID']]

        df_alsforms3 = df_alsforms2[~(df_alsforms2.loc[:,'OID'].str.contains('|'.join(ntreqfrms)))]

        if len(df_rulefiles)>0:
            df_rulefiles2 = df_rulefiles[['STUDY_NAME','FORM_NAME']]


            df_rf_reshaped = \
            (df_rulefiles2.set_index(df_rulefiles2.columns.drop('FORM_NAME',1).tolist())
               .FORM_NAME.str.split(',', expand=True)
               .stack()
               .reset_index()
               .rename(columns={0:'FORM_NAME'})
               .loc[:, df_rulefiles2.columns]
            )

        df_alsforms33 = df_alsforms3[df_alsforms3['Study ID'].isin(df_rf_reshaped['STUDY_NAME'].tolist())]
        df_dts2 = df_dts1[['STUDY_NAME', 'OID','table1','table2','LSH setup']]


        df_dts3 = df_dts2[df_dts2['STUDY_NAME'].isin(df_masterlist1['Study'].tolist()) ]

        ntreqdtsfrms = ['YD','YR']
        df_dts3 = df_dts3[~(df_dts3.loc[:,'OID'].str.contains('|'.join(ntreqdtsfrms)))]
        # df_dts3.to_excel(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS\df_dts3new_0.xlsx')
        # df_masterlist1.to_excel(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS\df_masterlist1new.xlsx')

        df_dts3 = df_dts3.merge(df_masterlist1, left_on=['STUDY_NAME'],right_on=['Study'], how='left', suffixes=['', '_'], indicator=True)
        # df_dts3.to_excel(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS\df_dts3new.xlsx')
        df_dts4 = df_dts3[['STUDY_NAME', 'OID', 'Study Status' ,'Lead DBD','Transition DBD','Old Rule File Developer','Rule File Developer', 'DB Go-Live Date', 'FPFV' , 'FPFV Actual','table1','table2','LSH setup','LSP', 'Actual DB Go-Live Date']]



        df_merge=df_dts4.merge(df_rf_reshaped, left_on=['STUDY_NAME','OID'],right_on=['STUDY_NAME','FORM_NAME'], how='left', suffixes=['', '_'], indicator=True)
        # df_merge.to_excel(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS\df_mergenew.xlsx')
        df_merge= df_merge[df_merge['_merge']=='left_only']
        df_merge = df_merge[['STUDY_NAME', 'OID', 'Study Status' ,'Lead DBD','Transition DBD','Old Rule File Developer','Rule File Developer', 'DB Go-Live Date', 'FPFV' , 'FPFV Actual','table1','table2','LSH setup' ,'LSP', 'Actual DB Go-Live Date']]


        df_merge1=df_alsforms33.merge(df_rf_reshaped, left_on=['Study ID','OID'],right_on=['STUDY_NAME','FORM_NAME'], how='left', suffixes=['', '_'], indicator=True)
        df_merge2=df_alsforms3.merge(df_rf_reshaped, left_on=['Study ID'],right_on=['STUDY_NAME'], how='left', suffixes=['', '_'], indicator=True)
        df_merge3=df_rf_reshaped.merge(df_alsforms3, left_on=['STUDY_NAME'],right_on=['Study ID'], how='left', suffixes=['', '_'], indicator=True)

        df3= df_merge1[df_merge1['_merge']=='left_only']
        df4 = df3[['Study ID', 'OID']]

        df4= df4.merge(df_masterlist1, left_on=['Study ID'],right_on=['Study'], how='left', suffixes=['', '_'], indicator=True)
        # df4= df4.merge(df_edcreasnnotlisted, left_on=['Study ID'],right_on=['STUDY_NAME'], how='left', suffixes=['', '_'], indicator=True)

        edcmiss = df4[['Study ID', 'OID', 'Study Status' ,'Lead DBD','Transition DBD','Old Rule File Developer','Rule File Developer', 'DB Go-Live Date', 'FPFV', 'FPFV Actual', 'LSP', 'Actual DB Go-Live Date']]

        df5= df_merge2[df_merge2['_merge']=='left_only']
        df_alsworf1 = df5[['Study ID']]
        df_alsworf = df_alsworf1.drop_duplicates()

        #print(df_merge3)
        df6= df_merge3[df_merge3['_merge']=='left_only']
        df_rfwoals1 = df6[['STUDY_NAME']]
        df_rfwoals = df_rfwoals1.drop_duplicates()

        # path_pout = SITE_ROOT1 + '\\EDMFR\\prev output'
        # lists = os.listdir(path_pout)
        list_filenames.extend(para4.filename)
        sub = '.xlsx'
        files_als = [mystr for mystr in lists if sub in mystr]
        # os.chdir(path_pout)
        df_sp1 = pd.read_excel(para4, sheet_name='missingEDCFormsList')
        df_sp2 = pd.read_excel(para4, sheet_name='DTSFormsmissing')
        df_sp3 = pd.read_excel(para4, sheet_name='StudiesPresentinALSnotinRF')
        df_sp4 = pd.read_excel(para4, sheet_name='StudiesPresentinRFnotinALS')
        # df_sp1 = pd.read_excel(files_als[0], sheet_name='missingEDCFormsList')
        # df_sp2 = pd.read_excel(files_als[0], sheet_name='DTSFormsmissing')
        # df_sp3 = pd.read_excel(files_als[0], sheet_name='StudiesPresentinALSnotinRF')
        # df_sp4 = pd.read_excel(files_als[0], sheet_name='StudiesPresentinRFnotinALS')

        if not 'Comment' in df_sp3.columns:
            df_sp3['Comment'] = ""

        if not 'Comment' in df_sp4.columns:
            df_sp4['Comment'] = ""

        df_sp1 = df_sp1[['Study ID','OID','Reason for Not Including in Rule File','Date Planned to be included','Action Required Yes/No']]

        df_edcmiss=edcmiss.merge(df_sp1, left_on=['Study ID','OID'],right_on=['Study ID','OID'], how='left', suffixes=['', '_'], indicator=True)
        df_edcmiss = df_edcmiss[['Study ID', 'OID', 'Study Status' ,'Lead DBD','Transition DBD','Old Rule File Developer','Rule File Developer', 'DB Go-Live Date', 'Actual DB Go-Live Date', 'FPFV', 'FPFV Actual', 'LSP','Reason for Not Including in Rule File','Date Planned to be included','Action Required Yes/No']]
        df_edcmiss = df_edcmiss.rename(columns={"Study Status": "DB Phase",'DB Go-Live Date':"Planned DB Go-Live Date", 'FPFV': 'FPFV Planned'})
        df_sp2 = df_sp2[['STUDY_NAME','OID','Reason for Not Including in Rule File','Date Planned to be included','Action Required Yes/No']]




        df_dtsmiss=df_merge.merge(df_sp2, left_on=['STUDY_NAME','OID'],right_on=['STUDY_NAME','OID'], how='left', suffixes=['', '_'], indicator=True)
        df_dtsmiss = df_dtsmiss[['STUDY_NAME', 'OID', 'Study Status' ,'Lead DBD','Transition DBD','Old Rule File Developer','Rule File Developer', 'DB Go-Live Date', 'Actual DB Go-Live Date', 'FPFV', 'FPFV Actual', 'LSP','Reason for Not Including in Rule File','Date Planned to be included','table1','table2','LSH setup','Action Required Yes/No']]
        df_dtsmiss = df_dtsmiss.rename(
            columns={"Study Status": "DB Phase", 'DB Go-Live Date': "Planned DB Go-Live Date", 'FPFV': 'FPFV Planned'})
        # df_dtsmiss.to_excel(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS\df_dtsmissnew.xlsx')



    # path_spvl = SITE_ROOT1 + '\\EDMFR\\studyprodverlist'
    # lists = os.listdir(para5)
    # lists = os.listdir(path_spvl)
    list_filenames.extend(para5.filename)
    # os.chdir(path_spvl)
    df_spvl=pd.DataFrame()
    if para5:
    # if len(lists) == 1:
        df_spvl = pd.read_csv(para5)
        # df_spvl = pd.read_csv(lists[0])
        # df_spvl_1['ALS Version'] = df_spvl['eCRF Version'].str.split('_').str[0]
        # df_spvl_1['ALS Version']
        # df_spvl_1 = df_spvl.loc[:, ('Study', 'eCRF Version')]
        df_spvl.loc[:,"date"] = pd.to_datetime(df_spvl["Timestamp"])
        df_spvl = df_spvl.sort_values(by=["date"])
        df_spvl.sort_values('date').groupby('Study').tail(1)
        df_spvl1 = df_spvl.sort_values('date').groupby('Study').tail(1)
        # df_spvl1_v1 = df_spvl1.loc[df_spvl1['eCRF Version'].astype(str).str.contains("^1.")]
        df_spvl1_v1 = df_spvl1.drop_duplicates(subset=['Study'])

        print("df_spvl1['eCRF Version']",df_spvl1['eCRF Version'])

        # df_spvl1_v1.to_excel(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS\output_df_spvl1_v1.xlsx')

        # df_als_study_latest = df_als_sp_1.groupby("Study").max().reset_index()

        # df_alsforms_draftname = df_alsforms1[df_alsforms1.duplicated(subset=['Study ID', 'DraftName'])]

        df_alsforms2 = df_alsforms1.loc[:,['Study ID', 'DraftName']]
        df_alsforms_draftname = df_alsforms2.drop_duplicates(subset=['Study ID', 'DraftName'])
        df_alsforms_draftname['newDraftName'] = df_alsforms_draftname['DraftName'].str.split('_').str[0]
        print("df_alsforms_draftname['newDraftName']",df_alsforms_draftname['DraftName'].tolist())
        print("df_alsforms_draftname['newDraftName']",df_alsforms_draftname['newDraftName'].tolist())
        # df_alsforms_draftname.loc[df_alsforms_draftname["DraftName"].str.contains('[0-9]$'), "newDraftName"] = df_alsforms_draftname["DraftName"].str.rsplit('_', 1).str.get(0)
        # df_als_study_draftissue['newDraftName'] = df_als_study_draftissue['newDraftName'].fillna('')

        df_als_study_draftissue = df_alsforms_draftname[
        ~df_alsforms_draftname.DraftName.astype(str).str.startswith(('0', '1', '2', '3', '4', '5', '6', '7', '8', '9'))]

        df_als_study_draftissue = df_als_study_draftissue.merge(df_dbd, left_on=['Study ID'], right_on=['Trial'], how='left',
                                                        suffixes=['', '_'], indicator=True)
        #
        # if not 'Comment' in df_alsworf.columns:
        #     df_alsworf['Comment'] = ""
        #
        # if not 'Comment' in df_rfwoals.columns:
        #     df_rfwoals['Comment'] = ""

        df_alsworf = df_alsworf.merge(df_dbd, left_on=['Study ID'], right_on=['Trial'], how='left',
                                                        suffixes=['', '_'], indicator=True)
        df_rfwoals = df_rfwoals.merge(df_dbd, left_on=['STUDY_NAME'], right_on=['Trial'], how='left',
                                                        suffixes=['', '_'], indicator=True)

        df_alsworf = df_alsworf.drop(['_merge'], axis=1)
        df_rfwoals = df_rfwoals.drop(['_merge'], axis=1)

        df_rfwoals = df_rfwoals.loc[~df_rfwoals['Trial'].isna()]

        # df_sp3.to_excel(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS\df_sp3.xlsx')
        # df_alsworf.to_excel(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS\df_alsworf.xlsx')


        df_alsworf = df_alsworf.merge(df_sp3, left_on=['Study ID'], right_on=['Study ID'], how='left',
                                                        suffixes=['', '_'], indicator=True
                                      )
        # df_alsworf.to_excel(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS\df_alsworf_m.xlsx')

        df_rfwoals = df_rfwoals.merge(df_sp4, left_on=['STUDY_NAME'], right_on=['STUDY_NAME'], how='left',
                                                        suffixes=['', '_'], indicator=True)

        # df_alsworf = df_alsworf[df_alsworf['_merge'] == 'left_only']
        df_alsworf = df_alsworf.drop_duplicates()
        # df_rfwoals = df_rfwoals[df_rfwoals['_merge'] == 'left_only']
        df_rfwoals = df_rfwoals.drop_duplicates()
        df_rfwoals = df_rfwoals.loc[~df_rfwoals['STUDY_NAME'].isna()]



        df_alsworf = df_alsworf.drop(['_merge'], axis=1)
        df_rfwoals = df_rfwoals.drop(['_merge'], axis=1)

        df_als_study_draftissue = df_als_study_draftissue.drop(['_merge'], axis=1)

        list_of_col_names1 = ['Study ID', 'Trial','Lead DBD', 'Transition DBD', 'Rule File Developer', 'Comment']
        list_of_col_names2 = ['STUDY_NAME','Lead DBD', 'Transition DBD', 'Rule File Developer', 'Comment']

        df_alsworf = df_alsworf.filter(list_of_col_names1)
        df_rfwoals = df_rfwoals.filter(list_of_col_names2)

        # df_als_study_draftissue = df_alsforms_draftname[
        #     (~df_alsforms_draftname.newDraftName.str.startswith(('0', '1', '2', '3', '4', '5', '6', '7', '8', '9'))) |
        #     (~df_alsforms_draftname.DraftName.str.startswith(('0', '1', '2', '3', '4', '5', '6', '7', '8', '9')))]

        df_als_study_latest = df_alsforms_draftname.loc[~(df_alsforms_draftname['DraftName'].isna())]
        # df_als_study_latest = df_als_study_latest.merge(df_alsforms_draftname, left_on=['Study'], right_on=['Study ID'],
        #                                                 how='left', suffixes=['', '_'],
        #                         indicator=True)
        df_als_study_latest = df_als_study_latest.merge(df_spvl1, left_on=['Study ID'], right_on=['Study'], how='left',
                                                        suffixes=['', '_'], indicator=True)
        df_spvl1_v1 = df_spvl1_v1.loc[(~df_spvl1_v1['Study'].isin(df_als_study_latest['Study ID'])) &
                                      (df_spvl1_v1['Study'].isin(df_masterlist1['Study']))]

        df_spvl1_v1.to_excel(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS\output_df_spvl1_v1_1.xlsx')
        df_als_study_latest = df_als_study_latest.drop(['_merge'], axis=1)
        df_als_study_latest1a = df_als_study_latest.loc[(df_als_study_latest['eCRF Version'].isna())]
        df_als_study_latest1c = df_als_study_latest.loc[~(df_als_study_latest['eCRF Version'].isna())]
        df_als_study_latest1b = df_als_study_latest1c.loc[(~(df_als_study_latest1c['eCRF Version'].str.contains('_')))]
        df_als_study_latest1a['Latest Version?'] = 'Study_Production_Version_List is not up to date'
        # df_als_study_latest1a.loc[:,'Latest Version?'] = 'Study_Production_Version_List is not up to date'
        df_als_study_latest1b['Latest Version?'] = 'Naming convention of eCRF Version  is not proper which might not be starting with version number follwed by underscore or Study_Production_Version_List is not up to date'
        df_als_study_latest2 = df_als_study_latest1c.loc[(((df_als_study_latest1c['eCRF Version'].str.contains('_'))))]

        df_als_study_latest2.loc[:,'sdver'] = df_als_study_latest2['eCRF Version'].str.split("_").str[0].tolist()
        df_als_study_latest2 = df_als_study_latest2.loc[(~(df_als_study_latest2['eCRF Version'].isna()) &
                                                         ((df_als_study_latest2['eCRF Version'].str.contains('.'))))]
        df_als_study_latest2.loc[:,'sdverint'] = df_als_study_latest2['sdver'].str.split(".").str[0].tolist()

        req = ["^1", "^2", "^3", "^4", "^5", "^6", "^7", "^8", "^9", "^0"]
        df_als_study_latest2 = df_als_study_latest2.loc[(df_als_study_latest2["DraftName"].str.contains("|".join(req)))]
        # df_als_study_draftissue = df_als_study_latest2.loc[~(df_als_study_latest2["DraftName"].str.contains("|".join(req)))]
        # df_als_study_draftissue = df_als_study_latest2[df_als_study_latest2.DraftName.str.startswith(('0', '1','2', '3','4', '5','6', '7','8', '9'))]
        # print(len(df_als_study_latest2))

        df_als_study_latest2.loc[:,'draftver'] = df_als_study_latest2['DraftName'].str.split("_").str[0].tolist()
        df_als_study_latest2 = df_als_study_latest2.loc[(~(df_als_study_latest2['DraftName'].isna()) &
                                                         ((df_als_study_latest2['DraftName'].str.contains('.'))))]
        df_als_study_latest2.loc[:,'draftverint'] = df_als_study_latest2['draftver'].str.split(".").str[0].tolist()

        # print(df_als_study_latest2.loc[df_als_study_latest2['draftverint'] == 'NEUROSCIENCE','draftverint'])
        # print(df_als_study_latest2.loc[df_als_study_latest2['sdverint'] == 'NEUROSCIENCE', 'sdverint'])

        df_als_study_latest2 = df_als_study_latest2.loc[df_als_study_latest2['draftverint'].str.contains('[0-9]')]
        df_als_study_latest2 = df_als_study_latest2.loc[df_als_study_latest2['sdverint'].str.contains('[0-9]')]


        df_als_study_latest2.to_excel(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS\op1.xlsx')


        # will update on 10:00 issue
        # df_als_study_latest2['draftverint?'] = True
        # df_als_study_latest2['sdverint?'] = True
        #
        # df_als_study_latest2.loc['draftverint?'] = df_als_study_latest2['draftverint'].astype(str).str.isnumeric()
        # df_als_study_latest2.loc['sdverint?'] = df_als_study_latest2['sdverint'].astype(str).str.isnumeric()
        df_als_study_latest2.to_excel(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS\output3.xlsx')
        # df_als_study_latest2 = df_als_study_latest2.loc[(df_als_study_latest2['draftverint?'] == 1) & (df_als_study_latest2['sdverint?'] == 1)]
        df_als_study_latest2.sort_values('draftverint').groupby('Study ID').tail(1)
        df_als_study_latest2 = df_als_study_latest2.sort_values('draftverint').groupby('Study ID').tail(1)

        print("df_als_study_latest2",df_als_study_latest2)
        df_als_study_latest2.loc[(((df_als_study_latest2['draftverint'].astype(int) < df_als_study_latest2['sdverint'].astype(int)))),
                                 'Latest Version?'] = 'ALS repository is not up to date'
        df_als_study_latest2.loc[(((df_als_study_latest2['draftverint'].astype(int) > df_als_study_latest2['sdverint'].astype(int)))),
                                 'Latest Version?'] = 'Study_Production_Version_List is not up to date'
        # df_als_study_latest2.loc[(((df_als_study_latest2['ALS Version'] < df_als_study_latest2['sdverint'].astype(int)))),
        #                          'Latest Version?'] = 'ALS repository is not up to date'
        # df_als_study_latest2.loc[(((df_als_study_latest2['ALS Version'] > df_als_study_latest2['sdverint'].astype(int)))),
        #                          'Latest Version?'] = 'Study_Production_Version_List is not up to date'
        df_als_study_latest2 = df_als_study_latest2.append(df_als_study_latest1a)
        df_als_study_latest2 = df_als_study_latest2.append(df_als_study_latest1b)
        df_spvl1_v1 = df_spvl1_v1.rename(columns={'Study': "Study ID"})
        # df_spvl1_v1 = df_spvl1_v1.drop(['_merge', 'Study_'], axis=1)
        df_spvl1_v1['Latest Version?'] = 'Study ALS is not placed in ALS repository'
        df_spvl1_v1.to_excel(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS\output_df_spvl1_v1_3.xlsx')
        print("df_spvl1_v1 len",len(df_spvl1_v1))
        print("df_als_study_latest2 len",len(df_als_study_latest2))
        df_als_study_latest2.to_excel(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS\output_df_als_study_latest2_1.xlsx')
        df_als_study_latest2 = df_als_study_latest2.append(df_spvl1_v1)
        print("df_als_study_latest2 len",len(df_als_study_latest2))
        df_als_study_latest2 = df_als_study_latest2.loc[~(df_als_study_latest2['Latest Version?'].isna())]
        df_als_study_latest2.to_excel(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS\output_df_als_study_latest2_2.xlsx')


        # df_als_sp_stdmatch = df_als_sp.loc[:['Name', 'Study']]

        # df_als_sp["Namesplit"] = df_als_sp['Name'].astype(str).apply(lambda x: x.split('_')[0] if len(x.split(' ')) > 1 else '')
        # df_als_sp["Namesplit"] = df_als_sp['Namesplit'].astype(str).apply(lambda x: x.split('_')[0] if len(x.split('_')) > 1 else '')
        df_als_sp['Namesplit'] = df_als_sp.apply(lambda x: x.Study.strip() in x.Name, axis=1)
        # df_als_sp_stdmatch = df_als_sp.loc[(df_als_sp['Namesplit'] != '')]
        df_als_sp_stdmatch = df_als_sp.loc[(df_als_sp['Namesplit'] == False)]
        df_als_sp_stdmatch1 = pd.DataFrame()

        # df_als_sp_stdmatch1 = df_als_sp_stdmatch.loc[~(df_als_sp_stdmatch['Name'].str.contains(df_als_sp_stdmatch['Study'].str()))]
        # df_als_sp_stdmatch2 = df_als_sp_stdmatch.loc[(df_als_sp_stdmatch['Name'].str.contains(df_als_sp_stdmatch['Study']))]
        # df_als_sp_stdmatch2 = df_als_sp_stdmatch.loc[~(df_als_sp_stdmatch['Namesplit'] == df_als_sp_stdmatch['Study'])]
        # df_als_sp_stdmatch1 = df_als_sp_stdmatch1.append(df_als_sp_stdmatch2)
        df_masterlist2 = df_masterlist1.loc[:, ['Study', 'Lead DBD', 'Transition DBD']]

        df_als_study_latest2 = df_als_study_latest2.merge(df_masterlist2, left_on=['Study ID'], right_on=['Study'],
                                                          how='left',
                                                          suffixes=['', '_'], indicator=True)
        df_als_study_latest2 = df_als_study_latest2.drop(['_merge','Study_'], axis=1)

        df_als_sp_dup = df_als_sp_dup.merge(df_masterlist2, left_on=['Study'], right_on=['Study'], how='left',
                                            suffixes=['', '_'], indicator=True)
        df_als_sp_dup = df_als_sp_dup.drop(['_merge'], axis=1)
        df_als_sp_stdmatch = df_als_sp_stdmatch.merge(df_masterlist2, left_on=['Study'], right_on=['Study'], how='left',
                                                      suffixes=['', '_'], indicator=True)
        df_als_sp_stdmatch = df_als_sp_stdmatch.drop(['_merge'], axis=1)
        print("processing....please wait")
    # xlprocess=process_exists('EXCEL.EXE')
    # print(xlprocess)
    # print(type(xlprocess))
    path_out = SITE_ROOT1 + '\\EDMFR\\output report'
    purgedir(para11)
    os.chdir(para11)
    # purgedir(path_out)
    # os.chdir(path_out)

    df_edcmiss = df_edcmiss.drop_duplicates()
    # df_edcmiss = df_edcmiss.drop_duplicates(subset=['Study ID', 'OID', 'DB Phase', 'Lead DBD', 'Transition DBD',
    #             'Old Rule File Developer', 'Rule File Developer',
    #             'Planned DB Go-Live Date', 'Actual DB Go-Live Date',
    #             'FPFV Planned', 'FPFV Actual', 'LSP',
    #             'Reason for Not Including in Rule File', 'Date Planned to be included', 'Action Required Yes/No'])

    df_dtsmiss = df_dtsmiss.drop_duplicates()
    # df_dtsmiss = df_dtsmiss.drop_duplicates(subset=['STUDY_NAME', 'OID', 'DB Phase', 'Lead DBD', 'Transition DBD',
    #                                                 'Old Rule File Developer',	'Rule File Developer',
    #                                                 'Planned DB Go-Live Date',	'Actual DB Go-Live Date',
    #                                                 'FPFV Planned', 'FPFV Actual', 'LSP',
    #                                                 'Reason for Not Including in Rule File',
    #                                                 'Date Planned to be included', 'table1', 'table2','LSH setup',
    #                                                 'Action Required Yes/No'])

    # writer = pd.ExcelWriter('EDCDTSMissingFormsReport.xlsx')


    df_filenames = pd.DataFrame(list_filenames, columns=['Uploaded_File_names'])
    df_filenames_xls = pd.DataFrame(files_crf1, columns=['File_names_with_.xls'])
    print(df_filenames)
    options = {}
    options['strings_to_formulas'] = False
    options['strings_to_urls'] = False
    with pd.ExcelWriter('EDCDTSMissingFormsReport.xlsx',engine='openpyxl', mode='w',
                        options=options) as writer:
        df_edcmiss.to_excel(writer, sheet_name='missingEDCFormsList', index=False)
        sheetname = 'missingEDCFormsList'
        worksheet = writer.sheets[sheetname]
        xcel_wraptext(ws=worksheet)
        col_bg_col(ws=worksheet)
        auto_format_cell_width1(worksheet)
        df_dtsmiss = df_dtsmiss.drop_duplicates()
        df_dtsmiss.to_excel(writer, sheet_name='DTSFormsmissing', index=False)
        sheetname = 'DTSFormsmissing'
        worksheet = writer.sheets[sheetname]
        xcel_wraptext(ws=worksheet)
        col_bg_col(ws=worksheet)
        auto_format_cell_width1(worksheet)

        df_alsworf.to_excel(writer, sheet_name='StudiesPresentinALSnotinRF', index=False)
        sheetname = 'StudiesPresentinALSnotinRF'
        worksheet = writer.sheets[sheetname]
        xcel_wraptext(ws=worksheet)
        col_bg_col(ws=worksheet)
        auto_format_cell_width1(worksheet)
        df_rfwoals.to_excel(writer, sheet_name='StudiesPresentinRFnotinALS', index=False)
        sheetname = 'StudiesPresentinRFnotinALS'
        worksheet = writer.sheets[sheetname]
        xcel_wraptext(ws=worksheet)
        col_bg_col(ws=worksheet)
        auto_format_cell_width1(worksheet)

        # print("df_als_study_draftissue",len(df_als_study_draftissue))
        if len(df_als_study_draftissue) > 0:
            df_als_study_draftissue.to_excel(writer, sheet_name='studies_draftnameissue', index=False)
            sheetname = 'studies_draftnameissue'
            worksheet = writer.sheets[sheetname]
            xcel_wraptext(ws=worksheet)
            col_bg_col(ws=worksheet)
            auto_format_cell_width1(worksheet)
        if len(df_als_study_latest2) > 0:
            df_als_study_latest2.to_excel(writer, sheet_name='Latest_studies_version_ALSs', index=False)
            sheetname = 'Latest_studies_version_ALSs'
            worksheet = writer.sheets[sheetname]
            xcel_wraptext(ws=worksheet)
            col_bg_col(ws=worksheet)
            auto_format_cell_width1(worksheet)
        if len(df_als_sp_dup) > 0:
            df_als_sp_dup.to_excel(writer, sheet_name='Duplicate_ALSs_Versions', index=False)
            sheetname = 'Duplicate_ALSs_Versions'
            worksheet = writer.sheets[sheetname]
            xcel_wraptext(ws=worksheet)
            col_bg_col(ws=worksheet)
            auto_format_cell_width1(worksheet)
        if len(df_als_sp_nul) > 0:
            df_als_sp_nul.to_excel(writer, sheet_name='Empty_ALSs_Study_Version', index=False)
            sheetname = 'Empty_ALSs_Study_Version'
            worksheet = writer.sheets[sheetname]
            xcel_wraptext(ws=worksheet)
            col_bg_col(ws=worksheet)
            auto_format_cell_width1(worksheet)
        if len(df_als_sp_stdmatch) > 0:
            df_als_sp_stdmatch.to_excel(writer, sheet_name='Mismatch_Study_Names', index=False)
            sheetname = 'Mismatch_Study_Names'
            worksheet = writer.sheets[sheetname]
            xcel_wraptext(ws=worksheet)
            col_bg_col(ws=worksheet)
            auto_format_cell_width1(worksheet)
        if len(df_filenames) > 0:
            df_filenames.to_excel(writer, sheet_name='File_Names', index=False)
            sheetname = 'File_Names'
            worksheet = writer.sheets[sheetname]
            xcel_wraptext(ws=worksheet)
            col_bg_col(ws=worksheet)
            auto_format_cell_width1(worksheet)
        if len(df_ALS_SP_u) > 0:
            df_ALS_SP_u.to_excel(writer, sheet_name='warnings', index=False)
            sheetname = 'warnings'
            worksheet = writer.sheets[sheetname]
            xcel_wraptext(ws=worksheet)
            col_bg_col(ws=worksheet)
            auto_format_cell_width1(worksheet)

            writer.save()

    # purgedir(path_crf)
    # purgedir(path_mp)
    # purgedir(path_mt)
    # purgedir(path_pout)


    # pythoncom.CoInitialize()
    # xlapp = client.dynamic.Dispatch("Excel.Application")
    # xlapp.Application.Quit()
#########################
    # if os.path.exists(EDMFR_dt_path):
    #     recursive_copy(path_out, EDMFR_dt_path)
    #     webbrowser.open(EDMFR_dt_path)
    #     os.chdir(SITE_ROOT1)
    # popupmsg("Please check the output report folder :" + EDMFR_dt_path)

# edmfr_prog3('EDMFR')