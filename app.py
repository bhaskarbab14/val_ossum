from flask import Flask, flash, request, render_template, send_file
from flask_pymongo import PyMongo
from pymongo import MongoClient
import pandas as pd
app = Flask(__name__)
import pymongo
import datetime, time
# import tkinter as tk
# import xlrd, sys
# import openpyxl
# import xlsxwriter
# import csv
import os
import pandas as pd
import numpy as np
from numpy import math
import random
# from tkinter import filedialog
import shutil
# import shutil, pyodbc
import subprocess
# from openpyxl.drawing.image import Image
from subprocess import check_output
import win32com.client as win32
# import accessdb
# path =''
# from meza import io
# import win32com.client
from win32com import client
# import win32
import win32com
import win32com.client
# import traceback
import tkinter as tk
from tkinter import messagebox, ttk
import webbrowser
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import copy
from openpyxl.utils import get_column_letter
import warnings
import linecache
from collections import Counter
import scipy.sparse as sp
import psutil
from werkzeug.utils import secure_filename

# from ALSsDownload_GVs import *
from Rada_Con_ALS_GVs import *
# import matplotlib.pyplot as plot
from RADADeskApp_comp import *
from EDMFR_new import *
from RAMS_dbprocs import *
from RAMS_dbconn import *

#
# import urllib.parse
#
# username = urllib.parse.quote_plus('pillibh2')
# password = urllib.parse.quote_plus("sairam@21")
#
# url = "mongodb+srv://{}:{}@clustertry1.mwxqe.mongodb.net".format(username, password)
# # url = "mongodb+srv://{}:{}@clustertry1.mwxqe.mongodb.net/ALSs_Rave?retryWrites=true&w=majority".format(username, password)
#
# # mongodb_client =  PyMongo(app, uri=url)
# mongodb_client = pymongo.MongoClient(url)
# db = mongodb_client.db_als
# als = db.alss
#
# personDocument = {
#   "name": { "first": "Alan", "last": "Turing" },
#   "birth": datetime.datetime(1912, 6, 23),
#   "death": datetime.datetime(1954, 6, 7),
#   "contribs": [ "Turing machine", "Turing test", "Turingery" ],
#   "views": 1250000
# }
# als.insert_one(personDocument)
# print("db test",db)
# print("db test",als)
#
# print("mongodb_client dbs",mongodb_client.list_databases())


#####################################functions#########
radaout_curr = ""
radaout_file =""

ALLOWED_EXTENSIONS = set(['xlsx'])
ALLOWED_EXTENSIONS1 = set(['xlsx','xlsm','.accdb'])
ALLOWED_EXTENSIONS2 = set(['csv'])

def allowed_file(filename,allow_ext):
        print("filename",filename)
        print(filename.rsplit('.', 1)[1].lower())
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in allow_ext

def unique_non_null(s):
    return s.dropna().astype(str).unique()


def unique1(strlist1):
    # intilize a null list

    unique_list = []

    # traverse for all elements
    for x in strlist1:
        for y in x.split(','):
            # check if exists in unique_list or not
            if y not in unique_list:
                unique_list.append(y)
    return unique_list


def common_member(a, b):
    a_set = set(a)
    b_set = set(b)
    if (a_set & b_set):
        return a_set & b_set
    else:
        return np.NaN


def keep_duplicates_list(l):
    dupes_list = []
    for e in l:
        num_of_occ = l.count(e)
        if num_of_occ % 2 == 0:
            dupes_list.append(e)
    return dupes_list


# #Check If Value of Column Is Contained in Another Column in the Same Row
# def find_value_column(row, fld1, fld2):
#     for keyword in row[fld1]:
#         if keyword in row[fld2].lower():
#             return True
#         else:
#             return False

def col_bg_col(ws):
    for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
        for cell in rows:
            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type="solid")


def auto_format_cell_width1(ws):
    for letter in range(1, ws.max_column):
        maximum_value = 0
        for cell in ws[get_column_letter(letter)]:
            val_to_check = len(str(cell.value))
            if val_to_check > maximum_value:
                if val_to_check < 50:
                    maximum_value = val_to_check
                else:
                    maximum_value = 50
        ws.column_dimensions[get_column_letter(letter)].width = maximum_value + 1


def xcel_wraptext(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            # alignment = copy.copy(cell.alignment)
            # alignment.wrapText = True
            # Alignment.wrap_text
            # cell.column = alignment
            # ws.cell(row=cell.row, column=col).alignment = alignment


def auto_format_cell_width(ws):
    for col in ws.columns:
        column = get_column_letter(col[0].column)
        width = len(str(col[0])) - 7
        ws.column_dimensions[column].width = width


def move_sheet(wb, from_loc=None, to_loc=None):
    sheets = wb._sheets

    # if no from_loc given, assume last sheet
    if from_loc is None:
        from_loc = len(sheets) - 1

    # if no to_loc given, assume first
    if to_loc is None:
        to_loc = 0

    sheet = sheets.pop(from_loc)
    sheets.insert(to_loc, sheet)


files_sht_als = []
files_sht_ssd = []
files_sht_ecs = []


def xcelread(filep, sheetname, usecolr):
    dfx7 = pd.DataFrame()
    dfx6 = pd.DataFrame()
    dfx5 = pd.DataFrame()
    files_sht_als = []
    xls = pd.ExcelFile(filep)
    if sheetname in xls.sheet_names:
        dfx6 = xls.parse(sheet_name=sheetname, index_col=None, usecols=usecolr, keep_default_na=False,
                         na_values=[''])
        dfx6.dropna(axis=0, how='all', inplace=True)
        dfx5 = xls.parse(sheet_name='CRFDraft', index_col=None, usecols='A:C', keep_default_na=False,
                         na_values=[''])
        dfx5.dropna(axis=0, how='all', inplace=True)
        if len(dfx6) > 0:
            dfx6.loc[:, 'DraftName'] = dfx5.iloc[0, 0]
            dfx6.loc[:, 'Study ID'] = dfx5.iloc[0, 2]
            dfx7 = dfx7.append(dfx6, ignore_index=True, sort=True)
        files_sht_als = files_sht_als.extend(sheetname)
    return dfx7


def RepresentsInt(s):
    try:
        int(s)
        return True
    except ValueError:
        return False


def is_nan(x):
    return isinstance(x, float) and math.isnan(x)


def specxlread_df_getcol_ind(filep, sheetname, colpar, usecolr=None):
    col_skiprows = 0
    df_ssd_sht = pd.DataFrame()
    xl = pd.ExcelFile(filep)
    files_sht_ssd = []
    files_sht_ecs = []
    if sheetname in xl.sheet_names:
        df_getcol_ind = pd.read_excel(filep, sheetname, nrows=20, usecols='A', keep_default_na=False)
        col_skiprows = 0
        files_sht_ssd = files_sht_ssd.extend(sheetname)
        files_sht_ecs = files_sht_ecs.extend(sheetname)

        for ind in range(len(df_getcol_ind)):
            if not (is_nan(df_getcol_ind.iloc[ind, 0]) | RepresentsInt(df_getcol_ind.iloc[ind, 0])):
                if df_getcol_ind.iloc[ind, 0].strip() == colpar:
                    col_skiprows = ind + 1
                    break
    return col_skiprows


def specxlread(filep, sheetname, colpar, usecolr=None):
    df_ssd_sht = pd.DataFrame()
    xl = pd.ExcelFile(filep)
    files_sht_ssd = []
    files_sht_ecs = []
    if sheetname in xl.sheet_names:
        df_getcol_ind = pd.read_excel(filep, sheetname, nrows=20, usecols='A', na_values=[''],
                                      keep_default_na=False)
        col_skiprows = 0
        col_skiprows_nd = 0
        files_sht_ssd = files_sht_ssd.extend(sheetname)
        files_sht_ecs = files_sht_ecs.extend(sheetname)

        for ind in range(len(df_getcol_ind)):
            if not (is_nan(df_getcol_ind.iloc[ind, 0]) | RepresentsInt(df_getcol_ind.iloc[ind, 0])):
                if df_getcol_ind.iloc[ind, 0].strip() == colpar:
                    col_skiprows = ind + 1
                    col_skiprows_nd = 1
                    break

        # print("col_skiprows_nd", col_skiprows)
        if col_skiprows_nd == 1:
            xls = pd.ExcelFile(filep)
            if sheetname in xls.sheet_names:
                df_ssd_sht = pd.read_excel(xls, sheetname, skiprows=col_skiprows, na_values=[''],
                                           keep_default_na=False)
                if usecolr:

                    if usecolr is not None:
                        # astype(int)(usecolr) > 0:
                        for col in usecolr:
                            if col in df_ssd_sht.columns:
                                df_ssd_sht[col] = df_ssd_sht[[col]].ffill()

            df_ssd_sht = df_ssd_sht.loc[:, ~df_ssd_sht.columns.str.contains('^Unnamed')]
    return df_ssd_sht


def recursive_copy(src, dest):
    """
    Copy each file from src dir to dest dir, including sub-directories.
    """
    for item in os.listdir(src):
        file_path = os.path.join(src, item)

        # if item is a file, copy it
        if os.path.isfile(file_path):
            shutil.copy(file_path, dest)

        # else if item is a folder, recurse
        elif os.path.isdir(file_path):
            new_dest = os.path.join(dest, item)
            os.mkdir(new_dest)
            recursive_copy(file_path, new_dest)


def clearfoldercontent(target_dir):
    with os.scandir(target_dir) as entries:
        for entry in entries:
            if entry.is_file() or entry.is_symlink():
                os.remove(entry.path)
            elif entry.is_dir():
                shutil.rmtree(entry.path)


def purgedir(parent):
    for root, dirs, files in os.walk(parent):
        for item in files:
            # Delete subordinate files
            filespec = os.path.join(root, item)
            # if filespec.endswith('.bak'):
            os.unlink(filespec)
        for item in dirs:
            # Recursively perform this operation for subordinate directories
            purgedir(os.path.join(root, item))


nan_value = float("NaN")

NORM_FONT = ("Verdana", 10)


def popupmsg(msg):
    popup = tk.Tk()
    popup.wm_title("!")
    label = ttk.Label(popup, text=msg, font=NORM_FONT)
    label.pack(side="top", fill="x", pady=10)
    B1 = ttk.Button(popup, text="Ok", command=popup.destroy)
    B1.pack()
    popup.mainloop()


def process_exists(process_name):
    call = 'TASKLIST', '/FI', 'imagename eq %s' % process_name
    processes = []
    for process in check_output(call).splitlines()[3:]:
        process = process.decode()
        processes.append(process.split())
    return processes


def proc_kill(procnam):
    for proc in psutil.process_iter():
        if proc.name() == procnam:
            proc.kill()
            time.sleep(20)
            # popupmsg("killed EXCEL.exe")




########################################routes#################

basedir = os.path.abspath(os.path.dirname(__file__))
DATABASE4 = ''
uploadfolder = basedir + '\\uploads'
uploadfolder_als = basedir + '\\uploads\\als'
uploadfolder_src = basedir + '\\uploads\\Source Documents'
consolidatedALS_path = basedir + '\\uploads\\consolidatedALS'
DATABASE1 = "Consolidated_ALS_Global.xlsx"
DATABASE = consolidatedALS_path + "\\" + DATABASE1

DATABASE2 = "Consolidated_ALS.xlsx"
DATABASE3 = consolidatedALS_path + "\\" + DATABASE2
DATABASE4 = "Consolidated_ALS_Selected_Tabs(CAST).xlsx"
DATABASE5 = consolidatedALS_path + "\\" + DATABASE4
DATABASE7 = "Study Summary.xlsx"
DATABASE8 = consolidatedALS_path + "\\" + DATABASE7

radaout_path = uploadfolder + '\\output\\rada'
edmfr_path = uploadfolder + '\\output\\EDMFR'
CONSD_RF_path = uploadfolder + '\\output\\EDMFR\\CONSD RF'
iBOW_path = uploadfolder + '\\output\\EDMFR\\iBOW'
iDTS_path = uploadfolder + '\\output\\EDMFR\\iDTS'
mappingfile_path = uploadfolder + '\\output\\EDMFR\\mappingfile'
output_report_path = uploadfolder + '\\output\\EDMFR\\output report'
prevoutput_path = uploadfolder + '\\output\\EDMFR\\prev output'
studyprodverlist_path = uploadfolder + '\\output\\EDMFR\\studyprodverlist'
tempmapfile_path = uploadfolder + '\\output\\EDMFR\\tempmapfile'

DATABASE5_1 = "EDCDTSMissingFormsReport.xlsx"
DATABASE6 = output_report_path + "\\" + DATABASE5_1
#templates
rada_temp = uploadfolder + '\\temp'
edmfr_templatefile_path = uploadfolder + '\\temp_edmfr'
edmfr_maptempfile_path = uploadfolder + '\\temp_edmfr2'

print("basedir",basedir)
sheetnames = ['CRFDraft', 'Forms', 'Fields', 'DataDictionaries', 'DataDictionaryEntries', 'Checks', 'CheckSteps',
              'CheckActions', 'Derivations', 'Matrices', 'DerivationSteps', 'LabVariableMappings', 'Folders',
              'CustomFunctions', 'CoderConfiguration', 'CoderSupplementalTerms']

if not os.path.exists(uploadfolder):
    os.makedirs(uploadfolder)

if not os.path.exists(uploadfolder_als):
    os.makedirs(uploadfolder_als)

if not os.path.exists(uploadfolder_src):
    os.makedirs(uploadfolder_src)

if not os.path.exists(consolidatedALS_path):
    os.makedirs(consolidatedALS_path)

if not os.path.exists(radaout_path):
    os.makedirs(radaout_path)
if not os.path.exists(edmfr_path):
    os.makedirs(edmfr_path)

if not os.path.exists(CONSD_RF_path):
    os.makedirs(CONSD_RF_path)

if not os.path.exists(iBOW_path):
    os.makedirs(iBOW_path)

if not os.path.exists(mappingfile_path):
    os.makedirs(mappingfile_path)

if not os.path.exists(output_report_path):
    os.makedirs(output_report_path)

if not os.path.exists(prevoutput_path):
    os.makedirs(prevoutput_path)
if not os.path.exists(studyprodverlist_path):
    os.makedirs(studyprodverlist_path)
if not os.path.exists(tempmapfile_path):
    os.makedirs(tempmapfile_path)

if not os.path.exists(radaout_path):
    os.makedirs(radaout_path)

UPLOAD_FOLDER ='/uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

@app.route('/')
def my_home():
    return render_template('home.html')


@app.route('/refuter1', methods=["GET", "POST"])
def refuter1(**kwargs):
        return render_template('index.html')
@app.route('/rada1', methods=["GET", "POST"])
def rada1(**kwargs):
        return render_template('rada.html')
@app.route('/consolid1', methods=["GET", "POST"])
def consolid1(**kwargs):
        return render_template('consolid.html')

@app.route('/uploads1', methods=["GET", "POST"])
def uploads1(**kwargs):
        return render_template('uploads.html')

@app.route('/EDMFR1', methods=["GET", "POST"])
def EDMFR1(**kwargs):
        return render_template('EDMFR.html')


@app.route('/RAMS1', methods=["GET", "POST"])
def RAMS1(**kwargs):

    studies = []
    activities = []
    subgroups = ['CDD','DAP','DRA','COE']
    timespenthrs = [*range(0, 11, 1)]
    timespentmins = [*range(0, 60, 1)]
    ramslist = Rams_proc("pre_rams")
    studies = ramslist[0]
    print("studies",studies)
    # studies = studies.loc[studies['Trial'].isna(), ["Trial"]]
    # studies = studies['Trial'].tolist()
    activities = ramslist[1]
    # df = ramslist[2]
    print("studies",type(studies))
    print("studies",studies)
    # if request.method == 'POST' and request.form['subgroups']. == 'CDD':
    activities = activities.loc[activities['Sub Group'] == 'CDD', ["Activity"]]
    activities = activities['Activity'].tolist()
    print("activities", studies)
    return render_template('RAMS_pre.html', studies=studies, activities=activities, subgroups=subgroups,
                           timespenthrs = timespenthrs, timespentmins = timespentmins
                           )
    #
    # return render_template('RAMS_pre.html', studies=studies, activities=activities, subgroups=subgroups,
    #                        timespenthrs = timespenthrs, timespentmins = timespentmins,
    #                        column_names=df.columns.values, row_data=list(df.values.tolist()),
    #                        link_column="id", zip=zip
    #                        )



@app.route('/RAMS', methods=["GET", "POST"])
def RAMS(**kwargs):



    # SDV = request.form.getlist('SDV select')
    msg =''

    study = request.form.get('studylist')
    subgroup = request.form.get('subgroups')
    activity = request.form.get('activities')
    Comments = request.form.get('Comments')
    timespenthr = request.form.get('timespenthrs')
    print("timespenthr",timespenthr)
    timespentmin = request.form.get('timespentmins')
    print("timespentmins",timespentmin)

    ramslist= Rams_proc("post_rams",study_p=study, subgroup_p=subgroup, activity_p=activity, Comments_p=Comments,
                   timespenthrs_p=timespenthr,timespentmins_p=timespentmin)
    studies = []
    activities = []
    subgroups = ['CDD', 'DAP', 'DRA', 'COE']
    timespenthrs = [*range(0, 11, 1)]
    timespentmins = [*range(0, 60, 1)]
    # ramslist = Rams_proc("post_rams")
    studies = ramslist[0]
    activities = ramslist[1]
    msg = ramslist[2]
    # df = ramslist[3]

    # msg = "Entry has been added successfully"
    print("ramslist", type(ramslist))
    print("ramslist", (ramslist))
    print("ramslist", (ramslist[1]))
    print("ramslist", (ramslist[2]))
    print("studies", type(studies))
    print("studies", studies)
    print("activities", activities)
    print(msg)
    print("msg",msg)
    print("msg",len(msg))
    # print("msg",len(msg[0]))

    # return render_template('RAMS_pre.html', msg = msg, studies=studies, activities=activities, subgroups=subgroups,
    #                        timespenthrs = timespenthrs, timespentmins = timespentmins,
    #                        column_names=df.columns.values, row_data=list(df.values.tolist()),
    #                        link_column="id", zip=zip
    #                        )

    return render_template('RAMS_pre.html', msg = msg, studies=studies, activities=activities, subgroups=subgroups,
                           timespenthrs = timespenthrs, timespentmins = timespentmins
                           )


# @app.route('/my_form')
@app.route('/index', methods=["GET", "POST"])
def index(**kwargs):
    # para1 = kwargs.get('para1', 'para1')
    # print("para1",para1)
    # if para1 == 'Refuter':
    #     return render_template('index.html')
    # elif para1 == 'RADA':
    #     return render_template('rada.html')
    # elif para1 == 'Consolid':
    #     return render_template('consolid.html')
    # elif para1 == 'uploads':
    #     return render_template('uploads.html')

    if request.method == "POST":

        if "ReFUTER" in request.form.keys() :
        # if request.form.values .isin(["ReFUTER"]):
            return render_template('index.html')
        elif "Rada" in request.form.keys():
            return render_template('rada.html')
        elif "Consolidated ALS" in request.form.keys():
            # als_cons_prog_rada(uploadfolder_als,'GV_Study',consolidatedALS_path)
            return render_template('consolid.html',  data = sheetnames)
        elif "Uploads" in request.form.keys():
            return render_template('uploads.html')

# @app.route('/RAMS', methods=[ "POST"])
# def RAMS():
#     sheetlist = request.form.getlist('sheet_names')
#

@app.route('/consolid', methods=[ "POST"])
def consolid():


    GV = request.form.getlist('ALS select')
    sheetlist = request.form.getlist('sheet_names')

    print(GV)
    # GV = "GV_study"
    sheetnames = ['CRFDraft', 'Forms', 'Fields', 'DataDictionaries', 'DataDictionaryEntries', 'Checks',
                  'CheckSteps',
                  'CheckActions', 'Derivations', 'Matrices', 'DerivationSteps', 'LabVariableMappings', 'Folders',
                  'CustomFunctions', 'CoderConfiguration', 'CoderSupplementalTerms']
    print("sheetnames",sheetlist)
    if not GV:
        print(GV)
        GV = "GV_study"
    else:
        GV = GV[0]

    if len(sheetlist) == 0:
        print("sheetnames",sheetlist)
        sheetnames = ['CRFDraft', 'Forms', 'Fields', 'DataDictionaries', 'DataDictionaryEntries', 'Checks',
                      'CheckSteps',
                      'CheckActions', 'Derivations', 'Matrices', 'DerivationSteps', 'LabVariableMappings', 'Folders',
                      'CustomFunctions', 'CoderConfiguration', 'CoderSupplementalTerms']
        # sheetnames = ['Folders']

    DATABASE = als_cons_prog_rada(uploadfolder_als, GV, consolidatedALS_path, sheetnames,sheetlist, uploadfolder_src,
                                  iBOW_path,iDTS_path)
    return render_template('uploadout.html', data=DATABASE)

@app.route('/edmfr', methods=[ "POST"])
def edmfr():

    global edmfrout_curr, edmfrout_file
    if request.method == "POST":
        Rulefile =request.files['Rulefile']
        iBOWfile =request.files['iBOWfile']
        mapfile =request.files['mapfile']
        prevfile =request.files['prevfile']
        stdprdfile =request.files['stdprdfile']
        # dwnfile =request.files['dwnfile']
        print("mapfile",mapfile)
        if mapfile:
            print("mapfile",mapfile)
            purgedir(tempmapfile_path)
            if mapfile and allowed_file(mapfile.filename,ALLOWED_EXTENSIONS2):
                filename = secure_filename(mapfile.filename)
                mapfile.save(os.path.join(tempmapfile_path, filename))

        edmfr_prog3(Rulefile,iBOWfile,mapfile,prevfile,stdprdfile,0, uploadfolder_src, consolidatedALS_path,
                    edmfr_templatefile_path, edmfr_maptempfile_path, output_report_path,tempmapfile_path)
    return render_template('edmfrout.html')

@app.route('/edmfr-file/')
def return_files_edmfr():
    try:
        # return send_file(consolidatedALS_path, attachment_filename='Rave Diagnostic Tool Report.xlsx')
        return send_file(DATABASE6, download_name='EDCDTSMissingFormsReport.xlsx')
        # return send_file(DATABASE6, attachment_filename='EDCDTSMissingFormsReport.xlsx')
    except Exception as e:
        return str(e)

@app.route('/consolidated-file/')
def return_files_con_als():
    try:
        # return send_file(consolidatedALS_path, attachment_filename='Rave Diagnostic Tool Report.xlsx')
        return send_file(DATABASE, download_name='Consolidated_ALS_Global.xlsx')
        # return send_file(DATABASE, attachment_filename='Consolidated_ALS_Global.xlsx')
    except Exception as e:
        return str(e)


@app.route('/consolidated-file_study/')
def return_files_con_als_study():
    try:
        # return send_file(consolidatedALS_path, attachment_filename='Rave Diagnostic Tool Report.xlsx')
        # return send_file(DATABASE3, attachment_filename='Consolidated_ALS.xlsx')
        return send_file(DATABASE3, download_name='Consolidated_ALS.xlsx')
    except Exception as e:
        return str(e)


@app.route('/consolidated-file_cast/')
def return_files_con_als_cast():
    try:
        # return send_file(consolidatedALS_path, attachment_filename='Rave Diagnostic Tool Report.xlsx')
        return send_file(DATABASE5, download_name='Consolidated_ALS_Selected_Tabs(CAST).xlsx')
        # return send_file(DATABASE5, attachment_filename='Consolidated_ALS_Selected_Tabs(CAST).xlsx')
    except Exception as e:
        return str(e)


@app.route('/consolidated-file_studysum/')
def return_files_con_als_studysum():
    try:
        # return send_file(consolidatedALS_path, attachment_filename='Rave Diagnostic Tool Report.xlsx')
        return send_file(DATABASE8, download_name='Study Summary.xlsx')
        # return send_file(DATABASE8, attachment_filename='Study Summary.xlsx')
    except Exception as e:
        return str(e)


@app.route('/rada', methods=["POST"])
def rada():
    global radaout_curr, radaout_file
    if request.method == "POST":
        # radaout = pd.DataFrame()
        print(request.files['alsfile'])
        alsfile =request.files['alsfile']
        ssdfile =request.files['ssdfile']
        ecsfile =request.files['ecsfile']
        print("alsfine",alsfile)

        SDV = request.form.getlist('SDV select')
        if SDV:
            SDV = SDV[0]
        else:
            SDV = 1
        # print("SDV", SDV)
        # print("SDV", SDV[0])
        # purgedir(radaout_path)
        df_list = rada_prog(SDV,alsfile,ssdfile,ecsfile,uploadfolder_src,consolidatedALS_path,radaout_path, rada_temp)
        radaout_curr = df_list[-1]
        radaout_file = df_list[-2]
        df_list = df_list[:-2]
        # df = xcelread(alsfile, 'Forms', 'A:O')
        # for file in request.files:
        # # for row in request.form.values():
        #     print(file)
        #     print(type(file))
        #     print((file))
        #     data.append(row)
        # if "Run_rada" in request.form.keys() :
        #     pass
        # if request.form.values .isin(["ReFUTER"]):
        # return render_template('radaout.html',column_names=df.columns.values, row_data=list(df.values.tolist()),
        #                        zip=zip)
        return render_template('radaout.html', fieldstable=[df_list[0].to_html(classes='data', index=False)],

                               formstable=[df_list[1].to_html(classes='data', index=False)]
                               , dictstable=[df_list[2].to_html(classes='data', index=False)],

                               multistable=[df_list[3].to_html(classes='data', index=False)]
                               , frmrestricts=[df_list[4].to_html(classes='data', index=False)],

                               echecks=[df_list[5].to_html(classes='data', index=False)],
                               alschecks=[df_list[6].to_html(classes='data', index=False)],

                               ssdchecks=[df_list[7].to_html(classes='data', index=False)]
                               ,ecschecks=[df_list[8].to_html(classes='data', index=False)],

                               warnings=[df_list[9].to_html(classes='data', index=False)]
                               , titles=['na', 'Results_FIELDS', 'Results_FORMS','Results_DICTIONARY','Results_Multiple'
                ,'Results_Restrictions','Results_EditChecks','Results_ALS','Results_SSD',
'Results_ECS','warnings'],value=['Results_FIELDS', 'Results_FORMS','Results_DICTIONARY','Results_Multiple',
                                 'Results_Restrictions','Results_EditChecks','Results_ALS','Results_SSD',
'Results_ECS','warnings'])


@app.route('/Radaout-file/')
def return_files_radaout():
    try:
        # return send_file(consolidatedALS_path, attachment_filename='Rave Diagnostic Tool Report.xlsx')
        return send_file(radaout_curr, download_name=radaout_file)
        # return send_file(radaout_curr, attachment_filename=radaout_file)
    except Exception as e:
        return str(e)


@app.route('/uploads', methods=["POST"])
def uploads():
    if request.method == "POST":

        alslistf =request.form.keys()
        metalist1 =request.files.getlist('metafile')
        alslist1 =request.files.getlist('alsfile')
        # iBowfile1 =request.files.getlist('iBowfile')
        # abcdfile1 =request.files.getlist('abcdfile')
        # ndpfile1 =request.files.getlist('ndpfile')
        
        
        iBowfile1 =request.files['iBowfile']
        abcdfile1 =request.files['abcdfile']
        ndpfile1 =request.files['ndpfile']
        idtsfile1 =request.files['idtsfile']
        print("alslistf",alslistf)

        # print("alslist",alslist1)
        # print("metalist",metalist1)
        alslist = []
        metalist = []
        ibowlist = []
        abcdlist = []
        ndplist = []

        for row in alslist1:
            print(row.filename)
            alslist.append(row.filename)
            break

        for row in metalist1:
            metalist.append(row.filename)
            break
        # 
        # for row in iBowfile1:
        #     ibowlist.append(row.filename)
        #     break
        # 
        # for row in ndpfile1:
        #     ndplist.append(row.filename)
        #     break
        # 
        # for row in abcdfile1:
        #     abcdlist.append(row.filename)
        #     break
        # 
        # 

        df_list = uploadfolder_als
        print("alslist",alslist)
        print("metalist",metalist)
        print(len(alslist) > 0)
        print((request.form.keys()))
        print(("Upload ALSs" in request.form.keys()))
        if (len(metalist) > 0) & ("Upload Metadata" in request.form.keys()):
            print("Upload metaaa")
            purgedir(uploadfolder_src)
            for file in metalist1:
                if file and allowed_file(file.filename,ALLOWED_EXTENSIONS1):
                    filename = secure_filename(file.filename)
                    file.save(os.path.join(uploadfolder_src, filename))
                    # file.save(os.path.join(basedir,app.config['UPLOAD_FOLDER'], filename))

            return render_template('home.html', msg="Documents are uploaded " + uploadfolder_src)

        elif (len(alslist) > 0) & ("Upload ALSs" in request.form.keys()):
            print("Upload ALSsss")
            purgedir(uploadfolder_als)
            for file in alslist1:
                if file and allowed_file(file.filename,ALLOWED_EXTENSIONS):
                    filename = secure_filename(file.filename)
                    file.save(os.path.join(uploadfolder_als, filename))
                    # file.save(os.path.join(basedir,app.config['UPLOAD_FOLDER'], filename))

            return render_template('home.html', msg="Documents are uploaded " + uploadfolder_als)

        elif ("Upload studies list" in request.form.keys()):

            purgedir(iBOW_path)
            if iBowfile1 and allowed_file(iBowfile1.filename, ALLOWED_EXTENSIONS):
            #     filename = secure_filename(iBowfile1.filename)
            #     iBowfile1.save(os.path.join(iBOW_path, filename))
                  iBowfile1.save(os.path.join(iBOW_path, iBowfile1.filename))
                
            Rams_uploads(iBowfile1,'ibow')


            return render_template('home.html', msg="Documents are uploaded" )

        elif  ("Upload activities list" in request.form.keys()):
            Rams_uploads(abcdfile1,'abcd')

            return render_template('home.html', msg="Documents are uploaded")

        elif  ("Upload NDPs list" in request.form.keys()):
            Rams_uploads(ndpfile1,'ndp')

            return render_template('home.html', msg="Documents are uploaded")


        elif  ("Upload iDTS file" in request.form.keys()):
            purgedir(iDTS_path)
            if idtsfile1 and allowed_file(idtsfile1.filename, ALLOWED_EXTENSIONS):
                filename = secure_filename(idtsfile1.filename)
                idtsfile1.save(os.path.join(iDTS_path, filename))

            return render_template('home.html', msg="Documents are uploaded")



            # flash('File(s) successfully uploaded')

        # return render_template('home.html', msg= "Documents are uploaded " + uploadfolder_als)
        # return render_template('uploadout.html',msg= "Documents are uploaded " + uploadfolder_als)

@app.route('/data', methods=["POST"])
def data():
    if request.method == "POST":
    #     if request.form.get() == "ReFUTER":
        data = []
        for row in request.form.values():
            data.append(row)
        return render_template('data.html', data=data)
    # elif request.method == 'GET':
    #     return render_template('index.html')
    # elif request.method == 'GET':
    #         data = []
    #         for row in request.form.values():
    #             print(row)
    #             print(type(data))
    #             print((data))
    #             data.append(row)
    #         return render_template('data.html', data=data)

#
# @app.route('/my_form1', methods=['GET','POST'])
# def my_form1():
#     if request.method == "POST":
#             data = []
#             for row in request.form.values():
#                 print(row)
#                 print(type(data))
#                 print((data))
#                 data.append(row)
#             return render_template('data.html', data=data)

#
#
# @app.route('/', methods=['GET','POST'])
# def parse_data():
#     if request.method == "POST":
#         if request.form["ReFUTER"] == "ReFUTER":
#             data = []
#             for row in request.form.values():
#                 print(row)
#                 print(type(data))
#                 print((data))
#                 data.append(row)
#             # d = request.form['text1']
#             # data = data.append(d)
#             # FTEs_SU_CDD = (float(data[0])*float(data[10]))/1600
#             # FTEs_SU_PPC = ((float(data[6])* sum(float(data[14]),float(data[15])))+(float(data[6])*0.6*float(data[18])))/1600
#             print(data)
#         return render_template('data.html', data=data)
if __name__ == "__main__":
    app.run(debug=True, port=8080, use_reloader=False)
#
#
# from flask import Flask,render_template, request
# import pandas as pd
# import csv
#
# app = Flask(__name__)
#
# @app.route('/',methods =['GET','POST'])
# def index():
#     return render_template('index.html')
#
# @app.route('/data',methods =['GET','POST'])
# def data():
#     if request.method == 'POST':
#         d = request.form['Upload']
#         data = []
#         with open(d) as file:
#             csvfile = csv.reader(file)
#             for row in csvfile:
#                 data = data.append(row)
#         data = pd.DataFrame(data)
#         print(data.head(1))
#         # return render_template('data.html')
#         return render_template('data.html', data = data)
#         # render_template('data.html',data = data.to_html(header = False, index = False))
#
# if __name__ == '__main__':
#     app.run(debug=True)
