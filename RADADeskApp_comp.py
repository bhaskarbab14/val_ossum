# purgedir(path_als)
def rada_prog(para1,para2,para3,para4,para5,para6,para7,para8):
    # sdv = 1
    sdv = para1
    import datetime, time
    # import tkinter as tk
    import xlrd, sys
    import openpyxl
    import xlsxwriter
    import csv
    import os
    import pandas as pd
    import numpy as np
    from numpy import math
    import random
    # from tkinter import filedialog
    import shutil
    # import shutil, pyodbc
    import subprocess
    from openpyxl.drawing.image import Image
    from subprocess import check_output
    import win32com.client as win32
    # import accessdb
    # path =''
    # from meza import io
    # import win32com.client
    from win32com import client
    # import win32
    import win32com
    import win32com.client
    # import traceback
    import tkinter as tk
    from tkinter import messagebox, ttk
    import webbrowser
    from openpyxl.drawing.image import Image
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
    import copy
    from openpyxl.utils import get_column_letter
    import warnings
    import linecache
    from collections import Counter
    import scipy.sparse as sp
    import psutil
    # from ALSsDownload_GVs import *
    # import matplotlib.pyplot as plot

    #####################################functions#########


    def unique_non_null(s):
        return s.dropna().astype(str).unique()

    def unique1(strlist1):
        # intilize a null list

        unique_list = []

        # traverse for all elements
        for x in strlist1:
            for y in x.split(','):
                # check if exists in unique_list or not
                if y not in unique_list:
                    unique_list.append(y)
        return unique_list

    def common_member(a, b):
        a_set = set(a)
        b_set = set(b)
        if (a_set & b_set):
            return a_set & b_set
        else:
            return np.NaN

    def keep_duplicates_list(l):
        dupes_list = []
        for e in l:
            num_of_occ = l.count(e)
            if num_of_occ % 2 == 0:
                dupes_list.append(e)
        return dupes_list

    # #Check If Value of Column Is Contained in Another Column in the Same Row
    # def find_value_column(row, fld1, fld2):
    #     for keyword in row[fld1]:
    #         if keyword in row[fld2].lower():
    #             return True
    #         else:
    #             return False

    def col_bg_col(ws):
        for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
            for cell in rows:
                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type="solid")

    def auto_format_cell_width1(ws):
        for letter in range(1, ws.max_column):
            maximum_value = 0
            for cell in ws[get_column_letter(letter)]:
                val_to_check = len(str(cell.value))
                if val_to_check > maximum_value:
                    if val_to_check < 50:
                        maximum_value = val_to_check
                    else:
                        maximum_value = 50
            ws.column_dimensions[get_column_letter(letter)].width = maximum_value + 1

    def xcel_wraptext(ws):
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                # alignment = copy.copy(cell.alignment)
                # alignment.wrapText = True
                # Alignment.wrap_text
                # cell.column = alignment
                # ws.cell(row=cell.row, column=col).alignment = alignment

    def auto_format_cell_width(ws):
        for col in ws.columns:
            column = get_column_letter(col[0].column)
            width = len(str(col[0])) - 7
            ws.column_dimensions[column].width = width

    def move_sheet(wb, from_loc=None, to_loc=None):
        sheets = wb._sheets

        # if no from_loc given, assume last sheet
        if from_loc is None:
            from_loc = len(sheets) - 1

        # if no to_loc given, assume first
        if to_loc is None:
            to_loc = 0

        sheet = sheets.pop(from_loc)
        sheets.insert(to_loc, sheet)

    files_sht_als = []
    files_sht_ssd = []
    files_sht_ecs = []

    def xcelread(filep, sheetname, usecolr):
        dfx7 = pd.DataFrame()
        dfx6 = pd.DataFrame()
        dfx5 = pd.DataFrame()
        files_sht_als = []
        xls = pd.ExcelFile(filep)
        if sheetname in xls.sheet_names:
            dfx6 = xls.parse(sheet_name=sheetname, index_col=None, usecols=usecolr, keep_default_na=False,
                             na_values=[''])
            dfx6.dropna(axis=0, how='all', inplace=True)
            dfx5 = xls.parse(sheet_name='CRFDraft', index_col=None, usecols='A:C', keep_default_na=False,
                             na_values=[''])
            dfx5.dropna(axis=0, how='all', inplace=True)
            if len(dfx6) > 0:
                dfx6.loc[:, 'DraftName'] = dfx5.iloc[0, 0]
                dfx6.loc[:, 'Study ID'] = dfx5.iloc[0, 2]
                dfx7 = dfx7.append(dfx6, ignore_index=True, sort=True)
            files_sht_als = files_sht_als.extend(sheetname)
        return dfx7

    def RepresentsInt(s):
        try:
            int(s)
            return True
        except ValueError:
            return False

    def is_nan(x):
        return isinstance(x, float) and math.isnan(x)

    def specxlread_df_getcol_ind(filep, sheetname, colpar, usecolr=None):
        col_skiprows = 0
        df_ssd_sht = pd.DataFrame()
        xl = pd.ExcelFile(filep)
        files_sht_ssd = []
        files_sht_ecs = []
        if sheetname in xl.sheet_names:
            df_getcol_ind = pd.read_excel(filep, sheetname, nrows=20, usecols='A', keep_default_na=False)
            col_skiprows = 0
            files_sht_ssd = files_sht_ssd.extend(sheetname)
            files_sht_ecs = files_sht_ecs.extend(sheetname)

            for ind in range(len(df_getcol_ind)):
                if not (is_nan(df_getcol_ind.iloc[ind, 0]) | RepresentsInt(df_getcol_ind.iloc[ind, 0])):
                    if df_getcol_ind.iloc[ind, 0].strip() == colpar:
                        col_skiprows = ind + 1
                        break
        return col_skiprows

    def specxlread(filep, sheetname, colpar, usecolr=None):
        df_ssd_sht = pd.DataFrame()
        xl = pd.ExcelFile(filep)
        files_sht_ssd = []
        files_sht_ecs = []
        if sheetname in xl.sheet_names:
            df_getcol_ind = pd.read_excel(filep, sheetname, nrows=20, usecols='A', na_values=[''],
                                          keep_default_na=False)
            col_skiprows = 0
            col_skiprows_nd = 0
            files_sht_ssd = files_sht_ssd.extend(sheetname)
            files_sht_ecs = files_sht_ecs.extend(sheetname)

            for ind in range(len(df_getcol_ind)):
                if not (is_nan(df_getcol_ind.iloc[ind, 0]) | RepresentsInt(df_getcol_ind.iloc[ind, 0])):
                    if df_getcol_ind.iloc[ind, 0].strip() == colpar:
                        col_skiprows = ind + 1
                        col_skiprows_nd = 1
                        break

            # print("col_skiprows_nd", col_skiprows)
            if col_skiprows_nd == 1:
                xls = pd.ExcelFile(filep)
                if sheetname in xls.sheet_names:
                    df_ssd_sht = pd.read_excel(xls, sheetname, skiprows=col_skiprows, na_values=[''],
                                               keep_default_na=False)
                    if usecolr:

                        if usecolr is not None:
                              # astype(int)(usecolr) > 0:
                            for col in usecolr:
                                if col in df_ssd_sht.columns:
                                    df_ssd_sht[col] = df_ssd_sht[[col]].ffill()

                df_ssd_sht = df_ssd_sht.loc[:, ~df_ssd_sht.columns.str.contains('^Unnamed')]
        return df_ssd_sht

    def recursive_copy(src, dest):
        """
        Copy each file from src dir to dest dir, including sub-directories.
        """
        for item in os.listdir(src):
            file_path = os.path.join(src, item)

            # if item is a file, copy it
            if os.path.isfile(file_path):
                shutil.copy(file_path, dest)

            # else if item is a folder, recurse
            elif os.path.isdir(file_path):
                new_dest = os.path.join(dest, item)
                os.mkdir(new_dest)
                recursive_copy(file_path, new_dest)

    def clearfoldercontent(target_dir):
        with os.scandir(target_dir) as entries:
            for entry in entries:
                if entry.is_file() or entry.is_symlink():
                    os.remove(entry.path)
                elif entry.is_dir():
                    shutil.rmtree(entry.path)

    def purgedir(parent):
        for root, dirs, files in os.walk(parent):
            for item in files:
                # Delete subordinate files
                filespec = os.path.join(root, item)
                # if filespec.endswith('.bak'):
                os.unlink(filespec)
            for item in dirs:
                # Recursively perform this operation for subordinate directories
                purgedir(os.path.join(root, item))

    nan_value = float("NaN")

    NORM_FONT = ("Verdana", 10)

    def popupmsg(msg):
        popup = tk.Tk()
        popup.wm_title("!")
        label = ttk.Label(popup, text=msg, font=NORM_FONT)
        label.pack(side="top", fill="x", pady=10)
        B1 = ttk.Button(popup, text="Ok", command=popup.destroy)
        B1.pack()
        popup.mainloop()

    def process_exists(process_name):
        call = 'TASKLIST', '/FI', 'imagename eq %s' % process_name
        processes = []
        for process in check_output(call).splitlines()[3:]:
            process = process.decode()
            processes.append(process.split())
        return processes

    def proc_kill(procnam):
        for proc in psutil.process_iter():
            if proc.name() == procnam:
                proc.kill()
                time.sleep(20)
                # popupmsg("killed EXCEL.exe")


    #######################################

    df_ecs_ec = pd.DataFrame()
    df_ecs_esae = pd.DataFrame()
    df_ecs_pd = pd.DataFrame()
    df_stage_domain = pd.DataFrame()
    df_Questions = pd.DataFrame()
    df_Test_Category = pd.DataFrame()
    list_ecs_ec_misscols = []
    list_ecs_esae_misscols = []
    list_ecs_pd_misscols = []
    df_ssd_colist = pd.DataFrame()
    df_ecs_study = pd.DataFrame()
    df_ssd_study = pd.DataFrame()
    df_ssd_frmattr = pd.DataFrame()
    df_ssd_newfrmattr = pd.DataFrame()
    df_ssd_esae1 = pd.DataFrame()
    df_ssd_esae2 = pd.DataFrame()
    df_ssd_colist_reshp1 = pd.DataFrame()
    formnames_s1 = []
    df_als_forms = pd.DataFrame()
    df_forms = pd.DataFrame()
    df_als_flds = pd.DataFrame()
    df_lbsettings = pd.DataFrame()
    df_dicts = pd.DataFrame()
    df_als_checks1 = pd.DataFrame()
    df_ca = pd.DataFrame()
    df_flds = pd.DataFrame()
    df_cf = pd.DataFrame()
    df_dictnams = pd.DataFrame()
    df_als_folders = pd.DataFrame()
    df_forms_non_ops_esae_pd = pd.DataFrame()
    df_flds_chks = pd.DataFrame()
    df_checksteps = pd.DataFrame()
    df_als_checks = pd.DataFrame()
    df_ca_oq = pd.DataFrame()
    df_flds_defaults_split = pd.DataFrame()
    df_dicts_defaults_dict = pd.DataFrame()
    df_flds_numeric = pd.DataFrame()
    df_flds_non_esae = pd.DataFrame()
    df_Derivations = pd.DataFrame()
    df_ca1 = pd.DataFrame()
    df_flds_inact = pd.DataFrame()
    df_forms_inact = pd.DataFrame()
    df_crfdraft = pd.DataFrame()
    df_lbsettings1 = pd.DataFrame()
    df_flds_non_ops_esae = pd.DataFrame()
    df_stage_domain2 = pd.DataFrame()
    df_Matrices = pd.DataFrame()
    df_flds_sdpv = pd.DataFrame()
    df_flds_default_sdp = pd.DataFrame()
    df_ssd_frmattr_df_getcol_ind = pd.DataFrame()
    df_ssd_esae1_df_getcol_ind = pd.DataFrame()
    df_ssd_esae2_df_getcol_ind = pd.DataFrame()
    df_ssd_newfrmattr_df_getcol_ind = pd.DataFrame()
    df_ssd_colist_df_getcol_ind = pd.DataFrame()
    df_ecs_esae_df_getcol_ind = pd.DataFrame()
    df_ecs_ec_df_getcol_ind = pd.DataFrame()
    df_ecs_pd_df_getcol_ind = pd.DataFrame()
    df_ecs_pd_reshp = pd.DataFrame()
    df_warnings = pd.DataFrame()
    df_forms_global_esae_no = pd.DataFrame()
    df_flds_glob_formyes_fieldno = pd.DataFrame()
    df_flds_glob_formyes_fieldyes = pd.DataFrame()
    df_chk129 = pd.DataFrame()
    df_chk130 = pd.DataFrame()
    df_radachecks2 = pd.DataFrame()
    df_codelist_StageDomain = pd.DataFrame(columns = [['Controlled Terminology', 'Sdtm',
                                                                            'Long Codelist']])

    #######################################
    # with warnings.catch_warnings(record=True) as w:
    #     # Cause all warnings to always be triggered.
    #     warnings.simplefilter("default")
    #
    # username = os.environ["USERNAME"]
    # ossum_path1 = 'C:\\Users\\' + username
    # osssum_dir = 'OSSUM'
    # ossum_path = ossum_path1 +'\\' + osssum_dir
    # ALSdownloadpath = 'C:\\Users\\' + username + '\\ALSdownloadstd'
    # cons_ALS_Global_path = 'Consolidated_ALS_Global'
    # ossum_cons_path = ossum_path + '\\' + cons_ALS_Global_path
    #
    # dir = os.getcwd()
    #
    # SITE_ROOT = os.getcwd()
    # PARENT_ROOT = os.path.abspath(os.path.join(SITE_ROOT, os.pardir))

    # path_sp = ossum_path + '\Source Documents'
    # path_sp = ossum_path + '\Source Documents'
    # path_sp = SITE_ROOT + '\Source Documents'
    # path_sp = PARENT_ROOT + '\Source Documents'
    # os.chdir(path_sp)
    # lists = os.listdir(path_sp)
##########################

    # if os.path.exists(ossum_path1):
    #     if not os.path.exists(ALSdownloadpath):
    #         os.makedirs(ALSdownloadpath)
    #
    # if os.path.exists(ossum_cons_path):
    #     DATABASE1 = "Consolidated_ALS_Global.xlsx"
    #     DATABASE = ossum_cons_path + "\\" + DATABASE1
    #     if os.path.isfile(DATABASE):
    #         xls = pd.ExcelFile(DATABASE)
    #         # df_gv_forms = pd.read_excel(xls, 'Summary')
    #         xls.close()
    #     else:
    #         popupmsg("Please download and consolidate standards ALS in SCAD at Home page")
    #         sys.exit("Please download and consolidate standards ALS in SCAD at Home page")
    # else:
    #     popupmsg("Please download and consolidate standards ALS in SCAD at Home page")
    #     sys.exit("Please download and consolidate standards ALS in SCAD at Home page")
    #


    ######################################
    # print("checking excel process")
    # xlprocess = process_exists('EXCEL.EXE')
    # if len(xlprocess) > 0:
    #     proc_kill("EXCEL.EXE")
    #     print("Closing EXCEL.EXE, please wait")
    #     time.sleep(10)
    #
    # filename = path_sp + '\\' + 'SharepointALSdwn.xlsm'
    # filename1 = path_sp + '\\' + 'StandardALSs.xlsm'
    # filename2 = path_sp + '\\' + 'StandardALSdwn.xlsm'
    # #######
    # #
    # # xlapp = win32.gencache.EnsureDispatch('Excel.Application')
    # # wb = xlapp.Workbooks.Open(Filename=filename2, ReadOnly=0)
    # # wb.RefreshAll()
    # # print("Refreshing global volumes, please wait")
    # # xlapp.CalculateUntilAsyncQueriesDone()
    # # xlapp.DisplayAlerts = False
    # # wb.Save()
    # # xlapp.Quit()
    # # os.chdir(path_sp)
    #
    # df_standard_als = pd.read_excel("StandardALSdwn.xlsm")
    #
    #
    #
    # df_standard_als = df_standard_als.filter(items=['Name'])
    #
    # df_standard_als.loc[df_standard_als['Name'].astype(str).str.contains('#'), 'Name'] = \
    #     df_standard_als['Name'].astype(str).str.split('#', 1).str.get(0)
    #

    ############################################################
    # query = 'select * from [ALS Repository]'
    #
    # df_ALS_SP = pd.read_sql(query, conn)
    # conn.close()
    ####################################################

    if os.path.exists(para6):
        DATABASE1 = "Consolidated_ALS_Global.xlsx"
        DATABASE = para6 + "\\" + DATABASE1
        if os.path.isfile(DATABASE):
            pass
            # xls = pd.ExcelFile(DATABASE)
            # # df_gv_forms = pd.read_excel(xls, 'CRFDraft')
            # df_gv_forms = pd.read_excel(xls, 'Summary')
            # xls.close()
            # listlen = os.listdir(ALSdownloadpath)

            # list_1 = df_gv_forms['DraftName'].tolist()
            # list_1 = df_gv_forms['Name'].tolist()
            # print("set(listlen)",(listlen) )
            # list_2 = df_standard_als['Name'].tolist()
            # print((list_2) )
            # if not (set(listlen) == set(list_2)):
            #     print(set(listlen) )
            #     print(set(list_2) )
            #     # ALSsDownload(para1='GV')
            #     # con_als(para1="Cons_Gv")
            #     xlprocess = process_exists('EXCEL.EXE')
            #     if len(xlprocess) > 0:
            #         proc_kill("EXCEL.EXE")
            #         print("Closing EXCEL.EXE, please wait")
            #     popupmsg("Go->SCAD->click on 'Download ALSs (Global Volumes)'->wait for while ->"
            #              "click on 'Consolidated ALSs (GVs)'")
                # popupmsg("click on 'Upload Codelist (if refresh)' and wait for while ")
                # popupmsg("Go->SCAD->click on 'Download ALSs (Global Volumes)'->wait for while ->"
                #          "click on 'Consolidated ALSs (GVs)'")
                # popupmsg("Please download and consolidate standards ALS in SCAD at Home page")
                # sys.exit("Please download and consolidate standards ALS in SCAD at Home page")
                # sys.exit("Go->SCAD->click on 'Download ALSs (Global Volumes)'->wait for while ->"
                #          "click on 'Consolidated ALSs (GVs)'")
        #
        # else:
        #     # ALSsDownload(para1='GV')
        #     # con_als(para1="Cons_Gv")
        #     df = pd.DataFrame()
        #     df.to_excel(DATABASE)
        #     popupmsg("Go & click on 'Download ALSs (Global Volumes)'->wait for while ->"
        #              "click on 'Consolidated ALSs (GVs)'")
        #     sys.exit("Go & click on 'Download ALSs (Global Volumes)'->wait for while ->"
        #              "click on 'Consolidated ALSs (GVs)'")
    # else:
        # ALSsDownload(para1='GV')
        # con_als(para1="Cons_Gv")


        #
        # os.makedirs(ossum_cons_path)
        #
        # popupmsg("Go->SCAD->click on 'Download ALSs (Global Volumes)'->wait for while ->"
        #          "click on 'Consolidated ALSs (GVs)'")
        # sys.exit("Go->SCAD->click on 'Download ALSs (Global Volumes)'->wait for while ->"
        #          "click on 'Consolidated ALSs (GVs)'")

    #######################

    # # for codelist from SP
    # try:
    #     DATABASE1 = "SG_Codelist.accdb"
    #     DATABASE = path_sp + "\\" + DATABASE1
    #     pyodbc.pooling = False
    #     conn = cursor = None
    #     conn = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=%s;' % (DATABASE))
    #     cursor = conn.cursor()
    #     query = 'select * from [CODELIST]'
    #     df_codelist_StageDomain = pd.read_sql(query, conn)
    #
    #     # query = 'select * from [RaveX Standards]'
    #     # df_standard_als = pd.read_sql(query, conn)
    #     #
    #     # df_standard_als = df_standard_als.loc[(df_standard_als["Research Area Info (optional)"] == "ALS") &
    #     #                                       (df_standard_als["Status"] == "Final")]
    #     # df_standard_als = df_standard_als.filter(items=['Name'])
    #     conn.close()
    # except:
    #     popupmsg("You have MS Access Engine issues. Please use other mode to upload codelist")
    #     sys.exit(sys.exc_info())
    # finally:
    #     print("finally")
    #     try:
    #         cursor.close()
    #     except:
    #         pass
    #     try:
    #         conn.close()
    #     except:
    #         pass
    # except:
    #     DATABASE1 = "SG_Codelist.accdb"
    #     DATABASE = path_sp + "\\" + DATABASE1
    #     try:
    #         conn = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=%s;' % (DATABASE))
    #         cursor = conn.cursor()
    #         query = 'select * from [CODELIST]'
    #         df_codelist_StageDomain = pd.read_sql(query, conn)
    #         # query = 'select * from [RaveX Standards]'
    #         # df_standard_als = pd.read_sql(query, conn)
    #         # conn.close()
    #         # df_standard_als = df_standard_als.loc[(df_standard_als["Research Area Info (optional)"] == "ALS") &
    #         #                                   (df_standard_als["Status"] == "Final")]
    #         # df_standard_als = df_standard_als.filter(items=['Name'])
    #         con.close()
    #     except:
    #         popupmsg("MS access files are opened, please close")
    #         sys.exit(sys.exc_info())


        # if not (set(df_gv_forms['Name'].tolist()) == set(df_standard_als['Name'].tolist())):
        #     xlprocess = process_exists('EXCEL.EXE')
        #     if len(xlprocess) > 0:
        #         proc_kill("EXCEL.EXE")
        #         print("Closing EXCEL.EXE, please wait")
        #     popupmsg("Go->SCAD->click on 'Download ALSs (Global Volumes)'->wait for while ->"
        #                  "click on 'Consolidated ALSs (GVs)'")
        #     sys.exit("Please download and consolidate standards ALS in SCAD at Home page")
        # # xls = pd.ExcelFile('Codelist_stagedomain.xlsx')
        # df_codelist_StageDomain = pd.read_excel(xls, 'Codelist', skiprows=1)

    ###################
    # xls = pd.ExcelFile('Stage Metadata.xlsx')
    # df_CODELIST = pd.read_excel(xls, 'Sheet1',skiprows=1)
    # xls.close()
    xls ='CODELIST.xlsx'
    CODELIST_file = para5 + '\\' + xls
    if os.path.exists(CODELIST_file):
        xls = pd.ExcelFile(CODELIST_file)
        df_codelist_StageDomain = pd.read_excel(xls, 'CODELIST')
        xls.close()
    xls ='NonDomainForms.xlsx'
    Non_DomainForms_file = para5 + '\\' + xls
    if os.path.exists(Non_DomainForms_file):
        xls = pd.ExcelFile(Non_DomainForms_file)
        df_non_domain_forms = pd.read_excel(xls, 'NonDomainForms')
        df_eSAE_forms = pd.read_excel(xls, 'eSAEForms')
        xls.close()

    ########################################################
    xls = 'ELE_CODE.xlsx'
    ELE_CODE_file = para5 + '\\' + xls
    if os.path.exists(ELE_CODE_file):
        modx = os.path.getmtime(ELE_CODE_file)
        xmod = datetime.datetime.fromtimestamp(modx)
        date = datetime.datetime.now()
        modTime = time.mktime(date.timetuple())
        fileLocation = os.path.abspath(xls)
    # if not datetime.datetime.today().date() == xmod.date():
    #     xlapp = win32com.client.DispatchEx("Excel.Application")
    #     xlapp.Quit()
    #     xlapp = win32com.client.DispatchEx("Excel.Application")
    #     # print("currentpath", path_sp)
    #     # print("currentpath", filename1)
    #     wb = xlapp.Workbooks.Open(filename1)
    #     wb.RefreshAll()
    #     xlapp.CalculateUntilAsyncQueriesDone()
    #     xlapp.DisplayAlerts = False
    #     wb.Save()
    #     xlapp.Quit()
    #     os.utime(fileLocation, (modTime, modTime))
    # print("currentpath",path_sp)
        xls = pd.ExcelFile(ELE_CODE_file)
        # xls = pd.ExcelFile('ELE_CODE.xlsx')
        df_stage_domain = pd.read_excel(xls, 'stage_domain')
        # df_ELEMENTS = pd.read_excel(xls, 'ELEMENTS')
        # # df_CODELIST = pd.read_excel(xls, 'CODELIST')
        df_Questions = pd.read_excel(xls, 'Questions')
        # df_Questionnaires = pd.read_excel(xls, 'Questionnaires')
        df_Test_Category = pd.read_excel(xls, 'Test Category')
        # df_Test_Units = pd.read_excel(xls, 'Test Units')
        # df_Units = pd.read_excel(xls, 'Units')
        # df_Conversion = pd.read_excel(xls, 'Conversion')
        # df_Precision = pd.read_excel(xls, 'Precision')
        xls.close()

    ########################################################
    xls = 'Reserved_Words.xlsx'
    RES_WORDS_file = para5 + '\\' + xls
    print("RES_WORDS_file ds",RES_WORDS_file)
    if os.path.exists(RES_WORDS_file):
        print("RES_WORDS_file ds", RES_WORDS_file)
        xls = pd.ExcelFile(RES_WORDS_file)
        df_reservedwords_forms = pd.read_excel(xls, 'Reserved Words')
        xls.close()
    xls = 'Rave_global_metadata.xlsx'
    Rave_global_file = para5 + '\\' + xls
    if os.path.exists(Rave_global_file):
        xls = pd.ExcelFile(Rave_global_file)
        df_global_stage_domain = pd.read_excel(xls, '4.STAGE_DOM_ELT', skiprows=2)
        xls.close()
    DATABASE1 = "Consolidated_ALS_GLOBAL.xlsx"
    DATABASE = para6 + "\\" + DATABASE1
    # df_globalforms = xcelread(DATABASE, 'Forms')
    print("df_globalforms DATABASE",DATABASE)
    if os.path.exists(DATABASE):
        print("df_globalforms ds")
        df_globalforms = xcelread(DATABASE, 'Forms', 'A:O')

        df_globalfields = xcelread(DATABASE, 'Fields','A:AT')
    # df_globalfields = xcelread(DATABASE, 'Fields', 'A,B,G,F,H,I,J,L,O,P,T,U,AC,AP,AQ')
    # print("df_globalfields.columns",df_globalfields.columns)
    # print("df_globalfields.columns",df_globalforms.columns)

        df_globalforms = df_globalforms.loc[(df_globalforms["DraftFormActive"] == True)]
        df_globalfields = df_globalfields.loc[(df_globalfields["DraftFieldActive"] == True)]

        df_globalfields = df_globalfields.merge(df_globalforms.loc[:, ['OID', 'DraftFormName']], left_on=['FormOID'],
                                                right_on=['OID'], how='left', suffixes=['', '_'],
                                                indicator=True)
        df_globalfields = df_globalfields.loc[df_globalfields['_merge'] == 'both']

        mycols = set(df_globalfields.columns)
        mycols.remove('_merge')

        ######## Fields from global ALS
        df_globalfields = df_globalfields[mycols]
        df_globalfields['form_field'] = df_globalfields['FormOID'].str.strip() + df_globalfields[
            'FieldOID'].str.strip()

    if len(df_global_stage_domain) > 0:
        df_global_stage_domain = df_global_stage_domain.iloc[2:]
        df_global_stage_domain2 = df_global_stage_domain.loc[(df_global_stage_domain["Data State"] == 'EDC')]
        df_global_stage_domain2 = df_global_stage_domain2.loc[(df_global_stage_domain2["Element Role"] != 'Identifier')]
        reqfields = ['REQ', 'EXP']
        # reqfields2=['QSTSTLG']
        # nreqfields=['SITEID','STUDYID','SUBJID','DOMAIN']
        nreqfields = ['SITEID', 'STUDYID', 'SUBJID', 'DOMAIN', '_']
        df_global_stage_domain2 = df_global_stage_domain2.loc[
            (df_global_stage_domain2['Element Core'].astype(str).str.contains('|'.join(reqfields)) &
             (
                 (df_global_stage_domain2['Target Data Element'].astype(str).str.contains('|'.join(nreqfields)) == False))) == True]
        df_global_stage_domain2.loc[~(df_global_stage_domain2['Target Data Element'].isna()), 'domain_ele'] = \
            df_global_stage_domain2['Data Domain'].str.strip() + \
            df_global_stage_domain2['Target Data Element'].str.strip()
        df_global_stage_domain2['Target Data Element'] = df_global_stage_domain2['Target Data Element'].str.strip()

    #####################
    ######################

    # path_ssd = SITE_ROOT + '\SSD'
    # path_ssd = PARENT_ROOT + '\SSD'
    # os.chdir(path_ssd)
    lists = para3
    # lists = os.listdir(path_ssd)
    # sub = '.xlsx'
    # files = [mystr for mystr in lists if sub in mystr ]
    files_ssd = lists
    ###################

    if files_ssd :
    # if len(files_ssd) == 1:

        df_ssd_colist = specxlread(files_ssd, 'Collation List', 'Epoch',
                                   ['Epoch', 'Visit Name \n(Folder Name)', 'Folder OID',
                                    'Matrix-Dynamics User requirement for folder(If any)', 'Validation Name'])
        df_ssd_colist.dropna(axis=0, how='all', inplace=True)
        # df_ssd_colist = df_ssd_colist.apply(lambda x: np.nan if isinstance(x, str) and
        #                                                         (x.isspace() or not x) else x)
        df_ssd_colist_df_getcol_ind = specxlread_df_getcol_ind(files_ssd, 'Collation List', 'Epoch',
                                                               ['Epoch', 'Visit Name \n(Folder Name)', 'Folder OID',
                                                                'Matrix-Dynamics User requirement for folder(If any)',
                                                                'Validation Name'])


        df_ssd_frmattr = specxlread(files_ssd, 'Form Attributes', 'Form Name \n(eCRF Description)')
        df_ssd_frmattr.dropna(axis=0, how='all', inplace=True)
        df_ssd_study = specxlread(files_ssd, 'Cover Page', 'Study')

        # df_ssd_frmattr1 = df_ssd_frmattr.apply(lambda x: np.NaN if isinstance(x, str) and
        #                                                         (x.isspace() or not x) else x)
        df_ssd_frmattr_df_getcol_ind = specxlread_df_getcol_ind(files_ssd, 'Form Attributes',
                                                                'Form Name \n(eCRF Description)')
        if 'Unique Form Annotation\nFormOID' in df_ssd_frmattr.columns:
            df_ssd_frmattr['Unique Form Annotation\nFormOID'] = \
                df_ssd_frmattr['Unique Form Annotation\nFormOID'].str.strip()
            df_ssd_frmattr['Unique Form Annotation\nFormOID'] = \
                df_ssd_frmattr['Unique Form Annotation\nFormOID'].str.strip('\n')

        if 'Form Name \n(eCRF Description)' in df_ssd_frmattr.columns:
            df_ssd_frmattr['Form Name \n(eCRF Description)'] = \
                df_ssd_frmattr['Form Name \n(eCRF Description)'].str.strip()

        df_ssd_newfrmattr = specxlread(files_ssd, 'New Form and Field Attributes', 'Request Type')
        df_ssd_newfrmattr.dropna(axis=0, how='all', inplace=True)
        # df_ssd_newfrmattr = df_ssd_newfrmattr1.apply(lambda x: np.nan if isinstance(x, str) and
        #                                                         (x.isspace() or not x) else x)
        df_ssd_newfrmattr_df_getcol_ind = specxlread_df_getcol_ind(files_ssd, 'New Form and Field Attributes',
                                                                   'Request Type')

        df_ssd_esae1 = specxlread(files_ssd, 'eSAE', 'STUDY')
        df_ssd_esae1.dropna(axis=0, how='all', inplace=True)
        # df_ssd_esae1 = df_ssd_esae1_1.apply(lambda x: np.nan if isinstance(x, str) and
        #                                                         (x.isspace() or not x) else x)
        df_ssd_esae1 = df_ssd_esae1.iloc[:2]
        df_ssd_esae1_df_getcol_ind = specxlread_df_getcol_ind(files_ssd, 'eSAE', 'STUDY')

        df_ssd_esae2 = specxlread(files_ssd, 'eSAE', 'STUDY NAME')
        df_ssd_esae2.dropna(axis=0, how='all', inplace=True)
        # df_ssd_esae2 = df_ssd_esae2_1.apply(lambda x: np.nan if isinstance(x, str) and
        #                                                         (x.isspace() or not x) else x)
        df_ssd_esae2 = df_ssd_esae2.iloc[:1]
        df_ssd_esae2_df_getcol_ind = specxlread_df_getcol_ind(files_ssd, 'eSAE', 'STUDY NAME')

        if len(df_ssd_frmattr) > 0:
            formnames_s1 = df_ssd_frmattr.iloc[:, 0].tolist()
            if len(df_ssd_newfrmattr) > 0:
                formnames_s2 = df_ssd_newfrmattr.iloc[:, 1].tolist()
                formnames_s1.extend(formnames_s2)

        if (len(df_ssd_colist) > 0) :
            if (('Validation Name' in df_ssd_colist.columns) & ((df_ssd_colist['Validation Name'].count()) > 0) ):

                df_ssd_colist_reshp1 = df_ssd_colist.copy()
                # df_ssd_colist_reshp.columns = [c.replace(' ', '_') for c in df_ssd_colist_reshp.columns]

                df_ssd_colist_reshp1 = df_ssd_colist_reshp1.rename(columns={"Validation Name": "Validation_Name"})
                if ((df_ssd_colist_reshp1['Validation_Name'].count()) > 0):
                    df_ssd_colist_reshp1 = \
                        (df_ssd_colist_reshp1.set_index(df_ssd_colist_reshp1.columns.drop('Validation_Name', 1).tolist())
                             .Validation_Name.str.split('\n', expand=True)
                             .stack()
                             .reset_index()
                             .rename(columns={0: 'Validation_Name'})
                             .loc[:, df_ssd_colist_reshp1.columns]
                             )

                    df_ssd_colist_reshp1 = \
                        (df_ssd_colist_reshp1.set_index(df_ssd_colist_reshp1.columns.drop('Validation_Name', 1).tolist())
                             .Validation_Name.str.split(',', expand=True)
                             .stack()
                             .reset_index()
                             .rename(columns={0: 'Validation_Name'})
                             .loc[:, df_ssd_colist_reshp1.columns]
                             )

                    df_ssd_colist_reshp1 = \
                        (df_ssd_colist_reshp1.set_index(df_ssd_colist_reshp1.columns.drop('Validation_Name', 1).tolist())
                             .Validation_Name.str.split(':', expand=True)
                             .stack()
                             .reset_index()
                             .rename(columns={0: 'Validation_Name'})
                             .loc[:, df_ssd_colist_reshp1.columns]
                             )

                    df_ssd_colist_reshp1 = \
                        (df_ssd_colist_reshp1.set_index(df_ssd_colist_reshp1.columns.drop('Validation_Name', 1).tolist())
                             .Validation_Name.str.split(';', expand=True)
                             .stack()
                             .reset_index()
                             .rename(columns={0: 'Validation_Name'})
                             .loc[:, df_ssd_colist_reshp1.columns]
                             )


                    df_ssd_colist_reshp1 = \
                        (df_ssd_colist_reshp1.set_index(df_ssd_colist_reshp1.columns.drop('Validation_Name', 1).tolist())
                             .Validation_Name.str.split('-', expand=True)
                             .stack()
                             .reset_index()
                             .rename(columns={0: 'Validation_Name'})
                             .loc[:, df_ssd_colist_reshp1.columns]
                             )

                    df_ssd_colist_reshp1 = \
                        (df_ssd_colist_reshp1.set_index(df_ssd_colist_reshp1.columns.drop('Validation_Name', 1).tolist())
                             .Validation_Name.str.split(' ', expand=True)
                             .stack()
                             .reset_index()
                             .rename(columns={0: 'Validation_Name'})
                             .loc[:, df_ssd_colist_reshp1.columns]
                             )
                    ##
                    df_ssd_colist_reshp1 = df_ssd_colist_reshp1.rename(columns={"Validation_Name": "Validation Name"})

    ######################

    # path_ecs = SITE_ROOT + '\ECS'
    # path_ecs = PARENT_ROOT + '\ECS'
    # os.chdir(path_ecs)
    lists = para4
    # lists = os.listdir(path_ecs)
    # sub = '.xlsm'
    files_ecs = lists
    # files = [mystr for mystr in lists if sub in mystr ]

    ###################

    if (files_ecs):
        df_ecs_study = specxlread(files_ecs, 'Cover Page', 'Study')
        df_ecs_ec = specxlread(files_ecs, 'Edit checks', 'User Parameter Modified(Y/N)')
        df_ecs_ec.dropna(axis=0, how='all', inplace=True)

        col_list_ecs_ec = ["User Parameter Modified(Y/N)", "FormOID", "Form Name", "Edit check category", \
                           "Query Field", "Validation Name", "Check Type", "User Requirement", \
                           "Special Instructions\n(For example exclusion of "
                           "certain folders, partial Dates, Study Arms etc..)", \
                           "Check Action", "Edit Check Message", "Requires Response?", "Requires  Manual Close?", \
                           "Marking Group"]

        [list_ecs_ec_misscols.append(lst) for lst in col_list_ecs_ec if not (lst in df_ecs_ec.columns)]

        if len(df_ecs_ec) == 0:

            df_ecs_ec = pd.DataFrame(columns=col_list_ecs_ec)

        elif len(df_ecs_ec) != 0 and len(list_ecs_ec_misscols) > 0:

            df_ecs_ec1 = pd.DataFrame(columns=list_ecs_ec_misscols)
            df_ecs_ec = df_ecs_ec.append(df_ecs_ec1, ignore_index=True)
            df_ecs_ec.dropna(axis=0, how='all', inplace=True)
            # for lst in list_ecs_ec_misscols:
            #     df_ecs_ec[lst] = ''

        # df_ecs_ec = df_ecs_ec1.apply(lambda x: np.nan if isinstance(x, str) and
        #                                                         (x.isspace() or not x) else x)
        df_ecs_ec_df_getcol_ind = specxlread_df_getcol_ind(files_ecs, 'Edit checks',
                                                           'User Parameter Modified(Y/N)')
        df_ecs_esae = specxlread(files_ecs, 'eSAE tab', 'FormOID')
        df_ecs_esae.dropna(axis=0, how='all', inplace=True)

        col_list_ecs_esae = ["FormOID", "Form Name", "Edit check category", \
                             "Query Field", "Validation Name", "User Requirement",
                             "Check Action", "Edit Check Message", "Requires Response?", "Requires  Manual Close?", \
                             "Marking Group"]

        [list_ecs_esae_misscols.append(lst) for lst in col_list_ecs_esae if not (lst in df_ecs_esae.columns)]

        if len(df_ecs_esae) == 0:

            df_ecs_esae = pd.DataFrame(columns=col_list_ecs_esae)

        elif len(df_ecs_esae) != 0 and len(list_ecs_esae_misscols) > 0:
            df_ecs_esae1 = pd.DataFrame(columns=list_ecs_esae_misscols)
            df_ecs_esae = df_ecs_esae.append(df_ecs_esae1, ignore_index=True)
            df_ecs_esae.dropna(axis=0, how='all', inplace=True)
        # df_ecs_esae = df_ecs_esae1.apply(lambda x: np.nan if isinstance(x, str) and
        #                                                         (x.isspace() or not x) else x)

        df_ecs_esae_df_getcol_ind = specxlread_df_getcol_ind(files_ecs, 'eSAE tab', 'FormOID')
        df_ecs_pd = specxlread(files_ecs, 'PD Specs', 'PD Identification Method')
        df_ecs_pd.dropna(axis=0, how='all', inplace=True)

        col_list_ecs_pd = ["PD Identification Method", "PD Short Description",
                           "Protocol Deviation Term for reporting", "PD Identifier", "PD start date",
                           "Protocol Deviation Coded Term", "Protocol Visit Number/Description",
                           "Requirement of Medical Review (Y/N)", "Blinded (Y/N)", "PD Obsolete",
                            "Query Field", "Validation Name", "User Requirement",
                           "Check Action", "PD Check Message", "Require Manual Close? (Y/N)",
                           "Comments"]

        df_ecs_pd_columns = [x.upper() for x in df_ecs_pd.columns.values.tolist()]

        [list_ecs_pd_misscols.append(lst) for lst in col_list_ecs_pd if not (lst.upper() in df_ecs_pd_columns)]

        if len(df_ecs_pd) == 0:

            df_ecs_pd = pd.DataFrame(columns=col_list_ecs_pd)

        elif len(df_ecs_pd) != 0 and len(list_ecs_pd_misscols) > 0:
            df_ecs_pd1 = pd.DataFrame(columns=list_ecs_pd_misscols)
            df_ecs_pd = df_ecs_pd.append(df_ecs_pd1, ignore_index=True)
            df_ecs_pd.dropna(axis=0, how='all', inplace=True)
        # df_ecs_pd = df_ecs_pd.apply(lambda x: np.nan if isinstance(x, str) and
        #                                                         (x.isspace() or not x) else x)
        df_ecs_pd_df_getcol_ind = specxlread_df_getcol_ind(files_ecs, 'PD Specs', 'PD Identification Method')



    df_ecs_ec_reshp1 = pd.DataFrame()
    df_ecs_ec_reshp2 = pd.DataFrame()
    df_ecs_ec_reshp3 = pd.DataFrame()
    if (len(df_ecs_ec) > 0) & ('Validation_Name' in df_ecs_ec.columns):
        df_ecs_ec_reshp1 = df_ecs_ec.copy()
        df_ecs_ec_reshp2 = df_ecs_ec.copy()
        df_ecs_ec_reshp3 = df_ecs_ec.copy()
        # df_ecs_ec_reshp.columns = [c.replace(' ', '_') for c in df_ecs_ec_reshp.columns]
        df_ecs_ec_reshp1 = df_ecs_ec_reshp1.rename(columns={"Validation Name": "Validation_Name"})
        df_ecs_ec_reshp3 = df_ecs_ec_reshp3.rename(columns={"Form Name": "Form_Name"})
        if ((df_ecs_ec_reshp1['Validation_Name'].count()) > 0):
            df_ecs_ec_reshp1 = \
                (df_ecs_ec_reshp1.set_index(df_ecs_ec_reshp1.columns.drop('Validation_Name', 1).tolist())
                     .Validation_Name.str.split('\n', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'Validation_Name'})
                     .loc[:, df_ecs_ec_reshp1.columns]
                     )

            df_ecs_ec_reshp1 = \
                (df_ecs_ec_reshp1.set_index(df_ecs_ec_reshp1.columns.drop('Validation_Name', 1).tolist())
                     .Validation_Name.str.split(',', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'Validation_Name'})
                     .loc[:, df_ecs_ec_reshp1.columns]
                     )

            # df_ecs_ec_reshp1.loc[df_ecs_ec_reshp1['Validation_Name'].isna(), 'Validation_Name'] =\
            #     df_ecs_ec_reshp1['Validation_Name']

            df_ecs_ec_reshp1 = \
                (df_ecs_ec_reshp1.set_index(df_ecs_ec_reshp1.columns.drop('Validation_Name', 1).tolist())
                     .Validation_Name.str.split(':', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'Validation_Name'})
                     .loc[:, df_ecs_ec_reshp1.columns]
                     )

            df_ecs_ec_reshp1 = \
                (df_ecs_ec_reshp1.set_index(df_ecs_ec_reshp1.columns.drop('Validation_Name', 1).tolist())
                     .Validation_Name.str.split(';', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'Validation_Name'})
                     .loc[:, df_ecs_ec_reshp1.columns]
                     )

            df_ecs_ec_reshp1 = \
                (df_ecs_ec_reshp1.set_index(df_ecs_ec_reshp1.columns.drop('Validation_Name', 1).tolist())
                     .Validation_Name.str.split('-', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'Validation_Name'})
                     .loc[:, df_ecs_ec_reshp1.columns]
                     )

            df_ecs_ec_reshp1 = \
                (df_ecs_ec_reshp1.set_index(df_ecs_ec_reshp1.columns.drop('Validation_Name', 1).tolist())
                     .Validation_Name.str.split(' ', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'Validation_Name'})
                     .loc[:, df_ecs_ec_reshp1.columns]
                     )
    if (len(df_ecs_ec_reshp2) > 0) & ('FormOID' in df_ecs_ec_reshp2.columns):
        if ((df_ecs_ec_reshp2['FormOID'].count()) > 0) :
            ##
            df_ecs_ec_reshp2 = \
                (df_ecs_ec_reshp2.set_index(df_ecs_ec_reshp2.columns.drop('FormOID', 1).tolist())
                     .FormOID.str.split('\n', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'FormOID'})
                     .loc[:, df_ecs_ec_reshp2.columns]
                     )

            df_ecs_ec_reshp2 = \
                (df_ecs_ec_reshp2.set_index(df_ecs_ec_reshp2.columns.drop('FormOID', 1).tolist())
                     .FormOID.str.split(',', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'FormOID'})
                     .loc[:, df_ecs_ec_reshp2.columns]
                     )

            df_ecs_ec_reshp2 = \
                (df_ecs_ec_reshp2.set_index(df_ecs_ec_reshp2.columns.drop('FormOID', 1).tolist())
                     .FormOID.str.split(':', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'FormOID'})
                     .loc[:, df_ecs_ec_reshp2.columns]
                     )
            df_ecs_ec_reshp2 = \
                (df_ecs_ec_reshp2.set_index(df_ecs_ec_reshp2.columns.drop('FormOID', 1).tolist())
                     .FormOID.str.split(';', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'FormOID'})
                     .loc[:, df_ecs_ec_reshp2.columns]
                     )
            df_ecs_ec_reshp2 = \
                (df_ecs_ec_reshp2.set_index(df_ecs_ec_reshp2.columns.drop('FormOID', 1).tolist())
                     .FormOID.str.split('-', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'FormOID'})
                     .loc[:, df_ecs_ec_reshp2.columns]
                     )

            df_ecs_ec_reshp2 = \
                (df_ecs_ec_reshp2.set_index(df_ecs_ec_reshp2.columns.drop('FormOID', 1).tolist())
                     .FormOID.str.split(' ', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'FormOID'})
                     .loc[:, df_ecs_ec_reshp2.columns]
                     )

            df_ecs_ec_reshp2 = df_ecs_ec_reshp2.loc[~df_ecs_ec_reshp2['FormOID'].isna()]
            df_ecs_ec_reshp2 = df_ecs_ec_reshp2.loc[df_ecs_ec_reshp2['FormOID'].astype(str).str.contains('[a-zA-Z0-9]')]
        ##
    if (len(df_ecs_ec_reshp3) > 0) & ('Form_Name' in df_ecs_ec_reshp3.columns):
        if (df_ecs_ec_reshp3['Form_Name'].count() ) > 0 :
            df_ecs_ec_reshp3 = \
                (df_ecs_ec_reshp3.set_index(df_ecs_ec_reshp3.columns.drop('Form_Name', 1).tolist())
                     .Form_Name.str.split('\n', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'Form_Name'})
                     .loc[:, df_ecs_ec_reshp3.columns]
                     )

            df_ecs_ec_reshp3 = \
                (df_ecs_ec_reshp3.set_index(df_ecs_ec_reshp3.columns.drop('Form_Name', 1).tolist())
                     .Form_Name.str.split(',', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'Form_Name'})
                     .loc[:, df_ecs_ec_reshp3.columns]
                     )

            df_ecs_ec_reshp3 = \
                (df_ecs_ec_reshp3.set_index(df_ecs_ec_reshp3.columns.drop('Form_Name', 1).tolist())
                     .Form_Name.str.split(':', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'Form_Name'})
                     .loc[:, df_ecs_ec_reshp3.columns]
                     )
            df_ecs_ec_reshp3 = \
                (df_ecs_ec_reshp3.set_index(df_ecs_ec_reshp3.columns.drop('Form_Name', 1).tolist())
                     .Form_Name.str.split(';', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'Form_Name'})
                     .loc[:, df_ecs_ec_reshp3.columns]
                     )
            df_ecs_ec_reshp3 = \
                (df_ecs_ec_reshp3.set_index(df_ecs_ec_reshp3.columns.drop('Form_Name', 1).tolist())
                     .Form_Name.str.split('-', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'Form_Name'})
                     .loc[:, df_ecs_ec_reshp3.columns]
                     )

            df_ecs_ec_reshp1 = df_ecs_ec_reshp1.rename(columns={"Validation_Name": "Validation Name"})
            df_ecs_ec_reshp3 = df_ecs_ec_reshp3.rename(columns={"Form_Name": "Form Name"})

    ###############
    df_ecs_esae_reshp = pd.DataFrame()
    if (len(df_ecs_esae) > 0) & ("Validation Name" in df_ecs_esae.columns):

        df_ecs_esae_reshp = df_ecs_esae.copy()

        # df_ecs_esae_reshp.columns = [c.replace(' ', '_') for c in df_ecs_esae_reshp.columns]
        df_ecs_esae_reshp = df_ecs_esae_reshp.rename(columns={"Validation Name": "Validation_Name"})
        df_ecs_esae_reshp = df_ecs_esae_reshp.rename(columns={"Form Name": "Form_Name"})
        if ((df_ecs_esae_reshp['Validation_Name'].count()) > 0) :
            df_ecs_esae_reshp = \
                (df_ecs_esae_reshp.set_index(df_ecs_esae_reshp.columns.drop('Validation_Name', 1).tolist())
                     .Validation_Name.str.split('\n', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'Validation_Name'})
                     .loc[:, df_ecs_esae_reshp.columns]
                     )

            df_ecs_esae_reshp = \
                (df_ecs_esae_reshp.set_index(df_ecs_esae_reshp.columns.drop('Validation_Name', 1).tolist())
                     .Validation_Name.str.split(',', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'Validation_Name'})
                     .loc[:, df_ecs_esae_reshp.columns]
                     )

            df_ecs_esae_reshp = \
                (df_ecs_esae_reshp.set_index(df_ecs_esae_reshp.columns.drop('Validation_Name', 1).tolist())
                     .Validation_Name.str.split(':', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'Validation_Name'})
                     .loc[:, df_ecs_esae_reshp.columns]
                     )
            df_ecs_esae_reshp = \
                (df_ecs_esae_reshp.set_index(df_ecs_esae_reshp.columns.drop('Validation_Name', 1).tolist())
                     .Validation_Name.str.split(';', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'Validation_Name'})
                     .loc[:, df_ecs_esae_reshp.columns]
                     )
            df_ecs_esae_reshp = \
                (df_ecs_esae_reshp.set_index(df_ecs_esae_reshp.columns.drop('Validation_Name', 1).tolist())
                     .Validation_Name.str.split('-', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'Validation_Name'})
                     .loc[:, df_ecs_esae_reshp.columns]
                     )
            df_ecs_esae_reshp = \
                (df_ecs_esae_reshp.set_index(df_ecs_esae_reshp.columns.drop('Validation_Name', 1).tolist())
                     .Validation_Name.str.split(' ', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'Validation_Name'})
                     .loc[:, df_ecs_esae_reshp.columns]
                     )
        if ((df_ecs_esae_reshp['FormOID'].count()) > 0):
            ##
            df_ecs_esae_reshp = \
                (df_ecs_esae_reshp.set_index(df_ecs_esae_reshp.columns.drop('FormOID', 1).tolist())
                     .FormOID.str.split('\n', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'FormOID'})
                     .loc[:, df_ecs_esae_reshp.columns]
                     )

            df_ecs_esae_reshp = \
                (df_ecs_esae_reshp.set_index(df_ecs_esae_reshp.columns.drop('FormOID', 1).tolist())
                     .FormOID.str.split(',', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'FormOID'})
                     .loc[:, df_ecs_esae_reshp.columns]
                     )

            df_ecs_esae_reshp = \
                (df_ecs_esae_reshp.set_index(df_ecs_esae_reshp.columns.drop('FormOID', 1).tolist())
                     .FormOID.str.split(':', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'FormOID'})
                     .loc[:, df_ecs_esae_reshp.columns]
                     )
            df_ecs_esae_reshp = \
                (df_ecs_esae_reshp.set_index(df_ecs_esae_reshp.columns.drop('FormOID', 1).tolist())
                     .FormOID.str.split(';', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'FormOID'})
                     .loc[:, df_ecs_esae_reshp.columns]
                     )
            df_ecs_esae_reshp = \
                (df_ecs_esae_reshp.set_index(df_ecs_esae_reshp.columns.drop('FormOID', 1).tolist())
                     .FormOID.str.split('-', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'FormOID'})
                     .loc[:, df_ecs_esae_reshp.columns]
                     )
            df_ecs_esae_reshp = \
                (df_ecs_esae_reshp.set_index(df_ecs_esae_reshp.columns.drop('FormOID', 1).tolist())
                     .FormOID.str.split(' ', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'FormOID'})
                     .loc[:, df_ecs_esae_reshp.columns]
                     )

            ##
        if ((df_ecs_esae_reshp['Form_Name'].count()) > 0):
            df_ecs_esae_reshp = \
                (df_ecs_esae_reshp.set_index(df_ecs_esae_reshp.columns.drop('Form_Name', 1).tolist())
                     .Form_Name.str.split('\n', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'Form_Name'})
                     .loc[:, df_ecs_esae_reshp.columns]
                     )

            df_ecs_esae_reshp = \
                (df_ecs_esae_reshp.set_index(df_ecs_esae_reshp.columns.drop('Form_Name', 1).tolist())
                     .Form_Name.str.split(',', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'Form_Name'})
                     .loc[:, df_ecs_esae_reshp.columns]
                     )

            df_ecs_esae_reshp = \
                (df_ecs_esae_reshp.set_index(df_ecs_esae_reshp.columns.drop('Form_Name', 1).tolist())
                     .Form_Name.str.split(':', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'Form_Name'})
                     .loc[:, df_ecs_esae_reshp.columns]
                     )
            df_ecs_esae_reshp = \
                (df_ecs_esae_reshp.set_index(df_ecs_esae_reshp.columns.drop('Form_Name', 1).tolist())
                     .Form_Name.str.split(';', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'Form_Name'})
                     .loc[:, df_ecs_esae_reshp.columns]
                     )
            df_ecs_esae_reshp = \
                (df_ecs_esae_reshp.set_index(df_ecs_esae_reshp.columns.drop('Form_Name', 1).tolist())
                     .Form_Name.str.split('-', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'Form_Name'})
                     .loc[:, df_ecs_esae_reshp.columns]
                     )

            df_ecs_esae_reshp = df_ecs_esae_reshp.rename(columns={"Validation_Name": "Validation Name"})
            df_ecs_esae_reshp = df_ecs_esae_reshp.rename(columns={"Form_Name": "Form Name"})
    #####
    if ((len(df_ecs_pd) > 0) & ("Validation Name" in df_ecs_pd.columns)):
        df_ecs_pd_reshp = df_ecs_pd.copy()

        # df_ecs_pd_reshp.columns = [c.replace(' ', '_') for c in df_ecs_pd_reshp.columns]
        df_ecs_pd_reshp = df_ecs_pd_reshp.rename(columns={"Validation Name": "Validation_Name"})
        df_ecs_pd_reshp = df_ecs_pd_reshp.rename(columns={"Primary Form Name": "Primary_Form_Name"})

        if ((len(df_ecs_pd_reshp) > 0) & ("Validation Name" in df_ecs_pd_reshp.columns)):
            if ((df_ecs_pd_reshp['Validation_Name'].count()) > 0):
                df_ecs_pd_reshp = \
                    (df_ecs_pd_reshp.set_index(df_ecs_pd_reshp.columns.drop('Validation_Name', 1).tolist())
                         .Validation_Name.astype(str).str.split('\n', expand=True)
                         .stack()
                         .reset_index()
                         .rename(columns={0: 'Validation_Name'})
                         .loc[:, df_ecs_pd_reshp.columns]
                         )

                df_ecs_pd_reshp = \
                    (df_ecs_pd_reshp.set_index(df_ecs_pd_reshp.columns.drop('Validation_Name', 1).tolist())
                         .Validation_Name.astype(str).str.split(',', expand=True)
                         .stack()
                         .reset_index()
                         .rename(columns={0: 'Validation_Name'})
                         .loc[:, df_ecs_pd_reshp.columns]
                         )

                df_ecs_pd_reshp = \
                    (df_ecs_pd_reshp.set_index(df_ecs_pd_reshp.columns.drop('Validation_Name', 1).tolist())
                         .Validation_Name.astype(str).str.split(':', expand=True)
                         .stack()
                         .reset_index()
                         .rename(columns={0: 'Validation_Name'})
                         .loc[:, df_ecs_pd_reshp.columns]
                         )


                df_ecs_pd_reshp = \
                    (df_ecs_pd_reshp.set_index(df_ecs_pd_reshp.columns.drop('Validation_Name', 1).tolist())
                         .Validation_Name.astype(str).str.split(';', expand=True)
                         .stack()
                         .reset_index()
                         .rename(columns={0: 'Validation_Name'})
                         .loc[:, df_ecs_pd_reshp.columns]
                         )


                df_ecs_pd_reshp = \
                    (df_ecs_pd_reshp.set_index(df_ecs_pd_reshp.columns.drop('Validation_Name', 1).tolist())
                         .Validation_Name.astype(str).str.split('-', expand=True)
                         .stack()
                         .reset_index()
                         .rename(columns={0: 'Validation_Name'})
                         .loc[:, df_ecs_pd_reshp.columns]
                         )

                df_ecs_pd_reshp = \
                    (df_ecs_pd_reshp.set_index(df_ecs_pd_reshp.columns.drop('Validation_Name', 1).tolist())
                         .Validation_Name.astype(str).str.split(' ', expand=True)
                         .stack()
                         .reset_index()
                         .rename(columns={0: 'Validation_Name'})
                         .loc[:, df_ecs_pd_reshp.columns]
                         )
        if ((len(df_ecs_pd_reshp) > 0) & ("Form" in df_ecs_pd_reshp.columns)):

            if ((df_ecs_pd_reshp['Form'].count()) > 0):
                ##
                df_ecs_pd_reshp = \
                    (df_ecs_pd_reshp.set_index(df_ecs_pd_reshp.columns.drop('Form', 1).tolist())
                         .Form.astype(str).str.split('\n', expand=True)
                         .stack()
                         .reset_index()
                         .rename(columns={0: 'Form'})
                         .loc[:, df_ecs_pd_reshp.columns]
                         )

                df_ecs_pd_reshp = \
                    (df_ecs_pd_reshp.set_index(df_ecs_pd_reshp.columns.drop('Form', 1).tolist())
                         .Form.astype(str).str.split(',', expand=True)
                         .stack()
                         .reset_index()
                         .rename(columns={0: 'Form'})
                         .loc[:, df_ecs_pd_reshp.columns]
                         )

                df_ecs_pd_reshp = \
                    (df_ecs_pd_reshp.set_index(df_ecs_pd_reshp.columns.drop('Form', 1).tolist())
                         .Form.astype(str).str.split(':', expand=True)
                         .stack()
                         .reset_index()
                         .rename(columns={0: 'Form'})
                         .loc[:, df_ecs_pd_reshp.columns]
                         )

                df_ecs_pd_reshp = \
                    (df_ecs_pd_reshp.set_index(df_ecs_pd_reshp.columns.drop('Form', 1).tolist())
                         .Form.astype(str).str.split(';', expand=True)
                         .stack()
                         .reset_index()
                         .rename(columns={0: 'Form'})
                         .loc[:, df_ecs_pd_reshp.columns]
                         )

                df_ecs_pd_reshp = \
                    (df_ecs_pd_reshp.set_index(df_ecs_pd_reshp.columns.drop('Form', 1).tolist())
                         .Form.astype(str).str.split('-', expand=True)
                         .stack()
                         .reset_index()
                         .rename(columns={0: 'Form'})
                         .loc[:, df_ecs_pd_reshp.columns]
                         )

                df_ecs_pd_reshp = \
                    (df_ecs_pd_reshp.set_index(df_ecs_pd_reshp.columns.drop('Form', 1).tolist())
                         .Form.astype(str).str.split(' ', expand=True)
                         .stack()
                         .reset_index()
                         .rename(columns={0: 'Form'})
                         .loc[:, df_ecs_pd_reshp.columns]
                         )
                ##
        if ((len(df_ecs_pd_reshp) > 0) & ("Primary_Form_Name" in df_ecs_pd_reshp.columns)):
            if ((df_ecs_pd_reshp['Primary_Form_Name'].count()) > 0):
                df_ecs_pd_reshp = \
                    (df_ecs_pd_reshp.set_index(df_ecs_pd_reshp.columns.drop('Primary_Form_Name', 1).tolist())
                         .Primary_Form_Name.astype(str).str.split('\n', expand=True)
                         .stack()
                         .reset_index()
                         .rename(columns={0: 'Primary_Form_Name'})
                         .loc[:, df_ecs_pd_reshp.columns]
                         )

                df_ecs_pd_reshp = \
                    (df_ecs_pd_reshp.set_index(df_ecs_pd_reshp.columns.drop('Primary_Form_Name', 1).tolist())
                         .Primary_Form_Name.astype(str).str.split(',', expand=True)
                         .stack()
                         .reset_index()
                         .rename(columns={0: 'Primary_Form_Name'})
                         .loc[:, df_ecs_pd_reshp.columns]
                         )

                df_ecs_pd_reshp = \
                    (df_ecs_pd_reshp.set_index(df_ecs_pd_reshp.columns.drop('Primary_Form_Name', 1).tolist())
                         .Primary_Form_Name.astype(str).str.split(';', expand=True)
                         .stack()
                         .reset_index()
                         .rename(columns={0: 'Primary_Form_Name'})
                         .loc[:, df_ecs_pd_reshp.columns]
                         )

                df_ecs_pd_reshp = \
                    (df_ecs_pd_reshp.set_index(df_ecs_pd_reshp.columns.drop('Primary_Form_Name', 1).tolist())
                         .Primary_Form_Name.astype(str).str.split('-', expand=True)
                         .stack()
                         .reset_index()
                         .rename(columns={0: 'Primary_Form_Name'})
                         .loc[:, df_ecs_pd_reshp.columns]
                         )

                df_ecs_pd_reshp = \
                    (df_ecs_pd_reshp.set_index(df_ecs_pd_reshp.columns.drop('Primary_Form_Name', 1).tolist())
                         .Primary_Form_Name.astype(str).str.split(':', expand=True)
                         .stack()
                         .reset_index()
                         .rename(columns={0: 'Primary_Form_Name'})
                         .loc[:, df_ecs_pd_reshp.columns]
                         )

                df_ecs_pd_reshp = \
                    (df_ecs_pd_reshp.set_index(df_ecs_pd_reshp.columns.drop('Primary_Form_Name', 1).tolist())
                         .Primary_Form_Name.astype(str).str.split(';', expand=True)
                         .stack()
                         .reset_index()
                         .rename(columns={0: 'Primary_Form_Name'})
                         .loc[:, df_ecs_pd_reshp.columns]
                         )

                df_ecs_pd_reshp = \
                    (df_ecs_pd_reshp.set_index(df_ecs_pd_reshp.columns.drop('Primary_Form_Name', 1).tolist())
                         .Primary_Form_Name.astype(str).str.split('-', expand=True)
                         .stack()
                         .reset_index()
                         .rename(columns={0: 'Primary_Form_Name'})
                         .loc[:, df_ecs_pd_reshp.columns]
                         )

                df_ecs_pd_reshp = df_ecs_pd_reshp.rename(columns={"Validation_Name": "Validation Name"})
                df_ecs_pd_reshp = df_ecs_pd_reshp.rename(columns={"Primary_Form_Name": "Primary Form Name"})
        #####################

    path_codelist = 'C:\\Bhasp\\NVTSonco-work\\NVTSonco-work\\RADA\\Radaprogram\\Virtual Environments\\03Nov2021' + '\\codelist'
    # path_codelist = SITE_ROOT + '\\codelist'
    os.chdir(path_codelist)

    lists = os.listdir(path_codelist)

    sub = '.xlsx'
    files_codelist = [mystr for mystr in lists if sub in mystr]

    path_codelist_up = 'C:\\Bhasp\\NVTSonco-work\\NVTSonco-work\\RADA\\Radaprogram\\Virtual Environments\\03Nov2021' +  '\\codelistup'


    lists = os.listdir(path_codelist_up)

    sub = '.csv'
    files_codelist_up = [mystr for mystr in lists if sub in mystr]

    # initiation of dataframes
    # print(path_codelist)
    # print(files_codelist)

    if ((len(files_codelist) == 1)):
        # print(files_codelist)
        xls = pd.ExcelFile(files_codelist[0])
        if 'Sheet1' in xls.sheet_names:
            df_codelist_StageDomain = pd.read_excel(xls, 'Sheet1', nrows=1)
            # print("df_codelist_StageDomain", df_codelist_StageDomain.columns)
        if 'Controlled Terminology' in df_codelist_StageDomain.columns:
            df_codelist_StageDomain = pd.read_excel(xls, 'Sheet1')
        else:
            df_codelist_StageDomain = pd.read_excel(xls, 'Sheet1',skiprows=1)
            df_codelist_StageDomain = df_codelist_StageDomain.loc[:, ~df_codelist_StageDomain.columns.str.contains('^Unnamed')]
            # # df_crfdraft.replace("", nan_value, inplace=True)
            # # df_crfdraft = df_crfdraft.dropna(how='all')
            # df_codelist_StageDomain.drop_duplicates(keep=False, inplace=True)
        # print("df_codelist_StageDomain columns",df_codelist_StageDomain.columns)
        # print("df_codelist_StageDomain",len(df_codelist_StageDomain))
        # print("df_codelist_StageDomain",(df_codelist_StageDomain))
        # print("df_codelist_StageDomain",(df_codelist_StageDomain))
        if 'Controlled Terminology' in df_codelist_StageDomain.columns:
            df_codelist_StageDomain = df_codelist_StageDomain.filter(items=['Controlled Terminology', 'Sdtm', 'Long Codelist'])


    elif (len(files_codelist_up) == 1):
        os.chdir(path_codelist_up)

        print(files_codelist_up)
        # xls = pd.ExcelFile(files_codelist_up[0])
        # if 'Sheet1' in xls.sheet_names:
        #     df_codelist_StageDomain = pd.read_excel(xls, 'Sheet1', nrows=1)

        df_codelist_StageDomain = pd.read_csv(files_codelist_up[0])

        #     print("df_codelist_StageDomain", df_codelist_StageDomain.columns)
        # if 'Controlled Terminology' in df_codelist_StageDomain.columns:
        #     df_codelist_StageDomain = pd.read_excel(xls, 'Sheet1')
        # else:
        #     df_codelist_StageDomain = pd.read_excel(xls, 'Sheet1', skiprows=1)
        #     df_codelist_StageDomain = df_codelist_StageDomain.loc[:,
        #                               ~df_codelist_StageDomain.columns.str.contains('^Unnamed')]
        df_codelist_StageDomain = df_codelist_StageDomain.loc[:,
                                      ~df_codelist_StageDomain.columns.str.contains('^Unnamed')]

        # print("df_codelist_StageDomain columns", df_codelist_StageDomain.columns)
        # print("df_codelist_StageDomain", len(df_codelist_StageDomain))
        # print("df_codelist_StageDomain", (df_codelist_StageDomain))
        # print("df_codelist_StageDomain", (df_codelist_StageDomain))
        if 'Controlled Terminology' in df_codelist_StageDomain.columns:
            df_codelist_StageDomain = df_codelist_StageDomain.filter(items=['Controlled Terminology', 'Sdtm',
                                                                            'Long Codelist'])

            #####################

    # path_als = SITE_ROOT + '\\ALS'
    # # path_als = PARENT_ROOT + '\ALS'
    # os.chdir(path_als)

    # lists = os.listdir(path_als)

    lists = para2

    sub = '.xlsx'
    files_als = lists
    # files_als = [mystr for mystr in lists if sub in mystr]


    # initiation of dataframes


    if (((files_als) )):
        xls = pd.ExcelFile(files_als)
        if 'CRFDraft' in xls.sheet_names:
            df_crfdraft = pd.read_excel(xls, 'CRFDraft').head(1)
            # df_crfdraft = pd.read_excel(files_als, 'CRFDraft').head(1)

            df_crfdraft = df_crfdraft.loc[:, ~df_crfdraft.columns.str.contains('^Unnamed')]
            # df_crfdraft.replace("", nan_value, inplace=True)
            # df_crfdraft = df_crfdraft.dropna(how='all')
            df_crfdraft.drop_duplicates(keep='last', inplace=True)
        if 'LabVariableMappings' in xls.sheet_names:
            df_lbsettings1 = pd.read_excel(files_als, 'LabVariableMappings').head(2)
            df_lbsettings1 = df_lbsettings1.loc[:, ~df_lbsettings1.columns.str.contains('^Unnamed')]
            # df_lbsettings1.replace("", nan_value, inplace=True)
            # df_lbsettings1 = df_lbsettings1.dropna(how='all')
            df_lbsettings1.drop_duplicates(keep='last', inplace=True)
        if 'Matrices' in xls.sheet_names:
            df_Matrices = pd.read_excel(files_als, 'Matrices')
        if 'Forms' in xls.sheet_names:
            df_als_forms = xcelread(files_als, 'Forms', 'A:O')


            df_forms = df_als_forms.loc[(df_als_forms["DraftFormActive"] == True) |
                                        (df_als_forms["DraftFormActive"] == 'TRUE')]
            # check#126

            df_forms['OID_split'] = ""
            # df_forms.loc[df_forms['OID'].str.contains('[0-9]$'), 'OID_split'] = \
            #     df_forms['OID'].str.rsplit('_', 1).str.get(0)

            df_forms["OID_split"] = df_forms['OID'].astype(str).apply(
                lambda x: x.rsplit('_')[0] if x[-1].isdigit() else x)
            # check 127
            df_forms_global_esae_no =  df_forms.loc[(~(df_forms['OID_split'].isin(df_globalforms['OID'])) &
                                                     ~(df_forms['OID'].isin(df_globalforms['OID'])) )
                                                    & (~(df_forms['OID'].isin(df_eSAE_forms['FormOID'])))]
            #
            df_forms_global_yes_esae_no =  df_forms.loc[((df_forms['OID_split'].isin(df_globalforms['OID'])) |
                                                     (df_forms['OID'].isin(df_globalforms['OID'])) )
                                                    & (~(df_forms['OID'].isin(df_eSAE_forms['FormOID'])))]

            # df_forms = df_als_forms.loc[(df_als_forms["DraftFormActive"] == True ) |
            #                             (df_als_forms["DraftFormActive"] == 'TRUE' )]
            df_forms_inact = df_als_forms.loc[(df_als_forms["DraftFormActive"] == False) |
                                              (df_als_forms["DraftFormActive"] == 'FALSE')]

        if 'Folders' in xls.sheet_names:
            df_als_folders = xcelread(files_als, 'Folders', 'A:K')
        if 'Fields' in xls.sheet_names:
            df_als_flds = xcelread(files_als, 'Fields', 'A:AS')
            df_als_flds.dropna(axis=0, how='all', inplace=True)

            df_flds = df_als_flds.loc[(df_als_flds["DraftFieldActive"] == True) |
                                      (df_als_flds["DraftFieldActive"] == 'TRUE')]
            # df_flds = df_als_flds.loc[(df_als_flds["DraftFieldActive"] == True) |
            #                           (df_als_flds["DraftFieldActive"] == 'TRUE')]
            df_flds_inact = df_als_flds.loc[(df_als_flds["DraftFieldActive"] == False) |
                                            (df_als_flds["DraftFieldActive"] == 'FALSE')]
            if 'FormOID' in df_flds.columns:
                df_flds = df_flds.loc[df_flds['FormOID'].isin(df_forms['OID'])]


                df_flds = df_flds.merge(df_forms.loc[:, ['OID', 'DraftFormName']], left_on=['FormOID'],
                                    right_on=['OID'], how='left', suffixes=['', '_'],
                                    indicator=True)

                df_flds.drop_duplicates(subset=['FormOID', 'FieldOID'], keep='first', inplace=True)

                df_flds = df_flds.loc[df_flds['_merge'] == 'both']


                mycols = set(df_flds.columns)
                mycols.remove('_merge')



            # df_flds = df_flds.apply(lambda x: np.nan if isinstance(x, str) and
            #                                                         (x.isspace() or not x) else x)
            ####### Fields from ALS
                df_flds = df_flds[mycols]
        # exclude eSAE forms
        if (len(df_flds) > 0) & (len(df_eSAE_forms) > 0) & (len(df_non_domain_forms) > 0):
            df_flds_non_esae = df_flds.loc[(~(df_flds['FormOID'].isin(df_eSAE_forms['FormOID'])))]
            df_flds_non_ops_esae = df_flds.loc[(~(df_flds['FormOID'].isin(df_non_domain_forms['FormOID'])))]
            df_flds_nonglob = df_flds.loc[~(df_flds['FieldOID'].isna())]
            # df_flds_nonglob.loc[df_flds_nonglob['FormOID'].str.contains('[0-9]$'), 'FormOID_split'] = \
            #     df_flds_nonglob['FormOID'].str.rsplit('_', 1).str.get(0)
            #fix 09aug2021
            df_flds_nonglob['FormOID_split'] = ""

            df_flds_nonglob["FormOID_split"] = df_flds_nonglob['FormOID'].astype(str).apply(
                lambda x: x.rsplit('_')[0] if x[-1].isdigit() else x)

            df_flds_nonglob['DataDictionaryName_split'] = ""

            df_flds_nonglob["DataDictionaryName_split"] = df_flds_nonglob['DataDictionaryName'].astype(str).apply(
                lambda x: x.rsplit('_')[0] if x[-1].isdigit() else x)




            df_flds_nonglob['form_field'] = df_flds_nonglob['FormOID_split'].str.strip() + \
                                            df_flds_nonglob['FieldOID'].str.strip()

            df_flds_nonglob = df_flds_nonglob.loc[
                (~(df_flds_nonglob['FormOID'].isin(df_non_domain_forms['FormOID'])))]

            # check 127
            df_flds_glob_formyes_fieldno = df_flds_nonglob.loc[
                ~(df_flds_nonglob['form_field'].isin(df_globalfields['form_field'])) &
                ((df_flds_nonglob['FormOID'].isin(df_forms_global_yes_esae_no['OID'])))]
            df_flds_glob_formyes_fieldno = df_flds_glob_formyes_fieldno.loc[
                ~df_flds_glob_formyes_fieldno['VariableOID'].isna()]


            df_globalfields_attributes = df_globalfields.filter(items=['form_field','DataFormat', 'PreText', 'DataDictionaryName'])
            df_globalfields_attributes['DataDictionaryName_split_global'] = ""

            df_globalfields_attributes["DataDictionaryName_split_global"] = df_globalfields_attributes['DataDictionaryName'].astype(str).apply(
                lambda x: x.rsplit('_')[0] if x[-1].isdigit() else x)
            df_globalfields_attributes = df_globalfields_attributes.rename(columns={'DataFormat':'DataFormat_global', 'PreText':'PreText_global',
                                                       'DataDictionaryName' :'DataDictionaryName_global'})
            df_flds_glob_formyes_fieldyes = df_flds_nonglob.merge(df_globalfields_attributes, left_on=['form_field'],
                                        right_on=['form_field'], how='inner',
                                        suffixes=['', '_'],
                                        indicator=True)
            mycols = set(df_flds_glob_formyes_fieldyes.columns)
            mycols.remove('_merge')
            df_flds_glob_formyes_fieldyes = df_flds_glob_formyes_fieldyes[mycols]

            # df_flds_glob_formyes_fieldyes.to_excel(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS_\df_flds_glob_formyes_fieldyes_m.xlsx')

            # check 128

            df_flds_glob_formyes_fieldyes = df_flds_glob_formyes_fieldyes.loc[
                (~(df_flds_glob_formyes_fieldyes['DataFormat'] == (df_flds_glob_formyes_fieldyes['DataFormat_global']))
                 | (~(df_flds_glob_formyes_fieldyes['PreText'] == (df_flds_glob_formyes_fieldyes['PreText_global']))))
                    | ( ~(df_flds_glob_formyes_fieldyes['DataDictionaryName'] ==
                          (df_flds_glob_formyes_fieldyes['DataDictionaryName_global'])))]
            df_flds_glob_formyes_fieldyes = df_flds_glob_formyes_fieldyes.loc[
                ~df_flds_glob_formyes_fieldyes['VariableOID'].isna()]

            # df_flds_glob_formyes_fieldyes.to_excel(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS_\outputdf_flds_glob_formyes_fieldyes.xlsx')

            # df_globalfields.to_excel(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS_\outputdf_globalfields.xlsx')
            # df_flds_glob_formyes_fieldyes = df_flds_glob_formyes_fieldyes.loc[
            #     ~(df_flds_glob_formyes_fieldyes['DataFormat'] == (df_flds_glob_formyes_fieldyes['DataFormat_global'])) |
            #     ~(df_flds_glob_formyes_fieldyes['PreText'] == (df_flds_glob_formyes_fieldyes['PreText_global'])) | (
            #     ~(df_flds_glob_formyes_fieldyes['DataDictionaryName_split'] == (df_flds_glob_formyes_fieldyes['DataDictionaryName_split_global']))]

            # df_flds_glob_formyes_fieldyes = df_flds_nonglob.loc[
            #     (df_flds_nonglob['form_field'].isin(df_globalfields['form_field'])) &
            #     ((df_flds_nonglob['FormOID'].isin(df_forms_global_yes_esae_no['OID'])))]


            # df_flds_nonglob = df_flds_nonglob.loc[~(df_flds_nonglob['FormOID'].str.contains('|'.join(df_globalforms['OID'])))]

            df_flds_glob = df_flds.loc[((df_flds['FormOID'].isin(df_globalforms['OID'])) &
                                        ((df_flds['FormOID'].isin(df_non_domain_forms['FormOID']))))]

            df_flds_glob = df_flds_nonglob.loc[
                (df_flds_nonglob['FormOID'].astype(str).str.contains('|'.join(df_globalforms['OID'])))]

            df_flds_defaults_split = df_flds.loc[(~(df_flds['DataDictionaryName'].isna()) &
                                                  (~(df_flds['DefaultValue'].isna())) & (
                                                      ~(df_flds['DefaultValue'] == '|')))]

            if len(df_flds_defaults_split) > 0:
                df_flds_defaults_split = \
                    (df_flds_defaults_split.set_index(
                        df_flds_defaults_split.columns.drop('DefaultValue', 1).tolist())
                         .DefaultValue.str.split('|', expand=True)
                         .stack()
                         .reset_index()
                         .rename(columns={0: 'DefaultValue'})
                         .loc[:, df_flds_defaults_split.columns]
                         )

                df_flds_defaults_split = df_flds_defaults_split.loc[
                    ((~(df_flds_defaults_split['DefaultValue'].isna())) &
                     ((df_flds_defaults_split['DefaultValue'] != '')))]
                df_flds_defaults_split['dictname_CodedData'] = df_flds_defaults_split['DataDictionaryName'] + '-' + \
                                                               df_flds_defaults_split['DefaultValue']

            df_flds_numeric = df_flds.loc[df_flds['ControlType'] == 'Text']

            df_flds_numeric = (df_flds_numeric
                               .drop(['DataFormat'], axis=1)
                               .join(df_flds_numeric[['DataFormat']].apply(pd.to_numeric, errors='coerce')))

            df_flds_numeric = df_flds_numeric[df_flds_numeric[['DataFormat']].notnull().all(axis=1)]
            df_forms_nonglob = pd.DataFrame()
            df_forms_glob = pd.DataFrame()
            df_forms_non_ops_esae_pd = pd.DataFrame()
            df_forms_non_ops_esae_pd_glob = pd.DataFrame()

            df_forms_non_ops_esae = df_forms.loc[(~(df_forms['OID'].isin(df_non_domain_forms['FormOID'])))]
            df_forms_non_ops_esae_pd = df_forms.loc[(~(df_forms['OID'].isin(df_non_domain_forms['FormOID']))) &
                                                    (~(df_forms['OID'].astype(str).str.contains('|'.join(['DVG001', 'DVG002']))))]
            df_forms_non_pd = df_forms.loc[(~(df_forms['OID'].astype(str).str.contains('|'.join(['DVG001', 'DVG002']))))]

            df_forms_non_glob = df_forms.loc[(~(df_forms['OID'].isin(df_globalforms['OID'])))]
            df_forms_non_glob = df_forms_non_glob.loc[~(df_forms_non_glob['OID'].
                                                        str.contains('|'.join(df_globalforms['OID'])))]

            df_forms_non_ops_esae_pd_glob = df_forms_non_glob.loc[
                (~(df_forms_non_glob['OID'].astype(str).str.contains('|'.join(df_globalforms['OID'])))) &
                (~(df_forms['OID'].isin(df_non_domain_forms['FormOID'])))]

            df_forms_non_ops_esae_pd.dropna(subset=['OID'], how='all', inplace=True)

            df_forms_glob1 = df_forms.loc[(df_forms['OID'].isin(df_globalforms['OID']) &
                                           (df_forms['OID'].isin(df_non_domain_forms['FormOID'])))]

            df_forms_glob2 = df_forms.loc[df_forms['OID'].astype(str).str.contains('|'.join(df_globalforms['OID']))]

            df_forms_glob = df_forms_glob.append(df_forms_glob1)
            df_forms_glob = df_forms_glob.append(df_forms_glob2)

            df_forms_glob.dropna(subset=['OID'], how='all', inplace=True)

        df_lbsettings = xcelread(files_als, 'LabVariableMappings', 'A:E')
        df_lbsettings.dropna(axis=0, how='all', inplace=True)
        df_lbsettings.dropna()
        df_lbsettings.drop_duplicates()
        if 'DataDictionaries' in xls.sheet_names:
            df_dictnams = xcelread(files_als, 'DataDictionaries', 'A')
            df_dictnams.dropna(axis=0, how='all', inplace=True)
            df_dictnams = df_dictnams.loc[:, ~df_dictnams.columns.str.contains('^Unnamed')]
            df_dictnams = df_dictnams.loc[~(df_dictnams['DataDictionaryName'].isna())]
            df_dictnams['DataDictionaryName_split'] = ""
            df_dictnams["DataDictionaryName_split"] = df_dictnams['DataDictionaryName'].astype(str).apply(
                lambda x: x.rsplit('_')[0] if x[-1].isdigit() else x)

        if 'DataDictionaryEntries' in xls.sheet_names:
            df_dicts = xcelread(files_als, 'DataDictionaryEntries', 'A:E')
            df_dicts.dropna(axis=0, how='all', inplace=True)
            df_dicts = df_dicts.loc[:, ~df_dicts.columns.str.contains('^Unnamed')]
            df_dicts_defaults_dict = df_dicts.loc[~(df_dicts['CodedData'].isna())]
            df_dicts_defaults_dict.loc[
                ~(df_dicts_defaults_dict['DataDictionaryName'].isna()), 'dictname_CodedData'] = \
                df_dicts_defaults_dict['DataDictionaryName'] + '-' + df_dicts_defaults_dict['CodedData']

            df_dicts['DataDictionaryName_split'] = ""
            df_dicts["DataDictionaryName_split"] = df_dicts['DataDictionaryName'].astype(str).apply(
                lambda x: x.rsplit('_')[0] if x[-1].isdigit() else x)

            df_dicts = df_dicts.loc[~(df_dicts['CodedData'].isna())]
            df_dicts.loc[
                ~(df_dicts['DataDictionaryName'].isna()), 'dictname_CodedData'] = \
                df_dicts['DataDictionaryName_split'] + '-' + df_dicts['CodedData']
            df_dicts = df_dicts.loc[~(df_dicts['UserDataString'].isna())]
            df_dicts.loc[
                ~(df_dicts['DataDictionaryName'].isna()), 'dictname_UserDataString'] = \
                df_dicts['DataDictionaryName_split'] + '-' + df_dicts['UserDataString'].astype(str)


            df_codelist_StageDomain = df_codelist_StageDomain.loc[
                (~(df_codelist_StageDomain['Controlled Terminology'].isna())) &
                                                                  (~(df_codelist_StageDomain['Sdtm'].isna()))&
                                                                  (~(df_codelist_StageDomain['Long Codelist'].isna()))]
            df_codelist_StageDomain['dictname_CodedData']=''
            df_codelist_StageDomain['dictname_UserDataString']=''

            df_codelist_StageDomain.loc[
                ~(df_codelist_StageDomain['Controlled Terminology'].isna()), 'dictname_CodedData'] = \
                df_codelist_StageDomain['Controlled Terminology'] + '-' + df_codelist_StageDomain['Sdtm'].astype(str)
            # df_codelist_StageDomain = df_codelist_StageDomain.loc[~(df_codelist_StageDomain['Long Codelist'].isna())]
            df_codelist_StageDomain.loc[
                ~(df_codelist_StageDomain['Controlled Terminology'].isna()), 'dictname_UserDataString'] = \
                df_codelist_StageDomain['Controlled Terminology'] + '-' +\
                df_codelist_StageDomain['Long Codelist'].astype(str)
        if 'Checks' in xls.sheet_names:
            df_als_checks1 = xcelread(files_als, 'Checks', 'A:G')
            df_als_checks1.dropna(axis=0, how='all', inplace=True)
            # df_als_checks = df_als_checks1.copy()
            df_als_checks = df_als_checks1.loc[df_als_checks1['CheckActive'] == True]

        if 'CheckActions' in xls.sheet_names:
            df_ca = xcelread(files_als, 'CheckActions', 'A:O')
            df_ca.dropna(axis=0, how='all', inplace=True)
            df_ca = df_ca.loc[:, ~ (df_ca.columns.str.contains('^Unnamed'))]
            df_ca1 = df_ca.loc[(df_ca['ActionType'] == 'SetDataPoint')]
            df_ca2 = df_ca.loc[(df_ca['ActionType'] == 'SetDynamicSearchList')]
            df_ca2 = df_ca2.loc[df_ca2['FormOID'].notnull() & df_ca2['FieldOID'].notnull()]
            df_ca3 = df_ca.loc[(df_ca['ActionType'] == 'SetDataPointVisible')]

            # df_flds_defaults = df_flds.loc[(~(df_flds['DefaultValue'].isnull()))]
            df_flds_sdp = df_flds.merge(df_ca1, left_on=['Study ID', 'FormOID', 'FieldOID'],
                                        right_on=['Study ID', 'FormOID', 'FieldOID'], how='left',
                                        suffixes=['', '_'],
                                        indicator=True)
            df_flds_sdpv = df_flds.merge(df_ca3, left_on=['Study ID', 'FormOID', 'FieldOID'],
                                         right_on=['Study ID', 'FormOID', 'FieldOID'], how='left',
                                         suffixes=['', '_'],
                                         indicator=True)
            df_flds_default_sdp = df_flds_sdp.loc[
                (((~(df_flds_sdp['DefaultValue'].isnull())) & (~(df_flds_sdp['DefaultValue'].astype(str) == '|'))) |
                 (df_flds_sdp['ActionType'] == 'SetDataPoint')) ]




        if 'CustomFunctions' in xls.sheet_names:
            df_cf = xcelread(files_als, 'CustomFunctions', 'A:P')
            df_cf = df_cf.loc[:, ~(df_cf.columns.str.contains('^Unnamed'))]

            df_cacf = df_ca2.merge(df_cf, left_on=['Study ID', 'ActionOptions'],
                                   right_on=['Study ID', 'FunctionName'],
                                   how='left', suffixes=['', '_'], indicator=True)

            mycols = set(df_cacf.columns)
            mycols.remove('_merge')
            df_cacf = df_cacf[mycols]
            df_fldcacf = df_cacf.merge(df_flds, left_on=['Study ID', 'FormOID', 'FieldOID'],
                                       right_on=['Study ID', 'FormOID', 'FieldOID'], how='left', suffixes=['', '_'],
                                       indicator=True)

        if 'CheckSteps' in xls.sheet_names:
            df_checksteps = xcelread(files_als, 'CheckSteps', 'A:P')
            df_checksteps.dropna(axis=0, how='all', inplace=True)
            df_checksteps = df_checksteps.loc[:, ~(df_checksteps.columns.str.contains('^Unnamed'))]

            df_checksteps['VariableOID'] = df_checksteps.groupby('CheckName')['VariableOID'].ffill()

        if 'DerivationSteps' in xls.sheet_names:
            df_DerivationSteps = xcelread(files_als, 'DerivationSteps', 'A:P')
            df_DerivationSteps.dropna(axis=0, how='all', inplace=True)
            df_DerivationSteps = df_DerivationSteps.loc[:, ~(df_DerivationSteps.columns.str.contains('^Unnamed'))]
        if 'Derivations' in xls.sheet_names:
            df_Derivations = xcelread(files_als, 'Derivations', 'A:O')
            df_Derivations.dropna(axis=0, how='all', inplace=True)
            df_Derivations = df_Derivations.loc[:, ~(df_Derivations.columns.str.contains('^Unnamed'))]

            df_Derivations.dropna(subset=['DerivationName'], how='all', inplace=True)


        xls.close()


        if len(df_ca) > 0:

            df_ca_oq = df_ca.loc[df_ca['ActionType'] == 'OpenQuery']

        if len(df_ca_oq) > 0:
            df_ca_oq_reshp = \
                (df_ca_oq.set_index(df_ca_oq.columns.drop('ActionOptions', 1).tolist())
                     .ActionOptions.str.split(',', expand=True)
                     .stack()
                     .reset_index()
                     .rename(columns={0: 'ActionOptions'})
                     .loc[:, df_ca_oq.columns]
                     )
            df_ca1 = df_ca[df_ca['ActionType'] == 'SetDynamicSearchList']
            df_ca1 = df_ca.loc[df_ca['FormOID'].notnull() & df_ca['FieldOID'].notnull()]
            #########################
        # path_sp = SITE_ROOT + '\SPMDR'
        # # path_sp = PARENT_ROOT + '\SPMDR'
        # os.chdir(path_sp)
        # lists = os.listdir(path_sp)

        # retrieving the stage domain & reference metadata

        ########################
        # xls = 'ELE_CODE.xlsx'
        # filename1 = path_sp + '\\' + xls
        # modx = os.path.getmtime(xls)
        # xmod = datetime.datetime.fromtimestamp(modx)
        # date = datetime.datetime.now()
        # modTime = time.mktime(date.timetuple())
        # fileLocation = os.path.abspath(xls)
        # if not datetime.datetime.today().date() == xmod.date():
        #     xlapp = win32com.client.DispatchEx("Excel.Application")
        #     xlapp.Quit()
        #     xlapp = win32com.client.DispatchEx("Excel.Application")
        #     # print("currentpath", path_sp)
        #     # print("currentpath", filename1)
        #     wb = xlapp.Workbooks.Open(filename1)
        #     wb.RefreshAll()
        #     xlapp.CalculateUntilAsyncQueriesDone()
        #     xlapp.DisplayAlerts = False
        #     wb.Save()
        #     xlapp.Quit()
        #     os.utime(fileLocation, (modTime, modTime))
        # # print("currentpath",path_sp)
        # xls = pd.ExcelFile('ELE_CODE.xlsx')
        # df_stage_domain = pd.read_excel(xls, 'stage_domain')
        # # df_ELEMENTS = pd.read_excel(xls, 'ELEMENTS')
        # # # df_CODELIST = pd.read_excel(xls, 'CODELIST')
        # df_Questions = pd.read_excel(xls, 'Questions')
        # # df_Questionnaires = pd.read_excel(xls, 'Questionnaires')
        # df_Test_Category = pd.read_excel(xls, 'Test Category')
        # # df_Test_Units = pd.read_excel(xls, 'Test Units')
        # # df_Units = pd.read_excel(xls, 'Units')
        # # df_Conversion = pd.read_excel(xls, 'Conversion')
        # # df_Precision = pd.read_excel(xls, 'Precision')
        # xls.close()

    ######################
    ########################
    # path_temp = 'C:\\Bhasp\\NVTSonco-work\\NVTSonco-work\\RADA\\Radaprogram\\Virtual Environments\\03Nov2021' + '\Templates'
    # path_temp = SITE_ROOT + '\Templates'
    path_temp = para8 
    os.chdir(path_temp)

    lists = os.listdir(path_temp)
    sub = '.xlsx'
    # sub1 = 'xls'
    # uat = 'uat'
    files = [mystr for mystr in lists if sub in mystr]

    DATABASE1 = "Rave Diagnostic Tool Report.xlsx"
    DATABASE = para8 + "\\" + DATABASE1



    CRFversion = ''
    global studyid
    studyid = ''
    df_flds_dom = pd.DataFrame()
    if (files_als):
    # if len(files_als) == 1:
        if len(df_forms) > 0:

            df_forms.reset_index(inplace=True, drop=True)

            CRFversion = df_forms.loc[0, 'DraftName']

            # ConfirmationMessage = df_forms.loc[0, 'ConfirmationMessage']
            # SignaturePrompt = df_forms.loc[0, 'SignaturePrompt']
            # CRFversion = df_forms['DraftName'][0]
            df_CRFversion = pd.DataFrame({'CRFversion': [CRFversion]})
            studyid = df_forms.loc[0, 'Study ID']

            # studyid = df_forms['Study ID'][0]

            #
            # df_list = [{'Check_number':[0]}]
            # df_series = {'Check_number':[0]}
            # my_cols_list=['Applicability','FormOID','FieldOID','DataDictionaryName','DefaultValue','AnalyteName','Check_number', 'Description']

            # df_checks.reindex(columns=[*df_checks.columns.tolist(), *my_cols_list])


            df_flds_chks = df_flds.loc[~df_flds['FieldOID'].isna()]
            # df_flds_chks = df_flds.dropna(how='all')

            df_flds_chks['FormOID'].replace('', np.nan, inplace=True)
            df_flds_chks['AnalyteName'].replace('', np.nan, inplace=True)
            df_flds_chks2 = df_flds_chks
            df_flds_chks2["FieldOID_spl"] = ""
            df_flds_chks2["FieldOID_spl"] = df_flds_chks2.loc[~df_flds_chks2["FieldOID"].isna(),"FieldOID"]

            # df_flds_chks2["FieldOID_spl"] = df_flds_chks2["FieldOID_spl"].str.split("_", n=1, expand=True)
            df_flds_chks2["FieldOID_spl"] = df_flds_chks2["FieldOID_spl"].str.split("_", expand=True)[0]

            df_flds_chks2["FieldOID_spl2"] = df_flds_chks2['FieldOID'].astype(str).apply(
                lambda x: x.split('_')[1] if len(x.split('_')) > 1 else '')

            # Nreqfrms =["PRIMARY002","UPVG001","UPVG002","SAERF", "SAE_DOS", "SAE_TEST", "SAE_DEATH", "SAEINFO", "INV_REV", "AUTO_02", "AUTO_03", "AUTO_04", "AUTO_05", "AUTO_06", "AUTO_07", "AUTO_08", "AUTO_09", "AUTO_01", "TRANSMIT", "OPG001"]
            df_flds_dom = df_flds_chks2.loc[(~(df_flds_chks['FormOID'].isna()))]

            df_flds_dom = df_flds_dom.loc[(~(df_flds_dom["FormOID"].isin(df_non_domain_forms['FormOID']))) == True]
            df_flds_dom = df_flds_dom.loc[
                (~((df_flds_dom['FieldOID_spl'].str.replace(r'\d+', '') == 'LBL'))) == True]
            df_flds_dom = df_flds_dom.loc[
                ~((df_flds_dom['FormOID'].astype(str).str.contains('|'.join(['SSG001', 'SSG002'])))) == True]

            df_flds_dom['domain'] = df_flds_dom['FormOID'].astype(str).apply(
                lambda x: x[:4] if x[:2] == 'FA' else x[:2])
            df_flds_dom['domain_ele'] = df_flds_dom['domain'].str.strip() + df_flds_dom['FieldOID_spl'].str.strip()
            # df_flds_dom.loc[:,'domain_frm'] = df_flds_dom['domain'] + ',' + df_flds_dom['FormOID']

#####################



    dir = os.getcwd()
    SITE_ROOT = os.getcwd()
    username = os.environ["USERNAME"]
    ossum_path1 = 'C:\\Users\\' + username
    osssum_dir = 'OSSUM'
    ossum_path = 'C:\\Users\\' + username + '\\' + osssum_dir
    rada_dir = 'RADA'
    rada_path = para7
    datestring = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M')
    rada_dt_path = rada_path + '\\' + datestring
    # os.rename()

    # print(rada_dt_path)
    if os.path.exists(rada_path):
        os.makedirs(rada_dt_path)
        recursive_copy(path_temp, rada_dt_path)
        DATABASE1 = "Rave Diagnostic Tool Report.xlsx"
        DATABASE3 = studyid + "_Rave Diagnostic Tool Report_" + datestring + ".xlsx"
        DATABASE2 = rada_dt_path + "\\" + DATABASE1
        DATABASE4 = rada_dt_path + "\\" + DATABASE3
        if (os.path.exists(DATABASE2) ):
            os.rename(DATABASE2,DATABASE4)

    path_out = rada_dt_path
    # clearfoldercontent(path_out)
    os.chdir(path_out)
    # df_instr = pd.read_excel('Rave Diagnostic Tool Report.xlsx', skiprows=3, sheet_name='Instructions')
    df_instr = specxlread(DATABASE, 'Instructions', 'Check_number')
    # df_instr['Check_number_1'] = df_instr['Check_number'].astype(str)

    # path_out = 'C:\\Bhasp\\NVTSonco-work\\NVTSonco-work\\RADA\\Radaprogram\\Virtual Environments\\03Nov2021' +  '\output report'
    # path_out = PARENT_ROOT + '\result'


    # shutil.copy( path_temp , path_out)

    ###############copy the template results to output
    # clearfoldercontent(path_out)
    # recursive_copy(path_temp, path_out)


#############


    if len(df_stage_domain) > 0:
        df_stage_domain2 = df_stage_domain.loc[:, ['DATA_STATE', 'DATA_DOMAIN', 'TARGET_DATA_ELEMENT', 'STATUS']]

        df_stage_domain2 = df_stage_domain2.loc[
            ((df_stage_domain2['DATA_STATE'] == 'EDC') & (df_stage_domain2['STATUS'] == 'A'))]
        df_stage_domain2['domain_ele'] = df_stage_domain2['DATA_DOMAIN'].str.strip() + \
                                         df_stage_domain2['TARGET_DATA_ELEMENT'].str.strip()

    ###for check#05
    def rada_chk5_fn(df1, df2, fld1, fld2, chkno, **kwargs):
        df_chk = pd.DataFrame()
        cond = kwargs.get('cond', '')
        fld3 = kwargs.get('fld3', None)
        fld4 = kwargs.get('fld4', None)
        if len(df1) > 0:
            df_chk = df1.copy()
        if ((len(df_chk) > 0) & (len(df2) > 0)):
            if cond == '~':
                df_chk.loc[(~((df_chk[fld1].isin(df2[fld2])))) == True, ('Check_number')] = chkno
            else:
                df_chk.loc[(((df_chk[fld1].isin(df2[fld2])))) == True, ('Check_number')] = chkno

        if (len(df_chk) > 0):
            df_chk = df_chk.loc[(df_chk['Check_number'] == chkno)]

            if fld3:
                df_chk.rename(columns={fld3: fld4})

        return df_chk

    def rada_chk6_fn(df1, df2, fld1, fld2, chkno, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        if (len(df1) > 0) & (len(df2) > 0):
            df_chk = df1.loc[:, labellist1]
            df2[fld2] = df2[fld2].str.strip().str.lower()
            df_chk.loc[df_chk[fld1].str.strip().str.lower().isin(df2[fld2]) == True, 'Check_number'] = chkno


        if (len(df_chk) > 0):
            df_chk = df_chk.loc[(df_chk['Check_number'] == chkno)]

        return df_chk

    ###for check#07
    def rada_chk7_fn(df1, fld1, chkno):
        df_chk = pd.DataFrame()
        if len(df1) > 0:
            df_chk = df1.copy()
        if len(df_chk) > 0:
            df_chk.loc[(df_chk[fld1].str.replace('[A-Z0-9_]', '') == '') == False, ('Check_number')] = chkno

        if (len(df_chk) > 0):
            df_chk = df_chk.loc[(df_chk['Check_number'] == chkno)]

        return df_chk

    def rada_chk50_fn(df1, df2, fld1, fld2, fld3, newfld1, newfld2, fld4, chkno, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        labellist2 = kwargs.get('labellist2', [])

        if ((len(df1) > 0) & (len(df2) > 0)):
            df = df1.loc[~(df1[fld2].isna()), labellist1]
            df.loc[((df[fld2].astype(str).str.contains('^[0-9]*$')) == True), newfld2] = df[fld2]
            df.loc[((df[fld2].astype(str).str.contains('^[0-9]*$')) == True), newfld1] = df[fld2]
            df.loc[((df[fld2].astype(str).str.contains('^[$]')) == True), newfld1] = df[fld2].str.replace('$', '')
            df.loc[df[newfld1].astype(str).str.contains('.', na=False), newfld1] = df[newfld1].str.rsplit('.', 1).str.get(0)
            df = df.loc[df[newfld1].astype(str).str.contains('^[0-9]*$', na=False)]
            df2 = df2.loc[:, labellist2]
            df3 = df2.merge(df, left_on=[fld1], right_on=[fld1], how='left',
                            suffixes=['', '_'], indicator=True)
            df3 = df3.loc[~(df3[fld3].isna())]
            df3 = df3.loc[~(df3[newfld1].isna())]
            df3 = df3.drop(['_merge'], axis=1)

            df3 = df3.loc[((df3[fld3].str.len() > df3[newfld1].astype(int)) |
                           ((df3[newfld2].astype(str).str.contains('^[0-9]*$')) & (df3[fld3].astype(str).str.contains('[a-zA-Z]'))))]
            df3 = df3.sort_values(by=[fld1, fld2, fld4])
            f = {'CodedData': ','.join, 'UserDataString': ','.join}
            df3 = df3.groupby([fld1, fld2, fld4], as_index=False).agg(f)

            if (len(df3) > 0):
                df3.loc[~df3[fld1].isna(), 'Check_number'] = chkno
                df_chk = df_chk.append(df3)

        return df_chk

    def rada_chk54_fn(df1, df2, df3, df4, fld1, fld2, newfld1, fld3, fld4, chkno, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        labellist2 = kwargs.get('labellist2', [])
        dict_cf_list = []
        dict_cf_list2 = []

        if ((len(df1) > 0) & (len(df2) > 0) & (len(df3) > 0) & (len(df4) > 0)):
            df1_1 = df1.loc[~df1[fld1].isna(), labellist1]
            df3 = df3.loc[~df3[fld2].isna()]
            [dict_cf_list.append(s) if s not in dict_cf_list else dict_cf_list for s in df1_1[fld1] if any(s in xs for xs in df3[fld2])]
            # df1_1.loc[~(df1_1[fld1].isna()), newfld1] = '"' + df1_1[fld1] + '"'
            # df2.loc[~(df2[fld1].isna()), newfld1] = '"' + df2[fld1] + '"'

            df1_1 = df1_1.loc[~(df1_1[fld1].isin(dict_cf_list))]

            # df5 = df3.loc[~(df3[fld2].str.contains('|'.join(df1_1[newfld1]), na=False))]
            df1_1.loc[(~(df1_1[fld1].isin(df2[fld1]))), 'Check_number'] = chkno
            # df1_1.loc[((~(df1_1[fld1].isin(df2[fld1]))) & (~(df1_1[fld1].isin(df3[fld2])))), 'Check_number'] = chkno
            #
            # df5 = df3.loc[~(df3[fld2].str.contains('PerformCustomFunction', na=False))]
            # df5 = \
            #     (df5.set_index(df5.columns.drop('SourceCode', 1).tolist())
            #          .SourceCode.str.split(';', expand=True)
            #          .stack()
            #          .reset_index()
            #          .rename(columns={0: 'SourceCode'})
            #          .loc[:, df5.columns]
            #          )
            # df5 = \
            #     (df5.set_index(df5.columns.drop('SourceCode', 1).tolist())
            #          .SourceCode.str.split('\n', expand=True)
            #          .stack()
            #          .reset_index()
            #          .rename(columns={0: 'SourceCode'})
            #          .loc[:, df5.columns]
            #          )
            # df5 = df5.loc[~(df5[fld2].str.contains('PerformCustomFunction'))]
            # df5.loc[~(df5[fld4].isna()), newfld1] = 'PerformCustomFunction' + '\(' + '"' + df5[fld4] + '"'
            # df5 = df5.loc[~(df5.duplicated([fld4]))]
            # df3 = df3.loc[~(df3[fld2].str.contains('|'.join(df5[newfld1]), na=False))]
            # df4 = df4.loc[~(df4[fld3].isna())]
            # df3.loc[~(df3[fld4].isin(df4[fld3])), 'Check_number'] = chkno

            if (len(df1_1) > 0):
                df1_1 = df1_1.loc[(df1_1['Check_number'] == chkno)]
                df_chk = df_chk.append(df1_1)
                f = {fld1: ','.join}
                df_chk = df_chk.groupby(['Check_number'], as_index=False).agg(f)

        return df_chk

    def rada_chk55_fn(df1, df2, fld1, fld2, chkno, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])

        if ((len(df1) > 0) & (len(df2) > 0)):
            df_chk = df1.loc[:, labellist1]
            df_chk.loc[~(df_chk[fld1].isin(df2[fld2])), 'Check_number'] = chkno

        if (len(df_chk) > 0):
            df_chk = df_chk.loc[(df_chk['Check_number'] == chkno)]

        return df_chk

    def check(list):
        return list.count(list[1]) == len(list)

    def rada_chk11_fn(df1, fld1, fld2, fld3, fld4, new_fld2, new_fld3, new_fld4, chkno, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        labellist2 = kwargs.get('labellist2', [])
        labeldict1 = kwargs.get('labeldict1', {})
        if len(df1) > 0:
            df1 = df1.sort_values(by=[fld1, fld4])
            df1 = df1.loc[~(df1[fld1].isna())]
            df1.loc[~df1[fld2].isna(), new_fld2] = df1[fld2]
            df1.loc[~df1[fld3].isna(), new_fld3] = df1[fld3]
            df1.loc[~df1[fld4].isna(), new_fld4] = df1[fld4]
            df1 = df1.groupby(fld1, as_index=False).agg(labeldict1)
            df1 = df1.loc[df1[new_fld4].astype(int) > 1]

        if (len(df1) > 0):
            df_chk = df1.loc[:, labellist1]
            df_chk.loc[~(df_chk[fld1].isna()), 'Check_number'] = chkno

        return df_chk

    def rada_chk49_fn(df1, fld1, fld2, value1, fld3, chkno, **kwargs):
        df_chk = pd.DataFrame()
        #v0.2
        df_chk1 = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        labellist2 = kwargs.get('labellist2', [])
        labeldict1 = kwargs.get('labeldict1', {})
        if len(df1) > 0:
            df_chk1 = df1.loc[df1[fld2] == value1, labellist1]
            df_chk2 = df1.loc[~(df1[fld2] == value1), labellist1]
            if (len(df_chk1) >0) & (len(df_chk2) >0):
                df_chk1.loc[df_chk1[fld3].isin(df_chk2[fld3]), 'Check_number'] = chkno
            # df1 = df1.groupby(labellist2, as_index=False).agg(labeldict1)
            # df_chk = df1.loc[df1.duplicated([fld1])]

        if (len(df_chk1) > 0):
            df_chk = df_chk1.loc[(df_chk1['Check_number'] == chkno)]

        return df_chk

    def rada_chk52_fn(df, fld1, fld2, chkno, reqstr, reqstr2,  reqstr3, newfld1, newfld2):
        df_chk = pd.DataFrame()
        df_chk2 = pd.DataFrame()
        if len(df) > 0:
            df = df.loc[df[fld2].astype(str).str.contains('|'.join(reqstr3), case=False)]
            # df = df.loc[df[fld2].str.endswith('_', na=False)]

            df[newfld1] = df[fld2].astype(str).apply(lambda x: x.split('_')[0] if len(x.split('_')) == 2 else '')
            df.loc[df[fld2].astype(str).str.contains('|'.join(reqstr), case=False), newfld1] = 'LBVALUE'

            df = df.loc[(df[newfld1] != '')]
            df = df.groupby([fld1, newfld1]).size().reset_index(name=newfld2)

            df = df.loc[(df[newfld2] > 1)]
            if (len(df) > 0):
                df.loc[~df[fld1].astype(str).str.contains(reqstr2, case=False), 'Check_number'] = chkno


        if (len(df) > 0):
            df_chk = df.loc[(df['Check_number'] == chkno)]
            df_chk['FormOID'].replace('nan', np.nan, regex=True, inplace=True)
            df_chk['FormOID'] = df_chk['FormOID'].fillna("")
            f = {'FormOID': lambda x: ','.join(unique1(x))}
            df_chk = df_chk.groupby(['Check_number'], as_index=False).agg(f)

        return df_chk

    def rada_chk53_fn(df1, fld1, fld2, chkno, reqstr, reqstr2, newfld, newfld1, newfld2):
        df_chk = pd.DataFrame()
        df_chk2 = pd.DataFrame()
        df = pd.DataFrame()
        if len(df1) > 0:
            df1 = df1.loc[df1[fld1].astype(str).str.contains(reqstr, case=False)]

            df = df1.loc[df1[fld2].astype(str).str.contains('|'.join(reqstr2), case=False)]

            df_chk2 = df1[fld2].astype(str).str.contains('|'.join(reqstr2), case=False).groupby(df1[fld1]).sum().reset_index(
                name=newfld)

            if len(df_chk2) > 0:
                df_chk2.loc[df_chk2[newfld] == False, ('Check_number')] = chkno
            if len(df) > 0:
                df.loc[~(df[fld2].isna()), newfld1] = df[fld2].astype(str).apply(
                    lambda x: x.split('_')[0] if len(x.split('_')) > 1 else '')
            if len(df) > 0:
                df.loc[df[fld2].astype(str).str.contains('|'.join(reqstr), case=False), newfld1] = 'LBVALUE'
            if len(df) > 0:
                df = df.loc[(df[newfld1] != '')]
            if len(df) > 0:
                df = df.groupby([fld1, newfld1]).size().reset_index(name=newfld2)
            if len(df) > 0:
                # < 2 is for one LBVALUE in fieldOID does not require _TR in FormOID
                df.loc[(df[newfld2] < 2), 'Check_number'] = chkno


        if (len(df_chk2) > 0) & ('Check_number' in df_chk2.columns):
            df_chk2 = df_chk2.loc[(df_chk2['Check_number'] == chkno)]
            df_chk = df_chk.append(df_chk2)

        if (len(df) > 0):
            df = df.loc[(df['Check_number'] == chkno)]
            df_chk = df_chk.append(df)

        return df_chk

    def rada_chk67_fn(df1, fld1, fld2, chkno, **kwargs):
        df_chk = pd.DataFrame()
        df_chk1 = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        if len(df1) > 0:
            df1 = df1.loc[:, labellist1]
            df_chk1 = df1.loc[(df1.duplicated([fld1], keep='first')) | (df1.duplicated([fld2], keep='first'))]

        if (len(df_chk1) > 0):
            df_chk1.loc[~(df_chk1[fld1].isna()), 'Check_number'] = chkno
            f = {'OID': ','.join, 'DraftFormName': ','.join}
            df_chk1 = df_chk1.groupby(['Check_number'], as_index=False).agg(f)
            df_chk = df_chk.append(df_chk1)

        return df_chk

    def rada_chk48_fn(df1, fld1, chkno, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        newfield = kwargs.get('newfield', '')
        if len(df1) > 0:
            df1 = df1.loc[:, labellist1]
            df1[newfield] = df1[fld1].str.replace("'", '\'')
            df1[newfield] = df1[newfield].str.replace("'", r'\'')
            df1[newfield] = df1[newfield].str.replace("'", r'"')
            df1[newfield] = df1[newfield].str.replace(r'\\', '')
            # df1[newfield] = df1[newfield].str.replace('[a-zA-Z0-9_~!@#$%^&*()/+,.<>:;{}[]|\=-''"''-]', '').str.strip()
            # df1[newfield] = df1[fld1].str.replace('', np.NaN)
        if len(df1) > 0:
            df1.loc[(df1[newfield].str.replace('[a-zA-Z0-9_\~|!|@|#$%^&*?()/\+,.<>:;{}[|\]=|\-|''"|"|]',
                                               '').str.strip() == '')
                    == False, 'Check_number'] = chkno
        if (len(df1) > 0):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk70_fn(df1, df2, df3, fld1, fld2, fld3, fld4, chkno, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        if len(df1) > 0:
            df1 = df1.loc[~df1[fld3].isna(), labellist1]
            df2 = df2.loc[~df2[fld3].isna()]
        if len(df2) > 0:
            df2.drop_duplicates(subset=[fld3], keep='last', inplace=True)
        if (len(df2) > 0) & (len(df3) > 0):
            # updated on 10aug21 with '~'
            df2 = df2.loc[~df2[fld4].isin(df3[fld4])]
        if (len(df2) > 0) & (len(df1) > 0):
            df1 = df1.loc[~df1[fld3].isin(df2[fld3])]
        if len(df1) > 0:
            df1.loc[(((df1[fld1].str.replace('[a-zA-Z0-9_\~|!|@|#$%^&*?()/\+,.<>:;{}[|\]=|\-|''"|"|]',
                                             '').str.strip() == '') == False) |
                     ((df1[fld2].str.replace('[a-zA-Z0-9_\~|!|@|#$%^&*?()/\+,.<>:;{}[|\]=|\-|''"|"|]',
                                             '').str.strip() == '') == False)), 'Check_number'] = chkno

        if (len(df1) > 0):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk22_fn(df1, fld1, chkno, reqnum):
        df_chk = pd.DataFrame()
        if len(df1) > 0:
            df_chk = df1.copy()
        if len(df_chk) > 0:
            df_chk.loc[df_chk[fld1].astype(str).str.contains('|'.join(reqnum)), ('Check_number')] = chkno

        if (len(df_chk) > 0):
            df_chk = df_chk.loc[(df_chk['Check_number'] == chkno)]

        return df_chk

    def rada_chk23_fn(df1, fld1, fld2, chkno, **kwargs):
        df_chk = pd.DataFrame()
        df2 = pd.DataFrame()
        value1 = kwargs.get('value1', None)
        value2 = kwargs.get('value2', None)
        labellist1 = kwargs.get('labellist1', [])
        if len(df1) > 0:
            df1 = df1.loc[:, labellist1]
            df2 = df1.loc[df1[fld1].astype(str).str.contains('^[0-9]*$')]
            df3 = df1.loc[~(df1[fld1].astype(str).str.contains('^[0-9]*$'))]
        if len(df2) > 0:
            #updated 16 to 25
            df2.loc[df2[fld2].apply(len) > 25, 'Check_number'] = chkno
            # df3.loc[df3[fld2].apply(len) > 25, 'Check_number'] = chkno

        if (len(df2) > 0) & ('Check_number' in df2.columns):
            df_chk = df2.loc[(df2['Check_number'] == chkno)]
        #     df_chk = df_chk.append(df2)
        # if (len(df3) > 0) :
        #     df3 = df3.loc[(df3['Check_number'] == chkno)]
        #     df_chk = df_chk.append(df3)

        return df_chk

    def rada_chk26_fn(df2, fld1, fld2, fld3, chkno, **kwargs):
        df_chk = pd.DataFrame()
        value1 = kwargs.get('value1', None)
        labellist1 = kwargs.get('labellist1', [])
        frmlist2 = kwargs.get('frmlist2', [])

        if len(df2) > 0:
            df2 = df2.loc[:, labellist1]
            # df2 = df2.loc[~(df2[fld2].isin(frmlist2))]
        if len(df2) > 0:
            df2.loc[((df2[fld1] == False) | ((df2[fld1] == True) &
                                             ((df2[fld2].astype(str).str.contains(value1, na=False)) |
                                              (df2[fld3].astype(str).str.contains(value1, na=False))))), 'Check_number'] = chkno

        if (len(df2) > 0) & ('Check_number' in df2.columns):
            df_chk = df2.loc[(df2['Check_number'] == chkno)]

        return df_chk

        # (df_checksteps, df_flds, "VariableOID", "FormOID", 16, fld3='RecordPosition',
        # fld4='IsLog',
        # labellist1=["CheckName", "VariableOID", "FormOID", 'RecordPosition'],
        # labellist2=["VariableOID", "FormOID", 'IsLog'])

    def rada_chk16_fn(df1_1, df2, fld1, fld2, chkno, **kwargs):
        df_chk = pd.DataFrame()
        df1 = pd.DataFrame()
        fld3 = kwargs.get('fld3', None)
        fld4 = kwargs.get('fld4', None)
        labellist1 = kwargs.get('labellist1', [])
        labellist2 = kwargs.get('labellist2', [])

        if ((len(df1_1) > 0) & (len(df2) > 0)):
            df1 = df1_1.loc[:, labellist1]
            df1 = df1.loc[~(df1[fld3].isna())]
        if (len(df1) > 0):
            df1 = df1.loc[~(df1[fld1].isna())]
            df2 = df2.loc[:, labellist2]
        if ((len(df1) > 0) & (len(df2) > 0)):
            df1 = df1.merge(df2, left_on=[fld1], right_on=[fld1], how='left',
                            suffixes=['', '_'], indicator=True)
        if (len(df1) > 0):

            df1.loc[df1['_merge'] == 'left_only', 'Check_number'] = chkno

            values = [str(i) for i in np.arange(1, 50)]
            reqnums = list(values)
        if (len(df1) > 0):
            df1.loc[((((df1[fld3] == 0) & (df1[fld4] == True))) & (df1['_merge'] == 'both')) == True,
                    'Check_number'] = chkno
        if (len(df1) > 0):
            df1.loc[((((((df1[fld3].astype(str).str.contains('|'.join(reqnums)))) &
                        (df1[fld4] == False)))) & (df1['_merge'] == 'both')) == True,
                    'Check_number'] = chkno

        if (len(df1) > 0) & ('Check_number' in df1.columns):
            df1 = df1.drop(['_merge'], axis=1)
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk17_fn(df1, df2, df3, fld1, fld2, chkno, **kwargs):
        df_chk = pd.DataFrame()
        cond = kwargs.get('cond', None)
        frm1 = kwargs.get('frm1', None)
        frm2 = kwargs.get('frm2', None)
        fld3 = kwargs.get('fld3', None)
        fld4 = kwargs.get('fld4', None)
        fld5 = kwargs.get('fld5', None)
        fld6 = kwargs.get('fld6', None)
        labellist1 = kwargs.get('labellist1', [])
        labellist2 = kwargs.get('labellist2', [])
        if len(df1) > 0:
            df_chk = df1.copy()

        if ((len(df_chk) > 0) & (len(df2) > 0) & (len(df3) > 0)):
            if set(labellist1).issubset(df_chk.columns):
                df_chk = df_chk.loc[:, labellist1]
        if (len(df_chk) > 0):
            df_chk.loc[((df_chk[fld2].isna()) | (df_chk[fld2].astype(str).str.contains('invalid')) |
                        ((~(df_chk[fld1].isin(df2[fld1]))) & ((df_chk[fld1].isin(df3[fld1])))) |
                        (((df_chk[fld1].isin(df2[fld1]))) & (~(df_chk[fld1].isin(df3[fld1]))))),
                       ('Check_number')] = chkno


        if (len(df_chk) > 0) & ('Check_number' in df_chk.columns):
            df_chk = df_chk.loc[(df_chk['Check_number'] == chkno)]

        return df_chk

    def rada_chk18_fn(df1, df2, df3, fld1, fld2, chkno, **kwargs):
        df_chk1 = pd.DataFrame()
        df_chk2 = pd.DataFrame()
        fld3 = kwargs.get('fld3', None)
        fld4 = kwargs.get('fld4', None)
        fld5 = kwargs.get('fld5', None)
        labellist1 = kwargs.get('labellist1', [])

        if ((len(df1) > 0) & (len(df2) > 0) & (len(df3) > 0)):
            df_chk = df1.loc[:, labellist1]
            # df2 = df2.mask(df2 == '')
            df2 = df2.loc[~(df2[fld3].isna())]
            if len(df2) > 0:
                df2 = df2.loc[~df2[fld3].astype(str).str.contains('[a-zA-Z1-9_~!@#$%^&*()/+,.<>:;{}[]|\=-''"]', na=False)]
            if len(df2) > 0:
                df2 = df2.loc[df2[fld3].astype(str).str.contains('0')]
                df3 = df3.loc[df3[fld2].isin(df2[fld2])]
                df_chk1 = df_chk.loc[df_chk[fld4].isin(df3[fld4])]
                df_chk2 = df_chk.loc[(df_chk[fld5].astype(str).str.contains('IsNotEmpty', na=False))]
                if (len(df_chk1) > 0):
                    df_chk1.loc[~(df_chk1[fld4].isin(df_chk2[fld4])), ('Check_number')] = chkno

        if (len(df_chk1) > 0) & ('Check_number' in df_chk1.columns):
            df_chk2 = df_chk1.loc[(df_chk1['Check_number'] == chkno)]
            if (len(df_chk2) > 0):
                f = {'FormOID': 'first', 'VariableOID': 'first', 'CheckFunction': 'last', 'Check_number': 'first'}
                df_chk2 = df_chk2.groupby(['CheckName'], as_index=False).agg(f)

        return df_chk2

    def rada_chk43_fn(df1, fld1, fld2, fld3, newfld1, chkno, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        if len(df1) > 0:
            df1 = df1.loc[~(df1[fld1].isna()), labellist1]
            # df[fld1].fillna("", inplace=True)
            df1[fld2].replace('nan', np.nan, regex=True, inplace=True)
            df1[fld3].replace('nan', np.nan, regex=True, inplace=True)
            df1[fld2].fillna("", inplace=True)
            df1[fld3].fillna("", inplace=True)
            df1 = df1.sort_values(by=[fld1, fld2, fld3])

            f = {'CodedData': ','.join, 'UserDataString': ','.join}
            df1 = df1.groupby([fld1], as_index=False).agg(f)
            df1.loc[df1[fld1].astype(str).str.contains('[0-9]$'), newfld1] = df1[fld1].str.rsplit('_', 1).str.get(0)
            df1.loc[(df1[newfld1].isna()), newfld1] = df1[fld1]
            df1 = df1.loc[df1.duplicated([newfld1, fld2, fld3], keep=False)]
            f = {'DataDictionaryName': ','.join, 'CodedData': 'first', 'UserDataString': 'first'}
            df1 = df1.groupby([newfld1], as_index=False).agg(f)

        if (len(df1) > 0):
            df_chk = df1.loc[~(df1[fld1].isna())]
        if (len(df_chk) > 0):
            df_chk = df_chk.loc[:, labellist1]
            df_chk.loc[~(df_chk[fld1].isna()), 'Check_number'] = chkno

        return df_chk

    def rada_chk47_fn(df1, fld1, fld2, fld3, chkno, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        labeldict1 = kwargs.get('labeldict1', [])
        if len(df1) > 0:
            df1 = df1.loc[~(df1[fld1].isna())]
            if len(df1) > 0:
                df1[fld2].replace('nan', np.nan, regex=True, inplace=True)
            if len(df1) > 0:
                df1[fld3].replace('nan', np.nan, regex=True, inplace=True)
            if len(df1) > 0:
                df1[fld2].fillna("", inplace=True)
            if len(df1) > 0:
                df1[fld3].fillna("", inplace=True)
            # df1.loc[~df1[fld2].isna(), new_fld2] = df1[fld2]
            # df1.loc[~df1[fld3].isna(), new_fld3] = df1[fld3]
            # df1.loc[~df1[fld2].isna(), new2_fld2] = df1[fld2]
            # df1.loc[~df1[fld3].isna(), new2_fld3] = df1[fld3]
            if len(df1) > 0:
                df1 = df1.groupby([fld1], as_index=False).agg(labeldict1)
            # df1 = df1.loc[((df1[new_fld2] != df1[new2_fld2]) | (df1[new_fld3] != df1[new2_fld3]))]
            if len(df1) > 0:
                df1.loc[~df1[fld2].isna(), fld2] = df1[fld2].str.split('`')
            if len(df1) > 0:
                df1.loc[~df1[fld3].isna(), fld3] = df1[fld3].str.split('`')
            if len(df1) > 0:
                df1[fld2] = df1[fld2].apply(lambda x: keep_duplicates_list(x))
            if len(df1) > 0:
                df1[fld3] = df1[fld3].apply(lambda x: keep_duplicates_list(x))
            if len(df1) > 0:
                df1[fld2] = df1[fld2].apply(lambda x: str(x).strip('[] '))
            if len(df1) > 0:
                df1[fld3] = df1[fld3].apply(lambda x: str(x).strip('[] '))
            if len(df1) > 0:
                df1 = df1.loc[((df1[fld2].str.len() > 0) | (df1[fld3].str.len() > 0))]

            # dict = Counter(df1[fld2])
            # dict1 = Counter(df1[fld3])
            # dupes_list = []
            # dupes_list_fld2 = []
            # dupes_list_fld3 = []
            # [dupes_list_fld2.append(lst) for lst in df1[fld2] if not (lst in dupes_list)]

            # df1[new2_fld2] = list(df1[new2_fld2])
            # df1[new2_fld3] = list(df1[new2_fld3])
            # df1.loc[~df1[new2_fld2].isna(), "new2_fld2"] = df1[new2_fld2].apply(lambda x: len(x))
            # df1.loc[~df1[new2_fld3].isna(),"new2_fld3"] = df1[new2_fld3].apply(lambda x: len(x))

            # df1 = df1.loc[(df1[new_fld2].str.len() != df1[fld2].str.len()) |
            #                (df1[new_fld3].str.len() != df1[fld3].str.len()) ]
            # df1.loc[~df1[fld3].isna(), new_fld3] = df1[fld3]
            # df1.loc[df1.duplicated([fld1, fld2, fld3], keep=False), 'Check_number'] = chkno
            # print("fld1, fld2, fld3", len(df1))
            # df1.loc[df1.duplicated([fld1, fld2], keep=False), 'Check_number'] = chkno
            # print("fld1, fld2", len(df1))
            # df1.loc[df1.duplicated([fld1, fld3], keep=False), 'Check_number'] = chkno
            # print("fld1, fld3", len(df1))
            # df.loc[df.duplicated([fld1, fld2, fld3, newfld2], keep='last'), 'Check_number'] = chkno

        if (len(df1) > 0):
            df_chk = df1.loc[:, labellist1]
            df_chk.loc[~(df_chk[fld1].isna()), 'Check_number'] = chkno
            # df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk40_fn(df1, df2, df3, fld1, fld2, chkno, **kwargs):
        df_chk = pd.DataFrame()
        df_chk1 = pd.DataFrame()
        df_chk2 = pd.DataFrame()
        df_chk3 = pd.DataFrame()
        newfield = kwargs.get('newfield', None)
        newfield2 = kwargs.get('newfield2', None)
        value1 = kwargs.get('value1', None)
        value2 = kwargs.get('value2', None)
        fld3 = kwargs.get('fld3', None)
        fld4 = kwargs.get('fld4', None)
        fld5 = kwargs.get('fld5', None)
        fld6 = kwargs.get('fld6', None)
        fld7 = kwargs.get('fld7', None)
        fld8 = kwargs.get('fld8', None)
        labellist1 = kwargs.get('labellist1', [])
        labellist2 = kwargs.get('labellist2', [])
        if len(df1) > 0:
            df_chk = df1.copy()

        if ((len(df_chk) > 0) & (len(df2) > 0) & (len(df3) > 0)):
            df_chk = df_chk.loc[(~(df_chk[fld6].isna())), labellist1]
            df3 = df3.loc[:, labellist2]
            df3 = df3.loc[(~(df3[fld2].isna()))]
            if len(df3) > 0:
                df3[newfield] = df3[fld5].str.replace('$', '')
                df3.dropna(subset=[newfield], inplace=True)
            if len(df_chk) > 0:
                df_chk = df_chk.loc[df_chk[fld7] == value1]
            if len(df_chk) > 0:
                df_chk.loc[~(df_chk[fld1] == value2), ('Check_number')] = chkno

            if (len(df_chk) > 0) & (len(df3) > 0):
                df_chk2 = df_chk.merge(df3, left_on=[fld4, fld8], right_on=[fld4, fld8], how='left',
                                   suffixes=['', '_'], indicator=True)
            if len(df_chk2) > 0:
                df_chk2[newfield2] = df_chk2[fld6].str.len()
                df_chk2.loc[
                    (df_chk2[newfield2].astype(float) > df_chk2[newfield].astype(float)), ('Check_number')] = chkno
                mycols = set(df_chk2.columns)
                mycols.remove('_merge')
                df_chk2 = df_chk2[mycols]
                df_chk3 = df_chk2.merge(df2, left_on=[fld2, fld6], right_on=[fld2, fld3], how='left',
                                    suffixes=['', '_'], indicator=True)
            if len(df_chk3) > 0:
                df_chk3.loc[df_chk3['_merge'] == 'left', ('Check_number')] = chkno

                mycols = set(df_chk3.columns)
                mycols.remove('_merge')
                df_chk3 = df_chk3[mycols]

        if (len(df_chk1) > 0) & ('Check_number' in df_chk1.columns):
            df_chk1 = df_chk1.loc[(df_chk1['Check_number'] == chkno)]
            df_chk = df_chk.append(df_chk1)

        if (len(df_chk2) > 0) & ('Check_number' in df_chk2.columns):
            df_chk2 = df_chk2.loc[(df_chk2['Check_number'] == chkno)]
            df_chk = df_chk.append(df_chk2)
        if (len(df_chk3) > 0) & ('Check_number' in df_chk3.columns):
            df_chk3 = df_chk3.loc[(df_chk3['Check_number'] == chkno)]
            df_chk = df_chk.append(df_chk3)

        return df_chk

    def rada_chk41_fn(df1, df2, df3, df4, fld1, fld2, chkno, **kwargs):
        df_chk = pd.DataFrame()
        df_chk1 = pd.DataFrame()
        df_chk2 = pd.DataFrame()
        fld3 = kwargs.get('fld3', None)
        fld4 = kwargs.get('fld4', None)
        fld5 = kwargs.get('fld5', None)
        labellist1 = kwargs.get('labellist1', [])

        if len(df4) > 0:
            df4 = df4.loc[(df4[fld4] == True)]

        if ((len(df1) > 0) ):
            df_chk1 = df1.loc[:, labellist1]
            if len(df4) > 0:
                df_chk1 = df_chk1.loc[(df_chk1[fld5].isin(df4[fld5])) == True]
        if ((len(df_chk1) > 0) & ((len(df2) > 0))):
            df_chk1.loc[(df_chk1[fld1].isin(df2[fld1])) == True, 'Check_number'] = chkno

        if ((len(df1) > 0) & ((len(df3) > 0))):
            df_chk2 = df1.loc[:, labellist1]
            df_chk2.loc[((df_chk2[fld2].isin(df3[fld3]))) == True, 'Check_number'] = chkno

        if (len(df_chk1) > 0) & ('Check_number' in df_chk1.columns):
            df_chk1 = df_chk1.loc[(df_chk1['Check_number'] == chkno)]
            df_chk = df_chk.append(df_chk1)
        if (len(df_chk2) > 0) & ('Check_number' in df_chk2.columns):
            df_chk2 = df_chk2.loc[(df_chk2['Check_number'] == chkno)]
            df_chk = df_chk.append(df_chk2)

        return df_chk

    def rada_chk75_fn(df1, chkno, str1, col1):
        df_chk = pd.DataFrame()
        if len(df1) > 0:
            df1.loc[((df1[col1] != str1)) == True, 'Check_number'] = chkno

        if (len(df1) > 0) & ('Check_number' in df1.columns):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk29_fn(df1, df2, col1, col2, col3, col4, col5, value1, chkno, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        labellist2 = kwargs.get('labellist2', [])
        labellist3 = kwargs.get('labellist3', [])
        labellist4 = kwargs.get('labellist4', [])

        if len(df1) > 0:
            df1.drop_duplicates(subset=[col1], keep='first', inplace=True)
            df1 = df1.loc[:, labellist4]
            df1.loc[((df1[col1] != value1)) == True, 'Check_number'] = chkno

        if len(df2) > 0:
            df2.drop_duplicates(keep='first', inplace=True)
            df2.loc[~(df2[col2].astype(str).str.contains('|'.join(labellist1), na=True)), 'Check_number'] = chkno
            df2.loc[~(df2[col3].astype(str).str.contains('|'.join(labellist2), na=True)), 'Check_number'] = chkno
            df2.loc[df2[col4].isna(), 'Check_number'] = chkno
            df2.loc[~(df2[col5].astype(str).str.contains('|'.join(labellist3), na=True)), 'Check_number'] = chkno
        else:
            df2 = pd.DataFrame({"Check_number": [chkno]})
            # df2['Check_number'] = chkno

        if (len(df1) > 0) & ('Check_number' in df1.columns):
            df1 = df1.loc[(df1['Check_number'] == chkno)]
            df_chk = df_chk.append(df1)
        if (len(df2) > 0) & ('Check_number' in df2.columns):
            df2 = df2.loc[(df2['Check_number'] == chkno)]
            df_chk = df_chk.append(df2)

        return df_chk

    def rada_chk79_fn(df1, df2, chkno):
        df_chk = pd.DataFrame()
        df_chk1 = pd.DataFrame()
        df_chk2 = pd.DataFrame()

        if len(df1) > 0:
            df_chk1 = df1.copy()
            df_chk1['Check_number'] = chkno
        if len(df2) > 0:
            df_chk2 = df2.copy()
            df_chk2['Check_number'] = chkno

        if (len(df_chk1) > 0):
            df_chk = df_chk.append(df_chk1)
        if (len(df_chk2) > 0):
            df_chk = df_chk.append(df_chk2)

        return df_chk

    def rada_chk19_fn(df2, df3, fld1, fld2, fld3, chkno, **kwargs):
        df_chk = pd.DataFrame()
        df_chk1 = pd.DataFrame()
        #v0.2
        df4 = pd.DataFrame()
        df5= pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        if len(df2) > 0:
            df_chk = df2.copy()
        if ((len(df_chk) > 0) & (len(df3) > 0)):
            df_chk = df_chk.loc[:, labellist1]
            df_chk = df_chk.loc[~((df_chk[fld2] == 'SetDataPoint') | (df_chk[fld2] == 'SetDataPointVisible')
                                  | (df_chk[fld2] == 'SetDynamicSearchList'))]
            df3[fld3] = df3[fld3].fillna('')
            df3['newfld1'] = df3[fld1] + df3[fld3]
            df_chk['newfld'] = df_chk[fld1] + df_chk[fld3]
            valuelist1 = unique_non_null(df_chk['newfld']).tolist()

            df_chk['count'] = 0
            f = {'newfld1': ','.join}
            # df3.to_excel(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS\output.xlsx')
            df3 = df3.groupby([fld1], as_index=False).agg(f)
            f = {'newfld': ','.join, 'VariableOID': ','.join, 'count': 'count'}
            df_chk = df_chk.groupby([fld1], as_index=False).agg(f)
            # df_chk = df_chk.loc[df_chk['count'] > 1]
            df_chk = df_chk.merge(df3, left_on=[fld1],
                                  right_on=[fld1], how='left',
                                  suffixes=['', '_'],
                                  indicator=True)
            df5 = df_chk.loc[df_chk['_merge'] == 'left_only']
            df4 = df_chk.loc[df_chk['_merge'] == 'both']
            df5 = df5.drop(['_merge'], axis=1)
            df4 = df4.drop(['_merge'], axis=1)
        if len(df4) >0 :
            df4['same_chkvar'] = df4.apply(lambda x: common_member(x['newfld'].split(','), x['newfld1'].split(',')), axis=1)
            # df_chk.loc[~(df_chk['newfld'].isin(df3['newfld'])), 'Check_number'] = chkno


        if (len(df4) > 0):
            df4.loc[df4['same_chkvar'].isna(), 'Check_number'] = chkno
            df_chk1 = df_chk1.append(df4)
        if (len(df5) > 0):
            df5['Check_number'] = chkno
            df_chk1 = df_chk1.append(df5)

        return df_chk1

    def rada_chk102_fn(df1, df2, df3, fld1, fld2, fld3, fld4, chkno, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        if ((len(df1) > 0) & (len(df2) > 0)):

            df1 = df1.loc[~(df1[fld3].isna()), labellist1]
            df1 = df1.loc[df1[fld3] == 0]
            df3 = df3.loc[df3[fld4] == True]
            df2 = df2.loc[df2[fld3].isna() & (df2[fld2].isin(df3[fld2]))]
        if (len(df1) > 0) & (len(df2) > 0):
            df1.loc[((df1[fld1].isin(df2[fld1]))), 'Check_number'] = chkno

        if (len(df1) > 0) & ('Check_number' in df1.columns):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk60_fn(df3, fld1, fld2, cond1, cond2, cond3, cond4, chkno, **kwargs):
        df_chk = pd.DataFrame()
        df1 = pd.DataFrame()
        df2 = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        if len(df3) > 0:
            df1 = df3.loc[:, labellist1]
            df2 = df1.loc[df1[fld2] == cond1]
            df2 = df2.loc[~df2[fld1].isna()]
            df2 = df2.loc[~df2[fld1].str.startswith(('RSG'))]
            if len(df2) >0 :
                df2.loc[~(df2[fld1].astype(str).str.contains(cond3)), 'Check_number'] = chkno
            df1 = df1.loc[df1[fld2] == cond2]
            df1 = df1.loc[~df1[fld1].isna()]
            df1 = df1.loc[~df1[fld1].str.startswith(('RSG'))]
            if  (len(df1) > 0):
                df1.loc[~(df1[fld1].astype(str).str.contains(cond4)), 'Check_number'] = chkno



        if (len(df1) > 0) & ('Check_number' in df1.columns):
            df1 = df1.loc[(df1['Check_number'] == chkno)]
            df_chk = df_chk.append(df1)
        if (len(df2) > 0) & ('Check_number' in df2.columns):
            df2 = df2.loc[(df2['Check_number'] == chkno)]
            df_chk = df_chk.append(df2)

        return df_chk

    def rada_chk125_fn(df1, fld1, fld2, fld3, fld4, chkno, reqstr1, reqstr2, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        start_date = ['Start', 'START', 'Start Date', 'START DATE']
        start_time = ['Start', 'START', 'Start Time', 'START TIME']
        end_date = ['End', 'END', 'End Date', 'END DATE']
        end_time = ['End', 'END', 'End Time', 'END TIME']
        if len(df1) > 0:
            df1 = df1.loc[~df1[fld2].isna(), labellist1]
            df1 = df1.loc[(df1[fld2].astype(str).str.contains('|'.join(reqstr1))) | (df1[fld2].astype(str).str.contains('|'.join(reqstr2)))]
            df1['start_date'] = 0
            df1['start_time'] = 0
            df1['end_date'] = 0
            df1['end_time'] = 0
            df1.loc[(df1[fld2].astype(str).str.contains('|'.join(reqstr1)) & df1[fld4].astype(str).str.contains('|'.join(start_date))),
                    'start_date'] = 1
            df1.loc[(df1[fld2].astype(str).str.contains('|'.join(reqstr2)) & df1[fld4].astype(str).str.contains('|'.join(start_time))),
                    'start_time'] = 1
            df1.loc[(df1[fld2].astype(str).str.contains('|'.join(reqstr1)) & df1[fld4].astype(str).str.contains('|'.join(end_date))),
                    'end_date'] = 1
            df1.loc[(df1[fld2].astype(str).str.contains('|'.join(reqstr2)) & df1[fld4].astype(str).str.contains('|'.join(end_time))),
                    'end_time'] = 1

            df1['DataFormat'].replace('nan', np.nan, regex=True, inplace=True)
            df1['PreText'].replace('nan', np.nan, regex=True, inplace=True)
            df1['DataFormat'] = df1['DataFormat'].fillna('')
            df1['PreText'] = df1['PreText'].fillna('')

            f = {'FieldOID': ','.join, 'DataFormat': ','.join, 'PreText': ','.join, 'start_date':'sum',
                 'start_time':'sum', 'end_date':'sum', 'end_time':'sum'}
            df1 = df1.groupby([fld3], as_index=False).agg(f)

            df1 = df1.loc[(df1['start_date'] < df1['start_time']) | (df1['end_date'] < df1['end_time']) ]


            if len(df1) > 0:
                df1['Check_number'] = chkno

        if (len(df1) > 0):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk



    def rada_chk126_fn(df1, fld1,  chkno, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        if (len(df1) > 0) :
            df1 = df1.loc[~df1[fld1].isna(), labellist1]
            df1['Check_number'] = chkno

        if (len(df1) > 0):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk


    def rada_chk124_fn(df1, df2, fld1, fld2, fld3, fld4, fld5, fld6, fld7, chkno, **kwargs):
        df_chk = pd.DataFrame()
        df3= pd.DataFrame()
        #v0.2
        labellist1 = kwargs.get('labellist1', [])
        labellist2 = kwargs.get('labellist2', [])
        if (len(df1) > 0) & (len(df2) > 0):
            df1 = df1.loc[~(df1[fld7].isna()) , labellist1]
            df3 = df1.loc[((df1[fld2] == 'LBFAST'))]
            df3['newfld1'] = df3[fld1] + df3[fld2]
            df2 = df2.loc[:, labellist2]

            if ((df1[fld4].count() > 0) & (df2[fld5].count() > 0)):
                df1 = df1.merge(df2, left_on=[fld4], right_on=[fld5],
                                            how='left', suffixes=['', '_'], indicator=True)
                wordlist = ['non-fasting']
                wordlist1 = ['fasting']
                if len(df1) > 0:
                    df1 = df1.loc[df1['_merge'] == 'both']
                    df1 = df1.drop(['_merge'], axis=1)
                    df2 = df1.loc[((~(df1[fld6].astype(str).str.contains('|'.join(wordlist))) &
                                    (df1[fld6].astype(str).str.contains('|'.join(wordlist1))) ))]
                    if len(df2) > 0:
                        df2['newfld2'] = df2[fld1].astype(str) + 'LBFAST'
                    # if len(df1) > 0:
                    #     df1['newfld1'] = df1[fld1].astype(str) + df1[fld2].astype(str)
                        # df1 = df1.loc[((df1[fld2] == 'LBFAST'))]

                    if (len(df2) > 0) & (len(df3) > 0):
                        df2.loc[~df2['newfld2'].isin(df3['newfld1']), 'Check_number'] = chkno
                    if (len(df3) > 0) & (len(df2) > 0):
                        df3.loc[~df3[fld1].isin(df2[fld1]), 'Check_number'] = chkno

        if (len(df3) > 0) & ('Check_number' in df3.columns):
            df3 = df3.loc[(df3['Check_number'] == chkno)]
            df_chk = df_chk.append(df3)
        if ((len(df2) > 0) & ('Check_number' in df2.columns)):
            df2 = df2.loc[(df2['Check_number'] == chkno)]
            df_chk = df_chk.append(df2)

        return df_chk

    def rada_chk123_fn(df3, fld1, fld2, cond1, cond2, condlist, chkno, **kwargs):
        df_chk = pd.DataFrame()
        df1 = pd.DataFrame()
        df2 = pd.DataFrame()
        # df_chk123 = rada_chk123_fn(df_als_checks, "CheckName", 'Infix', '_CF_', 'custom function', 123,
        #                            labellist1=["CheckName", 'Infix'])
        labellist1 = kwargs.get('labellist1', [])
        if len(df3) > 0:
            df1 = df3.loc[:, labellist1]
            df2 = df1.loc[df1[fld1].astype(str).str.contains(cond1, na=False)]
            if len(df2) >0 :
                df2.loc[~(df2[fld2].astype(str).str.contains(cond2, na=False)), 'Check_number'] = chkno
            df1 = df1.loc[~(df1[fld1].astype(str).str.contains('|'.join(condlist), na=False))]
            df1 = df1.loc[df1[fld2].astype(str).str.contains(cond2, na=False)]
            if len(df1) > 0:
                df1.loc[~(df1[fld1].astype(str).str.contains(cond1, na=False)), 'Check_number'] = chkno

        if (len(df1) > 0) & ('Check_number' in df1.columns):
            df1 = df1.loc[(df1['Check_number'] == chkno)]
            df_chk = df_chk.append(df1)
        if (len(df2) > 0) & ('Check_number' in df2.columns):
            df2 = df2.loc[(df2['Check_number'] == chkno)]
            df_chk = df_chk.append(df2)

        return df_chk

    def rada_chk123_1_fn(df3, fld1, fld2, cond1, cond2, cond3, chkno, **kwargs):
        df_chk = pd.DataFrame()
        df1 = pd.DataFrame()
        df2 = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        labellist2 = kwargs.get('labellist2', [])
        if len(df3) > 0:
            df1 = df3.loc[:, labellist1]
            df2 = df1.loc[df1[fld1].astype(str).str.contains(cond2, na=False)]
            if len(df2) > 0:
                df2.loc[~(df2[fld2] == cond3), 'Check_number'] = chkno
            df1 = df1.loc[df1[fld1].astype(str).str.contains(cond1, na=False)]
            if len(df1) > 0:
                df1.loc[~(df1[fld2].astype(str).str.contains('|'.join(labellist2), na=True)), 'Check_number'] = chkno

        if (len(df1) > 0)  & ('Check_number' in df1.columns):
            df1 = df1.loc[(df1['Check_number'] == chkno)]
            df_chk = df_chk.append(df1)
        if (len(df2) > 0)  & ('Check_number' in df2.columns):
            df2 = df2.loc[(df2['Check_number'] == chkno)]
            df_chk = df_chk.append(df2)

        return df_chk

    def rada_chk21_fn(df1, df2, fld1, fld2, chkno, **kwargs):
        df_chk = pd.DataFrame()
        value1 = kwargs.get('value1', None)
        value2 = kwargs.get('value2', None)
        value3 = kwargs.get('value3', None)
        fld3 = kwargs.get('fld3', None)
        fld4 = kwargs.get('fld4', None)
        labellist1 = kwargs.get('labellist1', [])
        if len(df1) > 0:
            df_chk = df1.copy()
        if ((len(df_chk) > 0) & (len(df2) > 0)):
            df_chk = df_chk.loc[:, labellist1]
            df2 = df2.loc[(df2[fld2] == value1)]
            df_chk = df_chk.loc[(df_chk[fld3] == value2) | (df_chk[fld3] == value3)]
            # df_chk1 = df_chk1.loc[df_chk1[fld1].isin(df2[fld1])]
            if len(df_chk) >0:
                df_chk.loc[df_chk[fld1].isin(df2[fld1]) == True, ('Check_number')] = chkno
        # df_chk21 = rada_chk21_fn(df_checksteps, df_flds, "VariableOID", "ControlType", 21, fld3='CheckFunction',
        #                          fld4='CheckName', value1='CheckBox', value2='IsEmpty', value3='IsNotEmpty',
        #                          labellist1=["CheckName", "VariableOID", "FormOID", 'CheckFunction'])
        if (len(df_chk) > 0)  & ('Check_number' in df_chk.columns):
            df_chk = df_chk.loc[(df_chk['Check_number'] == chkno)]

        return df_chk

    def rada_chk32_fn(df1, df2, fld1, fld2, fld3, fld4, fld5, fld6, chkno, **kwargs):
        df_chk = pd.DataFrame()
        value1 = kwargs.get('value1', None)
        labellist1 = kwargs.get('labellist1', [])
        valuelist1 = kwargs.get('valuelist1', [])
        valuelist2 = kwargs.get('valuelist2', [])
        valuelist3 = kwargs.get('valuelist3', [])

        if ((len(df1) > 0) & (len(df2) > 0)):
            df1 = df1.loc[:, labellist1]
            df1 = df1.loc[~(df1[fld1].isin(df2[fld1]))]
            if len(df1) >0:
                df1 = df1.loc[(df1[fld3].astype(str).str.contains('|'.join(valuelist1), na=False))]
            if len(df1) >0:
                df1 = df1.loc[~(df1[fld5].astype(str).str.contains('|'.join(valuelist3), na=False))]
            if len(df1) >0:
               df1 = df1.loc[~(df1[fld4].astype(str).str.contains('|'.join(valuelist2), na=False))]
            if len(df1) >0:
                df1 = df1.loc[~(df1[fld6].astype(str).str.contains('|'.join(valuelist2), na=False))]
            if len(df1) >0:
                df1.loc[((df1[fld2] == value1)) == True, ('Check_number')] = chkno

        if (len(df1) > 0) & ('Check_number' in df1.columns):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk65_fn(df1, fld1, fld2, fld3, fld4, fld5, fld6, chkno, **kwargs):
        df_chk = pd.DataFrame()
        value1 = kwargs.get('value1', None)
        labellist1 = kwargs.get('labellist1', [])
        valuelist1 = kwargs.get('valuelist1', [])

        if len(df1) > 0:
            df1 = df1.loc[~(df1[fld1].astype(str).str.contains('|'.join(valuelist1), na=False)), labellist1]
            if len(df1) > 0:
                df1.loc[((df1[fld2] == value1) | (df1[fld3] == value1) | (df1[fld4] == value1)
                     | (df1[fld5] == value1) | (df1[fld6] == value1)),
                    'Check_number'] = chkno

        if (len(df1) > 0)  & ('Check_number' in df1.columns):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk33_fn(df1, df2, fld1, fld2, fld3, chkno, **kwargs):
        df_chk = pd.DataFrame()
        value1 = kwargs.get('value1', None)
        labellist1 = kwargs.get('labellist1', [])
        valuelist1 = kwargs.get('valuelist1', [])

        if ((len(df1) > 0) & (len(df2) > 0)):
            df1 = df1.loc[:, labellist1]
            df1 = df1.loc[~(df1[fld1].isin(df2[fld1]))]
            if len(df1) > 0:
                df1 = df1.loc[~(df1[fld3].astype(str).str.contains('|'.join(valuelist1), na=False))]
            if len(df1) > 0:
                df1.loc[((df1[fld2] == value1)) == True, ('Check_number')] = chkno

        if (len(df1) > 0)  & ('Check_number' in df1.columns):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk71_fn(df1, fld1, fld2, fld3, chkno, **kwargs):
        df_chk = pd.DataFrame()
        value1 = kwargs.get('value1', None)
        labellist1 = kwargs.get('labellist1', [])
        valuelist1 = kwargs.get('valuelist1', [])

        if len(df1) > 0:
            df1 = df1.loc[df1[fld1] == value1, labellist1]
        if len(df1) > 0:
            df1.loc[((df1[fld2].astype(str).str.contains('|'.join(valuelist1), na=False))
                     | (df1[fld3].astype(str).str.contains('|'.join(valuelist1), na=False))), ('Check_number')] = chkno

        if (len(df1) > 0) & ('Check_number' in df1.columns):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk81_fn(df1, df2, df3, df4, fld1, fld2, fld3, chkno, **kwargs):
        df_chk = pd.DataFrame()
        df5 = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        valuelist1 = kwargs.get('valuelist1', [])

        if len(df1) > 0:
            df1 = df1.loc[:, labellist1]
            df1 = df1.loc[df1[fld1].astype(str).str.contains('|'.join(valuelist1), na=False)]
            df2 = df2.loc[~(df2[fld2].isna())]
            df3 = df3.loc[~(df3[fld2].isna())]
            df5 = df1.loc[((~(df1[fld2].isin(df2[fld2])) & (~(df1[fld2].isin(df3[fld2])))))]
            # df6 = df1.loc[~(df1[fld2].isin(df3[fld2]))]
            # df1.loc[~(df1[fld2].str.contains('|'.join(df2[fld2]), na=False)), 'Check_number'] = chkno
            # df4 = df1.merge(df3, left_on=[fld2], right_on=[fld2], how='left',
            #                             suffixes=['', '_'], indicator=True)
            # if len(df4) > 0:
            #     df4.loc[df4['_merge'] == 'left_only', 'Check_number'] = chkno
            # df1.loc[df1[fld2].str.contains('|'.join(valuelist1), na=False), 'Check_number'] = chkno

        if (len(df5) > 0):
            df5.loc[~(df5[fld2].isna()), 'Check_number'] = chkno
            df_chk = df_chk.append(df5)
        # if (len(df6) > 0) :
        #     df6['Check_number'] = chkno
        #     df_chk = df_chk.append(df6)

        return df_chk

    #

    def rada_chk25_fn(df1, df2, fld1, fld2, chkno, **kwargs):
        df_chk = pd.DataFrame()
        df_chk1 = pd.DataFrame()
        df_chk2 = pd.DataFrame()
        #v0.2
        df_chk3 = pd.DataFrame()
        value1 = kwargs.get('value1', None)
        fld3 = kwargs.get('fld3', None)
        fld4 = kwargs.get('fld4', None)
        fld5 = kwargs.get('fld5', None)
        labellist1 = kwargs.get('labellist1', [])

        if len(df1) > 0:
            df1 = df1.loc[:, labellist1]
            df_chk1 = df1.loc[(df1[fld2] == value1)]
            # df_chk1.sort_values(by=[fld1, fld4], inplace=True)
            # df_chk2 = df_chk1[df_chk1.duplicated([fld1, fld4]) == True]
            # fill null values with '' to not lose them during groupby
            # groupby all columns and calculate the length of the resulting groups
            # rename the series obtained with groupby to "group_count"
            # reset the index to get a DataFrame
            # replace '' with np.nan (this reverts our first operation)
            # sort DataFrame by "group_count" descending
            # df_chk2 = df_chk1.fillna('').groupby([fld1, fld4]).apply(len).rename('group_count').reset_index()\
            #     .replace('', np.nan).sort_values(by=['group_count'], ascending=False).a
            f = {'CheckName': ','.join, 'ActionOptions':'count'}
            df_chk2 = df_chk1.groupby([fld1, fld2], as_index=False).agg(f)

            f = {'CheckName': ','.join, 'ActionType':'count'}
            df_chk3 = df_chk1.groupby([fld1, fld3], as_index=False).agg(f)

            if (len(df2) > 0) & (len(df_chk1) > 0) & (fld3 in df_chk1.columns) & (fld5 in df2.columns):
                df_chk1.loc[(~(df_chk1[fld3].isin(df2[fld5]))) == True, 'Check_number'] = chkno
                # df_chk1.to_excel(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS\output_radachk1' + str(chkno) + '.xlsx')


        if (len(df_chk1) > 0) & ('Check_number' in df_chk1.columns):
            df_chk1 = df_chk1.loc[(df_chk1['Check_number'] == chkno)]
            df_chk = df_chk.append(df_chk1)

        if (len(df_chk2) > 0) :
            df_chk2 = df_chk2.loc[df_chk2['ActionOptions'] > 1]

            if (len(df_chk2) > 0):
                df_chk2[('Check_number')] = chkno
                df_chk = df_chk.append(df_chk2)

        if (len(df_chk3) > 0):
            df_chk3 = df_chk3.loc[df_chk3['ActionType'] > 1]
            if (len(df_chk3) > 0):
                df_chk3[('Check_number')] = chkno
                df_chk = df_chk.append(df_chk3)

        return df_chk

    def rada_chk27_fn(df1, fld1, chkno, reqstr):
        df_chk = pd.DataFrame()

        if len(df1) > 0:
            df_chk = df1.copy()
            df_chk.loc[df_chk[fld1].astype(str).str.contains(reqstr, na=False), 'Check_number'] = chkno

        if (len(df_chk) > 0):
            df_chk = df_chk.loc[(df_chk['Check_number'] == chkno)]

        return df_chk

    def rada_chk28_fn(df1, fld1, fld2, chkno, reqstrlist):
        df_chk = pd.DataFrame()

        if len(df1) > 0:
            df_chk = df1.copy()
            df_chk.loc[df_chk[fld1].astype(str).str.contains('|'.join(reqstrlist), na=False), 'Check_number'] = chkno

        if (len(df_chk) > 0) & ('Check_number' in df_chk.columns):
            df_chk = df_chk.loc[(df_chk['Check_number'] == chkno)]
            f = {fld2: ','.join}
            df_chk = df_chk.groupby(['Check_number'], as_index=False).agg(f)

        return df_chk


    def rada_chk30_fn(df1, fld1, chkno):
        df_chk = pd.DataFrame()

        if len(df1) > 0:
            df_chk = df1.loc[~df1[fld1].isna()]
        if len(df_chk) > 0:
            df_chk.loc[df_chk[fld1].apply(len) > 8000, 'Check_number'] = chkno

        if (len(df_chk) > 0)  & ('Check_number' in df_chk.columns):
            df_chk = df_chk.loc[(df_chk['Check_number'] == chkno)]
            f = {'FunctionName': '|'.join}
            df_chk = df_chk.groupby(['Check_number'], as_index=False).agg(f)

        return df_chk

    def rada_chk31_fn(df1, fld1, value1, chkno, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        if len(df1) > 0:
            df_chk = df1.copy()
        if len(df_chk) > 0:
            df_chk = df_chk.loc[:, labellist1]
            df_chk.loc[df_chk[fld1].astype(str).str.contains(value1), ('Check_number')] = chkno

        if (len(df_chk) > 0)  & ('Check_number' in df_chk.columns):
            df_chk = df_chk.loc[(df_chk['Check_number'] == chkno)]

        return df_chk

    #
    #
    def rada_chk1_fn(df1, fld1, fld2, chkno, **kwargs):
        df_chk = pd.DataFrame()
        newfld = kwargs.get('newfld', None)
        labellist1 = kwargs.get('labellist1', [])

        if len(df1) > 0:
            df_chk = df1.copy()

            # if cond == '~':
            df_chk = df_chk.loc[:, labellist1]
            df_chk = df_chk.loc[((df_chk[fld2] == 'LongText') | (df_chk[fld2] == 'Text'))]
            df_chk[newfld] = df_chk[fld1].str.replace('$', '')
            df_chk[newfld] = df_chk[newfld].str.replace('+', '')
            df_chk.dropna(subset=[newfld], inplace=True)

        if (len(df_chk) > 0):
            df_chk.loc[(((df_chk[newfld].astype(float) <= 40) & (df_chk[fld2] == 'LongText')) |
                        ((df_chk[newfld].astype(float) > 40) & (df_chk[fld2] == 'Text'))) == True,
                       ('Check_number')] = chkno

        if (len(df_chk) > 0) & ('Check_number' in df_chk.columns):
            df_chk = df_chk.loc[(df_chk['Check_number'] == chkno)]

        return df_chk

    def rada_chk2_fn(df1, fld1, fld2, fld3, chkno, length, **kwargs):
        df_chk = pd.DataFrame()
        newfld = kwargs.get('newfld', None)
        labellist1 = kwargs.get('labellist1', [])
        fldlist = kwargs.get('fldlist', [])

        if len(df1) > 0:
            df_chk = df1.loc[:, labellist1]
            df_chk = df_chk.loc[~(df_chk[fld3].astype(str).str.contains('|'.join(fldlist)))]
            # df_chk = df_chk.loc[((df_chk[fld2] == 'LongText') | (df_chk[fld2] == 'Text'))]
            df_chk[newfld] = df_chk[fld1].str.replace('$', '').str.strip()
            df_chk[newfld] = df_chk[newfld].str.replace('+', '').str.strip()
            df_chk.dropna(subset=[newfld], inplace=True)
            df_chk = df_chk.loc[df_chk[newfld].str.len() > 2]
            # df_chk.loc[df_chk[newfld].str.contains('.'), newfld] = df_chk[newfld].str.rsplit('.', 1).str.get(0)
            df_chk = df_chk.loc[df_chk[newfld].astype(str).str.contains('[0-9]')]
        if (len(df_chk) > 0):
            df_chk.loc[(df_chk[newfld].astype(float) > length) == True, 'Check_number'] = chkno

        if (len(df_chk) > 0) & ('Check_number' in df_chk.columns):
            df_chk = df_chk.loc[(df_chk['Check_number'] == chkno)]

        return df_chk

    def rada_chk3_fn(df1, df2, fld1, fld2, chkno, length, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        labellist2 = kwargs.get('labellist2', [])
        formlist = kwargs.get('formlist', [])

        if len(df1) > 0:

            df_chk = df1.loc[~df1[fld1].isin(df2[fld2]), labellist1]
            df_chk = df_chk.loc[~df_chk[fld1].isin(labellist2)]
            df_chk = df_chk.loc[~df_chk[fld1].astype(str).str.contains('|'.join(formlist))]
        if (len(df_chk) > 0):
            df_chk.loc[df_chk[fld1].apply(len) > length, 'Check_number'] = chkno

        if (len(df_chk) > 0) & ('Check_number' in df_chk.columns):
            df_chk = df_chk.loc[(df_chk['Check_number'] == chkno)]
            f = {fld1: '|'.join, 'FormOID': '|'.join}
            df_chk = df_chk.groupby(['Check_number'], as_index=False).agg(f)

        return df_chk

    def rada_chk12_fn(df1, fld1, fld2, fld3, chkno, reqstr1, reqstr2, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        formlist = kwargs.get('formlist', [])

        if len(df1) > 0:
            df1 = df1.loc[:, labellist1]
            df1 = df1.loc[~(df1[fld2].isna())]
            df1 = df1.loc[df1[fld2].astype(str).str.contains('|'.join(reqstr2))]
            df1 = df1.loc[~df1[fld3].astype(str).str.contains('|'.join(formlist))]
            if len(df1) > 0:
                df1 = df1.loc[~(df1[fld1].isna())]
            if len(df1) > 0:
                df1.loc[~(df1[fld1].astype(str).str.contains(reqstr1)), 'Check_number'] = chkno

        if (len(df1) > 0) & ('Check_number' in df1.columns):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk13_fn(df1, fld1, fld2, chkno, reqstr1, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])

        if len(df1) > 0:
            df1 = df1.loc[:, labellist1]
            df1 = df1.loc[df1[fld2].isna()]
            if len(df1) > 0:
                # df1 = df1.loc[~(df1[fld1].isna())]
                df1.loc[~(df1[fld1].astype(str).str.contains(reqstr1)), 'Check_number'] = chkno

        if (len(df1) > 0) & ('Check_number' in df1.columns):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk24_fn(df1, fld1, fld2, fld3, chkno, reqstr1, reqstr2, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])

        if len(df1) > 0:
            df1 = df1.loc[:, labellist1]
            df1 = df1.loc[df1[fld2].isna()]
        if len(df1) > 0:
            df1 = df1.loc[df1[fld1].astype(str).str.contains(reqstr1)]
            if len(df1) > 0:
                df1.loc[df1[fld3] != reqstr2, 'Check_number'] = chkno

        if (len(df1) > 0) & ('Check_number' in df1.columns):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    #
    def rada_chk39_fn(df1, fld1, chkno, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])

        if len(df1) > 0:
            df1 = df1.loc[:, labellist1]
            df1.loc[df1[fld1].isna(), 'Check_number'] = chkno

        if (len(df1) > 0) & ('Check_number' in df1.columns):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk51_fn(df1, fld1, chkno, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])

        if len(df1) > 0:
            df1 = df1.loc[:, labellist1]
            df1.loc[~(df1[fld1].isna()), 'Check_number'] = chkno


        if (len(df1) > 0) & ('Check_number' in df1.columns):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk69_fn(df1, fld1, fld2, fld3, fld4, fld5, chkno, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])

        if len(df1) > 0:
            df1 = df1.loc[:, labellist1]
            df1.loc[((~(df1[fld1].isna())) | (~(df1[fld2].isna())) | (~(df1[fld3].isna())) | (~(df1[fld4].isna())) |
                     (~(df1[fld5].isna()))), 'Check_number'] = chkno

        if (len(df1) > 0) & ('Check_number' in df1.columns):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk64_fn(df1, fld1, chkno, cond, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])

        if len(df1) > 0:
            df1 = df1.loc[:, labellist1]
            df1.loc[((df1[fld1] == cond) | (df1[fld1].isna())), 'Check_number'] = chkno

        if (len(df1) > 0) & ('Check_number' in df1.columns):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk80_fn(df1, fld1, chkno, cond, sdvvalue, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        if sdvvalue == 1:
            if len(df1) > 0:
                df1 = df1.loc[:, labellist1]
                df1.loc[((df1[fld1] == cond) | (df1[fld1].isna())), 'Check_number'] = chkno

            if (len(df1) > 0) & ('Check_number' in df1.columns):
                df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk36_fn(df1, fld1, fld2, chkno, cond, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])

        if len(df1) > 0:
            df1 = df1.loc[~(df1[fld2].isna()), labellist1]
        if len(df1) > 0:
            df1.loc[df1[fld1] == cond, 'Check_number'] = chkno

        if (len(df1) > 0) & ('Check_number' in df1.columns):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk45_fn(df1, fld1, chkno, cond):
        df_chk = pd.DataFrame()
        if len(df1) > 0:
            if ((df1[fld1]==cond).any() == False):
                df1['Check_number'] = chkno
                f = {fld1: ','.join}
                df_chk = df1.groupby(['Check_number'], as_index=False).agg(f)

        return df_chk

    def rada_chk38_fn(df1, fld1, fld2, fld3, chkno, cond, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        valuelist1 = kwargs.get('valuelist1', [])
        value1 = kwargs.get('value1', None)

        if len(df1) > 0:
            df1 = df1.loc[(~(df1[fld2].isna())), labellist1]
            # df1 = df1.loc[df1[fld2].str.contains(value1)]
        if len(df1) > 0:
            df1 = df1.loc[((df1[fld2].astype(str).str.contains(value1, na=False)) & (df1[fld3].astype(str).str.contains('|'.join(valuelist1), na=False)))]
        if len(df1) > 0:
            df1.loc[df1[fld1] == cond, 'Check_number'] = chkno

        if (len(df1) > 0) & ('Check_number' in df1.columns):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk56_fn(df1, fld1, fld2, chkno, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        value1 = kwargs.get('value1', None)

        if len(df1) > 0:
            df1 = df1.loc[(~(df1[fld1].isna())), labellist1]
            if len(df1) > 0:
                df1.loc[df1[fld2].apply(len) >= 20, 'Check_number'] = chkno

        if (len(df1) > 0) & ('Check_number' in df1.columns):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk57_fn(df1, fld1, fld2, chkno, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        value1 = kwargs.get('value1', None)

        if len(df1) > 0:
            df1 = df1.loc[(~(df1[fld1].isna())), labellist1]
            df2 = df1
            if len(df1) > 0:
                df1.loc[~df1[fld2].str.startswith(('$')), 'Check_number'] = chkno

        if (len(df1) > 0) & ('Check_number' in df1.columns):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk68_fn(df1, fld1, fld2, chkno, cond, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        valuelist1 = kwargs.get('valuelist1', [])

        if len(df1) > 0:
            df1 = df1.loc[df1[fld1] == cond, labellist1]
            if len(df1) > 0:
                df1.loc[~(df1[fld2].astype(str).str.contains('|'.join(valuelist1), na=False)), 'Check_number'] = chkno

        if (len(df1) > 0) & ('Check_number' in df1.columns):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk42_fn(df1, df2, fld1, fld2, fld3, chkno, cond, cond2, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])

        if len(df1) > 0:
            df1 = df1.loc[(df1[fld1] == cond), labellist1]
            df2 = df2.loc[df2[fld3] == cond2]
            if (len(df1) > 0) & (len(df2) > 0):
                df1.loc[~df1[fld2].isin(df2[fld2]), 'Check_number'] = chkno
            # df1.loc[df1[fld2].isin(df2[fld2]), 'Check_number'] = chkno

        if (len(df1) > 0) & ('Check_number' in df1.columns):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk34_fn(df1, fld1, fld2, chkno, cond, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])

        if len(df1) > 0:
            df1 = df1.loc[:, labellist1]
            df1.loc[((df1[fld1] == cond) & (df1[fld2] == cond)), 'Check_number'] = chkno

        if (len(df1) > 0) & ('Check_number' in df1.columns):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk35_fn(df1, df2, df3, fld1, fld2, fld3, fld4, chkno, **kwargs):
        #v0.2
        df_chk = pd.DataFrame()
        df_chk2 = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        formlist1 = kwargs.get('formlist1', [])

        if len(df1) > 0:

            df1 = df1.loc[~(df1[fld3].isin(df2[fld3])), labellist1]
        if len(df1) > 0:
            df1.loc[((~(df1[fld1].isna())) & (df1[fld2].isna())), 'Check_number'] = chkno
            df3 = df3.loc[~(df3[fld4].isna())]
            df2 = df1.loc[~df1[fld2].isna()]
            df_chk2 = df2.loc[df2[fld2].astype(str).str.contains('|'.join(df3[fld4]), na=False)]

        if (len(df1) > 0) & ('Check_number' in df1.columns):
            df1 = df1.loc[(df1['Check_number'] == chkno)]
            df_chk = df_chk.append(df1)

        if (len(df_chk2) > 0) & ('Check_number' in df_chk2.columns):
            df_chk2 = df_chk2.loc[(df_chk2['Check_number'] == chkno)]
            df_chk = df_chk.append(df_chk2)

        return df_chk

    def rada_chk46_fn(df1, df2, fld1, fld2, newfld1, chkno, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])

        if len(df1) > 0:
            df1 = df1.loc[:, labellist1]
            df1.loc[((df1[fld1].str.count('<i>') != df1[fld1].str.count('</i>')) |
                     (df1[fld1].str.count('<b>') != df1[fld1].str.count('</b>')) |
                     (df1[fld1].str.count('<sup>') != df1[fld1].str.count('</sup>')) |
                     (df1[fld1].str.count('<u>') != df1[fld1].str.count('</u>')) |
                     (df1[fld1].str.count('<sub>') != df1[fld1].str.count('</sub>'))), 'Check_number'] = chkno
            df1[newfld1] = df1[fld1].str.replace('<br/>|<i>|</i>|<b>|</b>|<u>|</u>|<sup>|</sup>|<sub>|</sub>', '')
            if (len(df2) > 0):
                df2 = df2.loc[~(df2[fld2].isna())]
            if (len(df2) > 0):
                df1.loc[df1[newfld1].astype(str).str.contains('|'.join(df2[fld2]), na=False), 'Check_number'] = chkno

        if (len(df1) > 0) & ('Check_number' in df1.columns):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk59_fn(df1, fld1, fld2, chkno, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])

        if len(df1) > 0:
            df1 = df1.loc[~(df1[fld2].isna()), labellist1]
            # df1[newfld1] = df1[fld3].str.rsplit('_', 1).str.get(0)
        if len(df1) > 0:
            df1.loc[~(df1[fld2].astype(str).str.contains('|'.join(df1[fld1]), na=False)), 'Check_number'] = chkno

        if (len(df1) > 0) & ('Check_number' in df1.columns):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk78_fn(df1, fld1, fld2, chkno, cond, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])

        if len(df1) > 0:
            df1 = df1.loc[:, labellist1]
            df1 = df1.loc[~(df1[fld2].isna())]
        if len(df1) > 0:
            df1.loc[((df1[fld1] == cond) | (df1[fld1].isna())), 'Check_number'] = chkno

        if (len(df1) > 0) & ('Check_number' in df1.columns):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk44_fn(df1, fld1, fld2, chkno, cond, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])

        if len(df1) > 0:
            df1 = df1.loc[~(df1[fld2].isna()), labellist1]
        if len(df1) > 0:
            df1.loc[((df1[fld1] == cond) & (df1[fld2].astype(str).str.contains('\|'))), 'Check_number'] = chkno

        if (len(df1) > 0) & ('Check_number' in df1.columns):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk73_fn(df1, fld1, fld2, fld3, chkno, cond, searchval1, **kwargs):
        df_chk = pd.DataFrame()
        df2 = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])

        if len(df1) > 0:
            # df1 = df1.loc[:, labellist1]
            df1 = df1.loc[(~df1[fld3].isna()) == True]

            # df1 = df1.loc[(df1[fld3].str.contains('^LBL')) == True]
            # df1 = df1.loc[(~((df1[fld3].str.replace(r'\d+', '') == 'LBL'))) == True]
            df1 = df1.loc[df1[fld1] == cond, labellist1]
            # df1 = df1.loc[~(df1[fld2].isna())]
            df2 = df1.loc[~(df1[fld2].isna())]
            if (len(df1) > 0) :
                df1.loc[df1[fld2].isna(), 'Check_number'] = chkno
            if (len(df2) > 0) :
                df2.loc[~(df2[fld2].astype(str).str.contains(searchval1)), 'Check_number'] = chkno

        if (len(df1) > 0) & ('Check_number' in df1.columns):
            df1 = df1.loc[(df1['Check_number'] == chkno)]
            df_chk = df_chk.append(df1)
        if (len(df2) > 0) & ('Check_number' in df2.columns):
            df2 = df2.loc[(df2['Check_number'] == chkno)]
            df_chk = df_chk.append(df2)

        return df_chk

    def rada_chk72_fn(df1, fld1, fld2, chkno, cond, searchval1, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])

        if len(df1) > 0:
            # df1 = df1.loc[:, labellist1]
            df1 = df1.loc[df1[fld1] == cond, labellist1]
            df1 = df1.loc[~(df1[fld2].isna())]
            if (len(df1) > 0):
                df1.loc[df1[fld2].astype(str).str.contains(searchval1), 'Check_number'] = chkno

        if (len(df1) > 0) & ('Check_number' in df1.columns):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk74_fn(df1, fld1, fld2, fld3, chkno, cond, searchval1, **kwargs):

        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])

        if len(df1) > 0:
            df1 = df1.loc[df1[fld3].isna(), labellist1]
            df1 = df1.loc[df1[fld2].astype(str).str.contains(searchval1)]
            if (len(df1) > 0):
                df1.loc[df1[fld1] == cond, 'Check_number'] = chkno

        if (len(df1) > 0)  & ('Check_number' in df1.columns):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk14_fn(df1, fld1, fld2, fld3, fld4, fld5, fld6, fld7, fld8, fld9, fld10, fld11, fld12, fld13, fld14,
                      fld15,
                      fld16, fld17, fld18, fld19, chkno, cond1, cond2, searchval1, **kwargs):

        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])


        # FieldOID'fld1	 'VariableOID'fld2	 'DataFormat'fld3	 'DataDictionaryName'fld4	'CodingDictionary'fld5
        # 'FixedUnit'fld6	 'HeaderText'fld7	 'SourceDocument'fld8	 'DefaultValue'fld9	 'SASLabel'fld10
        # 'IsRequired'fld11	'QueryFutureDate'fld12	 'IsVisible'fld13	 'AnalyteName'fld14
        # 'IsClinicalSignificance'fld15	'QueryNonConformance'fld16	 'ViewRestrictions'fld17
        # 'EntryRestrictions'fld18
        # 'ControlType'fld19
        if len(df1) > 0:
            df1 = df1.loc[df1[fld2].isna(), labellist1]
            df1 = df1.loc[df1[fld1].astype(str).str.contains(searchval1)]
            if len(df1) > 0:
                df1.loc[((~(df1[fld3].isna())) | (~(df1[fld4].isna())) | (~(df1[fld5].isna())) | (~(df1[fld6].isna())) |
                     (~(df1[fld7].isna())) | (df1[fld8] == cond1) | (~(df1[fld9].isna())) | (~(df1[fld10].isna())) |
                     (df1[fld11] == cond1) | (df1[fld12] == cond1) | (df1[fld13] == cond2) | (
                         ~(df1[fld14].isna())) |
                     (df1[fld15] == cond1) | (df1[fld16] == cond1) | (~(df1[fld17].isna())) | (
                         ~(df1[fld18].isna())) | (
                         ~(df1[fld19] == 'Text'))),
                    'Check_number'] = chkno

        if (len(df1) > 0) & ('Check_number' in df1.columns):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    #

    def rada_chk4_fn(df1, fld1, fld2, chkno, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])

        if len(df1) > 0:
            df_chk = df1.loc[:, labellist1]
            if len(df_chk) > 0:
                df_chk.loc[((df_chk[fld2] == 'CheckBox') & (df_chk[fld1] != "1")), 'Check_number'] = chkno

        if (len(df_chk) > 0) & ('Check_number' in df_chk.columns):
            df_chk = df_chk.loc[(df_chk['Check_number'] == chkno)]

        return df_chk

    def rada_chk20_fn(df1, df2, fld1, chkno, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])

        if len(df1) > 0:
            df1 = df1.loc[:, labellist1]
        if len(df1) > 0:
            df1.loc[~(df1[fld1].isin(df2[fld1])), 'Check_number'] = chkno

        if (len(df1) > 0) & ('Check_number' in df1.columns):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk82_fn(df1, fld1, fld2, chkno, **kwargs):
        df_chk = pd.DataFrame()
        searchvalue1 = kwargs.get('searchvalue1', None)
        fld3 = kwargs.get('fld3', None)
        labellist1 = kwargs.get('labellist1', [])
        valuelist1 = kwargs.get('valuelist1', [])

        if len(df1) > 0:
            df_chk = df1.loc[:, labellist1]
            df_chk.loc[(((~(df_chk[fld1].astype(str).str.contains(searchvalue1))) &
                         ((df_chk[fld2].apply(lambda x: not (set(valuelist1).issubset(set(str(x).split(',')))))) |
                          (df_chk[fld3].apply(lambda x: not (set(valuelist1).issubset(set(str(x).split(',')))))))) |
                        (((df_chk[fld1].astype(str).str.contains(searchvalue1))) &
                         ((df_chk[fld2].astype(str).str.contains('|'.join(valuelist1))) |
                          (df_chk[fld3].astype(str).str.contains('|'.join(valuelist1)))))), 'Check_number'] = chkno

        if (len(df_chk) > 0) & ('Check_number' in df_chk.columns):
            df_chk = df_chk.loc[(df_chk['Check_number'] == chkno)]

        return df_chk

    def rada_chk84_fn(df1, fld1, fld2, chkno,  **kwargs):
        df_chk = pd.DataFrame()
        cond = kwargs.get('cond', None)
        frm1 = kwargs.get('frm1', None)
        frm2 = kwargs.get('frm2', None)
        searchvalue1 = kwargs.get('searchvalue1', None)
        searchvalue2 = kwargs.get('searchvalue2', None)
        newfld = kwargs.get('newfld', None)
        fld3 = kwargs.get('fld3', None)
        fld4 = kwargs.get('fld4', None)
        fld5 = kwargs.get('fld5', None)
        fld6 = kwargs.get('fld6', None)
        labellist1 = kwargs.get('labellist1', [])
        labellist2 = kwargs.get('labellist2', [])
        valuelist1 = kwargs.get('valuelist1', [])
        valuelist2 = kwargs.get('valuelist2', [])
        frmlist1 = kwargs.get('frmlist1', [])
        frmlist2 = kwargs.get('frmlist2', [])
        searchlist1 = kwargs.get('searchlist1', [])

        if len(df1) > 0:
            df_chk = df1.copy()
            valuelist3 = valuelist1 + valuelist2
            frmlist3 = frmlist1 + frmlist2
            # if cond == '~':
            df_chk = df_chk.loc[:, labellist1]
            df_chk.loc[(((df_chk[fld1].astype(str).str.contains('|'.join(searchlist1))) &
                         ((df_chk[fld2].apply(lambda x: not (set(valuelist3).issubset(set(str(x).split(','))))))) &
                         (~(df_chk[fld4].astype(str).str.contains('|'.join(frmlist3))))) |
                        ((~(df_chk[fld1].astype(str).str.contains('|'.join(searchlist1)))) &
                         ((df_chk[fld2].astype(str).str.contains('|'.join(valuelist3))) |
                          (df_chk[fld3].astype(str).str.contains('|'.join(valuelist3)))) &
                         (~(df_chk[fld4].astype(str).str.contains('|'.join(frmlist3))))) |
                        ((~(df_chk[fld1].astype(str).str.contains('|'.join(searchlist1)))) &
                         ((df_chk[fld2].astype(str).str.contains('|'.join(valuelist3))) |
                          (
                              (df_chk[fld3].apply(lambda x: not (set(valuelist3).issubset(set(str(x).split(',')))))))) &
                         ((df_chk[fld4].astype(str).str.contains(frm1)))) |
                        ((~(df_chk[fld1].astype(str).str.contains('|'.join(searchlist1)))) &
                         ((df_chk[fld3].astype(str).str.contains('|'.join(valuelist1))) |
                          (
                              (df_chk[fld2].apply(lambda x: not (set(valuelist1).issubset(set(str(x).split(',')))))))) &
                         ((df_chk[fld4].astype(str).str.contains(frm2)))) |
                        ((~(df_chk[fld1].astype(str).str.contains('|'.join(searchlist1)))) &
                         ((df_chk[fld2].astype(str).str.contains('|'.join(valuelist2))) |
                          (
                              (df_chk[fld3].apply(lambda x: not (set(valuelist2).issubset(set(str(x).split(',')))))))) &
                         ((df_chk[fld4].astype(str).str.contains(frm2))))), ('Check_number')] = chkno

        if (len(df_chk) > 0) & ('Check_number' in df_chk.columns):
            df_chk = df_chk.loc[(df_chk['Check_number'] == chkno)]

        return df_chk

    def rada_chk86_fn(df1, fld1, fld2, chkno, **kwargs):
        df_chk = pd.DataFrame()
        cond = kwargs.get('cond', None)
        frm1 = kwargs.get('frm1', None)
        frm2 = kwargs.get('frm2', None)
        searchvalue1 = kwargs.get('searchvalue1', None)
        searchvalue2 = kwargs.get('searchvalue2', None)
        newfld = kwargs.get('newfld', None)
        fld3 = kwargs.get('fld3', None)
        fld4 = kwargs.get('fld4', None)
        fld5 = kwargs.get('fld5', None)
        fld6 = kwargs.get('fld6', None)
        labellist1 = kwargs.get('labellist1', [])
        labellist2 = kwargs.get('labellist2', [])
        valuelist1 = kwargs.get('valuelist1', [])
        valuelist2 = kwargs.get('valuelist2', [])
        valuelist3 = kwargs.get('valuelist3', [])
        frmlist1 = kwargs.get('frmlist1', [])
        frmlist2 = kwargs.get('frmlist2', [])
        searchlist1 = kwargs.get('searchlist1', [])

        if len(df1) > 0:
            df_chk = df1.copy()
            valuelist4 = valuelist3 + valuelist1 + valuelist2
            valuelist5 = valuelist3 + valuelist2
            valuelist6 = valuelist3 + valuelist1
            # frmlist3 = frmlist1 + frmlist2
            # if cond == '~':
            df_chk = df_chk.loc[:, labellist1]
            df_chk.loc[((((df_chk[fld3].apply(lambda x: not (set(valuelist4).issubset(set(str(x).split(','))))))
                          | (df_chk[fld2].astype(str).str.contains('|'.join(valuelist3)))) &
                         (~(df_chk[fld4].astype(str).str.contains('|'.join(frmlist1))))) |
                        (((df_chk[fld2].astype(str).str.contains('|'.join(valuelist4))) |
                          (
                              (df_chk[fld3].apply(lambda x: not (set(valuelist5).issubset(set(str(x).split(',')))))))) &
                         ((df_chk[fld4].astype(str).str.contains(frm1)))) |
                        (((df_chk[fld3].astype(str).str.contains('|'.join(valuelist4))) |
                          (
                              (df_chk[fld2].apply(lambda x: not (set(valuelist6).issubset(set(str(x).split(',')))))))) &
                         ((df_chk[fld4].astype(str).str.contains(frm2))))), ('Check_number')] = chkno

        if (len(df_chk) > 0) & ('Check_number' in df_chk.columns):
            df_chk = df_chk.loc[(df_chk['Check_number'] == chkno)]

        return df_chk

    def rada_chk85_fn(df1, fld1, fld2, chkno, **kwargs):
        df_chk = pd.DataFrame()
        cond = kwargs.get('cond', None)
        frm1 = kwargs.get('frm1', None)
        frm2 = kwargs.get('frm2', None)
        searchvalue1 = kwargs.get('searchvalue1', None)
        searchvalue2 = kwargs.get('searchvalue2', None)
        newfld = kwargs.get('newfld', None)
        fld3 = kwargs.get('fld3', None)
        fld4 = kwargs.get('fld4', None)
        fld5 = kwargs.get('fld5', None)
        fld6 = kwargs.get('fld6', None)
        labellist1 = kwargs.get('labellist1', [])
        labellist2 = kwargs.get('labellist2', [])
        valuelist1 = kwargs.get('valuelist1', [])
        valuelist2 = kwargs.get('valuelist2', [])
        valuelist3 = kwargs.get('valuelist3', [])
        frmlist1 = kwargs.get('frmlist1', [])
        frmlist2 = kwargs.get('frmlist2', [])
        searchlist1 = kwargs.get('searchlist1', [])

        if len(df1) > 0:
            df_chk = df1.copy()
            valuelist3 = valuelist1 + valuelist2
            # valuelist5 = valuelist3  + valuelist2
            # valuelist6 = valuelist3  + valuelist1
            # frmlist3 = frmlist1 + frmlist2
            # if cond == '~':
            df_chk = df_chk.loc[:, labellist1]
            df_chk.loc[((((df_chk[fld2].astype(str).str.contains('|'.join(valuelist3)))
                          | (df_chk[fld3].astype(str).str.contains('|'.join(valuelist3)))) &
                         (~(df_chk[fld4].astype(str).str.contains(frm2)))) |
                        # (~(df_chk[fld4].str.contains(frm2)))) |
                        # ((((df_chk[fld2].str.contains('|'.join(valuelist3)))
                        #   | (df_chk[fld3].str.contains('|'.join(valuelist3))))) &
                        #  ((df_chk[fld4].str.contains(frm1)))) |
                        (((df_chk[fld3].astype(str).str.contains('|'.join(valuelist3))) |
                          (
                              (df_chk[fld2].apply(lambda x: not (set(valuelist2).issubset(set(str(x).split(',')))))))) &
                         ((df_chk[fld4].astype(str).str.contains(frm2))))), ('Check_number')] = chkno

        if (len(df_chk) > 0) & ('Check_number' in df_chk.columns):
            df_chk = df_chk.loc[(df_chk['Check_number'] == chkno)]

        return df_chk

    def rada_chk87_fn(df1, fld1, fld2, chkno, **kwargs):
        df_chk = pd.DataFrame()
        df2 = kwargs.get('df2', None)
        fld3 = kwargs.get('fld3', None)
        fld4 = kwargs.get('fld4', None)
        fld5 = kwargs.get('fld5', None)
        fld6 = kwargs.get('fld6', None)
        labellist1 = kwargs.get('labellist1', [])
        frmlist1 = kwargs.get('frmlist1', [])

        if ((len(df1) > 0) & (len(df2) > 0)):
            df_chk = df1.copy()
            df_chk2 = df2.copy()
            df_chk = df_chk.loc[:, labellist1]
            df_chk.loc[
                (((~(df_chk[fld5].isin(df_chk2[fld5]))) & (~(df_chk[fld5].astype(str).str.contains('|'.join(frmlist1))))) &
                 ((~(df_chk[fld2].isna())) & (~(df_chk[fld3].isna())))),
                ('Check_number')] = chkno

        if (len(df_chk) > 0) & ('Check_number' in df_chk.columns):
            df_chk = df_chk.loc[(df_chk['Check_number'] == chkno)]

        return df_chk

    def rada_chk93_fn(df1, fld1, fld2, fld7, chkno, **kwargs):
        df_chk = pd.DataFrame()
        df2 = kwargs.get('df2', None)
        df3 = kwargs.get('df3', None)
        fld3 = kwargs.get('fld3', None)
        fld4 = kwargs.get('fld4', None)
        fld5 = kwargs.get('fld5', None)
        fld6 = kwargs.get('fld6', None)
        labellist1 = kwargs.get('labellist1', [])
        frmlist1 = kwargs.get('frmlist1', [])

        if ((len(df1) > 0) & (len(df2) > 0) & (len(df3) > 0)):
            df_chk = df1.copy()
            df_chk = df_chk.loc[(~df_chk[fld7].isna()) == True]
            # df_chk = df_chk.loc[(~df_chk[fld7].isna()) == True]
            df_chk2 = df2.copy()
            df_chk3 = df3.copy()
            df_chk = df_chk.loc[:, labellist1]
            df_chk.loc[
                (( (~(df_chk[fld5].astype(str).str.contains('|'.join(frmlist1))))) &
                 ((df_chk[fld6].isin(df_chk3[fld6]))) & ((df_chk[fld2].isna() & df_chk[fld3].isna()))),
                ('Check_number')] = chkno

        if (len(df_chk) > 0) & ('Check_number' in df_chk.columns):
            df_chk = df_chk.loc[(df_chk['Check_number'] == chkno)]

        return df_chk

    def rada_chk88_fn(df1, fld1, fld2, chkno, **kwargs):
        df_chk = pd.DataFrame()
        df2 = kwargs.get('df2', None)
        fld3 = kwargs.get('fld3', None)
        fld4 = kwargs.get('fld4', None)
        fld5 = kwargs.get('fld5', None)
        fld6 = kwargs.get('fld6', None)
        labellist1 = kwargs.get('labellist1', [])
        frmlist1 = kwargs.get('frmlist1', [])
        if len(df1) > 0:
            df1 = df1.loc[:, labellist1]
            df1 = df1.loc[(~(df1[fld2].isna())) & (~(df1[fld3].isna()))]
            # df1 = df1.loc[(((~(df1[fld5].str.contains('|'.join(frmlist1))))) &
            #                        ((~(df1[fld2].isna())) & (~(df1[fld3].isna()))))]
            if len(df1) > 0:
                df1['same_role'] = df1.apply(lambda x: common_member(x[fld2].split(','), x[fld3].split(',')), axis=1)

        if (len(df1) > 0) & ('Check_number' in df1.columns):
            df1.loc[~df1['same_role'].isna(), 'Check_number'] = chkno
            df_chk = df_chk.append(df1)

        return df_chk

    def rada_chk94_fn(df1, fld1, chkno, **kwargs):
        df_chk = pd.DataFrame()
        df_chk1 = pd.DataFrame()
        df_chk2 = pd.DataFrame()
        df2 = kwargs.get('df2', None)
        value2 = kwargs.get('value2', None)
        labellist1 = kwargs.get('labellist1', [])
        labellist2 = kwargs.get('labellist2', [])
        valuelist1 = kwargs.get('valuelist1', [])
        valuelist2 = kwargs.get('valuelist2', [])

        if ((len(df1) > 0) & (len(df2) > 0)):
            df_chk1 = df1.loc[:, labellist1]
            df_chk2 = df2.loc[:, labellist2]
            df_chk1.loc[(((df_chk1[fld1].astype(str).str.contains('|'.join(valuelist1), na=False))) &
                         (~(df_chk1[fld1].astype(str).str.contains(value2, na=False)))), 'Check_number'] = chkno
            df_chk2.loc[(((df_chk2[fld1].astype(str).str.contains('|'.join(valuelist1), na=False))) &
                         (~(df_chk2[fld1].astype(str).str.contains(value2, na=False)))), ('Check_number')] = chkno

        if (len(df_chk1) > 0) & ('Check_number' in df_chk1.columns):
            df_chk1 = df_chk1.loc[(df_chk1['Check_number'] == chkno)]
            df_chk = df_chk.append(df_chk1)

        if (len(df_chk2) > 0) & ('Check_number' in df_chk2.columns):

            df_chk2 = df_chk2.loc[(df_chk2['Check_number'] == chkno)]
            df_chk = df_chk.append(df_chk2)

        if (len(df_chk) > 0) & ('Check_number' in df_chk.columns):
            df_chk = df_chk.loc[(df_chk['Check_number'] == chkno)]

        return df_chk

    def rada_chk98_fn(df1, df2, fld1, fld2, fld3, frm1, chkno, **kwargs):
        df_chk1 = pd.DataFrame()
        #v0.2
        df_chk = pd.DataFrame()
        valuelist2 = []
        labellist2 = []
        labellist1 = kwargs.get('labellist1', [])
        valuelist1 = kwargs.get('valuelist1', [])
        # valuelist2 = kwargs.get('valuelist2', [])

        if (len(df1) > 0) & (len(df2) > 0):
            df_chk = df1.loc[:, labellist1]
            if frm1 == 'DVG001':
                df_chk = df_chk.loc[df_chk[fld1].astype(str).str.contains('DVG001', na=False)]
            elif frm1 == 'DVG002':
                df_chk = df_chk.loc[df_chk[fld1].astype(str).str.contains('DVG002', na=False)]

            df_chk2 = df2.loc[:, labellist1]
            df_chk2 = df_chk2.loc[(df_chk2[fld1] == frm1) & (~df_chk2[fld3].isna())]
            # df_chk['intersection'] = [list(set(a).intersection(set(b))) for a, b in zip(df_chk[fld2], df_chk2[fld2])]
            # df_chk1 = df_chk.loc[((df_chk[fld2].isna()))]
            if ((len(df_chk) > 0) & (len(df_chk2) > 0)):
                df_chk = df_chk.merge(df_chk2, left_on=['FormOID', 'FieldOID'],
                                      right_on=['FormOID', 'FieldOID'], how='inner',
                                      suffixes=['', '_'],
                                      indicator=True)
                mycols = set(df_chk.columns)
                mycols.remove('_merge')
                df_chk = df_chk[mycols]
            if len(df_chk) > 0:
                newfld = fld2 + '_'
                df_chk[fld2].fillna("", inplace=True)
                df_chk[newfld].fillna("", inplace=True)

                # df_chk['same_role'] = df_chk.apply(lambda x: common_member(x[fld2].split(','), x[newfld].split(',')), axis=1)
                df_chk.loc[df_chk.apply(lambda x: not (set(x[fld2].split(',')) == set(x[newfld].split(','))), axis=1),'Check_number'] = chkno

                # df_flds_dom_ufrm = df_flds_dom_ufrm.rename(columns={"domain_ele": "domain_ele2"})
                # if len(df_chk) > 0:
                    # df_chk.loc[((df_chk[fld3] == df_chk2[fld3]) & ((df_chk[fld2].apply(lambda x: not (set(df_chk2[fld2].iloc[0].split(',')) ==
                    #                                                 (set(str(x).split(',')))))))),
                    #            'Check_number'] = chkno
                    # df_chk = df_chk.loc[(df_chk['Check_number'] == chkno)]
            # df_chk.loc[((df_chk[fld1] == frm1) & (df_chk[fld2].str.contains('|'.join(valuelist1)))),
            #     'Check_number'] = chkno


        if (len(df_chk) > 0) & ('Check_number' in df_chk.columns):
            df_chk1 = df_chk.loc[(df_chk['Check_number'] == chkno)]

        return df_chk1

    def rada_chk95_1_fn(df1, df2, fld1, fld2, frm1, chkno, **kwargs):
        df_chk1 = pd.DataFrame()
        #v0.2
        df_chk = pd.DataFrame()
        valuelist2 = []
        labellist2 = []
        labellist1 = kwargs.get('labellist1', [])
        valuelist1 = kwargs.get('valuelist1', [])
        valuelist2 = kwargs.get('valuelist2', [])
        valuestr2 = ''

        if (len(df1) > 0) & (len(df2) > 0) :
            df_chk = df1.loc[:, labellist1]
            df_chk = df_chk.loc[df_chk[fld1].astype(str).str.contains('DVG001', na=False)]
            df_chk2 = df2.loc[:, labellist1]
            df_chk2 = df_chk2.loc[(df_chk2[fld1] == frm1)]
            # df_chk['intersection'] = [list(set(a).intersection(set(b))) for a, b in zip(df_chk[fld2], df_chk2[fld2])]
            # df_chk1 = df_chk.loc[((df_chk[fld2].isna()))]
            if (len(df_chk2) > 0) & (len(df_chk) > 0) & (fld2 in df_chk2.columns):
                if (len(df_chk2[fld2]) > 0) & ((df_chk2[fld2].count()) > 0) :
                    df_chk.loc[(((df_chk[fld2].apply(lambda x: not (set(df_chk2[fld2].iloc[0].split(',')) ==
                                                                    (set(str(x).split(',')))))))),
                               'Check_number'] = chkno
            if (len(df_chk) > 0) & ('Check_number' in df_chk.columns):
                df_chk = df_chk.loc[(df_chk['Check_number'] == chkno)]


            # df_chk.loc[((df_chk[fld1] == frm1) & (df_chk[fld2].str.contains('|'.join(valuelist1)))),
            #     'Check_number'] = chkno

        if (len(df_chk) > 0) & ('Check_number' in df_chk.columns):
            df_chk1 = df_chk.loc[(df_chk['Check_number'] == chkno)]

        return df_chk1

    def rada_chk96_fn(df1, df2, fld1, fld2, frm1, chkno, **kwargs):
        df_chk1 = pd.DataFrame()
        #v0.2
        df_chk = pd.DataFrame()
        valuelist2 = []
        labellist2 = []
        labellist1 = kwargs.get('labellist1', [])
        valuelist1 = kwargs.get('valuelist1', [])
        valuelist2 = kwargs.get('valuelist2', [])
        valuestr2 = ''

        if (len(df1) > 0) & (len(df2) > 0) :
            df_chk = df1.loc[:, labellist1]
            df_chk = df_chk.loc[df_chk[fld1].astype(str).str.contains('DVG002', na=False)]
            df_chk2 = df2.loc[:, labellist1]
            df_chk2 = df_chk2.loc[(df_chk2[fld1] == frm1)]
            # df_chk['intersection'] = [list(set(a).intersection(set(b))) for a, b in zip(df_chk[fld2], df_chk2[fld2])]
            # df_chk1 = df_chk.loc[((df_chk[fld2].isna()))]
            if (len(df_chk) > 0) & (len(df_chk2) > 0):
                if (len(df_chk2[fld2]) > 0):
                    df_chk.loc[(((df_chk[fld2].apply(lambda x: not (set(df_chk2[fld2].iloc[0].split(',')) ==
                                                                (set(str(x).split(',')))))))),
                           'Check_number'] = chkno
            if (len(df_chk) > 0) & ('Check_number' in df_chk.columns):
                    df_chk = df_chk.loc[(df_chk['Check_number'] == chkno)]
            # df_chk.loc[((df_chk[fld1] == frm1) & (df_chk[fld2].str.contains('|'.join(valuelist1)))),
            #     'Check_number'] = chkno

        if (len(df_chk) > 0) & ('Check_number' in df_chk.columns):
            df_chk1 = df_chk.loc[(df_chk['Check_number'] == chkno)]

        return df_chk1

    '''
    Chk#89
    No View or Entry restrictions (All Fields)
    Batch Upload
    Rave Web Services (RWS)
    OL Role
    Outputs Locked
    Outputs Standard
    Coder Import Role
    Data Manager 
    Data Manager - restricted
    Safety
    Adjudication Chair - Except ADJ pages
    Adjudication Committee - Except  ADJ pages
    Adjudication Coordinator - Except  ADJ pages
    Specialty Data Provider - Except  SDP Pages 
    PDR All Data    - For C3i Submission PDR generation
    RPA CQM
    
    
    Chk#90
    
    No Entry Restrictions (All Fields)
    
    Read Only
    Read Only - All Sites
    Trial Management
    Trial Management - Restricted
    PDR Site Data
    Batch Upload
    Rave Web Services (RWS)
    OL Role
    Outputs Locked
    Outputs Standard
    Coder Import Role
    Data Manager 
    Data Manager - restricted
    Safety
    Adjudication Chair - Except ADJ pages
    Adjudication Committee - Except  ADJ pages
    Adjudication Coordinator - Except  ADJ pages
    Specialty Data Provider - Except  SDP Pages 
    PDR All Data    - For C3i Submission PDR generation
    RPA CQM
    '''

    def rada_chk89_fn(df1, fld2, chkno, **kwargs):
        df_chk = pd.DataFrame()
        df_chk1 = pd.DataFrame()
        df_chk3 = pd.DataFrame()
        df2 = kwargs.get('df2', None)
        newfld = kwargs.get('newfld', None)
        fld3 = kwargs.get('fld3', None)
        fld5 = kwargs.get('fld5', None)
        labellist1 = kwargs.get('labellist1', [])
        valuelist1 = kwargs.get('valuelist1', [])
        valuelist2 = kwargs.get('valuelist2', [])
        valuelist3 = kwargs.get('valuelist3', [])
        valuelist4 = kwargs.get('valuelist4', [])
        frmlist1 = kwargs.get('frmlist1', [])

        valuelist5 = valuelist1 + valuelist3 + valuelist4

        if ((len(df1) > 0) & (len(df2) > 0)):

            df_chk1 = df1.loc[:, labellist1]

            df_chk1 = df_chk1.loc[(~(df_chk1[fld5].astype(str).str.contains('|'.join(frmlist1))))]

            df_chk1.loc[df_chk1[fld2].astype(str).str.contains('|'.join(valuelist5), na=False), 'Check_number'] = chkno
            # df_chk1.loc[(((~(df_chk1[fld5].isin(df_chk2[fld5]))) & (
            #     ~(df_chk1[fld5].str.contains('|'.join(frmlist1))))) &
            #              (df_chk1[fld2].str.contains('|'.join(valuelist5)))),
            #             ('Check_number')] = chkno


        if (len(df_chk1) > 0) & ('Check_number' in df_chk1.columns):
            df_chk1 = df_chk1.loc[(df_chk1['Check_number'] == chkno)]
            df_chk = df_chk.append(df_chk1)
        if (len(df_chk3) > 0) & ('Check_number' in df_chk3.columns):
            df_chk3 = df_chk3.loc[(df_chk3['Check_number'] == chkno)]
            df_chk = df_chk.append(df_chk3)

        return df_chk

    def rada_chk90_fn(df1, chkno, **kwargs):
        df_chk = pd.DataFrame()
        df_chk3 = pd.DataFrame()
        df2 = kwargs.get('df2', None)
        fld3 = kwargs.get('fld3', None)
        fld5 = kwargs.get('fld5', None)
        labellist1 = kwargs.get('labellist1', [])
        valuelist1 = kwargs.get('valuelist1', [])
        valuelist2 = kwargs.get('valuelist2', [])
        valuelist3 = kwargs.get('valuelist3', [])
        valuelist4 = kwargs.get('valuelist4', [])
        frmlist1 = kwargs.get('frmlist1', [])

        valuelist6 = valuelist1 + valuelist2 + valuelist3 + valuelist4

        if ((len(df1) > 0) & (len(df2) > 0)):

            df_chk2 = df2.copy()

            df_chk3 = df1.loc[:, labellist1]
            df_chk3.loc[(((~(df_chk3[fld5].isin(df_chk2[fld5]))) & (
                ~(df_chk3[fld5].astype(str).str.contains('|'.join(frmlist1))))) &
                         (df_chk3[fld3].astype(str).str.contains('|'.join(valuelist6), na=False))),
                        'Check_number'] = chkno

        if (len(df_chk3) > 0) & ('Check_number' in df_chk3.columns):

            df_chk3 = df_chk3.loc[(df_chk3['Check_number'] == chkno)]

            df_chk = df_chk.append(df_chk3)
           

        return df_chk

    def rada_chk91_fn(df1, fld2, chkno, **kwargs):
        df_chk = pd.DataFrame()
        df_chk3 = pd.DataFrame()
        df2 = kwargs.get('df2', None)
        df3 = kwargs.get('df3', None)
        newfld = kwargs.get('newfld', None)
        fld3 = kwargs.get('fld3', None)
        fld4 = kwargs.get('fld4', None)
        fld5 = kwargs.get('fld5', None)
        fld6 = kwargs.get('fld6', None)
        frmlist1 = kwargs.get('frmlist1', [])
        labellist1 = kwargs.get('labellist1', [])
        valuelist1 = kwargs.get('valuelist1', [])
        valuelist2 = kwargs.get('valuelist2', [])
        valuelist3 = kwargs.get('valuelist3', [])
        valuelist4 = kwargs.get('valuelist4', [])
        valuelist5 = kwargs.get('valuelist5', [])

        valuelist6 = valuelist1 + valuelist2 + valuelist3 + valuelist4 + valuelist5

        if ((len(df1) > 0) & (len(df2) > 0) & (len(df3) > 0)):
            df_chk = df3.copy()
            df_chk2 = df2.copy()
            df_chk3 = df3.copy()
            df_chk = df_chk.loc[:, labellist1]
            df_chk = df_chk.loc[(~(df_chk[fld2].isna()))]
            df_chk = df_chk.loc[(~(df_chk[fld3].isna()))]
        if (len(df_chk) >0) & (len(df_chk3) >0):
            df_chk.loc[
                (( (~(df_chk[fld5].astype(str).str.contains('|'.join(frmlist1))))) &
                 ((df_chk[fld6].isin(df_chk3[fld6]))) & (~(df_chk[fld2].isna())) &
                 ((df_chk[fld2].apply(lambda x: not (set(valuelist6).issubset(set(str(x).split(',')))))))),
                'Check_number'] = chkno

            # df_chk.to_excel(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS\output.xlsx')

        if (len(df_chk) > 0) & ('Check_number' in df_chk.columns):
            df_chk = df_chk.loc[(df_chk['Check_number'] == chkno)]

        return df_chk

    def rada_chk92_fn(df1, fld1, fld2, chkno,  **kwargs):
        df_chk = pd.DataFrame()
        df_chk3 = pd.DataFrame()
        df2 = kwargs.get('df2', None)
        df3 = kwargs.get('df3', None)
        newfld = kwargs.get('newfld', None)
        fld3 = kwargs.get('fld3', None)
        fld4 = kwargs.get('fld4', None)
        fld5 = kwargs.get('fld5', None)
        fld6 = kwargs.get('fld6', None)
        frmlist1 = kwargs.get('frmlist1', [])
        labellist1 = kwargs.get('labellist1', [])
        valuelist1 = kwargs.get('valuelist1', [])
        valuelist2 = kwargs.get('valuelist2', [])
        valuelist3 = kwargs.get('valuelist3', [])
        valuelist4 = kwargs.get('valuelist4', [])
        valuelist5 = kwargs.get('valuelist5', [])

        valuelist6 = valuelist1 + valuelist2 + valuelist3 + valuelist4 + valuelist5

        if ((len(df1) > 0) & (len(df2) > 0) & (len(df3) > 0)):
            df_chk = df3.copy()
            df_chk2 = df2.copy()
            df_chk3 = df3.copy()
            df_chk = df_chk.loc[:, labellist1]
            df_chk = df_chk.loc[(~(df_chk[fld2].isna()))]
            df_chk = df_chk.loc[(~(df_chk[fld3].isna()))]
        if (len(df_chk) >0) & (len(df_chk3) >0):
            df_chk.loc[
                (( (~(df_chk[fld5].astype(str).str.contains('|'.join(frmlist1))))) &
                 ((df_chk[fld6].isin(df_chk3[fld6]))) & (~(df_chk[fld3].isna())) &
                 ((df_chk[fld3].apply(lambda x: not (set(valuelist1).issubset(set(str(x).split(',')))))))),
                'Check_number'] = chkno
            # (~(df_chk[fld5].isin(df_chk2[fld5]))) &

        if (len(df_chk) > 0) & ('Check_number' in df_chk.columns):
            df_chk = df_chk.drop_duplicates(subset=['Check_number',fld5, fld6])
            df_chk = df_chk.loc[(df_chk['Check_number'] == chkno)]

        return df_chk

        # Check#99: Check for empty cells in Forms, New Form and Field Attributes and Collation List tabs

    def rada_chk99_fn(df1, chkno, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        if len(df1) > 0:
            df1 = df1.iloc[:, labellist1]
            df1.loc[(((df1.iloc[:, 0].isna())) | ((df1.iloc[:, 1].isna())) | (
                (df1.iloc[:, 2].isna())) & ((df1.iloc[:, 3].isna()))) == True, ('Check_number')] = chkno
            df1 = df1.loc[(df1['Check_number'] == chkno)]

        if (len(df1) > 0):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk100a_fn(df1, chkno, **kwargs):
        df_chk = pd.DataFrame()
        rowlist1 = kwargs.get('rowlist1', [])
        labellist1 = kwargs.get('labellist1', [])
        if len(df1) > 0:
            df1 = df1.iloc[rowlist1, labellist1]
            df1.loc[(((df1.iloc[:, 0].isna())) | ((df1.iloc[:, 1].isna())) | (
                (df1.iloc[:, 2].isna()))) == True, ('Check_number')] = chkno
            df1 = df1.loc[(df1['Check_number'] == chkno)]

        if (len(df1) > 0):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk100b_fn(df1, chkno, **kwargs):
        df_chk = pd.DataFrame()
        rowlist1 = kwargs.get('rowlist1', [])
        labellist1 = kwargs.get('labellist1', [])
        if len(df1) > 0:
            df1 = df1.iloc[rowlist1, labellist1]
            df1.loc[(((df1.iloc[:, 0].isna())) | ((df1.iloc[:, 1].isna()))) == True, ('Check_number')] = chkno
            df1 = df1.loc[(df1['Check_number'] == chkno)]

        if (len(df1) > 0):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk115_fn(df1, col_ind, fld1, chkno, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        collist1 = kwargs.get('collist1', [])
        if len(df1) > 0:
            # df1 = df1.iloc[:, labellist1]
            x, y = sp.coo_matrix(df1.isnull()).nonzero()
            if ((len(x) > 0) & (len(x) > 0)):
                list1 = list(zip(x, y))
                list2 = [col_ind + 2, 1]
                list_sum = (np.add(list1, list2)).tolist()
                list_string = [list(map(str, lst)) for lst in list_sum]
                for a in list_string:
                    a[0] = 'Row ' + a[0]
                    a[1] = 'Column ' + a[1]
                if len(list_string) > 0:
                    df_chk[fld1] = list_string
                    df_chk[['Row', 'Column']] = pd.DataFrame(df_chk[fld1].tolist(), index=df_chk.index)
                    df_chk[fld1] = df_chk[fld1].apply(lambda x: str(x).strip('[] '))
                    # df_chk = df_chk.loc[~df_chk[fld1].str.contains('|'.join(collist1))]
                    df_chk = df_chk.loc[~df_chk['Column'].isin(collist1)]
                    df_chk['Check_number'] = chkno
                    f = {fld1: '|'.join}
                    df_chk = df_chk.groupby(['Check_number'], as_index=False).agg(f)

        if (len(df_chk) > 0):
            df_chk = df_chk.loc[(df_chk['Check_number'] == chkno)]

        return df_chk

    def rada_chk120_1_fn(df1, col_ind, fld1, chkno, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        collist1 = kwargs.get('collist1', [])
        if len(df1) > 0:
            df1 = df1.iloc[:, labellist1]

            x, y = sp.coo_matrix(df1.isnull()).nonzero()

            if ((len(x) > 0) & (len(x) > 0)):
                list1 = list(zip(x, y))
                col_ind = col_ind + 2
                list2 = [col_ind, 1]
                list_sum = (np.add(list1, list2)).tolist()
                list_string = [list(map(str, lst)) for lst in list_sum]
                for a in list_string:
                    a[0] = 'Row ' + a[0]
                    a[1] = 'Column ' + a[1]


                if len(list_string) > 0:
                    df_chk[fld1] = list_string
                    df_chk[['Row', 'Column']] = pd.DataFrame(df_chk[fld1].tolist(), index=df_chk.index)
                    # print("df_chk[['Row'", df_chk['Row'])
                    df_chk[fld1] = df_chk[fld1].apply(lambda x: str(x).strip('[] '))
                    # df_chk = df_chk.loc[~df_chk[fld1].str.contains('|'.join(collist1))]
                    df_chk = df_chk.loc[~df_chk['Column'].isin(collist1)]
                    df_chk['Check_number'] = chkno
                    f = {fld1: '|'.join}
                    df_chk = df_chk.groupby(['Check_number'], as_index=False).agg(f)

        if (len(df_chk) > 0):
            df_chk = df_chk.loc[(df_chk['Check_number'] == chkno)]

        return df_chk

    def rada_chk120_2_fn(df1, col_ind, fld1, fld2, cond1, chkno, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        collist1 = kwargs.get('collist1', [])
        if (len(df1) > 0 ) & (fld2 in df1.columns):
            if df1[fld2].count() > 0:
                row_indexes = df1.index[df1[fld2].astype(str).str.contains('|'.join(cond1), na=False)].tolist()
                col_ind = col_ind + 2
                row_indexes_string = []
                row_indexes_string = [(col_ind + int(x)) for x in row_indexes]
                row_indexes_string1 = ['Row ' + str(x) for x in row_indexes_string]
                df1 = df1.iloc[:, labellist1]
                df1 = df1.replace('NA', np.NaN)
                df1 = df1.replace('', np.NaN)
                df1 = df1.replace(' ', np.NaN)
                df1 = df1.replace(r'', np.NaN)
                df1 = df1.replace(r' ', np.NaN)
                x, y = sp.coo_matrix(df1.isnull()).nonzero()
                if ((len(x) > 0) & (len(x) > 0)):
                    list1 = list(zip(x, y))
                    list2 = [col_ind, 1]
                    list_sum = (np.add(list1, list2)).tolist()
                    list_string = [list(map(str, lst)) for lst in list_sum]
                    for a in list_string:
                        a[0] = 'Row ' + a[0]
                        a[1] = 'Column ' + a[1]
                    if len(list_string) > 0:
                        df_chk[fld1] = list_string
                        df_chk[['Row', 'Column']] = pd.DataFrame(df_chk[fld1].tolist(), index=df_chk.index)
                        df_chk[fld1] = df_chk[fld1].apply(lambda x: str(x).strip('[] '))
                        # df_chk = df_chk.loc[~df_chk[fld1].str.contains('|'.join(collist1))]
                        df_chk = df_chk.loc[~df_chk['Column'].isin(collist1)]
                        df_chk = df_chk.loc[df_chk['Row'].isin(row_indexes_string1)]
            if len(df_chk) > 0:
                df_chk['Check_number'] = chkno

                f = {fld1: '|'.join}
                df_chk = df_chk.groupby(['Check_number'], as_index=False).agg(f)

        return df_chk

    def rada_chk99b_fn(df1, col_ind, fld1, fld_ind2, fld_ind3, chkno, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        collist1 = kwargs.get('collist1', [])
        if len(df1) > 0:
            row_indexes = df1.index[((df1.iloc[:, fld_ind2].isna()) & (df1.iloc[:, fld_ind3].isna()))].tolist()
            # print("chktst",df1.loc[df1[fld2].str.contains(cond1, na=True), fld2])
            # print("chktst",df1.loc[df1[fld2].str.contains(cond1, na=False), fld2])
            # print(col_ind)
            col_ind = col_ind + 2
            row_indexes_string = [(col_ind + int(x)) for x in row_indexes]
            row_indexes_string = ['Row ' + str(x) for x in row_indexes_string]
            df1 = df1.iloc[:, labellist1]
            # df1 = df1.replace('NA', np.NaN)
            # df1 = df1.replace('', np.NaN)
            x, y = sp.coo_matrix(df1.isnull()).nonzero()
            # print("x",x)
            # print("y",y)
            if ((len(x) > 0) & (len(x) > 0)):
                list1 = list(zip(x, y))
                list2 = [col_ind, 1]
                list_sum = (np.add(list1, list2)).tolist()
                list_string = [list(map(str, lst)) for lst in list_sum]
                for a in list_string:
                    a[0] = 'Row ' + a[0]
                    a[1] = 'Column ' + a[1]
                if len(list_string) > 0:
                    df_chk[fld1] = list_string
                    df_chk[['Row', 'Column']] = pd.DataFrame(df_chk[fld1].tolist(), index=df_chk.index)
                    df_chk[fld1] = df_chk[fld1].apply(lambda x: str(x).strip('[] '))
                    # df_chk = df_chk.loc[~df_chk[fld1].str.contains('|'.join(collist1))]
                    df_chk = df_chk.loc[~df_chk['Column'].isin(collist1)]
                    df_chk = df_chk.loc[df_chk['Row'].isin(row_indexes_string)]
            if len(df_chk) > 0:
                df_chk['Check_number'] = chkno

                f = {fld1: '|'.join}
                df_chk = df_chk.groupby(['Check_number'], as_index=False).agg(f)

        return df_chk

    def rada_chk116_fn(df1, chkno, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        if len(df1) > 0:
            # df1 = df1.iloc[:, labellist1]
            df1.loc[(((df1.iloc[:, 0].isna())) | ((df1.iloc[:, 1].isna())) | (
                (df1.iloc[:, 2].isna())) | ((df1.iloc[:, 3].isna())) | (
                         (df1.iloc[:, 4].isna())) | ((df1.iloc[:, 5].isna()))) == True, ('Check_number')] = chkno
            df1 = df1.loc[(df1['Check_number'] == chkno)]

        if (len(df1) > 0):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk101_fn(df1, df2, col1, col2, chkno, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        if (len(df1) > 0) & (len(df2) > 0):
            df1 = df1.iloc[:, labellist1]
            df1 = df1.loc[~df1.iloc[:, col1].isna()]
            # df1[col1] = df1[col1].str.strip().str.lower()
            # df2[col2] = df1[col2].str.strip().str.lower()
        if len(df1) > 0:
            df1.loc[((~(df1.iloc[:, col1].astype(str).str.strip().str.lower().isin(df2.iloc[:, col2].astype(str).str.strip().str.lower())))) == True, 'Check_number'] = chkno

            df1 = df1.loc[(df1['Check_number'] == chkno)]
            #v0.2
        if (len(df1) > 0) & ('Check_number' in df1.columns.values.tolist()):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    def rada_chk101_fn_new(df1, df2, col1, col2, chkno, **kwargs):
        df_chk = pd.DataFrame()
        labellist1 = kwargs.get('labellist1', [])
        labellist2 = kwargs.get('labellist2', [])
        if (len(df1) > 0) & (len(df2) > 0):
            # print("test columns",df1.iloc[:, labellist1].columns)
            if df1.iloc[:, labellist1].columns[0] == 'Form Name \n(eCRF Description)':
                df1 = df1.iloc[:, labellist1]
            elif df1.iloc[:, labellist2].columns[0] == 'Form Name \n(eCRF Description)':
                df1 = df1.iloc[:, labellist2]
            else:
                df1 = pd.DataFrame()

        if (len(df1) > 0) & (len(df2) > 0):
            df1 = df1.loc[~df1.iloc[:, col1].isna()]
            # df1[col1] = df1[col1].str.strip().str.lower()
            # df2[col2] = df1[col2].str.strip().str.lower()
            #28Jun21
            if len(df1) > 0:
                df1.loc[((~(df1.iloc[:, col1].astype(str).str.strip().str.lower().isin(df2.iloc[:, col2].astype(str).str.strip().str.lower())))) == True, 'Check_number'] = chkno

                df1 = df1.loc[(df1['Check_number'] == chkno)]
            #v0.2
        if (len(df1) > 0) & ('Check_number' in df1.columns.values.tolist()):
            df_chk = df1.loc[(df1['Check_number'] == chkno)]

        return df_chk

    ############################# Checks module ########################

    df_chk05 = rada_chk5_fn(df_forms, df_reservedwords_forms, "OID", 'Reserved by Windows_DOS', 5, fld3='OID',
                            fld4='FormOID')
    df_chk06 = rada_chk6_fn(df_flds, df_reservedwords_forms, "FieldOID", 'Lab view specific columns', 6,
                            labellist1=["FieldOID", "FormOID"])
    df_chk08 = rada_chk6_fn(df_flds, df_reservedwords_forms, "FieldOID", 'Clinical View Header Columns', 8,
                            labellist1=["FieldOID", "FormOID"])
    df_chk09 = rada_chk6_fn(df_flds, df_reservedwords_forms, "FieldOID", 'SQL Reserved Words', 9,
                            labellist1=["FieldOID", "FormOID"])

    df_chk07 = rada_chk7_fn(df_forms, "OID", 7)
    df_chk50 = rada_chk50_fn(df_flds, df_dicts, "DataDictionaryName", "DataFormat", "CodedData", "DataFormat_chr",
                             "DataFormat_int", "FieldOID", 50,
                             labellist2=["DataDictionaryName", "CodedData", "UserDataString"],
                             labellist1=["FieldOID", "DataDictionaryName", "DataFormat"])
    df_chk54 = rada_chk54_fn(df_dictnams, df_flds, df_cf, df_ca, "DataDictionaryName", 'SourceCode', "newsearchname",
                             "ActionOptions", "FunctionName", 54, labellist1=["DataDictionaryName"],
                             labellist2=["FunctionName"])

    df_chk55_1 = rada_chk55_fn(df_forms, df_flds, "OID", "FormOID", 55, labellist1=["DraftFormName", "OID"])
    if len(df_chk55_1) > 0:
        f = {'OID': ','.join}
        df_chk55_1 = df_chk55_1.groupby(['Check_number'], as_index=False).agg(f)
    df_chk55_2 = rada_chk55_fn(df_dictnams, df_dicts, "DataDictionaryName", "DataDictionaryName", 55,
                               labellist1=["DataDictionaryName"])
    if len(df_chk55_2) > 0:
        f = {'DataDictionaryName': ','.join}
        df_chk55_2 = df_chk55_2.groupby(['Check_number'], as_index=False).agg(f)
    df_chk10 = rada_chk7_fn(df_flds, "FieldOID", 10)

    df_chk11 = rada_chk11_fn(df_flds, 'DataDictionaryName', "FieldOID", "FormOID", 'DataFormat',
                             "FieldOIDs", "FormOIDs", 'DataFormats', 11,
                             labellist2=['DataFormat', 'DataDictionaryName'],
                             labeldict1={"FieldOID": ','.join, "FormOID": ','.join,
                                         'DataFormat': ','.join,
                                         'DataFormats': lambda x: len(set(x))},
                             labellist1=['DataDictionaryName', "FieldOID", "FormOID", 'DataFormat'])

    df_chk49 = rada_chk49_fn(df_flds_dom, 'FieldOID', 'ControlType', 'CheckBox', 'domain_ele', 49,
                             labellist1=["FieldOID", "FormOID", 'ControlType','domain_ele'],
                             labellist2=['FieldOID', 'ControlType'],
                             labeldict1={"FormOID": 'first'})
    # df_chk1.to_excel(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS\output_radachk1' + str(chkno) + '.xlsx')
    df_chk48 = rada_chk48_fn(df_forms, "DraftFormName", 48, newfield='Special_Characters',
                             labellist1=["OID", "DraftFormName"])

    df_chk70 = rada_chk70_fn(df_dicts, df_flds, df_eSAE_forms, "CodedData", "UserDataString", 'DataDictionaryName'
                             , "FormOID", 70,
                             labellist1=["DataDictionaryName", "CodedData", "UserDataString"])



    df_chk52 = rada_chk52_fn(df_flds, "FormOID", 'FieldOID', 52, ['lbvalue', 'lvalue','LBVALUE', 'LVALUE'], '_tr',
                             ['lbvalue', 'lvalue','LBVALUE', 'LVALUE', '\_\d+$','_tr$'],'FieldOID_split',
                             'counts')
    df_chk53 = rada_chk53_fn(df_flds, "FormOID", 'FieldOID', 53, '_tr', ['lbvalue', 'lvalue','LBVALUE', 'LVALUE', '_'], 'presnt',
                             'FieldOID_split', 'counts')

    df_chk67 = rada_chk67_fn(df_forms, "OID", "DraftFormName", 67, labellist1=["OID", "DraftFormName"])

    df_chk22 = rada_chk22_fn(df_forms, "OID", 22,
                             reqnum=['^1', '^2', '^3', '^4', '^5', '^6', '^7', '^8', '^9', '^0'])
    df_chk23 = rada_chk23_fn(df_als_folders, "OID", "FolderName", 23,
                             labellist1=["OID", "FolderName"], value1='UPV', value2='Unplanned')

    df_chk26 = rada_chk26_fn(df_forms_non_ops_esae_pd, "IsSignatureRequired", "ViewRestrictions",
                             'EntryRestrictions',
                             26, value1='Investigator - Add Subject',
                             labellist1=["OID", "IsSignatureRequired", "ViewRestrictions", "EntryRestrictions"],
                             frmlist2=['DVG001', 'DVG002'])
    #

    df_chk16 = rada_chk16_fn(df_checksteps, df_flds, "VariableOID", "FormOID", 16, fld3='RecordPosition',
                             fld4='IsLog',
                             labellist1=["CheckName", "VariableOID", "FormOID", 'RecordPosition'],
                             labellist2=["VariableOID", "FormOID", 'IsLog'])
    df_chk16_1 = rada_chk16_fn(df_ca, df_flds, "VariableOID", "FormOID", 16, fld3='RecordPosition', fld4='IsLog',
                               labellist1=["CheckName", "VariableOID", "FormOID", 'RecordPosition'],
                               labellist2=["VariableOID", "FormOID", 'IsLog'])

    df_chk17 = rada_chk17_fn(df_als_checks, df_ca, df_checksteps, "CheckName", "Infix", 17,
                             labellist1=["CheckName", "Infix"])
    df_chk18 = rada_chk18_fn(df_checksteps, df_dicts, df_flds, "CheckName", "DataDictionaryName",
                             18, fld3="CodedData", fld4="VariableOID", fld5="CheckFunction",
                             labellist1=["CheckName", "CheckFunction", "VariableOID", "FormOID"])

    df_chk43 = rada_chk43_fn(df_dicts, "DataDictionaryName", "CodedData", "UserDataString",
                             'DataDictionaryName_1', 43,
                             labellist1=["DataDictionaryName", "CodedData", "UserDataString"])

    df_chk47 = rada_chk47_fn(df_dicts, "DataDictionaryName", "CodedData", "UserDataString", 47,
                             labellist1=["DataDictionaryName", "CodedData", "UserDataString"],
                             labeldict1={"CodedData": '`'.join, "UserDataString": '`'.join})
    df_chk19 = rada_chk19_fn(df_ca, df_checksteps, "CheckName", 'ActionType', 'VariableOID', 19,
                             labellist1=["CheckName", 'ActionType','VariableOID'])
    df_chk102 = rada_chk102_fn(df_ca_oq, df_checksteps, df_flds, "CheckName", 'VariableOID', 'RecordPosition',
                               'IsLog', 102,
                               labellist1=["CheckName", 'ActionType', 'RecordPosition', "FormOID", "VariableOID"])
    df_chk60 = rada_chk60_fn(df_ca, "CheckName", 'ActionType', 'CustomFunction', 'OpenQuery', '_CF_', '_VAL_', 60,
                             labellist1=["CheckName", 'ActionType'])
    df_chk123 = rada_chk123_fn(df_als_checks, "CheckName", 'Infix', '_CF_', 'custom function', ['^SYS_', '^RSG_'],
                               123,
                               labellist1=["CheckName", 'Infix'])
    df_chk124 = rada_chk124_fn(df_flds_dom, df_Test_Category, 'FormOID', 'FieldOID', 'PreText','AnalyteName', 'PARM',
                               'PARMDES', 'VariableOID',
                               124,
                               labellist1=['FormOID', 'FieldOID', 'FieldOID_spl', 'PreText', 'AnalyteName','VariableOID'],
                               labellist2=['PARM', 'PARMDES', 'ALTERNATE_LAB_TEST_DESC'])
    df_chk125 = rada_chk125_fn(df_flds, 'FieldOID', 'DataFormat', "FormOID", 'PreText', 125,
                             ['dd MMM yyyy', 'dd- MMM- yyyy', 'yyyy'], ['HH:nn', 'HH:nn:ss', 'HH'],
                             labellist1=["FieldOID", "FormOID", 'DataFormat','PreText'])

    df_chk126 = rada_chk126_fn(df_forms_global_esae_no,"OID_split",  126,
                             labellist1=["OID", 'DraftName',"OID_split"])

    df_chk127 = rada_chk126_fn(df_flds_glob_formyes_fieldno ,"FieldOID",  127,
                             labellist1=["FieldOID", "FormOID", 'DraftName', 'PreText'])

    df_chk128 = rada_chk126_fn(df_flds_glob_formyes_fieldyes ,"FieldOID",  128,
                             labellist1=["FieldOID", "FormOID", 'DraftName','DataFormat','DataFormat_global',
                                         'PreText','PreText_global', 'DataDictionaryName' ,'DataDictionaryName_global'])

    df_chk123_1 = rada_chk123_1_fn(df_ca, "CheckName", 'ActionType', '_CF_', '_VAL_', 'OpenQuery', 123,
                                   labellist1=["CheckName", 'ActionType'],
                                   labelist2=['CustomFunction', 'IsPresent', 'SetDynamicSearchList'])
    df_chk20 = rada_chk20_fn(df_flds_defaults_split, df_dicts_defaults_dict, 'dictname_CodedData', 20,
                             labellist1=["FormOID", 'FieldOID', "DataDictionaryName", 'DefaultValue',
                                         'dictname_CodedData'])

    df_chk21 = rada_chk21_fn(df_checksteps, df_flds, "VariableOID", "ControlType", 21, fld3='CheckFunction',
                             fld4='CheckName', value1='CheckBox', value2='IsEmpty', value3='IsNotEmpty',
                             labellist1=["CheckName", "VariableOID", "FormOID", 'CheckFunction'])
    df_chk25 = rada_chk25_fn(df_ca, df_cf, "VariableOID", "ActionType", 25, fld3='ActionOptions',
                             fld4='CheckName', fld5='FunctionName', value1='SetDynamicSearchList',
                             labellist1=["CheckName", "VariableOID", 'ActionType', 'ActionOptions'])
    df_chk27 = rada_chk27_fn(df_cf, 'SourceCode', 27, 'AddComment()')
    df_chk28 = rada_chk28_fn(df_cf, 'SourceCode', 'FunctionName', 28, ['try','catch', 'Try','Catch'])
    df_chk30 = rada_chk30_fn(df_cf, 'SourceCode', 30)
    df_chk31 = rada_chk31_fn(df_als_checks,  "CheckName", '^SYS', 31, labellist1=["CheckName"])
    df_chk32 = rada_chk32_fn(df_flds, df_eSAE_forms, 'FormOID', 'QueryFutureDate', 'ControlType',
                             'ViewRestrictions', 'DataFormat', 'EntryRestrictions', 32, value1=False,
                             valuelist1=['DateTime', 'Date'],
                             labellist1=['FormOID', 'FieldOID', 'QueryFutureDate',
                                         'ControlType', 'ViewRestrictions', 'DataFormat', 'EntryRestrictions'],
                             valuelist2=['Clinical Research Coordinator',
                                         'Clinical Research Coordinator - Add Subject',
                                         'Clinical Research Coordinator - Restricted', 'Investigator',
                                         'Investigator - Add Subject',
                                         'Investigator - Restricted'], valuelist3=['HH:nn', 'HH:nn:ss', 'HH'])
    df_chk65 = rada_chk65_fn(df_flds, 'ControlType', 'CanSetRecordDate', 'CanSetDataPageDate', 'CanSetInstanceDate',
                             'CanSetSubjectDate', 'QueryFutureDate', 65, value1=True, valuelist1=['DateTime', 'Date'],
                             labellist1=['FormOID', 'FieldOID', 'ControlType', 'CanSetRecordDate',
                                         'CanSetDataPageDate',
                                         'CanSetInstanceDate', 'CanSetSubjectDate', 'QueryFutureDate'])

    df_chk33 = rada_chk33_fn(df_flds_numeric, df_eSAE_forms, 'FormOID', 'QueryNonConformance',
                             'ViewRestrictions', 33, value1=False,
                             labellist1=['FormOID', 'FieldOID', 'QueryNonConformance',
                                         'ControlType', 'ViewRestrictions'],
                             valuelist1=['Clinical Research Coordinator',
                                         'Clinical Research Coordinator - Add Subject',
                                         'Clinical Research Coordinator - Restricted', 'Investigator',
                                         'Investigator - Add Subject',
                                         'Investigator - Restricted'])

    df_chk71 = rada_chk71_fn(df_flds, 'IsRequired', 'ViewRestrictions', 'EntryRestrictions', 71, value1=True,
                             labellist1=['FormOID', 'FieldOID', 'EntryRestrictions', 'ViewRestrictions'],
                             valuelist1=['Clinical Research Coordinator',
                                         'Clinical Research Coordinator - Add Subject',
                                         'Clinical Research Coordinator - Restricted', 'Investigator',
                                         'Investigator - Add Subject',
                                         'Investigator - Restricted'])
    df_chk81 = rada_chk81_fn(df_flds_non_esae, df_Derivations, df_ca1, df_cf, 'ViewRestrictions', 'FieldOID',
                             'SourceCode', 81, labellist1=['FormOID', 'FieldOID', 'ViewRestrictions'],
                             labellist2=['FormOID', 'FieldOID', 'ViewRestrictions'],
                             valuelist1=['Clinical Research Coordinator',
                                         'Clinical Research Coordinator - Add Subject',
                                         'Clinical Research Coordinator - Restricted', 'Investigator',
                                         'Investigator - Add Subject',
                                         'Investigator - Restricted'])

    df_chk40 = rada_chk40_fn(df_ca, df_dicts, df_flds, "ActionOptions", "DataDictionaryName",
                             40, fld3="CodedData", fld7="ActionType", fld8='FormOID', fld4="VariableOID",
                             fld5="DataFormat", labellist1=["CheckName", 'FormOID', "VariableOID", 'ActionType',
                                                            'ActionOptions', 'ActionScript'],
                             value1='SetDataPoint', value2='DontEnterEmptyIfFalse', fld6='ActionScript',
                             labellist2=['FormOID', "VariableOID", 'DataDictionaryName', 'DataFormat'],
                             newfield='DataFormat_len', newfield2='ActionScript_len')
    df_chk41 = rada_chk41_fn(df_checksteps, df_flds_inact, df_forms_inact, df_als_checks, "VariableOID", "FormOID", 41,
                             fld3='OID', fld4='CheckActive',fld5='CheckName',
                             labellist1=["CheckName", "VariableOID", "FormOID"])
    str1 = 'Page successfully submitted.'
    str2 = 'I declare that all of the data within the Case Report Form(s) associated with this Protocol, covered by this Signature have been reviewed, and are accurate and complete. I intend for my electronic signatures to be the legally binding equivalent of my handwritten signature.'
    df_chk75 = rada_chk75_fn(df_crfdraft, 75, str1, 'ConfirmationMessage')
    df_chk76 = rada_chk75_fn(df_crfdraft, 76, str2, 'SignaturePrompt')
    labels_chk29 = ['FormOID', 'FieldOID', 'GlobalVariableOID', 'LocationMethod', 'LabStandardGroup']
    df_chk29 = rada_chk29_fn(df_crfdraft, df_lbsettings1, 'LabStandardGroup', 'GlobalVariableOID', 'FormOID',
                             'FieldOID', 'LocationMethod', 'PREFER_SI', 29, labellist1=['Age', 'SEX'],
                             labellist2=['OPG001', 'DM'], labellist3=['ClosestPriorToDate', 'EarliestDate'],
                             labellist4=['LabStandardGroup'])

    df_chk79 = rada_chk79_fn(df_flds_inact, df_forms_inact, 79)
    df_chk01 = rada_chk1_fn(df_flds, 'DataFormat', 'ControlType', 1, newfld='DataFormat_1',
                            labellist1=["FieldOID", "FormOID", 'DataFormat', 'ControlType'])
    df_chk02 = rada_chk2_fn(df_flds, 'DataFormat', 'ControlType', 'FieldOID', 2, 200, newfld='DataFormat_1',
                            labellist1=["FieldOID", "FormOID", 'DataFormat', 'ControlType'],
                            fldlist=["DVRATNL", "CONCAT", 'INV_RTNL'])
    df_chk03 = rada_chk3_fn(df_flds_non_ops_esae, df_stage_domain2, 'FieldOID', 'TARGET_DATA_ELEMENT', 3, 7,
                            labellist1=["FieldOID", "FormOID"],
                            formlist=['LBVALUE', 'LVALUE'],
                            labellist2=['E2B_COUNTRY', 'AEDSLTXT', 'MHDSLTXT', 'OPNTSIAE', 'DVHIDDEN'])
    df_chk12 = rada_chk12_fn(df_flds_non_esae, 'FieldOID', 'DataFormat', "FormOID", 12, 'DAT$',
                             ['dd MMM yyyy', 'dd- MMM- yyyy', 'yyyy'],
                             labellist1=["FieldOID", "FormOID", 'DataFormat'],
                             formlist=['PRIMARY002', 'SVG001', 'SVG002'])

    df_chk15 = rada_chk12_fn(df_flds_non_esae, 'FieldOID', 'DataFormat', "FormOID", 15, 'TIM$',
                             ['HH:nn', 'HH:nn:ss', 'HH'],
                             labellist1=["FieldOID", "FormOID", 'DataFormat'],
                             formlist=['PRIMARY002', 'SVG001', 'SVG002'])
    df_chk13 = rada_chk13_fn(df_flds, 'FieldOID', 'DataFormat', 13, '^LBL',
                             labellist1=["FieldOID", "FormOID", 'DataFormat'])
    df_chk24 = rada_chk24_fn(df_flds, 'FieldOID', 'DataFormat', 'ControlType', 24, '^LBL', 'Text',
                             labellist1=["FieldOID", "FormOID", 'DataFormat', 'ControlType'])
    df_chk39 = rada_chk39_fn(df_flds, 'PreText', 39,
                             labellist1=['FormOID', 'FieldOID', 'PreText'])
    df_chk51 = rada_chk51_fn(df_flds, 'SASFormat', 51, labellist1=['FormOID', 'FieldOID', 'SASFormat'])
    df_chk77 = rada_chk51_fn(df_flds, 'UnitDictionaryName', 77,
                             labellist1=['FormOID', 'FieldOID', 'UnitDictionaryName'])
    df_chk69 = rada_chk69_fn(df_flds, 'LowerRange', 'UpperRange', 'NCLowerRange', 'NCUpperRange',
                             'ReviewGroups',
                             69, labellist1=['FormOID', 'FieldOID', 'LowerRange', 'UpperRange', 'NCLowerRange',
                                             'NCUpperRange', 'ReviewGroups'])

    df_chk61 = rada_chk51_fn(df_flds, 'AcceptableFileExtensions', 61,
                             labellist1=['FormOID', 'FieldOID', 'AcceptableFileExtensions'])
    df_chk62 = rada_chk51_fn(df_flds, 'EproFormat', 62,
                             labellist1=['FormOID', 'FieldOID', 'EproFormat'])

    df_chk63 = rada_chk64_fn(df_flds, 'OtherVisits', 63, True,
                             labellist1=['FormOID', 'FieldOID', 'OtherVisits'])
    df_chk64 = rada_chk64_fn(df_flds, 'IsTranslationRequired', 64, True,
                             labellist1=['FormOID', 'FieldOID', 'IsTranslationRequired'])
    df_chk34 = rada_chk34_fn(df_flds, 'IsTranslationRequired', 'IsRequired', 34, True,
                             labellist1=['FormOID', 'FieldOID', 'IsTranslationRequired', 'IsRequired'])
    df_chk46 = rada_chk46_fn(df_flds, df_reservedwords_forms, 'PreText',
                             'HTML codes', 'PreText_1', 46, labellist1=['FormOID', 'FieldOID', 'PreText'])
    df_chk58 = rada_chk59_fn(df_flds, 'FieldOID', 'DraftFieldName', 58,
                             labellist1=['FormOID', 'FieldOID', 'DraftFieldName'])
    df_chk59 = rada_chk59_fn(df_flds, 'FieldOID', 'VariableOID', 59,
                             labellist1=['FormOID', 'FieldOID', 'VariableOID'])

    df_chk35 = rada_chk35_fn(df_flds, df_non_domain_forms, df_reservedwords_forms, 'VariableOID', 'SASLabel',
                             'FormOID', 'HTML codes', 35, labellist1=['FormOID', 'FieldOID', 'VariableOID', 'SASLabel'])

    df_chk36 = rada_chk36_fn(df_flds, 'SourceDocument', 'DefaultValue', 36, True,
                             labellist1=['FormOID', 'FieldOID', 'SourceDocument', 'DefaultValue'])

    df_chk45 = rada_chk45_fn(df_Matrices, 'OID', 45, "MASTERDASHBOARD")
    df_chk37 = rada_chk36_fn(df_flds, 'IsRequired', 'DefaultValue', 37, True,
                             labellist1=['FormOID', 'FieldOID', 'IsRequired', 'DefaultValue'])
    df_chk38 = rada_chk38_fn(df_flds, 'QueryNonConformance', 'DataFormat', 'ControlType', 38, True,
                             labellist1=['FormOID', 'FieldOID', 'QueryNonConformance', 'DataFormat', 'ControlType'],
                             valuelist1=['Text', 'DropDownList',  'RadioButton', 'LongText',
                                         'RadioButton (Vertical)'],
                             # valuelist1=['Text', 'DropDownList', 'Dynamic SearchList', 'RadioButton', 'LongText',
                             #             'RadioButton (Vertical)'],
                             value1='\$')

    df_chk56 = rada_chk56_fn(df_flds, 'CodingDictionary', 'FieldOID', 56,
                             labellist1=['FormOID', 'FieldOID', 'CodingDictionary'])
    df_chk57 = rada_chk57_fn(df_flds, 'CodingDictionary', 'DataFormat', 57,
                             labellist1=['FormOID', 'FieldOID', 'CodingDictionary', 'DataFormat'],
                             value1='$')
    df_chk68 = rada_chk68_fn(df_flds, 'DataFormat', 'ControlType', 68, '$200',
                             labellist1=['FormOID', 'FieldOID', 'DataFormat', 'ControlType'],
                             valuelist1=['DropDownList', 'Dynamic SearchList', 'RadioButton', 'LongText',
                                         'RadioButton (Vertical)'])
    df_chk42 = rada_chk42_fn(df_flds_sdpv, df_als_checks, 'IsVisible', 'CheckName', 'CheckActive', 42, False, True,
                             labellist1=['FormOID', 'FieldOID', 'IsVisible', 'CheckName'])

    df_chk66 = rada_chk64_fn(df_flds, 'CanSetRecordDate', 66, True,
                             labellist1=['FormOID', 'FieldOID', 'CanSetRecordDate'])
    df_chk80 = rada_chk80_fn(df_flds, 'SourceDocument', 80, True, sdv,
                             labellist1=['FormOID', 'FieldOID', 'SourceDocument'])

    df_chk78 = rada_chk78_fn(df_flds, 'IsLog', 'HeaderText', 78, False,
                             labellist1=['FormOID', 'FieldOID', 'IsLog', 'HeaderText'])
    df_chk44 = rada_chk44_fn(df_flds, 'IsLog', 'DefaultValue', 44, False,
                             labellist1=['FormOID', 'FieldOID', 'IsLog', 'DefaultValue'])

    df_chk73 = rada_chk73_fn(df_flds, 'DoesNotBreakSignature', 'ViewRestrictions', 'VariableOID', 73, True,
                             'Investigator',
                             labellist1=['FormOID', 'FieldOID', 'VariableOID', 'DoesNotBreakSignature',
                                         'ViewRestrictions'])
    df_chk72 = rada_chk72_fn(df_flds, 'DoesNotBreakSignature', 'ViewRestrictions', 72, False, 'Investigator',
                             labellist1=['FormOID', 'FieldOID', 'DoesNotBreakSignature', 'ViewRestrictions'])
    df_chk74 = rada_chk74_fn(df_flds, 'DoesNotBreakSignature', 'FieldOID', 'DataFormat', 74, False, '^LBL',
                             labellist1=['FormOID', 'FieldOID', 'DataFormat', 'DoesNotBreakSignature'])

    df_chk14 = rada_chk14_fn(df_flds, 'FieldOID', 'VariableOID', 'DataFormat', 'DataDictionaryName',
                             'CodingDictionary',
                             'FixedUnit', 'HeaderText', 'SourceDocument', 'DefaultValue', 'SASLabel', 'IsRequired',
                             'QueryFutureDate', 'IsVisible', 'AnalyteName', 'IsClinicalSignificance',
                             'QueryNonConformance',
                             'ViewRestrictions', 'EntryRestrictions', 'ControlType', 14, True, False, '^LBL',
                             labellist1=['FormOID', 'FieldOID', 'VariableOID', 'DataFormat', 'DataDictionaryName',
                                         'CodingDictionary', 'FixedUnit', 'HeaderText', 'SourceDocument',
                                         'DefaultValue',
                                         'SASLabel', 'IsRequired', 'QueryFutureDate', 'IsVisible', 'AnalyteName',
                                         'IsClinicalSignificance', 'QueryNonConformance', 'ViewRestrictions',
                                         'EntryRestrictions','ControlType'])

    df_chk04 = rada_chk4_fn(df_flds, 'DataFormat', 'ControlType', 4,
                            labellist1=["FieldOID", "FormOID", 'DataFormat', 'ControlType'])

    df_chk82 = rada_chk82_fn(df_forms, 'DraftFormName', 'ViewRestrictions', 82, fld3='EntryRestrictions',
                             searchvalue1='Adjudication',
                             labellist1=["OID", 'DraftFormName', 'ViewRestrictions', 'EntryRestrictions'],
                             valuelist1=["Adjudication Chair", "Adjudication Committee"])
    df_chk83 = rada_chk82_fn(df_forms, 'DraftFormName', 'ViewRestrictions', 83, fld3='EntryRestrictions',
                             searchvalue1='Specialty',
                             labellist1=["OID", 'DraftFormName', 'ViewRestrictions', 'EntryRestrictions'],
                             valuelist1=["Specialty Data Provider"])
    df_chk84 = rada_chk84_fn(df_forms, 'DraftFormName', 'ViewRestrictions', 84, fld3='EntryRestrictions',
                             fld4='OID',
                             searchvalue1='Specialty',
                             labellist1=["OID", 'DraftFormName', 'ViewRestrictions', 'EntryRestrictions'],
                             searchlist1=['Specialty', 'Adjudication'],
                             valuelist1=['Clinical Research Coordinator',
                                         'Clinical Research Coordinator - Add Subject',
                                         'Clinical Research Coordinator - Restricted',
                                         'Investigator', 'Investigator - Add Subject',
                                         'Investigator - Restricted'],
                             valuelist2=['Clinical Research Coordinator - Restricted', 'Investigator - Restricted'],
                             frmlist1=['OPG001', 'SAEINFO'], frmlist2=['DVG001', 'DVG002'], frm1='DVG001',
                             frm2='DVG002')

    df_chk86 = rada_chk86_fn(df_forms, 'DraftFormName', 'ViewRestrictions', 86, fld3='EntryRestrictions',
                             fld4='OID',
                             searchvalue1='Specialty',
                             labellist1=["OID", 'DraftFormName', 'ViewRestrictions', 'EntryRestrictions'],
                             searchlist1=['Adjudication'],
                             valuelist1=['Clinical Review', 'Field Monitor', 'Medical Review'],
                             valuelist2=['Clinical Review - Restricted', 'Field Monitor - Restricted',
                                         'Medical Review - Restricted'],
                             valuelist3=['Central Monitor'],
                             frmlist1=['DVG001', 'DVG002'], frm1='DVG001', frm2='DVG002')

    df_chk85 = rada_chk85_fn(df_forms, 'DraftFormName', 'ViewRestrictions', 85, fld3='EntryRestrictions',
                             fld4='OID',
                             searchvalue1='Specialty',
                             labellist1=["OID", 'DraftFormName', 'ViewRestrictions', 'EntryRestrictions'],
                             searchlist1=['Adjudication'],
                             valuelist1=['Batch Upload', 'Coder Import Role', 'Data Manager - restricted',
                                         'OL Role',
                                         'Outputs Locked', 'Outputs Standard', 'Power User',
                                         'Rave Web Services (RWS)',
                                         'Safety', 'Trial Management - Restricted', 'PDR All Data', 'RPA CQM'],
                             valuelist2=['Data Manager', 'Read Only', 'Read Only - All Sites', 'Trial Management',
                                         'PDR Site Data'],
                             # valuelist3=['Central Monitor'],
                             frmlist1=['DVG001', 'DVG002'], frm1='DVG001', frm2='DVG002')

    df_chk87 = rada_chk87_fn(df_flds, 'DraftFormName',
                             'ViewRestrictions', 87, df2=df_non_domain_forms, fld3='EntryRestrictions',
                             fld5='FormOID', fld6='FieldOID',
                             labellist1=['FormOID', 'FieldOID', 'DraftFormName', 'ViewRestrictions',
                                         'EntryRestrictions'],
                             frmlist1=['DVG001', 'DVG002'])
    df_chk88 = rada_chk88_fn(df_flds, 'DraftFormName',
                             'ViewRestrictions', 88, fld3='EntryRestrictions',
                             fld5='FormOID', fld6='FieldOID',
                             labellist1=['FormOID', 'FieldOID', 'DraftFormName', 'ViewRestrictions',
                                         'EntryRestrictions'],
                             frmlist1=['DVG001', 'DVG002'])

    df_chk93 = rada_chk93_fn(df_flds, 'DraftFormName',
                             'ViewRestrictions', 'VariableOID', 93, df2=df_non_domain_forms, df3=df_flds_default_sdp,
                             fld3='EntryRestrictions', fld4='DefaultValue',
                             fld5='FormOID', fld6='FieldOID',
                             labellist1=['FormOID', 'FieldOID', 'DraftFormName', 'ViewRestrictions',
                                         'EntryRestrictions',
                                         'DefaultValue'],
                             frmlist1=['DVG001', 'DVG002'])

    df_chk94 = rada_chk94_fn(df_forms, 'ViewRestrictions', 94, df2=df_flds,
                             labellist2=['FormOID', 'FieldOID', 'DraftFormName', 'ViewRestrictions'],
                             labellist1=["OID", 'DraftFormName', 'ViewRestrictions'],
                             valuelist1=['Clinical Research Coordinator - Add Subject'],
                             value2='PDR Site Data')

    '''
    
    Entry Restrictions for DVG001
    
    Adjudication Chair
    Adjudication Committee
    Adjudication Coordinator
    Batch Upload
    Clinical Review
    Field Monitor
    Medical Review
    Coder Import Role
    Data Manager
    Data Manager - restricted
    OL Role
    Outputs Locked
    Outputs Standard
    Power User
    Rave Web Services (RWS)
    Read Only
    Read Only - All Sites
    Safety
    Specialty Data Provider
    Trial Management
    Trial Management - Restricted
    PDR Site Data  
    PDR All Data   
    RPA CQM
    
    
    
    '''

    df_chk95 = rada_chk95_1_fn(df_forms, df_globalforms, 'OID', 'EntryRestrictions', 'DVG001', 95,
                               labellist1=["OID", 'DraftFormName', 'EntryRestrictions'],
                               valuelist1=['Adjudication Chair', 'Adjudication Committee', 'Adjudication Coordinator',
                                           'Batch Upload', 'Clinical Review','Field Monitor',
                                           'Medical Review', 'Coder Import Role','Data Manager',
                                           'Data Manager - restricted', 'OL Role','Outputs Locked',
                                           'Outputs Standard', 'Rave Web Services (RWS)','Safety',
                                           'Specialty Data Provider', 'PDR All Data',
                                           'RPA CQM', 'Read Only', 'Read Only - All Sites','Trial Management',
                                           'Trial Management - Restricted','PDR Site Data','Power User'])

    df_chk95_1 = rada_chk95_1_fn(df_forms, df_globalforms, 'OID', 'ViewRestrictions', 'DVG001', 95,
                                 labellist1=["OID", 'DraftFormName', 'ViewRestrictions'])

    '''
    #chk96
    View Restrictions for DVG001
    Batch Upload
    Clinical Research Coordinator
    Clinical Research Coordinator - Add Subject
    Clinical Research Coordinator - Restricted
    Central Monitor
    Clinical Review
    Clinical Review - Restricted
    Field Monitor
    Field Monitor - Restricted
    Medical Review
    Medical Review - Restricted
    Coder Import Role
    Data Manager
    Data Manager - restricted
    Investigator
    Investigator - Add Subject
    Investigator - Restricted
    OL Role
    Outputs Locked
    Outputs Standard
    Power User
    Rave Web Services (RWS)
    Read Only
    Read Only - All Sites
    Safety
    Trial Management
    Trial Management - Restricted
    PDR Site Data  
    PDR All Data   
    RPA CQM
    '''
    df_chk96 = rada_chk96_fn(df_forms,  df_globalforms, 'OID', 'ViewRestrictions','DVG002', 96,
                             labellist1=["OID", 'DraftFormName', 'ViewRestrictions'],
                             valuelist1=['Batch Upload', 'Clinical Research Coordinator',
                                         'Clinical Research Coordinator - Add Subject',
                                         'Clinical Research Coordinator - Restricted',
                                         'Investigator', 'Investigator - Add Subject',
                                         'Investigator - Restricted', 'Central Monitor', 'Clinical Review',
                                         'Field Monitor',
                                         'Medical Review', 'Clinical Review - Restricted',
                                         'Field Monitor - Restricted',
                                         'Medical Review - Restricted', 'Coder Import Role', 'Data Manager',
                                         'Data Manager - restricted', 'OL Role', 'Outputs Locked',
                                         'Outputs Standard', 'Rave Web Services (RWS)', 'Safety', 'PDR All Data',
                                         'RPA CQM', 'Read Only', 'Read Only - All Sites', 'Trial Management',
                                         'Trial Management - Restricted', 'PDR Site Data', 'Power User'])

    '''
    View Restrictions for DVG002
    
    Batch Upload
    Clinical Research Coordinator - Restricted
    Clinical Review - Restricted
    Field Monitor - Restricted
    Medical Review - Restricted
    Coder Import Role
    Data Manager - restricted
    Investigator - Restricted
    OL Role
    Outputs Locked
    Outputs Standard
    Power User
    Rave Web Services (RWS)
    Safety
    Trial Management - Restricted
    PDR All Data   
    RPA CQM
    '''
    df_chk98_1 = rada_chk98_fn(df_flds, df_globalfields,  'FormOID', 'ViewRestrictions', 'FieldOID', 'DVG001', 98,
                             labellist1=['FormOID', 'FieldOID', 'ViewRestrictions'])
    df_chk98_2 = rada_chk98_fn(df_flds, df_globalfields,  'FormOID', 'EntryRestrictions', 'FieldOID', 'DVG001', 98,
                             labellist1=['FormOID', 'FieldOID', 'EntryRestrictions'])
    df_chk98_3 = rada_chk98_fn(df_flds, df_globalfields,  'FormOID', 'ViewRestrictions', 'FieldOID', 'DVG002', 98,
                             labellist1=['FormOID', 'FieldOID', 'ViewRestrictions'])
    df_chk98_4 = rada_chk98_fn(df_flds, df_globalfields,  'FormOID', 'EntryRestrictions', 'FieldOID', 'DVG002', 98,
                             labellist1=['FormOID', 'FieldOID', 'EntryRestrictions'])

    '''
    Entry Restrictions for DVG002
    
    Adjudication Chair
    Adjudication Committee
    Adjudication Coordinator
    Batch Upload
    Clinical Research Coordinator
    Clinical Research Coordinator - Add Subject
    Central Monitor
    Clinical Review
    Clinical Review - Restricted
    Field Monitor
    Field Monitor - Restricted
    Medical Review
    Medical Review - Restricted
    Coder Import Role
    Data Manager
    Data Manager - restricted
    Investigator
    Investigator - Add Subject
    OL Role
    Outputs Locked
    Outputs Standard
    Power User
    Rave Web Services (RWS)
    Read Only
    Read Only - All Sites
    Safety
    Specialty Data Provider
    Trial Management
    Trial Management - Restricted
    PDR Site Data  
    PDR All Data   
    RPA CQM
    '''

    df_chk97 = rada_chk96_fn(df_forms, df_globalforms,  'OID', 'EntryRestrictions',  'DVG002', 97,
                             labellist1=["OID", 'DraftFormName', 'EntryRestrictions'],
                             valuelist1=["Adjudication Chair", "Adjudication Committee", "Adjudication Coordinator",
                                         'Batch Upload', 'Clinical Research Coordinator',
                                         'Clinical Research Coordinator - Add Subject',
                                         'Clinical Research Coordinator - Restricted',
                                         'Investigator', 'Investigator - Add Subject',
                                         'Investigator - Restricted', 'Central Monitor', 'Clinical Review',
                                         'Field Monitor',
                                         'Medical Review', 'Clinical Review - Restricted',
                                         'Field Monitor - Restricted',
                                         'Medical Review - Restricted', 'Coder Import Role', 'Data Manager',
                                         'Data Manager - restricted', 'OL Role', 'Outputs Locked',
                                         'Outputs Standard', 'Rave Web Services (RWS)', 'Safety',
                                         'Specialty Data Provider',
                                         'PDR All Data', 'RPA CQM', 'Read Only', 'Read Only - All Sites',
                                         'Trial Management', 'Trial Management - Restricted', 'PDR Site Data',
                                         'Power User'])

    df_chk89 = rada_chk89_fn(df_flds, 'ViewRestrictions', 89,
                             fld4='DefaultValue',
                             fld5='FormOID', fld6='FieldOID', df2=df_non_domain_forms,
                             labellist1=['FormOID', 'FieldOID', 'DraftFormName', 'ViewRestrictions',
                                         'EntryRestrictions',
                                         'DefaultValue'],
                             # searchlist1=['Adjudication'],
                             valuelist1=['Batch Upload', 'Rave Web Services (RWS)', 'OL Role', 'Outputs Locked',
                                         'Outputs Standard', 'Coder Import Role', 'Data Manager',
                                         'Data Manager - restricted', 'Safety', 'PDR All Data', 'RPA CQM'],
                             valuelist2=['Read Only', 'Read Only - All Sites', 'Trial Management',
                                         'Trial Management - Restricted', 'PDR Site Data'],
                             valuelist3=["Adjudication Chair", "Adjudication Committee",
                                         "Adjudication Coordinator"],
                             valuelist4=["Specialty Data Provider"],
                             # valuelist3=['Central Monitor'],
                             frmlist1=['DVG001', 'DVG002'], frm1='DVG001', frm2='DVG002')

    df_chk90 = rada_chk90_fn(df_flds, 90, fld3='EntryRestrictions',
                             fld4='DefaultValue',
                             fld5='FormOID', fld6='FieldOID', df2=df_eSAE_forms,
                             labellist1=['FormOID', 'FieldOID', 'DraftFormName', 'ViewRestrictions',
                                         'EntryRestrictions',
                                         'DefaultValue'],
                             # searchlist1=['Adjudication'],
                             valuelist1=['Batch Upload', 'Rave Web Services (RWS)', 'OL Role', 'Outputs Locked',
                                         'Outputs Standard', 'Coder Import Role', 'Data Manager',
                                         'Data Manager - restricted', 'Safety', 'Power User',
                                         'PDR All Data', 'RPA CQM'],
                             valuelist2=['Read Only', 'Read Only - All Sites', 'Trial Management',
                                         'Trial Management - Restricted', 'PDR Site Data'],
                             valuelist3=["Adjudication Chair", "Adjudication Committee",
                                         "Adjudication Coordinator"],
                             valuelist4=["Specialty Data Provider"],
                             # valuelist3=['Central Monitor'],
                             frmlist1=['DVG001', 'DVG002'], frm1='DVG001', frm2='DVG002')

    df_chk91 = rada_chk91_fn(df_flds,  'ViewRestrictions', 91, fld3='EntryRestrictions',
                             fld4='DefaultValue',
                             fld5='FormOID', fld6='FieldOID', df2=df_non_domain_forms, df3=df_flds_default_sdp,
                             labellist1=['FormOID', 'FieldOID', 'DraftFormName', 'ViewRestrictions',
                                         'EntryRestrictions',
                                         'DefaultValue'],
                             # searchlist1=['Adjudication'],
                             valuelist1=['Clinical Research Coordinator',
                                         'Clinical Research Coordinator - Add Subject',
                                         'Clinical Research Coordinator - Restricted',
                                         'Investigator', 'Investigator - Add Subject',
                                         'Investigator - Restricted'],
                             valuelist2=['Read Only', 'Read Only - All Sites', 'Trial Management',
                                         'Trial Management - Restricted', 'PDR Site Data'],
                             valuelist3=['Clinical Review', 'Field Monitor', 'Medical Review'],
                             valuelist4=['Clinical Review - Restricted', 'Field Monitor - Restricted',
                                         'Medical Review - Restricted'],
                             valuelist5=['Central Monitor'],
                             frmlist1=['DVG001', 'DVG002'], frm1='DVG001', frm2='DVG002')

    df_chk92 = rada_chk92_fn(df_flds, 'DraftFormName', 'ViewRestrictions', 92, fld3='EntryRestrictions',
                             fld4='DefaultValue',
                             fld5='FormOID', fld6='FieldOID', df2=df_non_domain_forms, df3=df_flds_default_sdp,
                             labellist1=['FormOID', 'FieldOID', 'DraftFormName', 'ViewRestrictions',
                                         'EntryRestrictions',
                                         'DefaultValue'],
                             # searchlist1=['Adjudication'],
                             valuelist1=['Clinical Research Coordinator',
                                         'Clinical Research Coordinator - Add Subject',
                                         'Clinical Research Coordinator - Restricted',
                                         'Investigator', 'Investigator - Add Subject',
                                         'Investigator - Restricted'],
                             valuelist2=['Read Only', 'Read Only - All Sites', 'Trial Management',
                                         'Trial Management - Restricted', 'PDR Site Data'],
                             valuelist3=['Clinical Review', 'Field Monitor', 'Medical Review'],
                             valuelist4=['Clinical Review - Restricted', 'Field Monitor - Restricted',
                                         'Medical Review - Restricted'],
                             valuelist5=['Central Monitor'],
                             frmlist1=['DVG001', 'DVG002'], frm1='DVG001', frm2='DVG002')

    # df_chk99 = rada_chk99_fn(df_ssd_frmattr, 99, labellist1=[0, 1, 3, 4])
    df_chk99a = rada_chk120_1_fn(df_ssd_frmattr, df_ssd_frmattr_df_getcol_ind, 'Missing_Cells', 99,
                                 labellist1=[0, 1], collist1=['Column 3'])
    df_chk99b = rada_chk99b_fn(df_ssd_frmattr, df_ssd_frmattr_df_getcol_ind, 'Missing_Cells', 3, 4, 99,
                               labellist1=[0, 1, 2, 3, 4], collist1=['Column 1', 'Column 2', 'Column 3'])
    # df_chk100a = rada_chk100a_fn(df_ssd_esae1, 100, rowlist1=[0], labellist1=[0, 1, 2])
    df_chk100a = rada_chk115_fn(df_ssd_esae1, df_ssd_esae1_df_getcol_ind, 'Missing_Cells', 100,
                                labellist1=[0, 1, 2], collist1=['nothing'])

    # df_chk100b = rada_chk100b_fn(df_ssd_esae1, 100, rowlist1=[0], labellist1=[0, 1])

    df_chk100b = rada_chk115_fn(df_ssd_esae2, df_ssd_esae2_df_getcol_ind, 'Missing_Cells', 100,
                                labellist1=[0, 1], collist1=['nothing'])

    df_chk101 = pd.DataFrame()
    df_chk101a = rada_chk101_fn_new(df_ssd_newfrmattr, df_ssd_frmattr, 0, 0, 101, labellist1=[1], labellist2=[2])
    if len(df_chk101a) > 0:
        df_chk101 = df_chk101.append(df_chk101a)
    df_chk101b = rada_chk101_fn_new(df_ssd_newfrmattr, df_ssd_colist, 0, 7, 101, labellist1=[1], labellist2=[2])
    if len(df_chk101b) > 0:
        df_chk101 = df_chk101.append(df_chk101b)
        df_chk101 = df_chk101.loc[df_chk101.duplicated(['Form Name \n(eCRF Description)'], keep='first')]

    if len(df_chk101) > 0:
        f = {'Form Name \n(eCRF Description)': ','.join}
        df_chk101 = df_chk101.groupby(['Check_number'], as_index=False).agg(f)

    df_chk117 = rada_chk101_fn(df_ssd_frmattr, df_ssd_colist, 0, 7, 117, labellist1=[0])
    if len(df_chk117) > 0:
        f = {'Form Name \n(eCRF Description)': ','.join}
        df_chk117 = df_chk117.groupby(['Check_number'], as_index=False).agg(f)
    df_chk118 = rada_chk101_fn(df_ssd_colist, df_ssd_frmattr, 0, 0, 118, labellist1=[7])
    if len(df_chk118) > 0:
        if ('Form Name \n(eCRF Description)\n(Select Forms from the drop down list in the order of their visibility for each visit)' in  df_chk118.columns):
        # df_chk118['Form Name \n(eCRF Description)\n(Select Forms from the drop down list in the order of ' \
        #           'their visibility for each visit)'] =\
        #     df_chk118['Form Name \n(eCRF Description)\n(Select Forms from the drop down list in the order of ' \
        #               'their visibility for each visit)'].astype(str)
            df_chk118['Form Name \n(eCRF Description)\n(Select Forms from the drop down list in the order of their visibility for each visit)'].replace('nan', np.nan, regex=True, inplace=True)
            df_chk118['Form Name \n(eCRF Description)\n(Select Forms from the drop down list in the order of their visibility for each visit)'] \
                = df_chk118['Form Name \n(eCRF Description)\n(Select Forms from the drop down list in the order of their visibility for each visit)'].fillna("")
            f = {
                "Form Name \n(eCRF Description)\n(Select Forms from the drop down list in the order of their visibility for each visit)":
                    lambda x: ','.join(unique1(x))}
            df_chk118 = df_chk118.groupby(['Check_number'], as_index=False).agg(f)

    df_chk115 = rada_chk115_fn(df_ssd_newfrmattr, df_ssd_newfrmattr_df_getcol_ind, 'Missing_Cells', 115,
                               labellist1=[0, 1, 2, 4, 5], collist1=['Column 4'])
    df_chk116 = rada_chk115_fn(df_ssd_colist, df_ssd_colist_df_getcol_ind, 'Missing_Cells', 116,
                               labellist1=[0, 1, 2, 4, 5], collist1=['Column 3', 'Column 6', 'Column 7',
                                                                     'Column 9', 'Column 11'])
    # df_chk116 = rada_chk116_fn(df_ssd_colist, 116, labellist1=[0, 1, 3, 4, 7, 9])

    # Check#103: For lab forms, Analyte mapped should match the Field OID and corresponding Lab Analyte in Pre text

    # def rada_chk103_fn(df1, df2, col1, col2, col3, col4, col5, col6, chkno, **kwargs):
    #     df_chk = pd.DataFrame()
    #     labellist1 = kwargs.get('labellist1', [])
    #     if len(df1) > 0:
    #         if df1[col4].count() > 0:
    #             df1 = df1.loc[(~(df1[col4].isna()))]
    #             df1 = df1.iloc[:, labellist1]
    #             df2 = df2.iloc[:, labellist2]
    #             if len(df1) > 0:
    #                 df1 = df1.merge(df2, left_on=['AnalyteName'], right_on=['PARM'],
    #                                             how='left', suffixes=['', '_'], indicator=True)
    #                 df1.loc[(~(df1['FieldOID_spl'] == df1['AnalyteName'])) == True, ('Check_number')] = chkno
    #             df1.loc[(((df1.iloc[:, 0].isna())) | ((df1.iloc[:, 1].isna())) | (
    #                 (df1.iloc[:, 2].isna())) | ((df1.iloc[:, 3].isna()))) == True, ('Check_number')] = chkno
    #             df1 = df1.loc[(df1['Check_number'] == chkno)]
    #
    #     if (len(df1) > 0):
    #         df_chk = df1.loc[(df1['Check_number'] == chkno)]
    #
    # return df_chk
    #
    #
    # df_chk103 = rada_chk103_fn(df_flds_chks, df_Test_Category, 'PreText', 'PARMDES', 'ALTERNATE_LAB_TEST_DESC',
    #                             'AnalyteName', 'FieldOID_spl', 'PARM', 103,
    #                             labellist1=['PARM','PARMDES','ALTERNATE_LAB_TEST_DESC'],
    #                            labellist2=['PARM','PARMDES','ALTERNATE_LAB_TEST_DESC'])

    df_chk103 = pd.DataFrame()
    df_chk103a = pd.DataFrame()
    df_chk103b = pd.DataFrame()

    labels_chk103 = ['FormOID', 'FieldOID', 'FieldOID_spl', 'PreText', 'AnalyteName']
    if len(df_flds_chks) > 0:
        df_flds_chks = df_flds_chks.loc[:, labels_chk103]
        if df_flds_chks['AnalyteName'].count() > 0:
            df_chk103 = df_flds_chks.copy()
            df_chk103 = df_chk103.loc[(~(df_flds_chks['AnalyteName'].isna()))]
            if not df_chk103.empty:
                # df_chk103["FieldOID_spl"] = df_chk103["FieldOID"]
                # df_chk103["FieldOID_spl"] = df_chk103["FieldOID_spl"].str.split("_", n=1, expand=True)
                df_chk103a = df_chk103.copy()
                if len(df_chk103a) > 0:
                    df_chk103a.loc[(~(df_chk103a['FieldOID_spl'] == df_chk103a['AnalyteName'])) == True, (
                        'Check_number')] = 103
                    df_chk103a = df_chk103a.loc[(df_chk103a['Check_number'] == 103)]
                # df_chk103b = df_chk103.copy()
                df_Test_Category = df_Test_Category.loc[:, ['PARM', 'PARMDES', 'ALTERNATE_LAB_TEST_DESC']]
                # df_chk103b.loc[((~(df_chk103b["AnalyteName"].isin(df_Test_Category['PARM']))) ) == True, ('Check_number')] = 103
                df_chk103 = df_chk103.merge(df_Test_Category, left_on=['AnalyteName'], right_on=['PARM'],
                                            how='left', suffixes=['', '_'], indicator=True)
                df_chk103b = df_chk103.copy()
                if len(df_chk103b) > 0:
                    df_chk103b.loc[df_chk103b['_merge'] == 'left_only', 'Check_number'] = 103
                    df_chk103b = df_chk103b.drop(['_merge'], axis=1)
                df_chk103 = df_chk103.loc[df_chk103['_merge'] == 'both']
                df_chk103 = df_chk103.drop(['_merge'], axis=1)

                if len(df_chk103) > 0:
                    df_chk103.loc[(~((df_chk103['PreText'].str.strip() == df_chk103['PARMDES'].str.strip()) |
                                     (df_chk103['PreText'].str.strip() == df_chk103[
                                         'ALTERNATE_LAB_TEST_DESC'].str.strip())))
                                  == True, ('Check_number')] = 103
                    df_chk103 = df_chk103.loc[(df_chk103['Check_number'] == 103)]

    # Check#104: Comparing Coded value of any TEST with stage metadata
    # Check#105: User data string and Coded value length
    df_chk104 = pd.DataFrame()
    df_chk104a = pd.DataFrame()
    df_chk105 = pd.DataFrame()
    df_chk105a = pd.DataFrame()
    reqfields = ['test', 'TEST','tst', 'TST']
    reqfields2 = ['QSTSTLG']
    nreqfields = ['TEST_UNIT']
    if len(df_dicts) > 0:
        df_chk104_5 = df_dicts.loc[(df_dicts['DataDictionaryName'].astype(str).str.contains('|'.join(reqfields)) & (
            (df_dicts['DataDictionaryName'].astype(str).str.contains('|'.join(nreqfields)) == False))) == True]
        df_chk104_5a = df_dicts.loc[(df_dicts['DataDictionaryName'].astype(str).str.contains('|'.join(reqfields2))) == True]

        if len(df_chk104_5) > 0:
            df_chk105 = df_chk104_5.copy()
            if len(df_chk104_5) > 0:
                df_chk104_5.loc[
                    (~(df_chk104_5["CodedData"].isin(df_Test_Category['PARM']))) == True, ('Check_number')] = 104
                df_chk104 = df_chk104_5.loc[(df_chk104_5['Check_number'] == 104)]
            if len(df_chk105) > 0:
                df_chk105.loc[((df_chk105["CodedData"].str.len() > 8) | (
                        df_chk105["UserDataString"].str.len() > 40)) == True, 'Check_number'] = 105

                df_chk105 = df_chk105.loc[(df_chk105['Check_number'] == 105)]

        if not df_chk104_5a.empty:
            df_chk105a = df_chk104_5a.copy()
            if len(df_chk104_5a) > 0:
                df_chk104_5a.loc[
                    (~(df_chk104_5a["CodedData"].isin(df_Questions['TESTCD']))) == True, ('Check_number')] = 104
                df_chk104a = df_chk104_5a.loc[(df_chk104_5a['Check_number'] == 104)]
            if len(df_chk105a) > 0:
                df_chk105a.loc[(df_chk105a["CodedData"].str.len() > 8) == True, 'Check_number'] = 105

                df_chk105a = df_chk105a.loc[(df_chk105a['Check_number'] == 105)]

        # df_dictnams, df_flds, df_eSAE_forms, df_codelist_StageDomain, "CodedData",
        # "UserDataString", 'DataDictionaryName'
        # , "FormOID", "DataDictionaryName_split", "Controlled Terminology", 129,
        # labellist1 = ["DataDictionaryName", "DataDictionaryName_split"],
        # reqfields = ['test', 'TEST', 'tst', 'TST'],
        # reqfields2 = ['QSTSTLG'],
        # nreqfields = ['TEST_UNIT']

        #check 129

        def rada_chk129_fn(df1, df2, df3,df4, fld1, fld2, fld3, fld4,  fld5,  fld6, chkno, **kwargs):
            df_chk = pd.DataFrame()
            labellist1 = kwargs.get('labellist1', [])
            reqfields = kwargs.get('reqfields', [])
            reqfields2 = kwargs.get('reqfields2', [])
            nreqfields = kwargs.get('nreqfields', [])
            if len(df1) > 0:
                df1 = df1.loc[~df1[fld3].isna(), labellist1]
                df2 = df2.loc[~df2[fld3].isna()]
                #
                # df1.to_excel(
                #     r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS_\outputdf1_0.xlsx')
                # df2.to_excel(
                #     r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS_\outputdf2_0.xlsx')
            if len(df2) > 0:
                df2.drop_duplicates(subset=[fld3], keep='last', inplace=True)
            if (len(df2) > 0) & (len(df3) > 0):
                df2 = df2.loc[~df2[fld4].isin(df3[fld4])]
                # df3.to_excel(
                #     r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS_\outputdf3.xlsx')
                # df2.to_excel(
                #     r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS_\outputdf2.xlsx')
            if (len(df2) > 0) & (len(df1) > 0):
                df1 = df1.loc[df1[fld3].isin(df2[fld3])]
                # df1.to_excel(
                #     r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS_\outputdf1.xlsx')
            if len(df1) > 0:
                # df1a = df1.loc[((~df1[fld3].str.contains('|'.join(reqfields)))) == True]
                #
                # df1b = df1.loc[( (
                #     (df1[fld3].str.contains('|'.join(nreqfields)) == True))) == True]
                #
                # df1c = df1.loc[(
                #                (~df1[fld3].str.contains('|'.join(reqfields2)))) == True]

                # df1a.to_excel(
                #     r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS_\outputdf1a.xlsx')
                # df1b.to_excel(
                #     r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS_\outputdf1b.xlsx')
                # df1c.to_excel(
                #     r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS_\outputdf1c.xlsx')

                df1 = df1.loc[((~df1[fld3].astype(str).str.contains('|'.join(reqfields))) | (
                    (df1[fld3].astype(str).str.contains('|'.join(nreqfields)) == True)) |
                               (~df1[fld3].astype(str).str.contains('|'.join(reqfields2)))) == True]
                # df1.to_excel(
                #     r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS_\outputdf1_1.xlsx')
            if len(df1) > 0:

                df1.loc[~df1[fld5].isin(df4[fld6] ), 'Check_number'] = chkno

            if (len(df1) > 0):
                df_chk = df1.loc[(df1['Check_number'] == chkno)]

            return df_chk

        # df_dicts, df_flds, df_eSAE_forms, df_codelist_StageDomain, "CodedData",
        # "UserDataString", 'DataDictionaryName'
        # , "FormOID", "DataDictionaryName_split", "Controlled Terminology",
        # "dictname_UserDataString", "dictname_CodedData",
        # 130,
        # labellist1 = ["DataDictionaryName", "CodedData", "UserDataString",
        #               'DataDictionaryName_split',
        #               "dictname_UserDataString", "dictname_CodedData"],
        # reqfields = ['test', 'TEST', 'tst', 'TST'],
        # reqfields2 = ['QSTSTLG'],
        # nreqfields = ['TEST_UNIT']
        # )

        def rada_chk130_fn(df1, df2, df3,df4, fld1, fld2, fld3, fld4,  fld5,  fld6,  fld7,  fld8,
                            chkno, **kwargs):
            df_chk = pd.DataFrame()
            labellist1 = kwargs.get('labellist1', [])
            reqfields = kwargs.get('reqfields', [])
            reqfields2 = kwargs.get('reqfields2', [])
            nreqfields = kwargs.get('nreqfields', [])
            if len(df1) > 0:
                df1 = df1.loc[~df1[fld3].isna(), labellist1]
                df2 = df2.loc[~df2[fld3].isna()]
            if (len(df2) > 0) & (len(df3) > 0):
                df2 = df2.loc[~df2[fld4].isin(df3[fld4])]
            if (len(df2) > 0) & (len(df1) > 0):
                df1 = df1.loc[df1[fld3].isin(df2[fld3])]

            if len(df1) > 0:
                # df1_1 = df1.loc[(~df1[fld3].str.contains('|'.join(reqfields)))]
                #                 # df1_2 = df1.loc[df1[fld3].str.contains('|'.join(nreqfields)) == True]
                #                 # df1_3 = df1.loc[~df1[fld3].str.contains('|'.join(reqfields2))]
                df1 = df1.loc[((~df1[fld3].astype(str).str.contains('|'.join(reqfields))) | (
                    (df1[fld3].astype(str).str.contains('|'.join(nreqfields)) == True))) == True]

                # df1.to_excel(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS_\df1_130.xlsx')
                # df1_1.to_excel(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS_\df1_1_130.xlsx')
                # df1_2.to_excel(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS_\df1_2_130.xlsx')
                # df1_3.to_excel(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS_\df1_3_130.xlsx')
            if ((len(df1) > 0) ):
                df1 = df1.loc[df1[fld5].isin(df4[fld6])]
                # df1 = df1.merge(df4, left_on=[fld5], right_on=[fld6], how='left',
                #                 suffixes=['', '_'], indicator=True)
                #
                # df1 = df1.loc[df1['_merge'] == 'both']
                # df1.to_excel(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS_\df1_4.xlsx')
            if ((len(df1) > 0) ):

                df1.loc[(~df1[fld7].isin(df4[fld7])) | (~df1[fld8].isin(df4[fld8])), 'Check_number'] = chkno
                # df1.to_excel(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\ALS_\df1_5.xlsx')
            if (len(df1) > 0) :
                df_chk = df1.loc[(df1['Check_number'] == chkno)]
                # df_chk = df_chk.rename(columns={"dictname_UserDataString_": "UserDataString_CT",
                #                                 "dictname_CodedData_": "dictname_CodedData_CT"})

            return df_chk

        df_chk129 = rada_chk129_fn(df_dictnams , df_flds, df_eSAE_forms,df_codelist_StageDomain, "CodedData",
                                   "UserDataString", 'DataDictionaryName'
                                   , "FormOID", "DataDictionaryName_split","Controlled Terminology", 129,
                                   labellist1=["DataDictionaryName" , "DataDictionaryName_split"],
                                   reqfields=['test', 'TEST', 'tst', 'TST'],
                                   reqfields2 = ['QSTSTLG'],
                                   nreqfields = ['TEST_UNIT']
                                   )
        df_chk130 = rada_chk130_fn(df_dicts, df_flds, df_eSAE_forms,df_codelist_StageDomain, "CodedData",
                                   "UserDataString", 'DataDictionaryName'
                                   , "FormOID", "DataDictionaryName_split","Controlled Terminology",
                                   "dictname_UserDataString","dictname_CodedData",

                                   130,
                                   labellist1=["DataDictionaryName", "CodedData", "UserDataString",
                                               'DataDictionaryName_split',
                                   "dictname_UserDataString","dictname_CodedData"],
                                   reqfields=['test', 'TEST', 'tst', 'TST'],
                                   reqfields2 = ['QSTSTLG'],
                                   nreqfields = ['TEST_UNIT']
                                   )
    # Check#106: Lab analytes mapped should not be duplicated
    df_chk106 = pd.DataFrame()
    df_chk106 = df_flds_chks.copy()
    labels_chk106 = ['FormOID', 'FieldOID', 'PreText', 'AnalyteName']
    if not df_chk106.empty:
        df_chk106 = df_chk106.loc[:, labels_chk106]
        df_chk106 = df_chk106.loc[~(df_chk106['AnalyteName'].isnull())]
        if len(df_chk106) > 0:
            df_chk106.loc[(df_chk106.duplicated(['FormOID', 'AnalyteName'])), ('Check_number')] = 106
            df_chk106 = df_chk106.loc[(df_chk106['Check_number'] == 106)]

    # Check#107:Compare Fields against stage metadata
    df_chk107 = pd.DataFrame()
    labels_chk107 = ['FormOID', 'FieldOID', 'FieldOID_spl2', 'domain_ele', 'PreText', 'domain']
    if ((len(df_flds_chks) > 0) & (len(df_flds_dom) > 0)):

        df_chk107 = df_flds_dom.loc[(~(df_flds_dom['FieldOID'].isna())), labels_chk107]
        df_chk107 = df_chk107.loc[(~(df_chk107['FieldOID'].isin(df_globalfields['FieldOID'])))]
        df_chk107 = df_chk107.loc[(((df_chk107['domain'] == 'LB')) &
                                   (((df_chk107['FieldOID_spl2'] == 'LBVALUE') |
                                     (df_chk107['FieldOID_spl2'] == 'LVALUE') |
                                     (df_chk107['FieldOID'] == 'LBVALUE')))) == False]

        if len(df_chk107) > 0:
            df_chk107.loc[
                (~(df_chk107["domain_ele"].isin(df_stage_domain2['domain_ele']))) == True, ('Check_number')] = 107

    # Check#108: Compare Edit Checks against ECS
    df_chk108 = pd.DataFrame()
    valnames_s1 = []
    labels_chk108 = ['CheckName']
    if len(df_als_checks) > 0:
        df_chk108 = df_als_checks.copy()
        df_chk108 = df_chk108.loc[:, labels_chk108]
    if len(df_chk108) > 0:

        if len(df_ecs_ec_reshp1) > 0:
            valnames_s1 = df_ecs_ec_reshp1.iloc[:, 5].tolist()
        if (len(df_ecs_esae_reshp) > 0) & (len(df_ecs_esae_reshp) > 0):
            valnames_s2 = df_ecs_esae_reshp.iloc[:, 4].tolist()
            valnames_s1.extend(valnames_s2)
        if (len(df_ecs_pd_reshp) > 0) & (len(df_ecs_pd_reshp) > 0):
            valnames_s3 = df_ecs_pd_reshp.iloc[:, 13].tolist()
            valnames_s1.extend(valnames_s3)
        if (len(df_ssd_colist) > 0) & (len(df_ssd_colist_reshp1) > 0):
            valnames_s4 = df_ssd_colist_reshp1.loc[:, 'Validation Name'].tolist()
            valnames_s1.extend(valnames_s4)
        # df_Derivations_list = df_d
        df_Derivations_list = df_Derivations['DerivationName'].tolist()


        df_chk108 = df_chk108.append(pd.DataFrame(df_Derivations_list, columns=['CheckName']), ignore_index=True)

        df_chk108.loc[(~(df_chk108['CheckName'].isin(valnames_s1))) == True, 'Check_number'] = 108
        # df_chk108.loc[(~(df_chk108.iloc[:,0].isin(valnames_s1))) == True, ('Check_number')] = 108
        # df_chk108.loc[(~(df_chk108["CheckName"].isin(valnames_ec_esae_pd))) == True, ('Check_number')] = 108
        df_chk108 = df_chk108.loc[(df_chk108['Check_number'] == 108)]
        df_chk108['CheckName'].replace('nan', np.nan, regex=True, inplace=True)
        df_chk108['CheckName'] = df_chk108['CheckName'].fillna("")
        f = {'CheckName': lambda x: ','.join(unique1(x))}
        df_chk108 = df_chk108.groupby(['Check_number'], as_index=False).agg(f)

    # Check#109: Compare Forms against SSD
    df_chk109 = pd.DataFrame()
    if len(df_als_forms) > 0:
        df_chk109 = df_forms.copy()
    if (len(df_chk109) > 0) & (len(formnames_s1) > 0):
        df_chk109.loc[((~(df_chk109['DraftFormName'].isin(formnames_s1)))) == True, 'Check_number'] = 109
        df_chk109 = df_chk109.loc[(df_chk109['Check_number'] == 109)]

        # df_chk109 = df_chk109.rename(columns={"OID": "FormOID"})
        f = {'DraftFormName': ','.join}
        df_chk109 = df_chk109.groupby(['Check_number'], as_index=False).agg(f)

    # Check#110: there should be no missing cells in an ECS Edit checks
    # Check#120: there should be no missing cells in an ECS eSAE tab
    # Check#121: there should be no missing cells in an ECS PD Specs
    df_chk110a = df_ecs_ec.copy()
    df_chk110 = pd.DataFrame()

    # if len(df_chk110a) > 0:
    #     df_chk110a = df_chk110a.loc[(((df_chk110a.iloc[:, 1].isna()) | (df_chk110a.iloc[:, 2].isna()) |
    #                                   (df_chk110a.iloc[:, 3].isna()) | (df_chk110a.iloc[:, 5].isna()) |
    #                                   (df_chk110a.iloc[:, 6].isna()) |
    #                                   (df_chk110a.iloc[:, 7].isna()) | (df_chk110a.iloc[:, 9].isna())) | (
    #                                          (df_chk110a['Check Action'].str.contains("Open|Query", na=False))
    #                                          & (
    #                                              (((df_chk110a.iloc[:, 4].isna())) |
    #                                               ((df_chk110a.iloc[:, 10].isna())) |
    #                                               ((df_chk110a.iloc[:, 11].isna())) |
    #                                               ((df_chk110a.iloc[:, 12].isna())) |
    #                                               ((df_chk110a.iloc[:, 13].isna())))
    #                                          ) | (((df_chk110a.iloc[:, 4] == 'NA')) |
    #                                               ((df_chk110a.iloc[:, 10] == 'NA')) |
    #                                               ((df_chk110a.iloc[:, 11] == 'NA')) |
    #                                               ((df_chk110a.iloc[:, 12] == 'NA')) |
    #                                               ((df_chk110a.iloc[:, 13] == 'NA'))
    #                                               ))) == True]
    #
    #
    # if len(df_chk110a) > 0:
    #     df_chk110 = df_chk110a.iloc[:, [1, 2, 3, 4, 5, 6, 7, 9, 10, 11, 12, 13]]
    #     df_chk110['Check_number'] = 110

    df_chk120a = df_ecs_esae.copy()
    df_chk120 = pd.DataFrame()
    df_chk120_1 = pd.DataFrame()

    # if len(df_chk120a) > 0:
    #     df_chk120a = df_chk120a.loc[(((df_chk120a.iloc[:, 0].isna()) | (df_chk120a.iloc[:, 1].isna()) |
    #                                   (df_chk120a.iloc[:,2].isna()) |
    #                                   (df_chk120a.iloc[:,4].isna()) | (df_chk120a.iloc[:,5].isna()) |
    #                                   (df_chk120a.iloc[:,6].isna()) | (df_chk120a.iloc[:,7].isna())) |
    #                                  ((df_chk120a['Check Action'].str.contains("Open|Query", na=False)) &
    #                                   (((df_chk120a.iloc[:, 3] == 'NA')) | ((df_chk120a.iloc[:, 8] == 'NA')) |
    #                                    ((df_chk120a.iloc[:, 9] == 'NA')) |
    #                                    ((df_chk120a.iloc[:, 10] == 'NA')) | ((df_chk120a.iloc[:, 11] == 'NA'))
    #                                    | (df_chk120a.iloc[:, 3].isna()) |
    #                                    (df_chk120a.iloc[:, 9].isna()) | (df_chk120a.iloc[:, 10].isna()) |
    #                                    (df_chk120a.iloc[:, 11].isna())))) == True]
    #     if len(df_chk120a) > 0:
    #         df_chk120 = df_chk120a.iloc[:, [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11]]
    #         df_chk120['Check_number'] = 120

    df_chk120_1 = rada_chk120_1_fn(df_ecs_esae, df_ecs_esae_df_getcol_ind, 'Missing_Cells', 120,
                                   labellist1=[0, 1, 2, 3, 4, 5, 6, 7], collist1=['Column 4'])
    df_chk120_2 = rada_chk120_2_fn(df_ecs_esae, df_ecs_esae_df_getcol_ind, 'Missing_Cells', 'Check Action',
                                   ["OpenQuery", "Open Query", "openquery", "open query"], 120,
                                   labellist1=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11],
                                   collist1=['Column 1', 'Column 2', 'Column 3', 'Column 5', 'Column 6', 'Column 7',
                                             'Column 8'
                                             ])
    df_chk110_1 = rada_chk120_1_fn(df_ecs_ec, df_ecs_ec_df_getcol_ind, 'Missing_Cells', 110,
                                   labellist1=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9],
                                   collist1=['Column 1', 'Column 5', 'Column 9'])
    df_chk110_2 = rada_chk120_2_fn(df_ecs_ec, df_ecs_ec_df_getcol_ind, 'Missing_Cells', 'Check Action',
                                   ["OpenQuery", "Open Query", "openquery", "open query"], 110,
                                   labellist1=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13],
                                   collist1=['Column 1', 'Column 2', 'Column 3', 'Column 4', 'Column 6', 'Column 7',
                                             'Column 8', 'Column 9', 'Column 10'])
    # df_chk121 = pd.DataFrame()
    df_chk121 = rada_chk115_fn(df_ecs_pd, df_ecs_pd_df_getcol_ind, 'Missing_Cells', 121,
                               labellist1=[0, 1, 2, 3, 4, 5, 6, 7, 8, 10, 11, 12, 13, 14, 15, 16, 17, 18, 20],
                               collist1=['Column 10', 'Column 20'])
    # df_chk120b = df_ecs_esae.copy()
    # if len(df_ecs_esae) > 0:
    #
    #     df_chk120b = df_chk120b.loc[((df_chk120b['Check Action'].str.strip() == 'OpenQuery') |
    #                                  ((df_chk120b['Check Action'].str.strip() == 'Open Query')))==True]
    # if len(df_chk120b) > 0:
    #     df_chk120b = df_chk120b.loc[(((df_chk120b.iloc[:, 3] == 'NA')) | ((df_chk120b.iloc[:, 8] == 'NA')) |
    #                     ((df_chk120b.iloc[:, 9] == 'NA')) |
    #                     ((df_chk120b.iloc[:, 10] == 'NA')) | ((df_chk120b.iloc[:, 11] == 'NA'))
    #                     | (df_chk120b.iloc[:, 3].isna()) |
    #                     (df_chk120b.iloc[:,9].isna()) | (df_chk120b.iloc[:,10].isna()) |
    #                     (df_chk120b.iloc[:,11].isna()))==True]
    #     if len(df_chk120b) > 0:
    #         df_chk120b = df_chk120b.iloc[:, [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11]]
    #         df_chk120b['Check_number'] = 120

    # df_chk121a = df_ecs_pd.copy()

    # if len(df_chk121a) > 0:
    #     df_chk121a = df_chk121a.loc[((df_chk121a.iloc[:,0].isna()) |(df_chk121a.iloc[:,1].isna()) |
    #                                  (df_chk121a.iloc[:,2].isna()) | (df_chk121a.iloc[:,3].isna()) |
    #                                  (df_chk121a.iloc[:,4].isna()) | (df_chk121a.iloc[:,5].isna()) |
    #                                  (df_chk121a.iloc[:,6].isna()) | (df_chk121a.iloc[:,7].isna()) |
    #                                  (df_chk121a.iloc[:,8].isna()) | (df_chk121a.iloc[:,10].isna()) |
    #                                  (df_chk121a.iloc[:,11].isna()) | (df_chk121a.iloc[:,12].isna()) |
    #                                  (df_chk121a.iloc[:,13].isna()) | (df_chk121a.iloc[:,14].isna()) |
    #                                  (df_chk121a.iloc[:,15].isna()) | (df_chk121a.iloc[:,16].isna()) |
    #                                  (df_chk121a.iloc[:,17].isna()) | (df_chk121a.iloc[:,18].isna()) |
    #                                  (df_chk121a.iloc[:,20].isna()))==True]
    #     if len(df_chk121a) > 0:
    #         df_chk121 = df_chk121a.iloc[:, [0, 1, 2, 3, 4, 5, 6, 7, 8, 10, 11, 12, 13, 14, 15, 16, 17, 18, 20]]
    #         df_chk121['Check_number'] = 121

    # Check#111:Compare the edit check properties like Required Response, Manual Close, Query Text, target field etc against ECS.
    # Check#119:Compare the edit check properties like Required Response, Manual Close, Query Text, target field etc against ECS.
    df_chk111 = pd.DataFrame()
    df_chk119 = pd.DataFrame()
    df_chk111b = pd.DataFrame()
    labels = ['Query Field', 'Validation Name',
              'Check Type', 'Check Action', 'Edit Check Message',
              'Requires Response?', 'Requires  Manual Close?', 'Marking Group']
    if len(df_ca) > 0:
        # print("len(df_ca)", len(df_ca))
        df_chk111 = df_ca_oq.copy()
        df_chk119 = df_ca_oq.copy()
        # valnames_s1 = df_ecs_ec_reshp.iloc[:,5].tolist()
        # valnames_s2 = df_ecs_esae_reshp.iloc[:,4].tolist()
        # valnames_s3 = df_ecs_pd_reshp.iloc[:,13].tolist()
        # valnames_s1.extend(valnames_s2)
        # valnames_s1.extend(valnames_s3)
        # print("df_ecs_ec_reshp.columns",df_ecs_ec_reshp.columns)
        labels = ['Query Field', 'Validation Name',
                  'Check Type', 'Check Action', 'Edit Check Message',
                  'Requires Response?', 'Requires  Manual Close?', 'Marking Group']
        if len(df_ecs_ec) > 0:
            df_chk111_1 = df_ecs_ec_reshp1.copy()
            labels1 = labels.copy()
            [labels.remove(lst) for lst in labels1 if not (lst in df_chk111_1.columns)]
            # for lst in labels1:
            #     if not (lst in df_chk111_1.columns):
            #          labels.remove(lst)

            df_chk111_1 = df_chk111_1.loc[:, labels]

            # print("df_chk111_1.columns",df_chk111_1.columns)
            if (len(df_chk111) > 0) & ('CheckName' in df_chk111.columns):
                df_chk111['CheckName_lower'] = df_chk111['CheckName'].str.strip().str.lower()
            if (len(df_chk111_1) > 0) & ('Validation Name' in df_chk111_1.columns):
                df_chk111_1['Validation Name_lower'] = df_chk111_1['Validation Name'].str.strip().str.lower()
            # df_chk111.loc[df_chk111['FieldOID'].isna(),'FieldOID'] = df_chk111['VariableOID']
            if (len(df_chk111) > 0) & (len(df_chk111_1) > 0) & ('Validation Name_lower' in df_chk111_1.columns)\
                    & ('CheckName_lower' in df_chk111.columns):
                df_chk111 = df_chk111.merge(df_chk111_1, left_on=['CheckName_lower'], right_on=['Validation Name_lower']
                                        , how='left', suffixes=['', '_'], indicator=True)
                df_chk111 = df_chk111.loc[df_chk111['_merge'] == 'both']

            if (len(df_chk111) > 0) & ('_merge' in df_chk111.columns):
                mycols = set(df_chk111.columns)
                mycols.remove('_merge')
                df_chk111 = df_chk111[mycols]
                df_chk111b = df_chk111.copy()
                # df_chk111b = df_chk111.loc[df_chk111['FieldOID'].isna()]

                df_chk111 = df_chk111.loc[~df_chk111['FieldOID'].isna()]
                df_chk111b.loc[~df_chk111b['FieldOID'].isna(), 'FieldOID_1'] = df_chk111b['FieldOID']
                df_chk111b.loc[~df_chk111b['Query Field'].isna(), 'Query Field_1'] = df_chk111b['Query Field']
                df_chk111b.loc[df_chk111b['FieldOID'].isna(), 'FieldOID_1'] = "Nothing"
                df_chk111b.loc[df_chk111b['Query Field'].isna(), 'Query Field_1'] = "Nothing"
                if len(df_chk111b) > 0:
                    df_chk111b.loc[[x[0] in x[1] for x in zip(df_chk111b['Query Field_1'], df_chk111b['VariableOID'])],
                               'Check_number'] = 111
                # df_chk111 = df_chk111.loc[~df_chk111['FieldOID'].isna()]
            if (len(df_chk111) > 0) & (set(labels1).issubset(df_chk111.columns)):
                df_chk111.loc[((df_chk111['CheckName'] == df_chk111['Validation Name']) &
                               (((df_chk111['ActionOptions'].astype(str).str.contains('RequiresResponse')) &
                                 ((df_chk111['Requires Response?'].isna()) |
                                  (df_chk111['Requires Response?'] == 'NA')))
                                | ((df_chk111['ActionOptions'].astype(str).str.contains('RequiresManualClose')) &
                                   ((df_chk111['Requires  Manual Close?'].isna()) |
                                    (df_chk111['Requires  Manual Close?'] == 'NA'))) |
                                ((df_chk111['ActionOptions'].astype(str).str.contains('Site from System')) &
                                 ((df_chk111['Marking Group'].isna()) | (df_chk111['Marking Group'] == 'NA'))) |
                                (~(df_chk111['ActionString'].str.strip().str.lower() ==
                                   df_chk111['Edit Check Message'].str.strip().str.lower())) |
                                (~(df_chk111['FieldOID'].str.strip().str.lower() ==
                                   df_chk111['Query Field'].str.strip().str.lower())))) == True,
                              'Check_number'] = 111
            if (len(df_chk111b) > 0) & (set(labels1).issubset(df_chk111b.columns)):
                df_chk111b.loc[((df_chk111b['CheckName'] == df_chk111b['Validation Name']) &
                                (((df_chk111b['ActionOptions'].astype(str).str.contains('RequiresResponse')) &
                                  ((df_chk111b['Requires Response?'].isna()) |
                                   (df_chk111b['Requires Response?'] == 'NA')))
                                 | ((df_chk111b['ActionOptions'].astype(str).str.contains('RequiresManualClose')) &
                                    ((df_chk111b['Requires  Manual Close?'].isna()) |
                                     (df_chk111b['Requires  Manual Close?'] == 'NA'))) |
                                 ((df_chk111b['ActionOptions'].astype(str).str.contains('Site from System')) &
                                  ((df_chk111b['Marking Group'].isna()) | (df_chk111b['Marking Group'] == 'NA'))) |
                                 (~(df_chk111b['ActionString'].str.strip().str.lower() ==
                                    df_chk111b['Edit Check Message'].str.strip().str.lower())))) == True,
                               'Check_number'] = 111
                df_chk111b = df_chk111b.loc[(df_chk111b['Check_number'] == 111)]

                if len(df_chk111b) > 0:
                    df_chk111 = df_chk111.append(df_chk111b)

        if len(df_ecs_esae_reshp) > 0:
            labels = ['Query Field', 'Validation Name',
                       'Check Action', 'Edit Check Message',
                      'Requires Response?', 'Requires  Manual Close?', 'Marking Group']
            df_chk119_1 = df_ecs_esae_reshp.copy()

            labels1 = labels.copy()

            [labels.remove(lst) for lst in labels1 if not (lst in df_chk119_1.columns)]
            # for lst in labels1:
            #     if not (lst in df_chk119_1.columns):
            #         labels.remove(lst)

            df_chk119_1 = df_chk119_1.loc[:, labels]

            if (len(df_chk119) > 0) & (len(df_chk119_1) > 0) & ('Validation Name' in df_chk119_1.columns)\
                    & ('CheckName' in df_chk119.columns):
                df_chk119 = df_chk119.merge(df_chk119_1, left_on=['CheckName'], right_on=['Validation Name'],
                                            how='left', suffixes=['', '_'], indicator=True)
                df_chk119 = df_chk119.loc[df_chk119['_merge'] == 'both']

                mycols = set(df_chk119.columns)
                mycols.remove('_merge')
                df_chk119 = df_chk119[mycols]

            if (len(df_chk119) > 0) & (set(labels1).issubset(df_chk119.columns)):
                df_chk119.loc[((df_chk119['CheckName'] == df_chk119['Validation Name']) & (((df_chk119[
                    'ActionOptions'].astype(str).str.contains(
                    'RequiresResponse')) & ((df_chk119['Requires Response?'].isna()) | (
                        df_chk119['Requires Response?'] == 'NA'))) | ((df_chk119['ActionOptions'].astype(str).str.contains(
                    'RequiresManualClose')) & ((df_chk119['Requires  Manual Close?'].isna()) | (
                        df_chk119['Requires  Manual Close?'] == 'NA'))) | ((df_chk119[
                    'ActionOptions'].astype(str).str.contains(
                    'Site from System')) & ((df_chk119['Marking Group'].isna()) | (
                        df_chk119['Marking Group'] == 'NA'))) | (~(
                        df_chk119['ActionString'] == df_chk119['Edit Check Message'])))) == True, (
                                  'Check_number')] = 119

                if (len(df_chk119) > 0 ) &  ('Check_number' in df_chk119.columns) :
                    df_chk119 = df_chk119.loc[(df_chk119['Check_number'] == 119)]


    # Check#112: visit date OID to be SVSTDT but not SVSTDAT
    df_chk112 = pd.DataFrame()
    reqfields = ["SVSTDAT"]
    if len(df_als_checks1) > 0:
        df_chk112 = df_flds_chks.copy()
    if len(df_chk112) > 0:
        df_chk112.loc[(df_chk112['FieldOID'].astype(str).str.contains('|'.join(reqfields))) == True, ('Check_number')] = 112
        df_chk112 = df_chk112.loc[(df_chk112['Check_number'] == 112)]

    # Check#113:Check for mandatory fields like topic variables, Test variables, QS variables
    df_chk113 = pd.DataFrame()
    df_chk113_c = pd.DataFrame()
    df_chk113_1 = pd.DataFrame()
    df_chk113_m = pd.DataFrame()
    if ((len(df_global_stage_domain) > 0) & (len(df_flds_dom) > 0)):
        df_flds_dom = df_flds_dom.loc[(~(df_flds_dom['FieldOID'].isna()))]
        df_chk113_c = df_global_stage_domain2.copy()

        if len(df_chk113_c) > 0:

            df_chk113 = df_chk113_c.loc[((df_chk113_c['Data Domain'].isin(df_flds_dom['domain'])))]

            df_dmn_ufrm = pd.DataFrame()
            df_dmn_ufrm = df_flds_dom.groupby(['domain', 'FormOID']).size().reset_index(name='Freq')
            mycols = set(df_dmn_ufrm.columns)
            mycols.remove('Freq')
            df_dmn_ufrm = df_dmn_ufrm[mycols]

            for i in range(0, len(df_dmn_ufrm)):
                # if df_dmn_ufrm.loc[i, 'domain'] == 'DM':
                df_flds_dom_ufrm = pd.DataFrame()
                df_chk113_m = pd.DataFrame()
                df_flds_dom_ufrm = df_flds_dom.loc[((df_flds_dom["FormOID"] == df_dmn_ufrm.loc[i, 'FormOID']) & (
                        df_flds_dom["domain"] == df_dmn_ufrm.loc[i, 'domain'])) == True]
                # print("df_flds_dom_ufrm sum", df_flds_dom_ufrm.duplicated( keep='first').sum())
                df_flds_dom_ufrm = df_flds_dom_ufrm.rename(columns={"domain_ele": "domain_ele2"})
                # print("df_flds_dom_ufrm2 sum", df_flds_dom_ufrm.duplicated( keep='first').sum())
                df_chk113_m = df_chk113.merge(df_flds_dom_ufrm, left_on=['Data Domain'], right_on=['domain'],
                                              how='left',
                                              suffixes=['', '_'], indicator=True)
                # print("df_chk113_m sum", df_chk113_m.duplicated( keep='first').sum())
                df_chk113_m = df_chk113_m.loc[df_chk113_m['_merge'] == 'both']
                # print("df_chk113_m sum", df_chk113_m.duplicated(keep='first').sum())
                if len(df_chk113_m) > 0:
                    df_chk113_m.loc[
                        (~(df_chk113_m["domain_ele"].isin(df_flds_dom_ufrm['domain_ele2']))) == True, (
                            'Check_number')] = 113
                    # print("df_chk113_m sum", df_chk113_m.duplicated( keep='first').sum())
                    # print(df_chk113_m)
                    df_chk113_1 = df_chk113_1.append(df_chk113_m)
                    # print("df_chk113_1 sum", df_chk113_1.duplicated( keep='first').sum())
                    # print(df_chk113_1)

            if len(df_chk113_1) > 0:
                df_chk113_1 = df_chk113_1.drop_duplicates()
                df_chk113 = df_chk113_1.loc[(df_chk113_1['Check_number'] == 113)]

            labels = ['Check_number', 'Data State', 'Data Domain', 'Target Data Element', 'Element Core',
                      'Element Role', 'FormOID']
            labels1 = labels.copy()
            [labels.remove(lst) for lst in labels1 if not (lst in df_chk113.columns)]
            # for lst in labels1:
            #     if not (lst in df_chk113.columns):
            #         labels.remove(lst)
            df_chk113 = df_chk113.loc[:, labels]
            df_chk113.drop_duplicates(keep="first", inplace=True)

            f = {'FormOID': ','.join, 'Data State': 'first', 'Data Domain': 'first', 'Element Core': 'first',
                 'Element Role': 'first', 'Check_number': 'first'}
            df_chk113 = df_chk113.groupby(['Target Data Element'], as_index=False).agg(f)
            # print(df_chk113)



    # Check#114: Compare ECS and SSD for form OIDs
    # Check#122: Compare ECS and SSD for form Names
    #v0.2
    df_chk122a = pd.DataFrame()
    df_chk122_1 = pd.DataFrame()
    df_chk114a = pd.DataFrame()
    df_chk114a_1 = pd.DataFrame()
    labels_114a = ['FormOID']
    labels_122a = ['Form Name']

    if len(df_ecs_ec) > 0:
        if len(df_ecs_ec_reshp2) > 0:
            df_chk114a_1 = df_ecs_ec_reshp2.copy()
        if len(df_ecs_ec_reshp3) > 0:
            df_chk122_1 = df_ecs_ec_reshp3.copy()
        # print("df_ecs_ec_reshp count", len(df_ecs_ec_reshp))
        # print("df_ecs_ec_reshp count", len(df_chk114a))
        # print("ECG001esf",df_ssd_frmattr.loc[df_ssd_frmattr['Unique Form Annotation\nFormOID']=='ECG001'], 'Form Name \n(eCRF Description)')
        # print("ECG001esf 2",df_ssd_frmattr.loc[df_ssd_frmattr['Unique Form Annotation\nFormOID'].str.contains('ECG001', na= False)], 'Form Name \n(eCRF Description)')
        # print(df_ssd_frmattr['Unique Form Annotation\nFormOID'])
        # print(df_ssd_frmattr['Unique Form Annotation\nFormOID'].values)
        # labels = ['Query Field','Validation Name',
        #           'Check Type', 'Check Action', 'Edit Check Message',
        #           'Requires Response?', 'Requires  Manual Close?', 'Marking Group']
        # labels = ['FormOID','Form Name','Query Field','Validation Name']

        # for lst in labels:
        #     if not lst in df_chk111_1.columns:
        #         print(lst)
        if len(df_chk114a_1) > 0:
            df_chk114a = df_chk114a_1.loc[:, labels_114a]
        if len(df_chk114a) > 0:
            df_chk114a = df_chk114a.loc[~(df_chk114a.duplicated(labels_114a))]
        if len(df_chk114a) > 0:
            df_chk114a = df_chk114a.loc[((~(df_chk114a['FormOID'].isna())))]
        if len(df_chk114a) > 0:
            df_chk114a['FormOID'] = df_chk114a['FormOID'].str.strip()
        if (len(df_chk114a) > 0) & (len(df_ssd_frmattr) > 0):
            # print(df_ssd_frmattr.columns)

            df_chk114a.loc[(~(df_chk114a['FormOID'].str.lower().isin(
                df_ssd_frmattr['Unique Form Annotation\nFormOID'].str.lower())
            )) == True, 'Check_number'] = 114

        if (len(df_chk114a) > 0) & ('Check_number' in df_chk114a.columns.values.tolist()):
            df_chk114a = df_chk114a.loc[(df_chk114a['Check_number'] == 114)]
            df_chk114a['FormOID'].replace('nan', np.nan, regex=True, inplace=True)
            df_chk114a['FormOID'] = df_chk114a['FormOID'].fillna("")

            f = {'FormOID': lambda x: ','.join(unique1(x))}
            df_chk114a = df_chk114a.groupby(['Check_number'], as_index=False).agg(f)
        #     # print("df_chk114a", df_chk114a.iloc[:, 1])
        #     # print("df_ecs_ec_reshp count", len(df_chk114a))
        #     # print("df_chk114a",df_ssd_frmattr['Form Name \n(eCRF Description)'].values)
        if len(df_chk122_1) > 0:
            df_chk122_1 = df_chk122_1.loc[((~(df_chk122_1['Form Name'].isna())))]
        if len(df_chk122_1) > 0:
            df_chk122_1 = df_chk122_1.loc[df_chk122_1['Form Name'].astype(str).str.contains('[a-zA-Z0-9]', na=False)]
        if len(df_chk122_1) > 0:
            df_chk122_1 = df_chk122_1.loc[:, labels_122a]
            df_chk122_1 = df_chk122_1.loc[~(df_chk122_1.duplicated(labels_122a))]
        if ((len(df_chk122_1) > 0) & (len(df_ssd_frmattr) > 0)):
            # print(df_ssd_frmattr.columns)
            df_chk122_1.loc[((~(df_chk122_1['Form Name'].str.lower().isin(
                df_ssd_frmattr['Form Name \n(eCRF Description)'].str.lower())))) == True, 'Check_number'] = 122
        if (len(df_chk122_1) > 0) & ('Check_number' in df_chk122_1.columns.values.tolist()):
            df_chk122a = df_chk122_1.loc[(df_chk122_1['Check_number'] == 122)]
            df_chk122a['Form Name'].replace('nan', np.nan,regex = True,inplace = True)
            df_chk122a['Form Name'] = df_chk122a['Form Name'].fillna("")
            f = {'Form Name': lambda x: ','.join(unique1(x))}
            df_chk122a = df_chk122a.groupby(['Check_number'], as_index=False).agg(f)

    ########
    df_chk114b = pd.DataFrame()
    df_chk122b = pd.DataFrame()
    if len(df_ecs_esae) > 0:
        if len(df_ecs_esae_reshp) > 0:
            df_chk114b_1 = df_ecs_esae_reshp.copy()

        # labels = ['FormOID','Form Name','Query Field','Validation Name']
        # for lst in labels:
        #     if not lst in df_chk111_1.columns:
        #         print(lst)
        if (len(df_chk114b_1) > 0) & ('FormOID' in df_chk114b_1.columns):
            df_chk114b_1['FormOID'] = df_chk114b_1['FormOID'].astype(str).str.strip()
            df_chk114b_1['FormOID'] = df_chk114b_1['FormOID'].astype(str).str.strip('\n')
        if (len(df_chk114b_1) > 0) & ('Form Name' in df_chk114b_1.columns):
            df_chk114b_1['Form Name'] = df_chk114b_1['Form Name'].astype(str).str.strip()
            df_chk114b_1['Form Name'] = df_chk114b_1['Form Name'].astype(str).str.strip('\n')
        if (len(df_chk114b_1) > 0) & ('FormOID' in df_chk114b_1.columns):
            df_chk114b = df_chk114b_1.loc[~(df_chk114b_1['FormOID'].isna())]
            if (len(df_chk114b) > 0) :
                df_chk114b = df_chk114b.loc[df_chk114b['FormOID'].astype(str).str.contains('[a-zA-Z0-9]', na=False)]
        if (len(df_chk114b_1) > 0) & ('Form Name' in df_chk114b_1.columns):
            df_chk122b = df_chk114b_1.loc[~(df_chk114b_1['Form Name'].isna())]
            if (len(df_chk122b) > 0):
                df_chk122b = df_chk122b.loc[df_chk122b['Form Name'].astype(str).str.contains('[a-zA-Z0-9]', na=False)]
        if (len(df_chk114b) > 0):
            df_chk114b = df_chk114b.loc[:, labels_114a]
        # df_chk114b = df_chk114b.loc[~(df_chk114b.duplicated(labels_114a))]
        if (len(df_chk122b) > 0):
            df_chk122b = df_chk122b.loc[:, labels_122a]
            df_chk122b = df_chk122b.loc[~(df_chk122b.duplicated(labels_122a))]
        #     print("df_ssd_frmattr length",len(df_ssd_frmattr))
        if (((len(df_chk114b) > 0) & (len(df_ssd_frmattr) > 0))):
            df_chk114b.loc[(~(df_chk114b['FormOID'].str.lower().isin(
                df_ssd_frmattr['Unique Form Annotation\nFormOID'].str.lower())
            )) == True, 'Check_number'] = 114
        if (len(df_chk114b) > 0) & ('Check_number' in df_chk114b.columns.values.tolist()):
            df_chk114b = df_chk114b.loc[(df_chk114b['Check_number'] == 114)]
            df_chk114b['FormOID'].replace('nan', np.nan, regex=True, inplace=True)
            df_chk114b['FormOID'] =  df_chk114b['FormOID'].fillna("")

            f = {'FormOID': lambda x: ','.join(unique1(x))}
            df_chk114b = df_chk114b.groupby(['Check_number'], as_index=False).agg(f)

        #print("df_ssd_frmattr length",len(df_ssd_frmattr))
        if (((len(df_chk122b) > 0) & (len(df_ssd_frmattr) > 0))):
            # print("df_chk122b['Form Name']", df_chk122b['Form Name'])
            df_chk122b.loc[(~(df_chk122b['Form Name'].str.lower().isin(
                df_ssd_frmattr['Form Name \n(eCRF Description)'].str.lower()))
                            ) == True, 'Check_number'] = 122
            # print("df_chk122b['Form Name']", df_chk122b['Form Name'].str.lower())
            # print("df_ssd_frmattr['Form Name \n(eCRF Description)']",
            #       df_ssd_frmattr['Form Name \n(eCRF Description)'].str.lower())
        if (len(df_chk122b) > 0) & ('Check_number' in df_chk122b.columns.values.tolist()):
            df_chk122b = df_chk122b.loc[df_chk122b['Check_number'] == 122]
            df_chk122b['Form Name'].replace('nan', np.nan,regex = True,inplace = True)
            df_chk122b['Form Name'] = df_chk122b['Form Name'].fillna("")
            f = {'Form Name': lambda x: ','.join(unique1(x))}
            # f = {'Form Name': ','.join}
            df_chk122b = df_chk122b.groupby(['Check_number'], as_index=False).agg(f)

    df_chk114c = pd.DataFrame()
    df_chk114c_1 = pd.DataFrame()
    df_chk122c = pd.DataFrame()
    labels_114c = ['Form']
    labels_122c = ['Primary Form Name']

    if len(df_ecs_pd) > 0:
        if len(df_ecs_pd_reshp) > 0:
            df_chk114c_1 = df_ecs_pd_reshp.copy()

            #     # labels = ['Form','Primary Form Name','Query Field','Validation Name']
            #     # for lst in labels:
            #     #     if not lst in df_chk111_1.columns:
            #     #         print(lst)
            if (len(df_chk114c_1) > 0) & ('Form' in df_chk114c_1.columns):
                if (df_chk114c_1['Form'].count() > 0 ):
                    df_chk114c_1['Form Name'] = df_chk114c_1['Form'].str.strip()
            if (len(df_chk114c_1) > 0) & ('Primary Form Name' in df_chk114c_1.columns):
                if (df_chk114c_1['Primary Form Name'].count() > 0):
                    df_chk114c_1['Primary Form Name'] = df_chk114c_1['Primary Form Name'].str.strip()
                    if (len(df_chk114c_1) > 0) & ('Form' in df_chk114c_1.columns):
                        df_chk114c = df_chk114c_1.loc[((~(df_chk114c_1['Form'].isna())))]
                if (len(df_chk114c) > 0) & ('Form' in df_chk114c.columns):
                    if (df_chk114c['Form'].count() > 0):
                        df_chk114c = df_chk114c.loc[df_chk114c['Form'].astype(str).str.contains('[a-zA-Z0-9]', na=False)]
            if (len(df_chk114c_1) > 0) & ('Primary Form Name' in df_chk114c_1.columns):
                df_chk122c = df_chk114c_1.loc[~(df_chk114c_1['Primary Form Name'].isna())]
                if len(df_chk122c) > 0:
                    df_chk122c = df_chk122c.loc[df_chk122c['Primary Form Name'].astype(str).str.contains('[a-zA-Z0-9]', na=False)]
            if (len(df_chk114c) > 0):
                df_chk114c = df_chk114c.loc[:, labels_114c]
                df_chk114c = df_chk114c.loc[~(df_chk114c.duplicated(labels_114c))]
            if (len(df_chk122c) > 0):
                df_chk122c = df_chk122c.loc[:, labels_122c]
                df_chk122c = df_chk122c.loc[~(df_chk122c.duplicated(labels_122c))]
        if ((len(df_chk114c) > 0) & (len(df_ssd_frmattr) > 0)):
            df_chk114c.loc[(~(
                df_chk114c['Form'].str.lower().isin(df_ssd_frmattr['Unique Form Annotation\nFormOID'].str.lower())))
                           == True, 'Check_number'] = 114
        if (len(df_chk114c) > 0) & ('Check_number' in df_chk114c.columns.values.tolist()):
            df_chk114c = df_chk114c.loc[(df_chk114c['Check_number'] == 114)]

            df_chk114c['Form'].replace('nan', np.nan,regex = True,inplace = True)

            df_chk114c['Form'] = df_chk114c['Form'].fillna("")

            f = {'Form': lambda x: ','.join(unique1(x))}
            df_chk114c = df_chk114c.groupby(['Check_number'], as_index=False).agg(f)

        if ((len(df_chk122c) > 0) & (len(df_ssd_frmattr) > 0)):
            df_chk122c.loc[(~(df_chk122c['Primary Form Name'].str.lower().isin(
                df_ssd_frmattr['Form Name \n(eCRF Description)'].str.lower())))
                           == True, 'Check_number'] = 122
        if (len(df_chk122c) > 0) & ('Check_number' in df_chk122c.columns.values.tolist()):
            df_chk122c = df_chk122c.loc[(df_chk122c['Check_number'] == 122)]
            df_chk122c['Primary Form Name'].replace('nan', np.nan, regex=True, inplace=True)
            df_chk122c['Primary Form Name'] = df_chk122c['Primary Form Name'].fillna("")
            f = {'Primary Form Name': lambda x: ','.join(unique1(x))}
            df_chk122c = df_chk122c.groupby(['Check_number'], as_index=False).agg(f)

    my_cols_list = ['ACTION TAKEN', 'IF NO, DBD COMMENTS', 'DBP 5-2-1 ID', 'DATE', 'DB LEAD APPROVAL']

    df_radachecks = pd.DataFrame(columns=my_cols_list)
    df_radachecks = df_radachecks.append(df_chk01)
    df_radachecks = df_radachecks.append(df_chk02)
    df_radachecks = df_radachecks.append(df_chk03)
    df_radachecks = df_radachecks.append(df_chk04)
    df_radachecks = df_radachecks.append(df_chk05)
    df_radachecks = df_radachecks.append(df_chk06)
    df_radachecks = df_radachecks.append(df_chk07)
    df_radachecks = df_radachecks.append(df_chk08)
    df_radachecks = df_radachecks.append(df_chk09)
    df_radachecks = df_radachecks.append(df_chk10)
    df_radachecks = df_radachecks.append(df_chk11)
    df_radachecks = df_radachecks.append(df_chk12)
    df_radachecks = df_radachecks.append(df_chk13)
    df_radachecks = df_radachecks.append(df_chk14)
    df_radachecks = df_radachecks.append(df_chk15)
    df_radachecks = df_radachecks.append(df_chk16)
    df_radachecks = df_radachecks.append(df_chk16_1)
    df_radachecks = df_radachecks.append(df_chk17)
    df_radachecks = df_radachecks.append(df_chk18)
    df_radachecks = df_radachecks.append(df_chk19)
    df_radachecks = df_radachecks.append(df_chk20)
    df_radachecks = df_radachecks.append(df_chk21)
    df_radachecks = df_radachecks.append(df_chk22)
    df_radachecks = df_radachecks.append(df_chk23)
    df_radachecks = df_radachecks.append(df_chk24)
    df_radachecks = df_radachecks.append(df_chk25)
    df_radachecks = df_radachecks.append(df_chk26)
    df_radachecks = df_radachecks.append(df_chk27)
    df_radachecks = df_radachecks.append(df_chk28)
    df_radachecks = df_radachecks.append(df_chk29)
    df_radachecks = df_radachecks.append(df_chk30)
    df_radachecks = df_radachecks.append(df_chk31)
    df_radachecks = df_radachecks.append(df_chk32)
    df_radachecks = df_radachecks.append(df_chk33)
    df_radachecks = df_radachecks.append(df_chk34)
    df_radachecks = df_radachecks.append(df_chk35)
    df_radachecks = df_radachecks.append(df_chk36)
    df_radachecks = df_radachecks.append(df_chk37)
    df_radachecks = df_radachecks.append(df_chk38)
    df_radachecks = df_radachecks.append(df_chk39)
    df_radachecks = df_radachecks.append(df_chk40)
    df_radachecks = df_radachecks.append(df_chk41)
    df_radachecks = df_radachecks.append(df_chk42)
    df_radachecks = df_radachecks.append(df_chk43)
    df_radachecks = df_radachecks.append(df_chk44)
    df_radachecks = df_radachecks.append(df_chk45)
    df_radachecks = df_radachecks.append(df_chk46)
    df_radachecks = df_radachecks.append(df_chk47)
    df_radachecks = df_radachecks.append(df_chk48)
    df_radachecks = df_radachecks.append(df_chk49)
    df_radachecks = df_radachecks.append(df_chk50)
    df_radachecks = df_radachecks.append(df_chk51)
    df_radachecks = df_radachecks.append(df_chk52)
    df_radachecks = df_radachecks.append(df_chk53)
    df_radachecks = df_radachecks.append(df_chk54)
    df_radachecks = df_radachecks.append(df_chk55_1)
    df_radachecks = df_radachecks.append(df_chk55_2)
    df_radachecks = df_radachecks.append(df_chk56)
    df_radachecks = df_radachecks.append(df_chk57)
    df_radachecks = df_radachecks.append(df_chk58)
    df_radachecks = df_radachecks.append(df_chk59)
    df_radachecks = df_radachecks.append(df_chk60)
    df_radachecks = df_radachecks.append(df_chk61)
    df_radachecks = df_radachecks.append(df_chk62)
    df_radachecks = df_radachecks.append(df_chk63)
    df_radachecks = df_radachecks.append(df_chk64)
    df_radachecks = df_radachecks.append(df_chk65)
    df_radachecks = df_radachecks.append(df_chk66)
    df_radachecks = df_radachecks.append(df_chk67)
    df_radachecks = df_radachecks.append(df_chk68)
    df_radachecks = df_radachecks.append(df_chk69)
    df_radachecks = df_radachecks.append(df_chk70)
    df_radachecks = df_radachecks.append(df_chk71)
    df_radachecks = df_radachecks.append(df_chk72)
    df_radachecks = df_radachecks.append(df_chk73)
    df_radachecks = df_radachecks.append(df_chk74)
    df_radachecks = df_radachecks.append(df_chk75)
    df_radachecks = df_radachecks.append(df_chk76)
    df_radachecks = df_radachecks.append(df_chk77)
    df_radachecks = df_radachecks.append(df_chk78)
    df_radachecks = df_radachecks.append(df_chk79)
    df_radachecks = df_radachecks.append(df_chk80)
    df_radachecks = df_radachecks.append(df_chk81)
    df_radachecks = df_radachecks.append(df_chk82)
    df_radachecks = df_radachecks.append(df_chk83)
    df_radachecks = df_radachecks.append(df_chk84)
    df_radachecks = df_radachecks.append(df_chk85)
    df_radachecks = df_radachecks.append(df_chk86)
    df_radachecks = df_radachecks.append(df_chk87)
    df_radachecks = df_radachecks.append(df_chk88)
    df_radachecks = df_radachecks.append(df_chk89)
    df_radachecks = df_radachecks.append(df_chk90)
    df_radachecks = df_radachecks.append(df_chk91)
    df_radachecks = df_radachecks.append(df_chk92)
    df_radachecks = df_radachecks.append(df_chk93)
    df_radachecks = df_radachecks.append(df_chk94)
    df_radachecks = df_radachecks.append(df_chk95)
    df_radachecks = df_radachecks.append(df_chk95_1)
    df_radachecks = df_radachecks.append(df_chk96)
    df_radachecks = df_radachecks.append(df_chk97)
    df_radachecks = df_radachecks.append(df_chk98_1)
    df_radachecks = df_radachecks.append(df_chk98_2)
    df_radachecks = df_radachecks.append(df_chk98_3)
    df_radachecks = df_radachecks.append(df_chk98_4)
    df_radachecks = df_radachecks.append(df_chk99a)
    df_radachecks = df_radachecks.append(df_chk99b)
    df_radachecks = df_radachecks.append(df_chk100a)
    df_radachecks = df_radachecks.append(df_chk100b)
    df_radachecks = df_radachecks.append(df_chk101)
    df_radachecks = df_radachecks.append(df_chk102)
    df_radachecks = df_radachecks.append(df_chk103)
    df_radachecks = df_radachecks.append(df_chk103a)
    df_radachecks = df_radachecks.append(df_chk103b)
    df_radachecks = df_radachecks.append(df_chk104)
    df_radachecks = df_radachecks.append(df_chk104a)
    df_radachecks = df_radachecks.append(df_chk105)
    df_radachecks = df_radachecks.append(df_chk105a)
    df_radachecks = df_radachecks.append(df_chk106)
    df_radachecks = df_radachecks.append(df_chk107)
    df_radachecks = df_radachecks.append(df_chk108)
    df_radachecks = df_radachecks.append(df_chk109)
    df_radachecks = df_radachecks.append(df_chk110_1)
    df_radachecks = df_radachecks.append(df_chk110_2)
    df_radachecks = df_radachecks.append(df_chk111)
    df_radachecks = df_radachecks.append(df_chk112)
    df_radachecks = df_radachecks.append(df_chk113)
    df_radachecks = df_radachecks.append(df_chk114a)
    df_radachecks = df_radachecks.append(df_chk114b)
    df_radachecks = df_radachecks.append(df_chk114c)
    df_radachecks = df_radachecks.append(df_chk115)
    df_radachecks = df_radachecks.append(df_chk116)
    df_radachecks = df_radachecks.append(df_chk117)
    df_radachecks = df_radachecks.append(df_chk118)
    df_radachecks = df_radachecks.append(df_chk119)
    df_radachecks = df_radachecks.append(df_chk120_1)
    df_radachecks = df_radachecks.append(df_chk120_2)
    df_radachecks = df_radachecks.append(df_chk121)
    df_radachecks = df_radachecks.append(df_chk122a)
    df_radachecks = df_radachecks.append(df_chk122b)
    df_radachecks = df_radachecks.append(df_chk122c)
    df_radachecks = df_radachecks.append(df_chk123)
    df_radachecks = df_radachecks.append(df_chk123_1)
    df_radachecks = df_radachecks.append(df_chk124)
    df_radachecks = df_radachecks.append(df_chk125)
    df_radachecks = df_radachecks.append(df_chk126)
    df_radachecks = df_radachecks.append(df_chk127)
    df_radachecks = df_radachecks.append(df_chk128)
    df_radachecks = df_radachecks.append(df_chk129)
    df_radachecks = df_radachecks.append(df_chk130)


    if len(df_radachecks) > 0:
        df_radachecks = df_radachecks.loc[:, ~df_radachecks.columns.str.contains('^Unnamed')]
        # df_radachecks.drop_duplicates(keep="first", inplace=True)
        # Merging with instructions
        # print(df_instr)
        if len(df_instr) > 0:
            df_instr1 = df_instr.filter(items=['Check_number', 'Query', 'Applicability'])
            # df_instr1 = df_instr[:, ['Check_number', 'Query']]
        # reqnum = ['^1', '^2', '^3', '^4', '^5', '^6', '^7', '^8', '^9', '^0']
        # df_radacheckssds=df_radachecks.loc[df_radachecks['Check_number'].str.contains('|'.join(reqnum))==False]
        # print(df_radacheckssds['Check_number'])
        # df_radachecks['Check_number_1'] = df_radachecks['Check_number'].astype(int)
        # print(df_radachecks.columns)
        # print(df_instr1.columns)
        mycols = set(df_radachecks.columns)
        if '_merge' in df_radachecks:
            mycols.remove('_merge')
            df_radachecks = df_radachecks[mycols]

        df_radachecks = df_radachecks.merge(df_instr1, left_on=['Check_number'], right_on=['Check_number'],
                                            how='left', suffixes=['', '_'], indicator=True)
        # df_radachecks=pd.concat([df_radachecks, df_instr], axis=1, sort=False)
        # print(df_radachecks)
        # df_flds_chks1=df_flds_chks
        df_radachecks = df_radachecks.loc[df_radachecks['_merge'] == 'both']

        if len(df_radachecks) > 0:
            # print(df_radachecks.columns)

            mycols = set(df_radachecks.columns)
            mycols.remove('_merge')
            df_radachecks2 = df_radachecks[mycols]


            labels = ['Applicability', 'Check_number', 'Query', 'FormOID', 'FieldOID',
                      'PreText', 'DataFormat', 'ControlType', 'QueryNonConformance','IsVisible',
                      'DoesNotBreakSignature', 'ViewRestrictions', 'CheckName',
                      'DataDictionaryName', 'DefaultValue', 'AnalyteName', 'DataFormat_global',
                                         'PreText_global', 'DataDictionaryName_global',
                      'ACTION TAKEN', 'IF NO, DBP COMMENTS',
                      'DBP 5-2-1 ID', 'DATE', 'DB LEAD APPROVAL']

            df_field_fchecks = df_radachecks2.loc[df_radachecks2['Applicability'] == 'Field']
            if not df_field_fchecks.empty:
                labels1 = labels.copy()
                [labels.remove(lst) for lst in labels1 if not (lst in df_field_fchecks.columns)]
                # for lst in labels1:
                #     if not (lst in df_field_fchecks.columns):
                #         labels.remove(lst)
            if not df_field_fchecks.empty:
                df_field_fchecks = df_field_fchecks.loc[:, labels]
            labels = ['Applicability', 'Check_number', 'Query', 'DataDictionaryName', "FieldOID", "CodedData",
                      "UserDataString", "DataFormat", 'SASLabel',
                      'ACTION TAKEN', 'IF NO, DBP COMMENTS',
                      'DBP 5-2-1 ID', 'DATE', 'DB LEAD APPROVAL']

            df_dict_fchecks = df_radachecks2.loc[df_radachecks2['Applicability'] == 'Data Dictionary']
            if not df_dict_fchecks.empty:
                labels1 = labels.copy()
                [labels.remove(lst) for lst in labels1 if not (lst in df_dict_fchecks.columns)]
                # for lst in labels1:
                #     if not (lst in df_dict_fchecks.columns):
                #         labels.remove(lst)
                df_dict_fchecks = df_dict_fchecks.loc[:, labels]

            df_als_fchecks = df_radachecks2.loc[df_radachecks2['Applicability'] == 'ALS']
            if not df_als_fchecks.empty:
                labels = ['Applicability', 'Check_number', 'Query', 'FormOID', 'DraftFormName', 'FieldOID',
                          'VariableOID', 'PreText', 'DataDictionaryName',
                          'DefaultValue', 'CheckName', 'AnalyteName', 'domain', 'Data State', 'Data Domain',
                          'ActionString', 'ActionOptions',
                          'Target Data Element', 'Element Core', 'Element Role',
                          'Query Field', 'Validation Name', 'Edit Check Message',
                          'Requires Response?', 'Requires  Manual Close?', 'Marking Group',
                          'ACTION TAKEN',
                          'IF NO, DBP COMMENTS', 'DBP 5-2-1 ID', 'DATE', 'DB LEAD APPROVAL']
                labels1 = labels.copy()
                [labels.remove(lst) for lst in labels1 if not (lst in df_als_fchecks.columns)]
                # for lst in labels1:
                #     if not (lst in df_als_fchecks.columns):
                #         labels.remove(lst)
                df_als_fchecks = df_als_fchecks.loc[:, labels]

            df_ssd_fchecks = df_radachecks2.loc[df_radachecks2['Applicability'] == 'SSD']
            if not df_ssd_fchecks.empty:
                labels = ['Applicability', 'Check_number', 'Query', 'Form Name \n(eCRF Description)',
                          'Missing_Cells',
                          'Unique Form Annotation\nFormOID', 'View Restrictions', 'Entry Restrictions',
                          'Request Type',
                          'Field Name (Applicable only for New field request type)',
                          'View Restrictions \n(Applicable only if different from Role restrictions guidance document)',
                          'Entry Restrictions\n(Applicable only if different from Role restrictions guidance document)'
                    , 'Epoch', 'Visit Name \n(Folder Name)', 'Visit Name \n(Sub Folder Name)',
                          'Folder OID', 'Matrix-Dynamics User requirement for folder(If any)',
                          'Target', 'Overdue',
                          'Form Name \n(eCRF Description)\n(Select Forms from the drop down list in the order of their visibility for each visit)',
                          'Matrix-Dynamics user requirements on form (if any)', 'Validation Name', 'STUDY',
                          'DRUG CODE FROM RaveX (ECTRT)', 'DRUG CODE FROM ARGUS', 'STUDY NAME',
                          'STUDY INDICATION', 'ACTION TAKEN', 'IF NO, DBP COMMENTS', 'DBP 5-2-1 ID', 'DATE',
                          'DB LEAD APPROVAL']
                labels1 = labels.copy()
                [labels.remove(lst) for lst in labels1 if not (lst in df_ssd_fchecks.columns)]
                # for lst in labels1:
                #     if not (lst in df_ssd_fchecks.columns):
                #         labels.remove(lst)
                df_ssd_fchecks = df_ssd_fchecks.loc[:, labels]

            df_ecs_fchecks = df_radachecks2.loc[df_radachecks2['Applicability'] == 'ECS']
            if not df_ecs_fchecks.empty:
                labels = ['Applicability', 'Check_number', 'Query', 'FormOID', 'Form Name', 'Edit check category',
                          'Missing_Cells',
                          'Query Field', 'Validation Name', 'Check Type',  'Check Action',
                          'User Requirement', 'Edit Check Message', 'Requires Response?', 'Requires  Manual Close?',
                          'Marking Group',
                          'PD Identification Method', 'PD Short Description',
                          'Protocol Deviation Term for reporting',
                          'PD Identifier', 'PD start date', 'Protocol Deviation Coded Term',
                          'Protocol Visit Number/Description', 'Requirement of Medical Review (Y/N)',
                          'Blinded (Y/N)', 'PD Obsolete', 'Form', 'Primary Form Name', 'PD Check Message',
                          'Require Manual Close? (Y/N)',
                          'ACTION TAKEN', 'IF NO, DBP COMMENTS', 'DBP 5-2-1 ID', 'DATE', 'DB LEAD APPROVAL']
                labels1 = labels.copy()
                [labels.remove(lst) for lst in labels1 if not (lst in df_ecs_fchecks.columns)]
                # for lst in labels1:
                #     if not (lst in df_ecs_fchecks.columns):
                #         labels.remove(lst)
                df_ecs_fchecks = df_ecs_fchecks.loc[:, labels]

            df_frmrestricts_fchecks = df_radachecks2.loc[df_radachecks2['Applicability'] == 'Restrictions']
            if not df_frmrestricts_fchecks.empty:
                labels = ['Applicability', 'Check_number', 'Query', 'OID',  'DraftFormName','FormOID', 'FieldOID'
                    , 'ViewRestrictions',
                          'EntryRestrictions',
                          'ACTION TAKEN', 'IF NO, DBP COMMENTS', 'DBP 5-2-1 ID', 'DATE', 'DB LEAD APPROVAL']
                labels1 = labels.copy()
                [labels.remove(lst) for lst in labels1 if not (lst in df_frmrestricts_fchecks.columns)]
                # for lst in labels1:
                #     if not (lst in df_ecs_fchecks.columns):
                #         labels.remove(lst)
                df_frmrestricts_fchecks = df_frmrestricts_fchecks.loc[:, labels]

            df_labset_fchecks = df_radachecks2.loc[df_radachecks2['Applicability'] == 'Lab settings/Forms/Fields']
            # if not df_labset_fchecks.empty:
            # labels = ['Applicability','Check_number', 'Query','GlobalVariableOID','OID','FormOID','FieldOID',
            #           'AnalyteName', 'PreText', 'FolderOID','LocationMethod','ACTION TAKEN',
            #           'IF NO, DBP COMMENTS','DBP 5-2-1 ID', 'DATE','DB LEAD APPROVAL']

            # labels1 = labels.copy()
            # [labels.remove(lst) for lst in labels1 if not (lst in df_labset_fchecks.columns)]
            # for lst in labels1:
            #     if not lst in df_labset_fchecks.columns:
            #          print(lst)
            #          labels.remove(lst)

            # df_labset_fchecks=df_labset_fchecks.loc[:, labels]

            df_folder_fchecks = df_radachecks2.loc[df_radachecks2['Applicability'] == 'Folder']
            df_cf_fchecks = df_radachecks2.loc[df_radachecks2['Applicability'] == 'Custom function']
            df_mat_fchecks = df_radachecks2.loc[df_radachecks2['Applicability'] == 'Matrices']
            df_ddcf_fchecks = df_radachecks2.loc[
                df_radachecks2['Applicability'] == 'Data Dictionary/custom function']
            df_frmdd_fchecks = df_radachecks2.loc[df_radachecks2['Applicability'] == 'Form/Data Dictionary']
            df_crfset_fchecks = df_radachecks2.loc[df_radachecks2['Applicability'] == 'CRF Draft Setting']
            df_fldfrm_fchecks = df_radachecks2.loc[df_radachecks2['Applicability'] == 'Field/Form']
            # df_forms_nonglob = df_forms_nonglob.append(df_folder_fchecks)
            df_multi_fchecks = pd.DataFrame()
            df_multi_fchecks = df_multi_fchecks.append(df_labset_fchecks)
            df_multi_fchecks = df_multi_fchecks.append(df_folder_fchecks)
            df_multi_fchecks = df_multi_fchecks.append(df_cf_fchecks)
            df_multi_fchecks = df_multi_fchecks.append(df_mat_fchecks)
            df_multi_fchecks = df_multi_fchecks.append(df_frmdd_fchecks)
            df_multi_fchecks = df_multi_fchecks.append(df_crfset_fchecks)
            df_multi_fchecks = df_multi_fchecks.append(df_fldfrm_fchecks)
            df_multi_fchecks = df_multi_fchecks.append(df_ddcf_fchecks)

            df_multi_fchecks = df_multi_fchecks.loc[:, ~df_multi_fchecks.columns.str.contains('^Unnamed')]

            if not df_multi_fchecks.empty:
                labels = ['Applicability', 'Check_number', 'Query', "OID", "FolderName", 'FormOID', 'FieldOID',
                          'PreText', 'FunctionName','SignaturePrompt',
                          'GlobalVariableOID', 'DataDictionaryName',
                          'AnalyteName', 'LocationMethod', 'LabStandardGroup', 'ACTION TAKEN',
                          'IF NO, DBP COMMENTS',
                          'DBP 5-2-1 ID', 'DATE', 'DB LEAD APPROVAL']
                labels1 = labels.copy()
                [labels.remove(lst) for lst in labels1 if not (lst in df_multi_fchecks.columns)]

                df_multi_fchecks = df_multi_fchecks.loc[:, labels]

            df_forms_fchecks = df_radachecks2.loc[df_radachecks2['Applicability'] == 'Form']
            if not df_forms_fchecks.empty:
                labels = ['Applicability', 'Check_number', 'Query', 'OID', 'DraftFormName', 'FormOID'
                    , 'FieldOID', 'PreText'
                    , 'DataDictionaryName', 'DefaultValue', 'AnalyteName', 'PARM', 'PARMDES',
                          'ALTERNATE_LAB_TEST_DESC',
                          "IsSignatureRequired", "ViewRestrictions", 'EntryRestrictions',
                          'ACTION TAKEN', 'IF NO, DBP COMMENTS', 'DBP 5-2-1 ID', 'DATE', 'DB LEAD APPROVAL']
                labels1 = labels.copy()
                [labels.remove(lst) for lst in labels1 if not (lst in df_forms_fchecks.columns)]
                # for lst in labels1:
                #     if not lst in df_forms_fchecks.columns:
                #         print(lst)
                #         labels.remove(lst)

                df_forms_fchecks = df_forms_fchecks.loc[:, labels]

            df_echecks_fchecks = df_radachecks2.loc[df_radachecks2['Applicability'] == 'Edit Checks']
            if len(df_echecks_fchecks):
                labels = ['Applicability', 'Check_number', 'Query', 'CheckName', 'FormOID', 'VariableOID',
                          'FieldOID',
                          'RecordPosition', 'CheckFunction', 'IsLog', 'ActionType', 'ActionOptions', 'ActionScript',
                          'DataFormat', 'DataDictionaryName',
                          'ACTION TAKEN', 'IF NO, DBP COMMENTS', 'DBP 5-2-1 ID', 'DATE', 'DB LEAD APPROVAL']
                labels1 = labels.copy()
                [labels.remove(lst) for lst in labels1 if not (lst in df_echecks_fchecks.columns)]
                # for lst in labels1:
                #     if not lst in df_forms_fchecks.columns:
                #         print(lst)
                #         labels.remove(lst)

                df_echecks_fchecks = df_echecks_fchecks.loc[:, labels]

            my_cols_list = ['OBJECT', 'FORM OID/DICTIONARY/EDITCHECK', 'FIELD OID/CODED DATA/USERDATA',
                            'CHECKPOINT', 'COMMENT', 'ACTION TAKEN', 'IF NO, DBP COMMENTS', 'DBP 5-2-1 ID', 'DATE',
                            'DB LEAD APPROVAL']
            # module = 'RADA'
            # DATABASE1 = "Rave Diagnostic Tool Report.xlsx"
            # DATABASE = para7 + "\\" + DATABASE1
            with pd.ExcelWriter(DATABASE4, engine='openpyxl', mode='a',options={'strings_to_formulas': False}) as writer:
                # if module == 'RADA' :
                #     workbook = writer.book
                #     # workbook.remove(workbook['Instructions_DOA'])
                #     del workbook['Instructions_DOA']
                if len(df_field_fchecks) > 0:
                    sheetname = 'Results_FIELDS'
                    df_field_fchecks.to_excel(writer, sheet_name=sheetname, index=False)
                    worksheet = writer.sheets[sheetname]
                    # col = worksheet.column_dimensions['C']
                    # col.alignment = Alignment(wrap_text=True)
                    xcel_wraptext(ws=worksheet)
                    auto_format_cell_width1(worksheet)
                    worksheet.column_dimensions['C'].width = 38
                    col_bg_col(ws=worksheet)
                if len(df_forms_fchecks) > 0:
                    sheetname = 'Results_FORMS'
                    df_forms_fchecks.to_excel(writer, sheet_name=sheetname, index=False)
                    worksheet = writer.sheets[sheetname]
                    # col = worksheet.column_dimensions['C']
                    # col.alignment = Alignment(wrap_text=True)
                    xcel_wraptext(ws=worksheet)
                    auto_format_cell_width1(worksheet)
                    worksheet.column_dimensions['C'].width = 38
                    col_bg_col(ws=worksheet)

                if len(df_dict_fchecks) > 0:
                    sheetname = 'Results_DICTIONARY'
                    df_dict_fchecks.to_excel(writer, sheet_name=sheetname, index=False)
                    worksheet = writer.sheets[sheetname]
                    # col = worksheet.column_dimensions['C']
                    # col.alignment = Alignment(wrap_text=True)
                    xcel_wraptext(ws=worksheet)
                    auto_format_cell_width1(worksheet)
                    worksheet.column_dimensions['C'].width = 38
                    col_bg_col(ws=worksheet)
                if len(df_multi_fchecks) > 0:
                    sheetname = 'Results_Multiple'
                    df_multi_fchecks.to_excel(writer, sheet_name=sheetname, index=False)
                    worksheet = writer.sheets[sheetname]
                    xcel_wraptext(ws=worksheet)
                    auto_format_cell_width1(worksheet)
                    worksheet.column_dimensions['C'].width = 38
                    col_bg_col(ws=worksheet)
                    # xcel_wraptext(ws=worksheet, col=2)
                if len(df_frmrestricts_fchecks) > 0:
                    sheetname = 'Results_Restrictions'
                    df_frmrestricts_fchecks.to_excel(writer, sheet_name=sheetname, index=False)
                    worksheet = writer.sheets[sheetname]
                    # col = worksheet.column_dimensions['C']
                    # col.alignment = Alignment(wrap_text=True)
                    xcel_wraptext(ws=worksheet)
                    auto_format_cell_width1(worksheet)
                    worksheet.column_dimensions['C'].width = 38
                    col_bg_col(ws=worksheet)
                if len(df_echecks_fchecks) > 0:
                    sheetname = 'Results_EditChecks'
                    df_echecks_fchecks.to_excel(writer, sheet_name=sheetname, index=False)
                    worksheet = writer.sheets[sheetname]
                    # col = worksheet.column_dimensions['C']
                    # col.alignment = Alignment(wrap_text=True)
                    xcel_wraptext(ws=worksheet)
                    auto_format_cell_width1(worksheet)
                    worksheet.column_dimensions['C'].width = 38
                    col_bg_col(ws=worksheet)
                if len(df_als_fchecks) > 0:
                    sheetname = 'Results_ALS'
                    df_als_fchecks.to_excel(writer, sheet_name=sheetname, index=False)
                    # print(writer.sheets)
                    # print(writer.sheets[sheetname])
                    # writer.save()
                    # writer.close()
                    worksheet = writer.sheets[sheetname]
                    xcel_wraptext(ws=worksheet)
                    col_bg_col(ws=worksheet)
                    # xcel_wraptext(ws=worksheet,col=2)
                    # auto_format_cell_width(ws=worksheet)
                    auto_format_cell_width1(worksheet)
                    worksheet.column_dimensions['C'].width = 38
                    worksheet.freeze_panes = "A1"

                    # col = worksheet.column_dimensions['C']
                    # col.alignment = Alignment(wrap_text=True)
                    # col.font = Font(wrap_text=True)
                    # col.font = Font(bold=True)

                if len(df_ssd_fchecks) > 0:
                    sheetname = 'Results_SSD'
                    df_ssd_fchecks.to_excel(writer, sheet_name=sheetname, index=False)
                    worksheet = writer.sheets[sheetname]
                    # col = worksheet.column_dimensions['C']
                    # col.alignment = Alignment(wrap_text=True)
                    xcel_wraptext(ws=worksheet)
                    auto_format_cell_width1(worksheet)
                    worksheet.column_dimensions['C'].width = 38
                    col_bg_col(ws=worksheet)
                if len(df_ecs_fchecks) > 0:
                    sheetname = 'Results_ECS'
                    df_ecs_fchecks.to_excel(writer, sheet_name=sheetname, index=False)
                    worksheet = writer.sheets[sheetname]
                    # col = worksheet.column_dimensions['C']
                    # col.alignment = Alignment(wrap_text=True)
                    xcel_wraptext(ws=worksheet)
                    auto_format_cell_width1(worksheet)
                    worksheet.column_dimensions['C'].width = 38
                    col_bg_col(ws=worksheet)
                    worksheet.freeze_panes = "A1"

                    #
                    # for sheet in writer.sheets:
                    #     worksheet = writer.sheets[sheet]
                    #     col = worksheet.column_dimensions['C']
                    #     # col.alignment = Alignment(wrap_text=True)
                    #     col.font = Font(bold=True)

                if len(df_ssd_study) > 0:

                    if studyid.upper() != df_ssd_study.columns[1].upper():
                        new_row = pd.DataFrame({'Warnings': ['The study names in ALS and SSD are mismatch, please '
                                                             'check if the file placed is correct']})
                        df_warnings = pd.concat([new_row, df_warnings]).reset_index(drop=True)
                if len(df_ecs_study) > 0:

                    if studyid.upper() != df_ecs_study.columns[1].upper():
                        new_row = pd.DataFrame({'Warnings': ['The study names in ALS and ECS are mismatch, please '
                                                         'check if the file placed is correct']})
                        df_warnings = pd.concat([new_row, df_warnings]).reset_index(drop=True)
                if len(list_ecs_ec_misscols) > 0:
                    new_row = pd.DataFrame({'Warnings': ['These are expected columns [' +
                                                         ','.join([str(elem) for elem in list_ecs_ec_misscols]) +
                                                         '] from ECS Edit checks tab which are missing, please '
                                                         'check if the file placed is correct']})
                    df_warnings = pd.concat([new_row, df_warnings]).reset_index(drop=True)
                if len(list_ecs_esae_misscols) > 0:
                    new_row = pd.DataFrame({'Warnings': ['These are expected columns [' +
                                                         ','.join([str(elem) for elem in list_ecs_esae_misscols]) +
                                                         '] from ECS esae tab which are missing, please '
                                                         'check if the file placed is correct']})
                    df_warnings = pd.concat([new_row, df_warnings]).reset_index(drop=True)
                if len(list_ecs_pd_misscols) > 0:
                    new_row = pd.DataFrame({'Warnings': ['These are expected columns [' +
                                                         ','.join([str(elem) for elem in list_ecs_pd_misscols]) +
                                                         '] from ECS PD tab which are missing, please '
                                                         'check if the file placed is correct']})
                    df_warnings = pd.concat([new_row, df_warnings]).reset_index(drop=True)
                if (files_als) :
                    new_row = pd.DataFrame({'Warnings': ['Please check if ALS is open or running the process behind']})
                    df_warnings = pd.concat([new_row, df_warnings]).reset_index(drop=True)
                if (files_ecs) :
                    new_row = pd.DataFrame({'Warnings': ['Please check if ECS is open or running the process behind']})
                    df_warnings = pd.concat([new_row, df_warnings]).reset_index(drop=True)
                if (files_ssd) :
                    new_row = pd.DataFrame({'Warnings': ['Please check if SSD is open or running the process behind']})
                    df_warnings = pd.concat([new_row, df_warnings]).reset_index(drop=True)
                if (files_als):
                    new_row = pd.DataFrame({'Warnings': ['no ALS placed in ALS folder']})
                    df_warnings = pd.concat([new_row, df_warnings]).reset_index(drop=True)
                    # df_warnings = pd.concat([new_row, df_warnings[:]]).reset_index(drop=True)
                if (files_ecs):
                    new_row = pd.DataFrame({'Warnings': ['no ECS placed in ECS folder']})
                    df_warnings = pd.concat([new_row, df_warnings]).reset_index(drop=True)
                if (files_ssd):
                    new_row = pd.DataFrame({'Warnings': ['no SSD placed in SSD folder']})
                    df_warnings = pd.concat([new_row, df_warnings]).reset_index(drop=True)

                if ((len(files_codelist) == 0) & (len(files_codelist_up) == 0)):
                    new_row = pd.DataFrame({'Warnings': ['neither codelist refresh nor manual upload is done']})
                    df_warnings = pd.concat([new_row, df_warnings]).reset_index(drop=True)

                if len(df_flds) == 0:
                    new_row = pd.DataFrame({'Warnings': ['No records are extracted from ALS, please check if the file '
                                                         'placed is correct with expected tabs/columns']})
                    df_warnings = pd.concat([new_row, df_warnings]).reset_index(drop=True)
                if ((len(df_ssd_frmattr) == 0) & (len(df_ssd_newfrmattr) == 0) & (len(df_ssd_colist) == 0)):
                    new_row = pd.DataFrame({'Warnings': ['No records are extracted from SSD, please check if the file '
                                                         'placed is correct with expected tabs/columns']})
                    df_warnings = pd.concat([new_row, df_warnings]).reset_index(drop=True)
                if ((len(df_ecs_ec) == 0) & (len(df_ecs_esae) == 0) & (len(df_ecs_pd) == 0)):
                    new_row = pd.DataFrame({'Warnings': ['No records are extracted from ECS, please check if the file '
                                                         'placed is correct with expected tabs/columns']})
                    df_warnings = pd.concat([new_row, df_warnings]).reset_index(drop=True)
                if len(list_ecs_ec_misscols) > 0:
                    new_row = pd.DataFrame({'Warnings': ['These are expected columns [' +
                                                         ','.join([str(elem) for elem in list_ecs_ec_misscols]) +
                                                         '] from ECS which are missing, please '
                                                         'check if the file placed is correct']})
                    df_warnings = pd.concat([new_row, df_warnings]).reset_index(drop=True)

                writer.save()

                # writer.close()

    else:

        if (files_als) :
            new_row = pd.DataFrame({'Warnings': ['Please check if ALS is open']})
            df_warnings = pd.concat([new_row, df_warnings]).reset_index(drop=True)
        if (files_ecs) :
            new_row = pd.DataFrame({'Warnings': ['Please check if ECS is open']})
            df_warnings = pd.concat([new_row, df_warnings]).reset_index(drop=True)
        if (files_ssd) :
            new_row = pd.DataFrame({'Warnings': ['Please check if SSD is open']})
            df_warnings = pd.concat([new_row, df_warnings]).reset_index(drop=True)
        if (files_als):
            new_row = pd.DataFrame({'Warnings': ['no ALS placed in ALS folder']})
            df_warnings = pd.concat([new_row, df_warnings]).reset_index(drop=True)
            # df_warnings = pd.concat([new_row, df_warnings[:]]).reset_index(drop=True)
        if (files_ecs):
            new_row = pd.DataFrame({'Warnings': ['no ECS placed in ECS folder']})
            df_warnings = pd.concat([new_row, df_warnings]).reset_index(drop=True)
        if (files_ssd):
            new_row = pd.DataFrame({'Warnings': ['no SSD placed in SSD folder']})
            df_warnings = pd.concat([new_row, df_warnings]).reset_index(drop=True)

        if len(df_flds) == 0:
            new_row = pd.DataFrame({'Warnings': ['No records are extracted from ALS, please check if the file '
                                                 'placed is correct with expected tabs/columns']})
            df_warnings = pd.concat([new_row, df_warnings]).reset_index(drop=True)
        if ((len(df_ssd_frmattr) == 0) & (len(df_ssd_newfrmattr) == 0) & (len(df_ssd_colist) == 0)):
            new_row = pd.DataFrame({'Warnings': ['No records are extracted from SSD, please check if the file '
                                                 'placed is correct with expected tabs/columns']})
            df_warnings = pd.concat([new_row, df_warnings]).reset_index(drop=True)
        if ((len(df_ecs_ec) == 0) & (len(df_ecs_esae) == 0) & (len(df_ecs_pd) == 0)):
            new_row = pd.DataFrame({'Warnings': ['No records are extracted from ECS, please check if the file '
                                                 'placed is correct with expected tabs/columns']})
            df_warnings = pd.concat([new_row, df_warnings]).reset_index(drop=True)
        if len(list_ecs_ec_misscols) > 0:
            new_row = pd.DataFrame({'Warnings': ['These are expected columns [' +
                                                 ','.join([str(elem) for elem in list_ecs_ec_misscols]) +
                                                 '] from ECS which are missing, please '
                                                 'check if the file placed is correct']})
            df_warnings = pd.concat([new_row, df_warnings]).reset_index(drop=True)

    if len(df_warnings) > 0:
        #
        # DATABASE1 = "Rave Diagnostic Tool Report.xlsx"
        # DATABASE = para7 + "\\" + DATABASE1
        with pd.ExcelWriter(DATABASE4, engine='openpyxl', mode='a') as writer:
            df_warnings.to_excel(writer, sheet_name='warnings', index=False)

            writer.save()
            # writer.close()

    # with pd.ExcelWriter('Rave Diagnostic Tool Report.xlsx', engine='openpyxl', mode='a') as writer:
    #     # workbook = writer.book
    #     for sheet in writer.sheets:
    #         worksheet = writer.sheets[sheet]
    #         col = worksheet.column_dimensions['C']
    #         col.alignment = Alignment(wrap_text=True)
    #
    # writer.save()
    # writer.close()

    #
    #
    #
    #
    # dir = os.getcwd()
    # SITE_ROOT = os.getcwd()
    # username = os.environ["USERNAME"]
    # ossum_path1 = 'C:\\Users\\' + username
    # osssum_dir = 'OSSUM'
    # ossum_path = 'C:\\Users\\' + username + '\\' + osssum_dir
    # rada_dir = 'RADA'
    # rada_path = para7
    # datestring = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M')
    # rada_dt_path = rada_path + '\\' + datestring
    # os.rename()
    #
    # # print(rada_dt_path)
    # if os.path.exists(rada_path):
    #     os.makedirs(rada_dt_path)

    # purgedir(path_als)
    # purgedir(path_ssd)
    # purgedir(path_ecs)

    # os.chdir(dir)

    if os.path.exists(path_out):
        # recursive_copy(path_out, rada_dt_path)
        os.chdir(rada_dt_path)
        df = pd.DataFrame()
        writer = pd.ExcelWriter(DATABASE4, engine='openpyxl', mode='a')
        workbook = writer.book
        # workbook.remove(workbook['Instructions_DOA'])
        df.to_excel(writer, sheet_name='Cover Page', index=False, startcol=3)
        worksheet = writer.sheets['Cover Page']
        worksheet['B1'].font = Font(bold=True, size=18, color="8a2be2")
        # worksheet['C1'].font = Font(bold=True)
        # worksheet['C1'].font = Font(size=20)
        # worksheet.merge_cells('A1:B3')
        worksheet.column_dimensions['A'].width = 23
        worksheet.column_dimensions['B'].width = 40
        worksheet.merge_cells('B1:C3')
        worksheet.column_dimensions['C'].width = 22
        worksheet.column_dimensions['D'].width = 40
        img = Image('NvtsIcon.png')
        # img.width = 111.21
        # img.height = 77.69
        worksheet.add_image(img, 'A1')
        c1 = worksheet.cell(row=1, column=2)
        c1.value = "  Rave Advanced Diagnostic Application"
        # worksheet['B1'].font = Font(bold=True)
        c2 = worksheet.cell(row=5, column=1)
        c2.value = "Study:"
        c3 = worksheet.cell(row=5, column=2)
        c3.value = studyid
        c4 = worksheet.cell(row=6, column=1)
        c4.value = "CRF Version:"
        c5 = worksheet.cell(row=6, column=2)
        c5.value = CRFversion
        c6 = worksheet.cell(row=7, column=1)
        c6.value = "Report Generated By:"
        c7 = worksheet.cell(row=7, column=2)
        c7.value = os.environ["USERNAME"]
        c8 = worksheet.cell(row=8, column=1)
        c8.value = "Report Generation Date:"
        c9 = worksheet.cell(row=8, column=2)
        c9.value = datetime.datetime.today()
        c10 = worksheet.cell(row=9, column=1)
        c10.value = "tSDV (1) / Non-tSDV (0)"
        c11 = worksheet.cell(row=9, column=2)
        print("sdv",sdv)
        c11.value = sdv
        move_sheet(workbook)
        if len(df_radachecks2):
            df_chk_bar = df_radachecks2.groupby(['Check_number']).size().reset_index(name='Freq')
            df_chs = pd.DataFrame([*range(1, 124, 1)], columns=['Check_number'])
            df_chs = df_chs.merge(df_chk_bar, left_on=['Check_number'],
                                                    right_on=['Check_number'], how='left', suffixes=['', '_'],
                                                    indicator=True)

        # df_chk_bar.to_excel(writer, sheet_name='Checks bar plot', index=False, startcol=3)
        # worksheet = writer.sheets['TitlePage']
        # writer = pd.ExcelWriter('farm_data.xlsx', engine='xlsxwriter')
        # df.to_excel(writer, sheet_name='chkbar')
        # workbook = writer.book
        # worksheet = writer.sheets['Checks bar plot']
        # chart = workbook.add_chart({'type': 'column'})

        module = 'RADA'
        if module == 'RADA':
            del workbook['Instructions_DOA']
        elif module == 'RADA_DOA':
            del workbook['Instructions']
        writer.save()
        # CLEAR the folder after report generation
        # purgedir('C:\\Bhasp\\NVTSonco-work\\NVTSonco-work\\RADA\\Radaprogram\\Virtual Environments\\03Nov2021\\ALS')
        # purgedir('C:\\Bhasp\\NVTSonco-work\\NVTSonco-work\\RADA\\Radaprogram\\Virtual Environments\\03Nov2021\\SSD')
        # purgedir('C:\\Bhasp\\NVTSonco-work\\NVTSonco-work\\RADA\\Radaprogram\\Virtual Environments\\03Nov2021\\ECS')
        # webbrowser.open(rada_dt_path)
        os.chdir(dir)

        df_list = []

        df_list.append(df_field_fchecks)
        df_list.append(df_forms_fchecks)
        df_list.append(df_dict_fchecks)
        df_list.append(df_multi_fchecks)
        df_list.append(df_frmrestricts_fchecks)
        df_list.append(df_echecks_fchecks)
        df_list.append(df_als_fchecks)
        df_list.append(df_ssd_fchecks)
        df_list.append(df_ecs_fchecks)
        df_list.append(df_warnings)
        df_list.append(DATABASE3)
        df_list.append(DATABASE4)

        return df_list

    # popupmsg("Please check the output report folder :" + rada_dt_path)


    # for wi in w:
    #     if wi.line is None:tSDV (0) / Non-tSDV (1)
    #         wi.line = linecache.getline(wi.filename, wi.lineno)
    #     print('line number {}:'.format(wi.lineno))
    #     print('line: {}'.format(wi.line))



#
#
# import datetime, time
# import datetime, time
# # import tkinter as tk
# import xlrd
# import openpyxl
# import xlsxwriter
# import csv
# import os
# import pandas as pd
# import numpy as np
# from numpy import math
# import random
# # from tkinter import filedialog
# import shutil
# import subprocess
# from openpyxl.drawing.image import Image
# # import accessdb
# # path =''
# # from meza import io
# import win32com.client
# # import traceback
# import tkinter as tk
# from tkinter import messagebox, ttk
# import webbrowser
# from openpyxl.drawing.image import Image
# from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
# import copy
# from openpyxl.utils import get_column_letter
# import warnings
# import linecache
# from collections import Counter
# import scipy.sparse as sp
#
# def recursive_copy(src, dest):
#     """
#     Copy each file from src dir to dest dir, including sub-directories.
#     """
#     for item in os.listdir(src):
#         file_path = os.path.join(src, item)
#
#         # if item is a file, copy it
#         if os.path.isfile(file_path):
#             shutil.copy(file_path, dest)
#
#         # else if item is a folder, recurse
#         elif os.path.isdir(file_path):
#             new_dest = os.path.join(dest, item)
#             os.mkdir(new_dest)
#             recursive_copy(file_path, new_dest)
#
# def clearfoldercontent(target_dir):
#     with os.scandir(target_dir) as entries:
#         for entry in entries:
#             if entry.is_file() or entry.is_symlink():
#                 os.remove(entry.path)
#             elif entry.is_dir():
#                 shutil.rmtree(entry.path)
#

#
# # Loop for SSDs
# SITE_ROOT = os.getcwd()
# lists = os.listdir(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\Consolidated SSD\SSD')
# for lst in lists:
#     clearfoldercontent(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\Radaprogram\OSSUM\SSD')
#     recursive_copy(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\Consolidated SSD\SSD', r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\Radaprogram\OSSUM\SSD')
#     files_in_directory = os.listdir(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\Consolidated SSD\SSD')
#     filtered_files = [file for file in files_in_directory if not file.endswith(lst)]
#     for file in filtered_files:
#         path_to_file = os.path.join( r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\Radaprogram\OSSUM\SSD', file)
#         os.remove(path_to_file)
#     rada_prog(1)
#     os.chdir(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\Radaprogram\OSSUM')


# Loop for ECSs
# SITE_ROOT = os.getcwd()
# lists = os.listdir(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\Consolidated ECS\ECS')
# for lst in lists:
#     clearfoldercontent(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\Radaprogram\OSSUM\ECS')
#     recursive_copy(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\Consolidated ECS\ECS', r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\Radaprogram\OSSUM\ECS')
#     files_in_directory = os.listdir(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\Radaprogram\OSSUM\ECS')
#     filtered_files = [file for file in files_in_directory if not file.endswith(lst)]
#     for file in filtered_files:
#         path_to_file = os.path.join( r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\Radaprogram\OSSUM\ECS', file)
#         os.remove(path_to_file)
#     rada_prog(1)
#     os.chdir(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\Radaprogram\OSSUM')


# Loop for ALSs
# lists = os.listdir(r'C:\Users\pillibh2\ALSdownload')
# for lst in lists:
#     clearfoldercontent(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\Radaprogram\OSSUM\ALS')
#     recursive_copy(r'C:\Users\pillibh2\ALSdownload', r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\Radaprogram\OSSUM\ALS')
#     files_in_directory = os.listdir(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\Radaprogram\OSSUM\ALS')
#     filtered_files = [file for file in files_in_directory if not file.endswith(lst)]
#     for file in filtered_files:
#         path_to_file = os.path.join( r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\Radaprogram\OSSUM\ALS', file)
#         os.remove(path_to_file)
#     rada_prog(1)
#     os.chdir(r'C:\Bhasp\NVTSonco-work\NVTSonco-work\RADA\Radaprogram\OSSUM')

# rada_prog(1)